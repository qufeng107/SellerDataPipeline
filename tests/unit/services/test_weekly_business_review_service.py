from __future__ import annotations

from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from seller_data_pipeline.services.weekly_business_review_service import (
    WeeklyBusinessReviewService,
    parse_week_start,
)


def test_parse_week_start_requires_monday() -> None:
    assert parse_week_start("2026-04-06") == date(2026, 4, 6)
    try:
        parse_week_start("2026-04-07")
    except ValueError as exc:
        assert "Monday" in str(exc)
    else:
        raise AssertionError("Expected non-Monday week_start to fail")


def test_calculate_weekly_business_review_core_metrics() -> None:
    result = WeeklyBusinessReviewService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        week_start=date(2026, 4, 6),
        generated_at_utc=datetime(2026, 4, 15),
        sales_traffic_rows=[
            _sales_row(date(2026, 4, 6), Decimal("100.00"), 4, 100),
            _sales_row(date(2026, 4, 7), Decimal("50.00"), 2, 50),
            _sales_row(date(2026, 4, 8), Decimal("0.00"), 0, 0),
            _sales_row(date(2026, 4, 9), Decimal("0.00"), 0, 0),
            _sales_row(date(2026, 4, 10), Decimal("0.00"), 0, 0),
            _sales_row(date(2026, 4, 11), Decimal("0.00"), 0, 0),
            _sales_row(date(2026, 4, 12), Decimal("0.00"), 0, 0),
        ],
        previous_sales_traffic_rows=[
            _sales_row(date(2026, 3, 30), Decimal("100.00"), 4, 100),
            _sales_row(date(2026, 3, 31), Decimal("0.00"), 0, 0),
            _sales_row(date(2026, 4, 1), Decimal("0.00"), 0, 0),
            _sales_row(date(2026, 4, 2), Decimal("0.00"), 0, 0),
            _sales_row(date(2026, 4, 3), Decimal("0.00"), 0, 0),
            _sales_row(date(2026, 4, 4), Decimal("0.00"), 0, 0),
            _sales_row(date(2026, 4, 5), Decimal("0.00"), 0, 0),
        ],
        order_item_rows=[
            _order_row("SKU-1", date(2026, 4, 6), 4, Decimal("100.00")),
            _order_row("SKU-1", date(2026, 4, 7), 2, Decimal("50.00")),
        ],
        previous_order_item_rows=[
            _order_row("SKU-1", date(2026, 3, 30), 4, Decimal("100.00"))
        ],
        sku_cost_rows=[_cost_row("SKU-1", Decimal("10.00"), Decimal("2.00"))],
        ads_campaign_rows=[
            _ads_campaign_row(date(2026, 4, 6), Decimal("30.00"), Decimal("100.00"), 2),
        ],
        previous_ads_campaign_rows=[
            _ads_campaign_row(date(2026, 3, 30), Decimal("10.00"), Decimal("50.00"), 1),
        ],
        ads_product_rows=[_ads_product_row("SKU-1", Decimal("30.00"), Decimal("100.00"), 2)],
        inventory_rows=[_inventory_row("SKU-1", date(2026, 4, 12), 42)],
        listing_rows=[_listing_row("SKU-1")],
    )

    assert result.status == "partial"
    assert result.sales_traffic_summary.ordered_product_sales == Decimal("150.00")
    assert result.sales_traffic_summary.units_ordered == 6
    assert result.ads_summary.spend == Decimal("30.00")
    assert result.estimated_cogs == Decimal("72.00")
    assert result.gross_margin_before_ads == Decimal("78.00")
    assert result.contribution_after_ads == Decimal("48.00")
    assert result.sku_performance[0].contribution_after_ads == Decimal("48.00")
    assert result.sku_performance[0].inventory_risk == "healthy"
    assert any(metric.metric == "ordered_product_sales" for metric in result.kpi_summary)


def test_missing_sku_cost_marks_needs_review() -> None:
    result = WeeklyBusinessReviewService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id=None,
        week_start=date(2026, 4, 6),
        generated_at_utc=datetime(2026, 4, 15),
        sales_traffic_rows=[_sales_row(date(2026, 4, 6), Decimal("20.00"), 1, 10)],
        order_item_rows=[_order_row("SKU-MISSING", date(2026, 4, 6), 1, Decimal("20.00"))],
        sku_cost_rows=[],
    )

    assert result.status == "needs_review"
    assert result.sku_performance[0].status == "missing_cost"
    assert any(warning.warning_code == "missing_sku_cost" for warning in result.warnings)


def test_missing_ads_context_is_partial_not_needs_review() -> None:
    result = WeeklyBusinessReviewService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        week_start=date(2026, 4, 6),
        generated_at_utc=datetime(2026, 4, 15),
        sales_traffic_rows=[
            _sales_row(date(2026, 4, 6), Decimal("20.00"), 1, 10),
            _sales_row(date(2026, 4, 7), Decimal("0.00"), 0, 0),
            _sales_row(date(2026, 4, 8), Decimal("0.00"), 0, 0),
            _sales_row(date(2026, 4, 9), Decimal("0.00"), 0, 0),
            _sales_row(date(2026, 4, 10), Decimal("0.00"), 0, 0),
            _sales_row(date(2026, 4, 11), Decimal("0.00"), 0, 0),
            _sales_row(date(2026, 4, 12), Decimal("0.00"), 0, 0),
        ],
        order_item_rows=[_order_row("SKU-1", date(2026, 4, 6), 1, Decimal("20.00"))],
        sku_cost_rows=[_cost_row("SKU-1", Decimal("5.00"), Decimal("0.00"))],
        ads_campaign_rows=[],
    )

    assert result.status == "partial"
    assert any(warning.warning_code == "ads_api_context_missing" for warning in result.warnings)
    assert result.contribution_after_ads == Decimal("15.00")


def test_write_report_files_creates_json_and_xlsx(tmp_path: Path) -> None:
    result = WeeklyBusinessReviewService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        week_start=date(2026, 4, 6),
        generated_at_utc=datetime(2026, 4, 15),
        sales_traffic_rows=[_sales_row(date(2026, 4, 6), Decimal("10.00"), 1, 10)],
        order_item_rows=[_order_row("SKU-1", date(2026, 4, 6), 1, Decimal("10.00"))],
        sku_cost_rows=[_cost_row("SKU-1", Decimal("2.00"), Decimal("1.00"))],
    )

    written = WeeklyBusinessReviewService().write_report_files(result=result, output_root=tmp_path)

    assert set(written.output_files) == {"json", "xlsx"}
    json_path = tmp_path / "ATVPDKIKX0DER/2026-04-06_2026-04-12/weekly_business_review.json"
    xlsx_path = tmp_path / "ATVPDKIKX0DER/2026-04-06_2026-04-12/weekly_business_review.xlsx"
    assert json_path.exists()
    assert xlsx_path.exists()
    workbook = load_workbook(xlsx_path, read_only=True)
    assert workbook.sheetnames == [
        "01_Executive_Summary",
        "02_Daily_Trend",
        "03_Sales_Traffic",
        "04_SKU_Performance",
        "05_Ads_Overview",
        "06_Inventory_Risk",
        "07_Alerts_Actions",
        "08_Reconciliation_Checks",
        "09_Raw_Metadata",
    ]


def test_run_uses_repo_and_writes_report(tmp_path: Path) -> None:
    repo = FakeWeeklyRepo()
    result = WeeklyBusinessReviewService(repo=repo).run(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        week_start=date(2026, 4, 6),
        output_root=tmp_path,
    )

    assert result.status in {"ok", "partial"}
    assert repo.calls == [
        "sales",
        "sales",
        "orders",
        "orders",
        "costs",
        "ads_campaign",
        "ads_campaign",
        "ads_product",
        "inventory",
        "listing",
        "settlement",
    ]
    assert result.output_files["xlsx"].endswith("weekly_business_review.xlsx")


def _sales_row(report_date: date, sales: Decimal, units: int, sessions: int) -> dict[str, object]:
    return {
        "report_date": report_date,
        "ordered_product_sales_amount": sales,
        "ordered_product_sales_currency": "USD",
        "units_ordered": units,
        "total_order_items": units,
        "sessions": sessions,
        "page_views": sessions * 2,
        "units_refunded": 0,
    }


def _order_row(
    sku: str,
    purchase_date: date,
    quantity: int,
    item_price: Decimal,
) -> dict[str, object]:
    return {
        "amazon_order_id": f"ORDER-{sku}-{purchase_date.isoformat()}",
        "purchase_date": purchase_date,
        "product_name": "Test Product",
        "seller_sku": sku,
        "asin": "B000TEST",
        "quantity": quantity,
        "currency": "USD",
        "item_price": item_price,
        "shipping_price": Decimal("0.00"),
        "item_promotion_discount": Decimal("0.00"),
        "ship_promotion_discount": Decimal("0.00"),
    }


def _cost_row(sku: str, product_cost: Decimal, first_mile_cost: Decimal) -> dict[str, object]:
    return {
        "marketplace_id": "ATVPDKIKX0DER",
        "seller_sku": sku,
        "asin": "B000TEST",
        "product_cost": product_cost,
        "first_mile_cost": first_mile_cost,
        "packaging_cost": Decimal("0.00"),
        "other_unit_cost": Decimal("0.00"),
        "currency": "USD",
        "effective_from": date(2026, 1, 1),
        "effective_to": None,
        "remark": "fixture",
    }


def _ads_campaign_row(
    report_date: date,
    cost: Decimal,
    sales_7d: Decimal,
    purchases_7d: int,
) -> dict[str, object]:
    return {
        "profile_id": "3917953989967300",
        "marketplace_id": "ATVPDKIKX0DER",
        "report_date": report_date,
        "campaign_id": "C1",
        "campaign_name": "Campaign 1",
        "campaign_status": "enabled",
        "impressions": 1000,
        "clicks": 50,
        "cost": cost,
        "sales_7d": sales_7d,
        "purchases_7d": purchases_7d,
        "units_sold_clicks_7d": purchases_7d,
    }


def _ads_product_row(
    sku: str,
    cost: Decimal,
    sales_7d: Decimal,
    purchases_7d: int,
) -> dict[str, object]:
    return {
        "profile_id": "3917953989967300",
        "marketplace_id": "ATVPDKIKX0DER",
        "report_date": date(2026, 4, 6),
        "campaign_id": "C1",
        "campaign_name": "Campaign 1",
        "advertised_asin": "B000TEST",
        "advertised_sku": sku,
        "impressions": 1000,
        "clicks": 50,
        "cost": cost,
        "sales_7d": sales_7d,
        "purchases_7d": purchases_7d,
        "units_sold_clicks_7d": purchases_7d,
    }


def _inventory_row(sku: str, snapshot_date: date, fulfillable: int) -> dict[str, object]:
    return {
        "snapshot_date": snapshot_date,
        "seller_sku": sku,
        "asin": "B000TEST",
        "product_name": "Test Product",
        "afn_fulfillable_quantity": fulfillable,
        "afn_reserved_quantity": 0,
        "afn_unsellable_quantity": 0,
        "afn_total_quantity": fulfillable,
    }


def _listing_row(sku: str) -> dict[str, object]:
    return {
        "snapshot_date": date(2026, 4, 12),
        "seller_sku": sku,
        "asin": "B000TEST",
        "item_name": "Test Product",
        "status": "Active",
    }


class FakeWeeklyRepo:
    def __init__(self) -> None:
        self.calls: list[str] = []

    def fetch_sales_traffic_daily_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("sales")
        return [
            _sales_row(start_date, Decimal("20.00"), 1, 10),
            _sales_row(start_date + timedelta(days=1), Decimal("0.00"), 0, 0),
            _sales_row(start_date + timedelta(days=2), Decimal("0.00"), 0, 0),
            _sales_row(start_date + timedelta(days=3), Decimal("0.00"), 0, 0),
            _sales_row(start_date + timedelta(days=4), Decimal("0.00"), 0, 0),
            _sales_row(start_date + timedelta(days=5), Decimal("0.00"), 0, 0),
            _sales_row(start_date + timedelta(days=6), Decimal("0.00"), 0, 0),
        ]

    def fetch_order_item_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("orders")
        return [_order_row("SKU-1", start_date, 1, Decimal("20.00"))]

    def fetch_sku_cost_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("costs")
        return [_cost_row("SKU-1", Decimal("5.00"), Decimal("0.00"))]

    def fetch_ads_campaign_daily_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("ads_campaign")
        assert profile_id == "3917953989967300"
        return []

    def fetch_ads_advertised_product_daily_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("ads_product")
        return []

    def fetch_latest_inventory_rows(
        self,
        *,
        marketplace_id: str,
        as_of_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("inventory")
        return [_inventory_row("SKU-1", as_of_date, 20)]

    def fetch_latest_listing_rows(
        self,
        *,
        marketplace_id: str,
        as_of_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("listing")
        return [_listing_row("SKU-1")]

    def fetch_settlement_preview_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("settlement")
        return []
