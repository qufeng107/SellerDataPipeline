from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from seller_data_pipeline.services.monthly_financial_close_service import (
    MonthlyFinancialCloseService,
    month_to_date_range,
)


def test_month_to_date_range_handles_month_end() -> None:
    assert month_to_date_range("2026-03") == (date(2026, 3, 1), date(2026, 3, 31))
    assert month_to_date_range("2026-02") == (date(2026, 2, 1), date(2026, 2, 28))


def test_calculate_monthly_financial_close_core_metrics() -> None:
    result = MonthlyFinancialCloseService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        month="2026-03",
        start_date=date(2026, 3, 1),
        end_date=date(2026, 3, 31),
        settlement_rows=[
            _settlement_row(1, "SKU-1", Decimal("100.00"), "product_sales", "revenue", 2),
            _settlement_row(2, "SKU-1", Decimal("-15.00"), "referral_fee", "amazon_fee", 2),
            _settlement_row(3, "SKU-1", Decimal("-20.00"), "fba_fee", "fba_fee", 2),
            _settlement_row(
                4,
                None,
                Decimal("-10.00"),
                "advertising_fee",
                "advertising_cost",
                None,
            ),
            _settlement_row(5, None, Decimal("5.00"), "reimbursement", "reimbursement", None),
        ],
        sku_cost_rows=[_cost_row("SKU-1", Decimal("20.00"), Decimal("5.00"), date(2026, 1, 1))],
        orders_summary={"order_count": 1, "ordered_units": 2},
        ads_summary={"ads_cost": Decimal("10.00"), "ads_clicks": 9},
        sales_traffic_summary={
            "units_ordered": 2,
            "sessions": 100,
            "ordered_product_sales_amount": Decimal("100.00"),
        },
    )

    assert result.status == "ok"
    assert result.financial_summary.settlement_net_amount == Decimal("60.00")
    assert result.financial_summary.product_sales_amount == Decimal("100.00")
    assert result.financial_summary.product_sales_units == 2
    assert result.financial_summary.product_cost_cogs == Decimal("40.00")
    assert result.financial_summary.first_mile_cogs == Decimal("10.00")
    assert result.financial_summary.packaging_cogs == Decimal("0.00")
    assert result.financial_summary.other_unit_cogs == Decimal("0.00")
    assert result.financial_summary.internal_cogs == Decimal("50.00")
    assert result.financial_summary.estimated_operating_profit == Decimal("10.00")
    assert result.financial_summary.profit_margin == Decimal("0.1000")
    assert result.financial_summary.settlement_led_estimated_profit == Decimal("10.00")
    assert result.financial_summary.management_estimated_profit_report_date_ads == Decimal("10.00")
    assert result.financial_summary.ads_timing_difference == Decimal("0.00")
    assert result.sku_profitability[0].unit_product_cost == Decimal("20.00")
    assert result.sku_profitability[0].unit_first_mile_cost == Decimal("5.00")
    assert result.sku_profitability[0].unit_standard_cost == Decimal("25.00")
    assert result.sku_profitability[0].product_cost_cogs == Decimal("40.00")
    assert result.sku_profitability[0].first_mile_cogs == Decimal("10.00")
    assert result.sku_profitability[0].estimated_profit_after_cogs == Decimal("15.00")
    assert result.executive_summary()["headline"].startswith(
        "2026-03 management operating profit"
    )

    payload = result.to_dict()
    assert payload["version"] == "v1.5-natural-month-finances"
    assert payload["financial_summary"]["management_operating_profit"] == "10.00"
    assert payload["financial_summary"]["landed_cogs"] == "50.00"
    assert "accountant_pack" in payload
    assert payload["accountant_pack"]["bookkeeping_summary"][0]["accounting_item"].startswith(
        "Product Sales Revenue"
    )


def test_missing_sku_cost_marks_needs_review() -> None:
    result = MonthlyFinancialCloseService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id=None,
        month="2026-03",
        start_date=date(2026, 3, 1),
        end_date=date(2026, 3, 31),
        settlement_rows=[
            _settlement_row(
                1,
                "SKU-MISSING",
                Decimal("30.00"),
                "product_sales",
                "revenue",
                1,
            )
        ],
        sku_cost_rows=[],
    )

    assert result.status == "needs_review"
    assert result.financial_summary.internal_cogs == Decimal("0.00")
    assert result.sku_profitability[0].status == "missing_cost"
    assert any(warning.warning_code == "missing_sku_cost" for warning in result.warnings)


def test_missing_ads_api_context_adds_warning_but_keeps_settlement_profit_ok() -> None:
    result = MonthlyFinancialCloseService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        month="2026-03",
        start_date=date(2026, 3, 1),
        end_date=date(2026, 3, 31),
        settlement_rows=[
            _settlement_row(1, "SKU-1", Decimal("100.00"), "product_sales", "revenue", 1),
            _settlement_row(
                2,
                None,
                Decimal("-10.00"),
                "advertising_fee",
                "advertising_cost",
                None,
            ),
        ],
        sku_cost_rows=[_cost_row("SKU-1", Decimal("20.00"), Decimal("5.00"), date(2026, 1, 1))],
        ads_summary={"ads_cost": Decimal("0.00"), "ads_row_count": 0},
        sales_traffic_summary={"ordered_product_sales_amount": Decimal("100.00")},
    )

    assert result.status == "ok"
    assert result.financial_summary.estimated_operating_profit == Decimal("65.00")
    assert any(warning.warning_code == "ads_api_context_missing" for warning in result.warnings)
    assert "Reconciliation warnings:" in " ".join(result.executive_summary()["key_points"])


def test_monthly_cost_matching_uses_effective_date_per_unit() -> None:
    result = MonthlyFinancialCloseService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id=None,
        month="2026-03",
        start_date=date(2026, 3, 1),
        end_date=date(2026, 3, 31),
        settlement_rows=[
            _settlement_row(
                1,
                "SKU-1",
                Decimal("30.00"),
                "product_sales",
                "revenue",
                1,
                posted_date=date(2026, 3, 1),
                order_item_code="ITEM-1",
            ),
            _settlement_row(
                2,
                "SKU-1",
                Decimal("30.00"),
                "product_sales",
                "revenue",
                1,
                posted_date=date(2026, 3, 20),
                order_item_code="ITEM-2",
            ),
        ],
        sku_cost_rows=[
            _cost_row(
                "SKU-1",
                Decimal("10.00"),
                Decimal("0.00"),
                date(2026, 1, 1),
                date(2026, 3, 10),
            ),
            _cost_row("SKU-1", Decimal("20.00"), Decimal("0.00"), date(2026, 3, 11)),
        ],
    )

    assert result.status == "ok"
    assert result.financial_summary.internal_cogs == Decimal("30.00")
    assert result.sku_profitability[0].unit_standard_cost == Decimal("15.00")
    assert "multiple cost rows matched" in "; ".join(result.sku_profitability[0].notes)


def test_write_report_files_creates_json_and_xlsx(tmp_path: Path) -> None:
    result = MonthlyFinancialCloseService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        month="2026-03",
        start_date=date(2026, 3, 1),
        end_date=date(2026, 3, 31),
        generated_at_utc=datetime(2026, 4, 1),
        settlement_rows=[
            _settlement_row(1, "SKU-1", Decimal("10.00"), "product_sales", "revenue", 1)
        ],
        sku_cost_rows=[_cost_row("SKU-1", Decimal("2.00"), Decimal("1.00"), date(2026, 1, 1))],
    )

    written = MonthlyFinancialCloseService().write_report_files(result=result, output_root=tmp_path)

    assert set(written.output_files) == {"json", "xlsx", "operating_xlsx", "accounting_xlsx"}
    json_path = tmp_path / "ATVPDKIKX0DER/2026-03/monthly_financial_close_2026-03.json"
    xlsx_path = tmp_path / "ATVPDKIKX0DER/2026-03/monthly_financial_close_2026-03.xlsx"
    operating_path = tmp_path / "ATVPDKIKX0DER/2026-03/monthly_operating_report_2026-03.xlsx"
    accounting_path = tmp_path / "ATVPDKIKX0DER/2026-03/accountant_monthly_workbook_2026-03.xlsx"
    assert json_path.exists()
    assert xlsx_path.exists()
    assert operating_path.exists()
    assert accounting_path.exists()
    workbook = load_workbook(xlsx_path, read_only=True)
    assert workbook.sheetnames == [
        "00_Readme_说明",
        "01_Summary",
        "02_Management_PnL",
        "03_Ads_Timing_Recon",
        "04_Settlement_Buckets",
        "05_Amount_Categories",
        "06_SKU_Profit",
        "07_Operational_Context",
        "08_Reconciliation_Checks",
        "09_Warnings",
        "10_Raw_Metadata",
        "09_Accounting_Summary",
        "10_Journal_Entries",
        "11_Quarter_Rollup",
        "12_FX_Rates",
        "13_Source_Doc_Index",
        "14_Payout_Recon",
        "15_Adjustments",
    ]
    operating_workbook = load_workbook(operating_path, read_only=True)
    assert operating_workbook.sheetnames == [
        "01_月度经营总览",
        "02_经营损益",
        "03_核验与口径",
    ]
    accounting_workbook = load_workbook(accounting_path, read_only=True)
    assert accounting_workbook.sheetnames == [
        "01_会计汇总",
        "02_分类明细",
        "03_源交易明细",
        "04_核验与说明",
    ]

    accounting_sheet = workbook["09_Accounting_Summary"]
    assert accounting_sheet["A1"].value == "Sheet Purpose / 本表用途"
    assert accounting_sheet["B1"].value.startswith("按会计做账视角汇总")
    assert accounting_sheet["A7"].value == "Line No. / 行号"
    assert accounting_sheet["B7"].value == "Accounting Item / 会计项目"
    assert accounting_sheet["H8"].value == "='12_FX_Rates'!$D$8"
    assert accounting_sheet["I8"].value == '=IFERROR(F8*H8,"")'
    fx_sheet = workbook["12_FX_Rates"]
    assert fx_sheet["A1"].value == "Sheet Purpose / 本表用途"
    assert fx_sheet["D7"].value == "FX Rate / 汇率"

    summary_sheet = workbook["01_Summary"]
    summary_metrics = {summary_sheet[f"A{row}"].value for row in range(2, summary_sheet.max_row + 1)}
    assert "Management Operating Profit / 经营利润" in summary_metrics
    assert "First-Mile Freight COGS / 头程海运COGS" in summary_metrics
    assert "Estimated Operating Profit / 估算经营利润" not in summary_metrics

    management_sheet = workbook["02_Management_PnL"]
    management_metrics = {
        management_sheet[f"A{row}"].value for row in range(2, management_sheet.max_row + 1)
    }
    assert "Management Operating Profit / 经营利润" in management_metrics
    assert "Less First-Mile Freight COGS / 减：头程海运成本" in management_metrics


def test_run_uses_repo_and_writes_report(tmp_path: Path) -> None:
    repo = FakeMonthlyRepo()
    result = MonthlyFinancialCloseService(repo=repo).run(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        month="2026-03",
        output_root=tmp_path,
    )

    assert result.status == "ok"
    assert repo.calls == [
        "settlement",
        "costs",
        "orders",
        "ads",
        "sales_traffic",
        "coupon",
        "promotion",
        "reimbursement",
    ]
    assert result.output_files["xlsx"].endswith("monthly_financial_close_2026-03.xlsx")


def _settlement_row(
    row_id: int,
    seller_sku: str | None,
    amount: Decimal,
    amount_category: str,
    profit_bucket: str,
    quantity: int | None,
    *,
    posted_date: date = date(2026, 3, 1),
    order_item_code: str | None = "ITEM-1",
) -> dict[str, object]:
    return {
        "id": row_id,
        "marketplace_id": "ATVPDKIKX0DER",
        "posted_date": posted_date,
        "amount": amount,
        "currency": "USD",
        "settlement_id": "SETTLEMENT-1",
        "transaction_type": "Order",
        "order_id": "ORDER-1",
        "order_item_code": order_item_code if seller_sku else None,
        "seller_sku": seller_sku,
        "quantity_purchased": quantity,
        "amount_category": amount_category,
        "profit_bucket": profit_bucket,
        "is_settlement_summary": False,
    }


def _cost_row(
    sku: str,
    product_cost: Decimal,
    first_mile_cost: Decimal,
    effective_from: date,
    effective_to: date | None = None,
) -> dict[str, object]:
    return {
        "marketplace_id": "ATVPDKIKX0DER",
        "seller_sku": sku,
        "asin": "B000TEST",
        "product_cost": product_cost,
        "first_mile_cost": first_mile_cost,
        "packaging_cost": Decimal("0.00"),
        "other_unit_cost": Decimal("0.00"),
        "currency": "USD",
        "effective_from": effective_from,
        "effective_to": effective_to,
        "remark": "fixture",
    }


class FakeMonthlyRepo:
    def __init__(self) -> None:
        self.calls: list[str] = []

    def fetch_settlement_profit_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("settlement")
        return [_settlement_row(1, "SKU-1", Decimal("20.00"), "product_sales", "revenue", 1)]

    def fetch_sku_cost_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("costs")
        return [_cost_row("SKU-1", Decimal("5.00"), Decimal("0.00"), date(2026, 1, 1))]

    def fetch_orders_period_summary(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> dict[str, object]:
        self.calls.append("orders")
        return {"order_count": 1, "ordered_units": 1}

    def fetch_ads_period_summary(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> dict[str, object]:
        self.calls.append("ads")
        assert profile_id == "3917953989967300"
        return {"ads_cost": Decimal("0.00")}

    def fetch_sales_traffic_period_summary(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> dict[str, object]:
        self.calls.append("sales_traffic")
        return {"units_ordered": 1, "ordered_product_sales_amount": Decimal("20.00")}

    def fetch_coupon_period_summary(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> dict[str, object]:
        self.calls.append("coupon")
        return {"coupon_count": 0}

    def fetch_promotion_period_summary(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> dict[str, object]:
        self.calls.append("promotion")
        return {"promotion_count": 0}

    def fetch_fba_reimbursement_period_summary(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> dict[str, object]:
        self.calls.append("reimbursement")
        return {"reimbursement_count": 0}


def test_large_ads_timing_difference_is_warning_not_delivery_blocker() -> None:
    result = MonthlyFinancialCloseService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        month="2026-06",
        start_date=date(2026, 6, 1),
        end_date=date(2026, 6, 30),
        settlement_rows=[
            _settlement_row(1, "SKU-1", Decimal("100.00"), "product_sales", "revenue", 1),
            _settlement_row(2, None, Decimal("-10.00"), "advertising_fee", "advertising_cost", None),
        ],
        sku_cost_rows=[_cost_row("SKU-1", Decimal("20.00"), Decimal("0.00"), date(2026, 1, 1))],
        ads_summary={"ads_cost": Decimal("500.00"), "ads_row_count": 10},
        sales_traffic_summary={"ordered_product_sales_amount": Decimal("100.00")},
    )

    timing = next(
        check for check in result.reconciliation_checks
        if check.check_name == "settlement_ads_fee_vs_ads_api_spend"
    )
    assert timing.status == "warning"
    assert timing.severity == "warning"
    assert result.status == "ok"


def test_natural_month_finances_drive_management_profit_when_present() -> None:
    finance_rows = [
        {
            "marketplace_id": "ATVPDKIKX0DER",
            "transaction_id": "order-1",
            "transaction_status": "DEFERRED_RELEASED",
            "transaction_type": "Shipment",
            "posted_date_local": date(2026, 3, 5),
            "marketplace_timezone": "America/Los_Angeles",
            "amount": Decimal("100.00"),
            "currency": "USD",
            "management_role": "operating",
            "management_include": True,
            "management_replace_with_ads_api": False,
            "review_required": False,
            "product_sales_amount": Decimal("120.00"),
            "unit_events_json": '[{"seller_sku":"SKU-1","quantity":2,"posted_date":"2026-03-05"}]',
        },
        {
            "marketplace_id": "ATVPDKIKX0DER",
            "transaction_id": "refund-1",
            "transaction_status": "RELEASED",
            "transaction_type": "Refund",
            "posted_date_local": date(2026, 3, 10),
            "marketplace_timezone": "America/Los_Angeles",
            "amount": Decimal("-20.00"),
            "currency": "USD",
            "management_role": "operating",
            "management_include": True,
            "management_replace_with_ads_api": False,
            "review_required": False,
            "product_sales_amount": Decimal("0.00"),
            "unit_events_json": "[]",
        },
        {
            "marketplace_id": "ATVPDKIKX0DER",
            "transaction_id": "fee-1",
            "transaction_status": "RELEASED",
            "transaction_type": "ServiceFee",
            "posted_date_local": date(2026, 3, 11),
            "marketplace_timezone": "America/Los_Angeles",
            "amount": Decimal("-5.00"),
            "currency": "USD",
            "management_role": "operating",
            "management_include": True,
            "management_replace_with_ads_api": False,
            "review_required": False,
            "product_sales_amount": Decimal("0.00"),
            "subscription_fee": Decimal("-3.00"),
            "coupon_fee": Decimal("-1.00"),
            "deal_fee": Decimal("0.00"),
            "storage_fee": Decimal("-1.00"),
            "customer_return_fee": Decimal("0.00"),
            "other_service_fee": Decimal("0.00"),
            "unit_events_json": "[]",
        },
        {
            "marketplace_id": "ATVPDKIKX0DER",
            "transaction_id": "ads-1",
            "transaction_status": "RELEASED",
            "transaction_type": "ProductAdsPayment",
            "posted_date_local": date(2026, 3, 12),
            "marketplace_timezone": "America/Los_Angeles",
            "amount": Decimal("-30.00"),
            "currency": "USD",
            "management_role": "ads_charge_reference",
            "management_include": False,
            "management_replace_with_ads_api": True,
            "review_required": False,
            "product_sales_amount": Decimal("0.00"),
            "unit_events_json": "[]",
        },
    ]
    result = MonthlyFinancialCloseService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        month="2026-03",
        start_date=date(2026, 3, 1),
        end_date=date(2026, 3, 31),
        settlement_rows=[
            _settlement_row(1, "SKU-1", Decimal("200.00"), "product_sales", "revenue", 2),
        ],
        finances_natural_month_rows=finance_rows,
        sku_cost_rows=[_cost_row("SKU-1", Decimal("20.00"), Decimal("5.00"), date(2026, 1, 1))],
        ads_summary={"ads_cost": Decimal("10.00"), "ads_row_count": 1},
        sales_traffic_summary={"ordered_product_sales_amount": Decimal("200.00")},
    )

    assert result.natural_month_finance is not None
    assert result.natural_month_finance.product_sales_amount == Decimal("120.00")
    assert result.natural_month_finance.operating_net_before_ads_replacement == Decimal("75.00")
    assert result.natural_month_finance.subscription_fee == Decimal("-3.00")
    assert result.natural_month_finance.coupon_fee == Decimal("-1.00")
    assert result.natural_month_finance.storage_fee == Decimal("-1.00")
    assert result.natural_month_finance.service_fee_total == Decimal("-5.00")
    assert result.natural_month_finance.landed_cogs == Decimal("50.00")
    assert result.natural_month_finance.management_operating_profit == Decimal("15.00")
    assert result.financial_summary.management_estimated_profit_report_date_ads == Decimal("15.00")
    assert next(
        c for c in result.reconciliation_checks if c.check_name == "finances_natural_month_coverage"
    ).status == "ok"


def test_zero_value_released_shipment_units_are_costed_without_adding_revenue() -> None:
    finance_rows = [
        {
            "marketplace_id": "ATVPDKIKX0DER",
            "transaction_id": "order-paid",
            "transaction_status": "DEFERRED_RELEASED",
            "transaction_type": "Shipment",
            "posted_date_local": date(2026, 6, 3),
            "marketplace_timezone": "America/Los_Angeles",
            "amount": Decimal("20.00"),
            "currency": "USD",
            "management_role": "operating",
            "management_include": True,
            "management_replace_with_ads_api": False,
            "review_required": False,
            "product_sales_amount": Decimal("25.00"),
            "unit_events_json": '[{"seller_sku":"SKU-1","quantity":1,"posted_date":"2026-06-03"}]',
        },
        {
            "marketplace_id": "ATVPDKIKX0DER",
            "transaction_id": "order-zero",
            "transaction_status": "RELEASED",
            "transaction_type": "Shipment",
            "posted_date_local": date(2026, 6, 4),
            "marketplace_timezone": "America/Los_Angeles",
            "amount": Decimal("0.00"),
            "currency": "USD",
            "management_role": "zero_value_unit_cogs_reference",
            "management_include": False,
            "management_replace_with_ads_api": False,
            "review_required": False,
            "product_sales_amount": Decimal("0.00"),
            "unit_events_json": '[{"seller_sku":"SKU-1","quantity":1,"posted_date":"2026-06-04"}]',
        },
    ]

    result = MonthlyFinancialCloseService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        month="2026-06",
        start_date=date(2026, 6, 1),
        end_date=date(2026, 6, 30),
        settlement_rows=[
            _settlement_row(1, "SKU-1", Decimal("20.00"), "product_sales", "revenue", 1),
        ],
        finances_natural_month_rows=finance_rows,
        sku_cost_rows=[_cost_row("SKU-1", Decimal("4.00"), Decimal("1.00"), date(2026, 1, 1))],
        ads_summary={"ads_cost": Decimal("0.00"), "ads_row_count": 1},
        sales_traffic_summary={"ordered_product_sales_amount": Decimal("25.00")},
    )

    assert result.natural_month_finance is not None
    assert result.natural_month_finance.product_sales_amount == Decimal("25.00")
    assert result.natural_month_finance.order_total == Decimal("20.00")
    assert result.natural_month_finance.product_sales_units == 2
    assert result.natural_month_finance.costed_units == 2
    assert result.natural_month_finance.landed_cogs == Decimal("10.00")
    assert result.natural_month_finance.management_operating_profit == Decimal("10.00")


def test_natural_month_cost_uses_unique_fnsku_to_seller_sku_fallback() -> None:
    finance_rows = [
        {
            "marketplace_id": "ATVPDKIKX0DER",
            "transaction_id": "liq-1",
            "transaction_status": "DEFERRED_RELEASED",
            "transaction_type": "RemovalShipment",
            "posted_date_local": date(2026, 7, 3),
            "marketplace_timezone": "America/Los_Angeles",
            "amount": Decimal("0.31"),
            "currency": "USD",
            "management_role": "operating",
            "management_include": True,
            "management_replace_with_ads_api": False,
            "review_required": False,
            "product_sales_amount": Decimal("0.00"),
            "unit_events_json": (
                '[{"seller_sku":"X004WU7DSH","quantity":1,'
                '"posted_date":"2026-07-03"}]'
            ),
        }
    ]

    result = MonthlyFinancialCloseService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id=None,
        month="2026-07",
        start_date=date(2026, 7, 1),
        end_date=date(2026, 7, 31),
        settlement_rows=[],
        finances_natural_month_rows=finance_rows,
        sku_cost_rows=[
            _cost_row(
                "SC-9HC3-5TFL",
                Decimal("4.17"),
                Decimal("0.5271"),
                date(2025, 1, 1),
            )
        ],
        inventory_cost_identity_rows=[
            {
                "marketplace_id": "ATVPDKIKX0DER",
                "fnsku": "X004WU7DSH",
                "seller_sku": "SC-9HC3-5TFL",
                "asin": "B0G1YF2ZBB",
            }
        ],
        ads_summary={"ads_cost": Decimal("0.00"), "ads_row_count": 1},
    )

    assert result.natural_month_finance is not None
    assert result.natural_month_finance.source_status == "ok"
    assert result.natural_month_finance.costed_units == 1
    assert result.natural_month_finance.missing_cost_skus == ()
    assert result.natural_month_finance.cost_identity_resolutions == (
        "X004WU7DSH->SC-9HC3-5TFL",
    )
    assert result.natural_month_finance.product_cost_cogs == Decimal("4.17")
    assert result.natural_month_finance.first_mile_cogs == Decimal("0.53")
    assert result.natural_month_finance.landed_cogs == Decimal("4.70")


def test_natural_month_cost_fnsku_fallback_fails_closed_when_mapping_is_ambiguous() -> None:
    finance_rows = [
        {
            "marketplace_id": "ATVPDKIKX0DER",
            "transaction_id": "liq-ambiguous",
            "transaction_status": "DEFERRED",
            "transaction_type": "RemovalShipment",
            "posted_date_local": date(2026, 5, 14),
            "marketplace_timezone": "America/Los_Angeles",
            "amount": Decimal("0.62"),
            "currency": "USD",
            "management_role": "operating",
            "management_include": True,
            "management_replace_with_ads_api": False,
            "review_required": False,
            "product_sales_amount": Decimal("0.00"),
            "unit_events_json": (
                '[{"seller_sku":"X004Q3AKFX","quantity":1,'
                '"posted_date":"2026-05-14"}]'
            ),
        }
    ]

    result = MonthlyFinancialCloseService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id=None,
        month="2026-05",
        start_date=date(2026, 5, 1),
        end_date=date(2026, 5, 31),
        settlement_rows=[],
        finances_natural_month_rows=finance_rows,
        sku_cost_rows=[
            _cost_row("SKU-A", Decimal("5.00"), Decimal("0.00"), date(2025, 1, 1)),
            _cost_row("SKU-B", Decimal("6.00"), Decimal("0.00"), date(2025, 1, 1)),
        ],
        inventory_cost_identity_rows=[
            {
                "marketplace_id": "ATVPDKIKX0DER",
                "fnsku": "X004Q3AKFX",
                "seller_sku": "SKU-A",
                "asin": "ASIN-A",
            },
            {
                "marketplace_id": "ATVPDKIKX0DER",
                "fnsku": "X004Q3AKFX",
                "seller_sku": "SKU-B",
                "asin": "ASIN-B",
            },
        ],
        ads_summary={"ads_cost": Decimal("0.00"), "ads_row_count": 1},
    )

    assert result.natural_month_finance is not None
    assert result.natural_month_finance.source_status == "needs_review"
    assert result.natural_month_finance.costed_units == 0
    assert result.natural_month_finance.missing_cost_skus == ("X004Q3AKFX",)
    assert result.natural_month_finance.cost_identity_resolutions == ()


def test_natural_month_direct_cost_wins_over_fnsku_alias() -> None:
    finance_rows = [
        {
            "marketplace_id": "ATVPDKIKX0DER",
            "transaction_id": "liq-direct",
            "transaction_status": "DEFERRED",
            "transaction_type": "RemovalShipment",
            "posted_date_local": date(2026, 5, 14),
            "marketplace_timezone": "America/Los_Angeles",
            "amount": Decimal("0.62"),
            "currency": "USD",
            "management_role": "operating",
            "management_include": True,
            "management_replace_with_ads_api": False,
            "review_required": False,
            "product_sales_amount": Decimal("0.00"),
            "unit_events_json": (
                '[{"seller_sku":"FNSKU-1","quantity":1,'
                '"posted_date":"2026-05-14"}]'
            ),
        }
    ]

    result = MonthlyFinancialCloseService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id=None,
        month="2026-05",
        start_date=date(2026, 5, 1),
        end_date=date(2026, 5, 31),
        settlement_rows=[],
        finances_natural_month_rows=finance_rows,
        sku_cost_rows=[
            _cost_row("FNSKU-1", Decimal("8.00"), Decimal("0.00"), date(2025, 1, 1)),
            _cost_row("CANONICAL", Decimal("5.00"), Decimal("0.00"), date(2025, 1, 1)),
        ],
        inventory_cost_identity_rows=[
            {
                "marketplace_id": "ATVPDKIKX0DER",
                "fnsku": "FNSKU-1",
                "seller_sku": "CANONICAL",
                "asin": "ASIN-1",
            }
        ],
        ads_summary={"ads_cost": Decimal("0.00"), "ads_row_count": 1},
    )

    assert result.natural_month_finance is not None
    assert result.natural_month_finance.product_cost_cogs == Decimal("8.00")
    assert result.natural_month_finance.cost_identity_resolutions == ()
