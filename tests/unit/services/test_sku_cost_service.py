from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from seller_data_pipeline.db.repositories.sku_cost_repo import (
    SkuCandidateRecord,
    SkuCostRecord,
    SkuCostWriteRecord,
)
from seller_data_pipeline.services.sku_cost_service import (
    SKU_COST_INPUT_SHEET,
    SkuCostWorkbookService,
    build_sku_cost_workbook,
    build_template_rows,
    read_sku_cost_workbook,
)


class FakeSkuCostRepo:
    def __init__(
        self,
        *,
        candidates: list[SkuCandidateRecord] | None = None,
        costs: dict[str, SkuCostRecord] | None = None,
        existing_keys: set[tuple[str, str, date]] | None = None,
    ) -> None:
        self.candidates = candidates or []
        self.costs = costs or {}
        self.existing_keys = existing_keys or set()
        self.inserted: list[SkuCostWriteRecord] = []
        self.updated: list[SkuCostWriteRecord] = []
        self.closed: list[tuple[str, str, date]] = []
        self.committed = False

    def fetch_sku_candidates(self, *, marketplace_id: str) -> list[SkuCandidateRecord]:
        return [row for row in self.candidates if row.marketplace_id == marketplace_id]

    def fetch_latest_sku_costs(self, *, marketplace_id: str) -> dict[str, SkuCostRecord]:
        return {
            sku: row
            for sku, row in self.costs.items()
            if row.marketplace_id == marketplace_id
        }

    def sku_cost_exists(
        self,
        *,
        marketplace_id: str,
        seller_sku: str,
        effective_from: date,
    ) -> bool:
        return (marketplace_id, seller_sku, effective_from) in self.existing_keys

    def close_previous_open_cost(
        self,
        *,
        marketplace_id: str,
        seller_sku: str,
        new_effective_from: date,
    ) -> int:
        self.closed.append((marketplace_id, seller_sku, new_effective_from))
        return 1

    def insert_sku_cost(self, record: SkuCostWriteRecord) -> None:
        self.inserted.append(record)

    def update_sku_cost(self, record: SkuCostWriteRecord) -> None:
        self.updated.append(record)

    def commit(self) -> None:
        self.committed = True


def test_build_template_rows_merges_current_cost_reference() -> None:
    candidates = [
        SkuCandidateRecord(
            marketplace_id="ATVPDKIKX0DER",
            seller_sku="SKU-A",
            asin="ASIN1",
            product_name="Neck Wallet",
            sku_sources="listing,orders",
            latest_source_date=date(2026, 5, 10),
        )
    ]
    current_cost = SkuCostRecord(
        marketplace_id="ATVPDKIKX0DER",
        seller_sku="SKU-A",
        asin="ASIN1",
        product_cost=Decimal("2.1000"),
        first_mile_cost=Decimal("0.4500"),
        packaging_cost=Decimal("0.1200"),
        other_unit_cost=Decimal("0.0000"),
        currency="USD",
        effective_from=date(2026, 1, 1),
    )

    rows = build_template_rows(candidates=candidates, current_costs={"SKU-A": current_cost})

    assert len(rows) == 1
    assert rows[0].current_cost == current_cost
    assert rows[0].product_name == "Neck Wallet"


def test_export_template_replaces_existing_file_and_writes_reference_columns(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "sku_cost_template.xlsx"
    output_path.write_text("old file", encoding="utf-8")
    repo = FakeSkuCostRepo(
        candidates=[
            SkuCandidateRecord(
                marketplace_id="ATVPDKIKX0DER",
                seller_sku="SKU-A",
                asin="ASIN1",
                product_name="Neck Wallet",
                sku_sources="listing",
                latest_source_date=date(2026, 5, 10),
            )
        ],
        costs={
            "SKU-A": SkuCostRecord(
                marketplace_id="ATVPDKIKX0DER",
                seller_sku="SKU-A",
                asin="ASIN1",
                product_cost=Decimal("2.1000"),
                first_mile_cost=Decimal("0.4500"),
                packaging_cost=Decimal("0.1200"),
                other_unit_cost=Decimal("0.0000"),
                currency="USD",
                effective_from=date(2026, 1, 1),
                remark="initial",
            )
        },
    )

    result = SkuCostWorkbookService().export_template(
        repo=repo,
        marketplace_id="ATVPDKIKX0DER",
        output_path=output_path,
    )

    assert result.row_count == 1
    workbook = load_workbook(output_path, data_only=True)
    sheet = workbook[SKU_COST_INPUT_SHEET]
    headers = [cell.value for cell in sheet[1]]
    row = {
        header: sheet.cell(row=2, column=index + 1).value
        for index, header in enumerate(headers)
    }
    assert row["seller_sku"] == "SKU-A"
    assert row["current_product_cost"] == 2.1
    assert row["current_remark"] == "initial"
    assert row["new_currency"] == "USD"


def test_read_sku_cost_workbook_parses_only_filled_new_cost_rows(tmp_path: Path) -> None:
    workbook = build_sku_cost_workbook(
        rows=[
            _template_row("SKU-A"),
            _template_row("SKU-B"),
        ],
        marketplace_id="ATVPDKIKX0DER",
        generated_at_utc=datetime(2026, 5, 18, tzinfo=UTC),
        default_currency="USD",
    )
    path = tmp_path / "sku_cost_template.xlsx"
    sheet = workbook[SKU_COST_INPUT_SHEET]
    headers = [cell.value for cell in sheet[1]]
    header_map = {header: index + 1 for index, header in enumerate(headers)}
    sheet.cell(row=2, column=header_map["new_product_cost"]).value = 2.1
    sheet.cell(row=2, column=header_map["new_first_mile_cost"]).value = 0.45
    sheet.cell(row=2, column=header_map["new_packaging_cost"]).value = 0.12
    sheet.cell(row=2, column=header_map["new_effective_from"]).value = date(2026, 1, 1)
    sheet.cell(row=2, column=header_map["purchase_or_batch_note"]).value = "first batch"
    workbook.save(path)

    candidates, issues = read_sku_cost_workbook(path)

    assert issues == []
    assert len(candidates) == 1
    record = candidates[0].record
    assert record.seller_sku == "SKU-A"
    assert record.product_cost == Decimal("2.1000")
    assert record.first_mile_cost == Decimal("0.4500")
    assert record.packaging_cost == Decimal("0.1200")
    assert record.other_unit_cost == Decimal("0")
    assert record.effective_from == date(2026, 1, 1)
    assert "first batch" in (record.remark or "")


def test_read_sku_cost_workbook_requires_product_cost_currency_and_effective_date(
    tmp_path: Path,
) -> None:
    workbook = build_sku_cost_workbook(
        rows=[_template_row("SKU-A")],
        marketplace_id="ATVPDKIKX0DER",
        generated_at_utc=datetime(2026, 5, 18, tzinfo=UTC),
        default_currency="USD",
    )
    path = tmp_path / "invalid_template.xlsx"
    sheet = workbook[SKU_COST_INPUT_SHEET]
    headers = [cell.value for cell in sheet[1]]
    header_map = {header: index + 1 for index, header in enumerate(headers)}
    sheet.cell(row=2, column=header_map["new_first_mile_cost"]).value = 0.45
    sheet.cell(row=2, column=header_map["new_currency"]).value = None
    workbook.save(path)

    candidates, issues = read_sku_cost_workbook(path)

    assert candidates == []
    messages = {issue.message for issue in issues}
    assert "new_product_cost is required." in messages
    assert "new_currency is required." in messages
    assert "new_effective_from is required." in messages


def test_import_workbook_execute_inserts_and_closes_previous_open_cost(tmp_path: Path) -> None:
    path = _filled_workbook(tmp_path, seller_sku="SKU-A")
    repo = FakeSkuCostRepo()

    result = SkuCostWorkbookService().import_workbook(
        repo=repo,
        workbook_path=path,
        dry_run=False,
    )

    assert result.status == "executed"
    assert result.inserted_rows == 1
    assert result.closed_previous_rows == 1
    assert len(repo.inserted) == 1
    assert repo.inserted[0].seller_sku == "SKU-A"
    assert repo.committed is True


def test_import_workbook_skips_existing_key_by_default(tmp_path: Path) -> None:
    path = _filled_workbook(tmp_path, seller_sku="SKU-A")
    repo = FakeSkuCostRepo(existing_keys={("ATVPDKIKX0DER", "SKU-A", date(2026, 1, 1))})

    result = SkuCostWorkbookService().import_workbook(
        repo=repo,
        workbook_path=path,
        dry_run=False,
    )

    assert result.skipped_existing_rows == 1
    assert repo.inserted == []
    assert result.issues[0].severity == "warning"


def test_import_workbook_update_existing_when_requested(tmp_path: Path) -> None:
    path = _filled_workbook(tmp_path, seller_sku="SKU-A")
    repo = FakeSkuCostRepo(existing_keys={("ATVPDKIKX0DER", "SKU-A", date(2026, 1, 1))})

    result = SkuCostWorkbookService().import_workbook(
        repo=repo,
        workbook_path=path,
        dry_run=False,
        update_existing=True,
    )

    assert result.updated_rows == 1
    assert len(repo.updated) == 1
    assert repo.inserted == []


def _template_row(seller_sku: str) -> SkuCandidateRecord:
    candidate = SkuCandidateRecord(
        marketplace_id="ATVPDKIKX0DER",
        seller_sku=seller_sku,
        asin="ASIN1",
        product_name="Neck Wallet",
        sku_sources="listing",
        latest_source_date=date(2026, 5, 10),
    )
    return build_template_rows(candidates=[candidate], current_costs={})[0]


def _filled_workbook(tmp_path: Path, *, seller_sku: str) -> Path:
    workbook = build_sku_cost_workbook(
        rows=[_template_row(seller_sku)],
        marketplace_id="ATVPDKIKX0DER",
        generated_at_utc=datetime(2026, 5, 18, tzinfo=UTC),
        default_currency="USD",
    )
    path = tmp_path / f"{seller_sku}.xlsx"
    sheet = workbook[SKU_COST_INPUT_SHEET]
    headers = [cell.value for cell in sheet[1]]
    header_map = {header: index + 1 for index, header in enumerate(headers)}
    sheet.cell(row=2, column=header_map["new_product_cost"]).value = 2.1
    sheet.cell(row=2, column=header_map["new_first_mile_cost"]).value = 0.45
    sheet.cell(row=2, column=header_map["new_packaging_cost"]).value = 0.12
    sheet.cell(row=2, column=header_map["new_currency"]).value = "USD"
    sheet.cell(row=2, column=header_map["new_effective_from"]).value = date(2026, 1, 1)
    workbook.save(path)
    return path
