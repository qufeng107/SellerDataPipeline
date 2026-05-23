from __future__ import annotations

from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from seller_data_pipeline.services.weekly_ads_optimization_service import (
    WeeklyAdsOptimizationService,
    WeeklyAdsOptimizationThresholds,
    parse_week_start,
)


def test_parse_week_start_requires_monday() -> None:
    assert parse_week_start("2026-05-11") == date(2026, 5, 11)
    try:
        parse_week_start("2026-05-12")
    except ValueError as exc:
        assert "Monday" in str(exc)
    else:
        raise AssertionError("Expected non-Monday week_start to fail")


def test_calculate_core_metrics_and_actions_for_2026_05_11_style_week() -> None:
    week_start = date(2026, 5, 11)
    campaign_rows = [_campaign_row(week_start + timedelta(days=i)) for i in range(7)]
    search_rows = []
    for offset in range(7):
        current_date = week_start + timedelta(days=offset)
        search_rows.extend(
            [
                _search_row(
                    report_date=current_date,
                    search_term="rfid passport wallet",
                    keyword="passport wallet",
                    match_type="broad",
                    cost=Decimal("2.00"),
                    sales=Decimal("0.00"),
                    purchases=0,
                    clicks=15,
                ),
                _search_row(
                    report_date=current_date,
                    search_term="travel neck wallet",
                    keyword="passport wallet",
                    match_type="broad",
                    cost=Decimal("13.00"),
                    sales=Decimal("60.00"),
                    purchases=3,
                    clicks=20,
                ),
            ]
        )
    targeting_rows = [
        _targeting_row(
            report_date=week_start + timedelta(days=i),
            keyword="passport wallet",
            match_type="broad",
            cost=Decimal("15.00"),
            sales=Decimal("30.00"),
            purchases=3,
            clicks=35,
        )
        for i in range(7)
    ]
    product_rows = [
        _product_row(
            report_date=week_start + timedelta(days=i),
            sku="SKU-1",
            cost=Decimal("15.00"),
            sales=Decimal("30.00"),
            purchases=3,
            units=3,
        )
        for i in range(7)
    ]
    sales_rows = [_sales_row(week_start + timedelta(days=i), Decimal("100.00"), 4, 100) for i in range(7)]

    result = WeeklyAdsOptimizationService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        week_start=week_start,
        generated_at_utc=datetime(2026, 5, 23),
        campaign_rows=campaign_rows,
        targeting_rows=targeting_rows,
        search_term_rows=search_rows,
        advertised_product_rows=product_rows,
        sales_traffic_rows=sales_rows,
        sku_cost_rows=[_cost_row("SKU-1", Decimal("4.00"), Decimal("0.17"))],
        settlement_advertising_summary={
            "settlement_advertising_fee": Decimal("-30.00"),
            "settlement_row_count": 12,
            "currency": "USD",
        },
    )

    assert result.status == "ok"
    assert result.overall_summary.ads_spend == Decimal("105.00")
    assert result.overall_summary.ads_sales_7d == Decimal("210.00")
    assert result.overall_summary.acos == Decimal("0.5000")
    assert result.overall_summary.tacos == Decimal("0.1500")
    assert result.currency == "USD"
    assert len(result.campaign_performance) == 1
    assert any(
        row.action_label == "negative_candidate"
        for row in result.search_term_action_candidates
    )
    assert any(
        row.action_label == "harvest_to_exact_candidate"
        for row in result.search_term_action_candidates
    )
    product = result.advertised_product_performance[0]
    assert product.unit_standard_cost == Decimal("4.17")
    assert product.estimated_ads_cogs == Decimal("87.57")
    assert result.action_items


def test_no_ads_data_status_requires_backfill() -> None:
    result = WeeklyAdsOptimizationService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        week_start=date(2026, 5, 11),
        generated_at_utc=datetime(2026, 5, 23),
        campaign_rows=[],
        search_term_rows=[],
    )

    assert result.status == "no_ads_data"
    assert any(warning.warning_code == "ads_campaign_data_missing" for warning in result.warnings)


def test_missing_search_term_coverage_marks_needs_backfill() -> None:
    week_start = date(2026, 5, 11)
    result = WeeklyAdsOptimizationService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        week_start=week_start,
        generated_at_utc=datetime(2026, 5, 23),
        campaign_rows=[_campaign_row(week_start + timedelta(days=i)) for i in range(7)],
        search_term_rows=[
            _search_row(
                report_date=week_start,
                cost=Decimal("1.00"),
                sales=Decimal("0.00"),
                purchases=0,
                clicks=1,
            )
        ],
    )

    assert result.status == "needs_backfill"
    assert any(check.check_name == "ads_search_term_coverage" for check in result.reconciliation_checks)


def test_spend_sanity_check_marks_large_difference() -> None:
    week_start = date(2026, 5, 11)
    result = WeeklyAdsOptimizationService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        week_start=week_start,
        generated_at_utc=datetime(2026, 5, 23),
        campaign_rows=[_campaign_row(week_start + timedelta(days=i), cost=Decimal("10.00")) for i in range(7)],
        targeting_rows=[
            _targeting_row(
                report_date=week_start + timedelta(days=i),
                cost=Decimal("1.00"),
                sales=Decimal("1.00"),
                purchases=1,
                clicks=1,
            )
            for i in range(7)
        ],
        search_term_rows=[
            _search_row(
                report_date=week_start + timedelta(days=i),
                cost=Decimal("10.00"),
                sales=Decimal("10.00"),
                purchases=1,
                clicks=1,
            )
            for i in range(7)
        ],
    )

    assert any(
        check.check_name == "campaign_vs_targeting_spend"
        and check.status == "needs_review"
        for check in result.reconciliation_checks
    )
    assert result.status == "reviewable_with_warnings"


def test_write_report_files_creates_json_and_xlsx(tmp_path: Path) -> None:
    week_start = date(2026, 5, 11)
    result = WeeklyAdsOptimizationService().calculate_from_rows(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        week_start=week_start,
        generated_at_utc=datetime(2026, 5, 23),
        campaign_rows=[_campaign_row(week_start + timedelta(days=i)) for i in range(7)],
        targeting_rows=[_targeting_row(report_date=week_start + timedelta(days=i)) for i in range(7)],
        search_term_rows=[_search_row(report_date=week_start + timedelta(days=i)) for i in range(7)],
        advertised_product_rows=[_product_row(report_date=week_start + timedelta(days=i)) for i in range(7)],
        sales_traffic_rows=[_sales_row(week_start + timedelta(days=i), Decimal("10.00"), 1, 10) for i in range(7)],
    )

    written = WeeklyAdsOptimizationService().write_report_files(result=result, output_root=tmp_path)

    assert set(written.output_files) == {"json", "xlsx"}
    json_path = tmp_path / "3917953989967300/2026-05-11_2026-05-17/weekly_ads_optimization.json"
    xlsx_path = tmp_path / "3917953989967300/2026-05-11_2026-05-17/weekly_ads_optimization.xlsx"
    assert json_path.exists()
    assert xlsx_path.exists()
    workbook = load_workbook(xlsx_path, read_only=True)
    assert workbook.sheetnames == [
        "01_Executive_Summary",
        "02_Daily_Trend",
        "03_Campaigns",
        "04_Targeting",
        "05_Search_Terms",
        "06_Search_Term_Actions",
        "07_Advertised_Products",
        "08_Action_Items",
        "09_Reconciliation_Checks",
        "10_Warnings",
        "11_Raw_Metadata",
    ]


def test_run_uses_repo_and_writes_report(tmp_path: Path) -> None:
    repo = FakeWeeklyAdsRepo()
    result = WeeklyAdsOptimizationService(repo=repo).run(
        marketplace_id="ATVPDKIKX0DER",
        profile_id="3917953989967300",
        week_start=date(2026, 5, 11),
        output_root=tmp_path,
    )

    assert result.output_files["xlsx"].endswith("weekly_ads_optimization.xlsx")
    assert repo.calls == [
        "campaign",
        "targeting",
        "search",
        "product",
        "sales",
        "costs",
        "settlement",
    ]


def _campaign_row(
    report_date: date,
    cost: Decimal = Decimal("15.00"),
    sales: Decimal = Decimal("30.00"),
    purchases: int = 1,
    clicks: int = 10,
) -> dict[str, object]:
    return {
        "profile_id": "3917953989967300",
        "marketplace_id": "ATVPDKIKX0DER",
        "report_date": report_date,
        "campaign_id": "C1",
        "campaign_name": "Campaign 1",
        "campaign_status": "enabled",
        "impressions": 1000,
        "clicks": clicks,
        "cost": cost,
        "sales_7d": sales,
        "purchases_7d": purchases,
        "units_sold_clicks_7d": purchases,
    }


def _targeting_row(
    report_date: date = date(2026, 5, 11),
    keyword: str = "passport wallet",
    match_type: str = "broad",
    cost: Decimal = Decimal("15.00"),
    sales: Decimal = Decimal("30.00"),
    purchases: int = 1,
    clicks: int = 10,
) -> dict[str, object]:
    return {
        "profile_id": "3917953989967300",
        "marketplace_id": "ATVPDKIKX0DER",
        "report_date": report_date,
        "campaign_id": "C1",
        "campaign_name": "Campaign 1",
        "ad_group_id": "AG1",
        "ad_group_name": "Ad Group 1",
        "keyword_id": "K1",
        "keyword": keyword,
        "match_type": match_type,
        "targeting": keyword,
        "impressions": 1000,
        "clicks": clicks,
        "cost": cost,
        "sales_7d": sales,
        "purchases_7d": purchases,
        "units_sold_clicks_7d": purchases,
    }


def _search_row(
    report_date: date = date(2026, 5, 11),
    search_term: str = "passport wallet",
    keyword: str = "passport wallet",
    match_type: str = "broad",
    cost: Decimal = Decimal("15.00"),
    sales: Decimal = Decimal("30.00"),
    purchases: int = 1,
    clicks: int = 10,
) -> dict[str, object]:
    return {
        "profile_id": "3917953989967300",
        "marketplace_id": "ATVPDKIKX0DER",
        "report_date": report_date,
        "campaign_id": "C1",
        "campaign_name": "Campaign 1",
        "ad_group_id": "AG1",
        "ad_group_name": "Ad Group 1",
        "keyword_id": "K1",
        "keyword": keyword,
        "match_type": match_type,
        "targeting": keyword,
        "search_term": search_term,
        "impressions": 1000,
        "clicks": clicks,
        "cost": cost,
        "sales_7d": sales,
        "purchases_7d": purchases,
        "units_sold_clicks_7d": purchases,
    }


def _product_row(
    report_date: date = date(2026, 5, 11),
    sku: str = "SKU-1",
    cost: Decimal = Decimal("15.00"),
    sales: Decimal = Decimal("30.00"),
    purchases: int = 1,
    units: int = 1,
) -> dict[str, object]:
    return {
        "profile_id": "3917953989967300",
        "marketplace_id": "ATVPDKIKX0DER",
        "report_date": report_date,
        "campaign_id": "C1",
        "campaign_name": "Campaign 1",
        "ad_group_id": "AG1",
        "ad_group_name": "Ad Group 1",
        "advertised_asin": "B000TEST",
        "advertised_sku": sku,
        "impressions": 1000,
        "clicks": 10,
        "cost": cost,
        "sales_7d": sales,
        "purchases_7d": purchases,
        "units_sold_clicks_7d": units,
    }


def _sales_row(report_date: date, sales: Decimal, units: int, sessions: int) -> dict[str, object]:
    return {
        "report_date": report_date,
        "ordered_product_sales_amount": sales,
        "ordered_product_sales_currency": "USD",
        "units_ordered": units,
        "total_order_items": units,
        "sessions": sessions,
        "page_views": sessions * 2,
    }


def _cost_row(sku: str, product_cost: Decimal, first_mile_cost: Decimal) -> dict[str, object]:
    return {
        "marketplace_id": "ATVPDKIKX0DER",
        "seller_sku": sku,
        "asin": "B000TEST",
        "product_cost": product_cost,
        "first_mile_cost": first_mile_cost,
        "packaging_cost": Decimal("0.00"),
        "other_unit_cost": Decimal("0.00"),
        "currency": "USD",
        "effective_from": date(2026, 1, 1),
        "effective_to": None,
        "remark": "fixture",
    }


class FakeWeeklyAdsRepo:
    def __init__(self) -> None:
        self.calls: list[str] = []

    def fetch_campaign_daily_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("campaign")
        return [_campaign_row(start_date + timedelta(days=i)) for i in range(7)]

    def fetch_targeting_daily_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("targeting")
        return [_targeting_row(report_date=start_date + timedelta(days=i)) for i in range(7)]

    def fetch_search_term_daily_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("search")
        return [_search_row(report_date=start_date + timedelta(days=i)) for i in range(7)]

    def fetch_advertised_product_daily_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("product")
        return [_product_row(report_date=start_date + timedelta(days=i)) for i in range(7)]

    def fetch_sales_traffic_daily_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("sales")
        return [_sales_row(start_date + timedelta(days=i), Decimal("10.00"), 1, 10) for i in range(7)]

    def fetch_sku_cost_rows(
        self,
        *,
        marketplace_id: str,
        as_of_date: date,
    ) -> list[dict[str, object]]:
        self.calls.append("costs")
        return [_cost_row("SKU-1", Decimal("4.00"), Decimal("0.17"))]

    def fetch_settlement_advertising_summary(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> dict[str, object]:
        self.calls.append("settlement")
        return {
            "settlement_advertising_fee": Decimal("-105.00"),
            "settlement_row_count": 10,
            "currency": "USD",
        }
