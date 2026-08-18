from __future__ import annotations

import json
from collections import defaultdict
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass, field
from datetime import UTC, date, datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Protocol

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from seller_data_pipeline.reports.accountant_monthly_workbook_writer import (
    build_accountant_monthly_workbook,
)
from seller_data_pipeline.reports.monthly_operating_report_writer import (
    build_monthly_operating_report_workbook,
)
from seller_data_pipeline.services.calculate_profit_service import (
    PRODUCT_SALES_CATEGORIES,
    SettlementProfitLine,
    SkuCostRecord,
)
from seller_data_pipeline.services.monthly_accountant_pack import (
    add_accountant_pack_sheets,
    build_accountant_pack_payload,
)
from seller_data_pipeline.services.report_bilingual import (
    add_bilingual_readme_sheet,
    bilingual_metric_label,
    xlsx_header_label,
)

MONEY_QUANT = Decimal("0.01")
RATIO_QUANT = Decimal("0.0001")
ZERO = Decimal("0")
ZERO_VALUE_UNIT_COGS_ROLE = "zero_value_unit_cogs_reference"
REPORT_TYPE = "monthly_financial_close"
REPORT_VERSION = "v1.6-natural-month-accounting"
DEFAULT_OUTPUT_ROOT = "runtime/analysis_reports/monthly_financial_close"
REVIEW_BUCKETS = {"unknown", "unclassified"}
REVIEW_CATEGORIES = {"unknown", "unclassified"}
SKU_PROFIT_SCOPE_NOTE = (
    "SKU profit is before allocation of account-level expenses such as advertising fees, "
    "subscription fees, coupon fees, storage fees, and other non-SKU settlement rows."
)
SETTLEMENT_LED_POLICY_NOTE = (
    "Settlement-led profit remains the accounting/close source of truth; operational sources "
    "are used for management analysis and timing reconciliation."
)
MANAGEMENT_PNL_POLICY_NOTE = (
    "Management P&L uses the Finances API marketplace-local natural-month ledger for operating "
    "transactions, excludes cash Transfer rows, replaces Finances/Amazon posted advertising "
    "charges with Ads API report-date spend, deducts natural-month landed COGS, and separately "
    "writes off verified warehouse-lost inventory at effective-date landed cost."
)


@dataclass(frozen=True)
class WarningEntry:
    warning_code: str
    severity: str
    message: str
    related_sku: str | None = None
    related_source: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "warning_code": self.warning_code,
            "severity": self.severity,
            "message": self.message,
            "related_sku": self.related_sku,
            "related_source": self.related_source,
        }


@dataclass(frozen=True)
class ReconciliationCheck:
    check_name: str
    status: str
    severity: str
    expected: str | None
    actual: str | None
    diff: str | None = None
    diff_pct: str | None = None
    message: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "check_name": self.check_name,
            "status": self.status,
            "severity": self.severity,
            "expected": self.expected,
            "actual": self.actual,
            "diff": self.diff,
            "diff_pct": self.diff_pct,
            "message": self.message,
        }


@dataclass(frozen=True)
class SettlementBucketBreakdownRow:
    profit_bucket: str
    amount: Decimal
    row_count: int
    share_of_product_sales: Decimal | None = None
    share_of_settlement_net: Decimal | None = None
    notes: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "profit_bucket": self.profit_bucket,
            "amount": _decimal_to_string(self.amount),
            "absolute_amount": _decimal_to_string(abs(self.amount)),
            "share_of_product_sales": _optional_ratio_to_string(self.share_of_product_sales),
            "share_of_settlement_net": _optional_ratio_to_string(self.share_of_settlement_net),
            "row_count": self.row_count,
            "notes": self.notes,
        }


@dataclass(frozen=True)
class AmountCategoryBreakdownRow:
    amount_category: str
    profit_bucket: str
    amount: Decimal
    row_count: int
    notes: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "amount_category": self.amount_category,
            "profit_bucket": self.profit_bucket,
            "amount": _decimal_to_string(self.amount),
            "absolute_amount": _decimal_to_string(abs(self.amount)),
            "row_count": self.row_count,
            "notes": self.notes,
        }


@dataclass(frozen=True)
class MonthlySkuProfitRow:
    seller_sku: str
    units: int
    product_sales_amount: Decimal
    settlement_net_amount: Decimal
    unit_product_cost: Decimal | None
    unit_first_mile_cost: Decimal | None
    unit_packaging_cost: Decimal | None
    unit_other_cost: Decimal | None
    unit_standard_cost: Decimal | None
    product_cost_cogs: Decimal
    first_mile_cogs: Decimal
    packaging_cogs: Decimal
    other_unit_cogs: Decimal
    internal_cogs: Decimal
    estimated_profit_after_cogs: Decimal
    profit_margin: Decimal | None
    revenue_share: Decimal | None
    currency: str | None
    cost_currency: str | None
    status: str
    notes: tuple[str, ...] = ()

    def to_dict(self) -> dict[str, Any]:
        return {
            "seller_sku": self.seller_sku,
            "units": self.units,
            "product_sales_amount": _decimal_to_string(self.product_sales_amount),
            "settlement_net_amount": _decimal_to_string(self.settlement_net_amount),
            "unit_product_cost": _optional_decimal_to_string(self.unit_product_cost),
            "unit_first_mile_cost": _optional_decimal_to_string(self.unit_first_mile_cost),
            "unit_packaging_cost": _optional_decimal_to_string(self.unit_packaging_cost),
            "unit_other_cost": _optional_decimal_to_string(self.unit_other_cost),
            "unit_standard_cost": _optional_decimal_to_string(self.unit_standard_cost),
            "product_cost_cogs": _decimal_to_string(self.product_cost_cogs),
            "first_mile_cogs": _decimal_to_string(self.first_mile_cogs),
            "packaging_cogs": _decimal_to_string(self.packaging_cogs),
            "other_unit_cogs": _decimal_to_string(self.other_unit_cogs),
            "internal_cogs": _decimal_to_string(self.internal_cogs),
            "estimated_profit_after_cogs": _decimal_to_string(self.estimated_profit_after_cogs),
            "profit_margin": _optional_ratio_to_string(self.profit_margin),
            "revenue_share": _optional_ratio_to_string(self.revenue_share),
            "currency": self.currency,
            "cost_currency": self.cost_currency,
            "status": self.status,
            "notes": list(self.notes),
        }


@dataclass(frozen=True)
class SkuCogsCalculation:
    total_cogs: Decimal
    product_cost_cogs: Decimal
    first_mile_cogs: Decimal
    packaging_cogs: Decimal
    other_unit_cogs: Decimal
    cost_currency: str | None
    unit_product_cost: Decimal | None
    unit_first_mile_cost: Decimal | None
    unit_packaging_cost: Decimal | None
    unit_other_cost: Decimal | None
    unit_standard_cost: Decimal | None
    status: str
    notes: tuple[str, ...]
    costed_units: int


@dataclass(frozen=True)
class InventoryLossCalculation:
    status: str
    finance_reimbursement_amount: Decimal
    reimbursement_report_amount: Decimal
    inventory_loss_units: int
    costed_units: int
    missing_cost_skus: tuple[str, ...]
    cost_identity_resolutions: tuple[str, ...]
    product_cost: Decimal
    first_mile_cost: Decimal
    packaging_cost: Decimal
    other_unit_cost: Decimal
    landed_cost: Decimal
    details: tuple[dict[str, Any], ...] = ()
    notes: tuple[str, ...] = ()


@dataclass(frozen=True)
class OperationalMetric:
    metric_group: str
    metric_name: str
    value: Decimal | int | str | None
    currency: str | None
    source: str
    notes: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "metric_group": self.metric_group,
            "metric_name": self.metric_name,
            "value": _json_value(self.value),
            "currency": self.currency,
            "source": self.source,
            "notes": self.notes,
        }


@dataclass(frozen=True)
class MonthlyFinancialSummary:
    settlement_net_amount: Decimal
    product_sales_amount: Decimal
    product_sales_units: int
    product_cost_cogs: Decimal
    first_mile_cogs: Decimal
    packaging_cogs: Decimal
    other_unit_cogs: Decimal
    internal_cogs: Decimal
    estimated_operating_profit: Decimal
    profit_margin: Decimal | None
    advertising_cost: Decimal
    fba_fee: Decimal
    amazon_fee: Decimal
    refund: Decimal
    promotion_cost: Decimal
    promotion_fee: Decimal
    reimbursement: Decimal
    settlement_led_estimated_profit: Decimal
    settlement_led_profit_margin: Decimal | None
    settlement_advertising_fee: Decimal
    settlement_advertising_fee_abs: Decimal
    settlement_net_excluding_posted_ads: Decimal
    ads_api_report_date_spend: Decimal
    ads_timing_difference: Decimal
    ads_timing_difference_pct: Decimal | None
    management_estimated_profit_report_date_ads: Decimal
    management_profit_margin_report_date_ads: Decimal | None

    def to_dict(self) -> dict[str, Any]:
        return {
            "settlement_net_amount": _decimal_to_string(self.settlement_net_amount),
            "product_sales_amount": _decimal_to_string(self.product_sales_amount),
            "product_sales_units": self.product_sales_units,
            "product_cost_cogs": _decimal_to_string(self.product_cost_cogs),
            "first_mile_cogs": _decimal_to_string(self.first_mile_cogs),
            "packaging_cogs": _decimal_to_string(self.packaging_cogs),
            "other_unit_cogs": _decimal_to_string(self.other_unit_cogs),
            "landed_cogs": _decimal_to_string(self.internal_cogs),
            "internal_cogs": _decimal_to_string(self.internal_cogs),
            "estimated_operating_profit": _decimal_to_string(self.estimated_operating_profit),
            "profit_margin": _optional_ratio_to_string(self.profit_margin),
            "advertising_cost": _decimal_to_string(self.advertising_cost),
            "fba_fee": _decimal_to_string(self.fba_fee),
            "amazon_fee": _decimal_to_string(self.amazon_fee),
            "refund": _decimal_to_string(self.refund),
            "promotion_cost": _decimal_to_string(self.promotion_cost),
            "promotion_fee": _decimal_to_string(self.promotion_fee),
            "reimbursement": _decimal_to_string(self.reimbursement),
            "settlement_led_estimated_profit": _decimal_to_string(
                self.settlement_led_estimated_profit
            ),
            "settlement_led_profit_margin": _optional_ratio_to_string(
                self.settlement_led_profit_margin
            ),
            "settlement_advertising_fee": _decimal_to_string(self.settlement_advertising_fee),
            "settlement_advertising_fee_abs": _decimal_to_string(
                self.settlement_advertising_fee_abs
            ),
            "settlement_net_excluding_posted_ads": _decimal_to_string(
                self.settlement_net_excluding_posted_ads
            ),
            "ads_api_report_date_spend": _decimal_to_string(self.ads_api_report_date_spend),
            "ads_timing_difference": _decimal_to_string(self.ads_timing_difference),
            "ads_timing_difference_pct": _optional_ratio_to_string(
                self.ads_timing_difference_pct
            ),
            "management_estimated_profit_report_date_ads": _decimal_to_string(
                self.management_estimated_profit_report_date_ads
            ),
            "management_profit_margin_report_date_ads": _optional_ratio_to_string(
                self.management_profit_margin_report_date_ads
            ),
            "management_operating_profit": _decimal_to_string(
                self.management_estimated_profit_report_date_ads
            ),
            "management_operating_margin": _optional_ratio_to_string(
                self.management_profit_margin_report_date_ads
            ),
            "settlement_close_profit": _decimal_to_string(
                self.settlement_led_estimated_profit
            ),
        }


@dataclass(frozen=True)
class NaturalMonthFinancialSummary:
    source_status: str
    marketplace_timezone: str | None
    ledger_row_count: int
    review_required_count: int
    review_required_amount: Decimal
    product_sales_amount: Decimal
    order_total: Decimal
    refund_total: Decimal
    liquidation_total: Decimal
    service_fee_total: Decimal
    subscription_fee: Decimal
    coupon_fee: Decimal
    deal_fee: Decimal
    storage_fee: Decimal
    customer_return_fee: Decimal
    other_service_fee: Decimal
    reimbursement_total: Decimal
    warehouse_lost_reimbursement_amount: Decimal
    warehouse_lost_reimbursement_report_amount: Decimal
    adjustment_total: Decimal
    finances_ads_charge_reference: Decimal
    transfer_reference: Decimal
    operating_net_before_ads_replacement: Decimal
    product_sales_units: int
    liquidation_units: int
    costed_units: int
    missing_cost_skus: tuple[str, ...]
    cost_identity_resolutions: tuple[str, ...]
    product_cost_cogs: Decimal
    first_mile_cogs: Decimal
    packaging_cogs: Decimal
    other_unit_cogs: Decimal
    landed_cogs: Decimal
    inventory_loss_status: str
    inventory_loss_units: int
    inventory_loss_costed_units: int
    inventory_loss_missing_cost_skus: tuple[str, ...]
    inventory_loss_cost_identity_resolutions: tuple[str, ...]
    inventory_loss_product_cost: Decimal
    inventory_loss_first_mile_cost: Decimal
    inventory_loss_packaging_cost: Decimal
    inventory_loss_other_unit_cost: Decimal
    inventory_loss_landed_cost: Decimal
    inventory_loss_details: tuple[dict[str, Any], ...]
    inventory_loss_notes: tuple[str, ...]
    ads_api_report_date_spend: Decimal
    management_operating_profit: Decimal
    management_operating_margin: Decimal | None

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_status": self.source_status,
            "marketplace_timezone": self.marketplace_timezone,
            "ledger_row_count": self.ledger_row_count,
            "review_required_count": self.review_required_count,
            "review_required_amount": _decimal_to_string(self.review_required_amount),
            "product_sales_amount": _decimal_to_string(self.product_sales_amount),
            "order_total": _decimal_to_string(self.order_total),
            "refund_total": _decimal_to_string(self.refund_total),
            "liquidation_total": _decimal_to_string(self.liquidation_total),
            "service_fee_total": _decimal_to_string(self.service_fee_total),
            "subscription_fee": _decimal_to_string(self.subscription_fee),
            "coupon_fee": _decimal_to_string(self.coupon_fee),
            "deal_fee": _decimal_to_string(self.deal_fee),
            "storage_fee": _decimal_to_string(self.storage_fee),
            "customer_return_fee": _decimal_to_string(self.customer_return_fee),
            "other_service_fee": _decimal_to_string(self.other_service_fee),
            "reimbursement_total": _decimal_to_string(self.reimbursement_total),
            "warehouse_lost_reimbursement_amount": _decimal_to_string(
                self.warehouse_lost_reimbursement_amount
            ),
            "warehouse_lost_reimbursement_report_amount": _decimal_to_string(
                self.warehouse_lost_reimbursement_report_amount
            ),
            "adjustment_total": _decimal_to_string(self.adjustment_total),
            "finances_ads_charge_reference": _decimal_to_string(self.finances_ads_charge_reference),
            "transfer_reference": _decimal_to_string(self.transfer_reference),
            "operating_net_before_ads_replacement": _decimal_to_string(self.operating_net_before_ads_replacement),
            "product_sales_units": self.product_sales_units,
            "liquidation_units": self.liquidation_units,
            "costed_units": self.costed_units,
            "missing_cost_skus": list(self.missing_cost_skus),
            "cost_identity_resolutions": list(self.cost_identity_resolutions),
            "product_cost_cogs": _decimal_to_string(self.product_cost_cogs),
            "first_mile_cogs": _decimal_to_string(self.first_mile_cogs),
            "packaging_cogs": _decimal_to_string(self.packaging_cogs),
            "other_unit_cogs": _decimal_to_string(self.other_unit_cogs),
            "landed_cogs": _decimal_to_string(self.landed_cogs),
            "inventory_loss_status": self.inventory_loss_status,
            "inventory_loss_units": self.inventory_loss_units,
            "inventory_loss_costed_units": self.inventory_loss_costed_units,
            "inventory_loss_missing_cost_skus": list(self.inventory_loss_missing_cost_skus),
            "inventory_loss_cost_identity_resolutions": list(
                self.inventory_loss_cost_identity_resolutions
            ),
            "inventory_loss_product_cost": _decimal_to_string(
                self.inventory_loss_product_cost
            ),
            "inventory_loss_first_mile_cost": _decimal_to_string(
                self.inventory_loss_first_mile_cost
            ),
            "inventory_loss_packaging_cost": _decimal_to_string(
                self.inventory_loss_packaging_cost
            ),
            "inventory_loss_other_unit_cost": _decimal_to_string(
                self.inventory_loss_other_unit_cost
            ),
            "inventory_loss_landed_cost": _decimal_to_string(
                self.inventory_loss_landed_cost
            ),
            "inventory_loss_details": [
                _json_safe_mapping(detail) for detail in self.inventory_loss_details
            ],
            "inventory_loss_notes": list(self.inventory_loss_notes),
            "ads_api_report_date_spend": _decimal_to_string(self.ads_api_report_date_spend),
            "management_operating_profit": _decimal_to_string(self.management_operating_profit),
            "management_operating_margin": _optional_ratio_to_string(self.management_operating_margin),
        }


@dataclass(frozen=True)
class MonthlyFinancialCloseResult:
    marketplace_id: str
    profile_id: str | None
    month: str
    start_date: date
    end_date: date
    generated_at_utc: datetime
    status: str
    currency: str | None
    settlement_row_count: int
    financial_summary: MonthlyFinancialSummary
    settlement_bucket_breakdown: tuple[SettlementBucketBreakdownRow, ...]
    amount_category_breakdown: tuple[AmountCategoryBreakdownRow, ...]
    sku_profitability: tuple[MonthlySkuProfitRow, ...]
    operational_context: tuple[OperationalMetric, ...]
    reconciliation_checks: tuple[ReconciliationCheck, ...]
    warnings: tuple[WarningEntry, ...]
    output_files: dict[str, str] = field(default_factory=dict)
    raw_metadata: dict[str, Any] = field(default_factory=dict)
    natural_month_finance: NaturalMonthFinancialSummary | None = None

    def to_dict(self) -> dict[str, Any]:
        payload = {
            "report_type": REPORT_TYPE,
            "version": REPORT_VERSION,
            "marketplace_id": self.marketplace_id,
            "profile_id": self.profile_id,
            "period": {
                "month": self.month,
                "start_date": self.start_date.isoformat(),
                "end_date": self.end_date.isoformat(),
            },
            "generated_at_utc": self.generated_at_utc.isoformat(),
            "status": self.status,
            "currency": self.currency,
            "settlement_row_count": self.settlement_row_count,
            "executive_summary": self.executive_summary(),
            "financial_summary": self.financial_summary.to_dict(),
            "natural_month_finance": (
                self.natural_month_finance.to_dict() if self.natural_month_finance else None
            ),
            "settlement_bucket_breakdown": [
                row.to_dict() for row in self.settlement_bucket_breakdown
            ],
            "amount_category_breakdown": [row.to_dict() for row in self.amount_category_breakdown],
            "sku_profitability": [row.to_dict() for row in self.sku_profitability],
            "operational_context": [metric.to_dict() for metric in self.operational_context],
            "reconciliation_checks": [check.to_dict() for check in self.reconciliation_checks],
            "warnings": [warning.to_dict() for warning in self.warnings],
            "raw_metadata": _json_safe_mapping(self.raw_metadata),
            "accountant_pack": build_accountant_pack_payload(self),
            "methodology_notes": {
                "settlement_led_policy": SETTLEMENT_LED_POLICY_NOTE,
                "management_pnl_policy": MANAGEMENT_PNL_POLICY_NOTE,
                "sku_profit_scope": SKU_PROFIT_SCOPE_NOTE,
            },
            "output_files": dict(self.output_files),
        }
        return payload

    def with_output_files(
        self,
        output_files: Mapping[str, str],
    ) -> MonthlyFinancialCloseResult:
        return MonthlyFinancialCloseResult(
            marketplace_id=self.marketplace_id,
            profile_id=self.profile_id,
            month=self.month,
            start_date=self.start_date,
            end_date=self.end_date,
            generated_at_utc=self.generated_at_utc,
            status=self.status,
            currency=self.currency,
            settlement_row_count=self.settlement_row_count,
            financial_summary=self.financial_summary,
            settlement_bucket_breakdown=self.settlement_bucket_breakdown,
            amount_category_breakdown=self.amount_category_breakdown,
            sku_profitability=self.sku_profitability,
            operational_context=self.operational_context,
            reconciliation_checks=self.reconciliation_checks,
            warnings=self.warnings,
            raw_metadata=self.raw_metadata,
            output_files=dict(output_files),
            natural_month_finance=self.natural_month_finance,
        )

    def executive_summary(self) -> dict[str, Any]:
        fs = self.financial_summary
        settlement_profit = fs.settlement_led_estimated_profit
        management_profit = fs.management_estimated_profit_report_date_ads
        margin = fs.management_profit_margin_report_date_ads
        inventory_loss_cost = (
            self.natural_month_finance.inventory_loss_landed_cost
            if self.natural_month_finance
            else ZERO
        )
        inventory_loss_phrase = (
            " and verified warehouse-lost inventory write-offs"
            if inventory_loss_cost != ZERO
            else ""
        )
        headline = (
            f"{self.month} management operating profit after report-date ads, landed COGS"
            f"{inventory_loss_phrase} was "
            f"{_format_money(management_profit, self.currency)}; settlement close profit was "
            f"{_format_money(settlement_profit, self.currency)}."
        )
        management_sales = (
            self.natural_month_finance.product_sales_amount
            if self.natural_month_finance
            else fs.product_sales_amount
        )
        management_cogs = (
            self.natural_month_finance.landed_cogs
            if self.natural_month_finance
            else fs.internal_cogs
        )
        management_first_mile = (
            self.natural_month_finance.first_mile_cogs
            if self.natural_month_finance
            else fs.first_mile_cogs
        )
        key_points = [
            (
                "Management natural-month product sales amount was "
                f"{_format_money(management_sales, self.currency)}."
            ),
            (
                "Management natural-month landed COGS was "
                f"{_format_money(management_cogs, self.currency)}, including first-mile freight "
                f"of {_format_money(management_first_mile, self.currency)}."
            ),
            (
                "Ads timing difference was "
                f"{_format_money(fs.ads_timing_difference, self.currency)} "
                "(Ads API report-date spend minus settlement posted-date advertising fee)."
            ),
            (
                "Settlement net amount was "
                f"{_format_money(fs.settlement_net_amount, self.currency)}."
            ),
            f"Report status is {self.status}.",
        ]
        if self.natural_month_finance and inventory_loss_cost != ZERO:
            key_points.insert(
                2,
                (
                    "Verified warehouse-lost inventory write-off was "
                    f"{_format_money(inventory_loss_cost, self.currency)} for "
                    f"{self.natural_month_finance.inventory_loss_units} unit(s)."
                ),
            )
        if margin is not None:
            key_points.insert(1, f"Management profit margin was {_format_ratio(margin)}.")
        reconciliation_warning_count = sum(
            1 for check in self.reconciliation_checks if check.status == "warning"
        )
        reconciliation_needs_review_count = sum(
            1 for check in self.reconciliation_checks if check.status == "needs_review"
        )
        non_info_warning_count = sum(1 for warning in self.warnings if warning.severity != "info")
        if reconciliation_warning_count or reconciliation_needs_review_count:
            key_points.append(
                "Reconciliation warnings: "
                f"{reconciliation_warning_count}; needs_review checks: "
                f"{reconciliation_needs_review_count}."
            )
        if non_info_warning_count:
            key_points.append(f"Non-info warning count: {non_info_warning_count}.")
        return {"headline": headline, "key_points": key_points}


class MonthlyFinancialCloseDataRepo(Protocol):
    def fetch_settlement_profit_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_finances_natural_month_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_sku_cost_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_inventory_cost_identity_rows(
        self,
        *,
        marketplace_id: str,
        as_of_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_orders_period_summary(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> dict[str, Any]: ...

    def fetch_ads_period_summary(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> dict[str, Any]: ...

    def fetch_sales_traffic_period_summary(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> dict[str, Any]: ...

    def fetch_coupon_period_summary(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> dict[str, Any]: ...

    def fetch_promotion_period_summary(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> dict[str, Any]: ...

    def fetch_fba_reimbursement_period_summary(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> dict[str, Any]: ...

    def fetch_fba_reimbursement_period_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...


class MonthlyFinancialCloseService:
    def __init__(self, repo: MonthlyFinancialCloseDataRepo | None = None) -> None:
        self.repo = repo

    def run(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        month: str,
        output_root: str | Path | None = DEFAULT_OUTPUT_ROOT,
        generated_at_utc: datetime | None = None,
    ) -> MonthlyFinancialCloseResult:
        if self.repo is None:
            raise ValueError("MonthlyFinancialCloseService.run requires a repo")

        current_inputs = self._fetch_month_inputs(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            month=month,
        )
        result = self.calculate_from_rows(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            month=month,
            generated_at_utc=generated_at_utc,
            **current_inputs,
        )
        if output_root is None:
            return result

        recent_results: list[MonthlyFinancialCloseResult] = [result]
        finance_rows_by_month: dict[str, Sequence[Mapping[str, Any]]] = {
            month: current_inputs["finances_natural_month_rows"] or ()
        }
        # Three-month trend is a presentation concern. Only load it when the
        # normalized Finances ledger is available; legacy/fake repos keep the
        # historical single-month execution path unchanged.
        if hasattr(self.repo, "fetch_finances_natural_month_rows"):
            for previous_month in _previous_month_keys(month, count=2):
                previous_inputs = self._fetch_month_inputs(
                    marketplace_id=marketplace_id,
                    profile_id=profile_id,
                    month=previous_month,
                )
                previous_result = self.calculate_from_rows(
                    marketplace_id=marketplace_id,
                    profile_id=profile_id,
                    month=previous_month,
                    generated_at_utc=generated_at_utc,
                    **previous_inputs,
                )
                recent_results.append(previous_result)
                finance_rows_by_month[previous_month] = (
                    previous_inputs["finances_natural_month_rows"] or ()
                )

        return self.write_report_files(
            result=result,
            output_root=output_root,
            recent_results=recent_results,
            finance_rows_by_month=finance_rows_by_month,
        )

    def _fetch_month_inputs(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        month: str,
    ) -> dict[str, Any]:
        if self.repo is None:
            raise ValueError("MonthlyFinancialCloseService requires a repo")
        start_date, end_date = month_to_date_range(month)
        return {
            "start_date": start_date,
            "end_date": end_date,
            "settlement_rows": self.repo.fetch_settlement_profit_rows(
                marketplace_id=marketplace_id,
                start_date=start_date,
                end_date=end_date,
            ),
            "finances_natural_month_rows": (
                self.repo.fetch_finances_natural_month_rows(
                    marketplace_id=marketplace_id,
                    start_date=start_date,
                    end_date=end_date,
                )
                if hasattr(self.repo, "fetch_finances_natural_month_rows")
                else None
            ),
            "sku_cost_rows": self.repo.fetch_sku_cost_rows(
                marketplace_id=marketplace_id,
                start_date=start_date,
                end_date=end_date,
            ),
            "inventory_cost_identity_rows": (
                self.repo.fetch_inventory_cost_identity_rows(
                    marketplace_id=marketplace_id,
                    as_of_date=end_date,
                )
                if hasattr(self.repo, "fetch_inventory_cost_identity_rows")
                else ()
            ),
            "orders_summary": self.repo.fetch_orders_period_summary(
                marketplace_id=marketplace_id,
                start_date=start_date,
                end_date=end_date,
            ),
            "ads_summary": self.repo.fetch_ads_period_summary(
                marketplace_id=marketplace_id,
                profile_id=profile_id,
                start_date=start_date,
                end_date=end_date,
            ),
            "sales_traffic_summary": self.repo.fetch_sales_traffic_period_summary(
                marketplace_id=marketplace_id,
                start_date=start_date,
                end_date=end_date,
            ),
            "coupon_summary": self.repo.fetch_coupon_period_summary(
                marketplace_id=marketplace_id,
                start_date=start_date,
                end_date=end_date,
            ),
            "promotion_summary": self.repo.fetch_promotion_period_summary(
                marketplace_id=marketplace_id,
                start_date=start_date,
                end_date=end_date,
            ),
            "fba_reimbursement_summary": self.repo.fetch_fba_reimbursement_period_summary(
                marketplace_id=marketplace_id,
                start_date=start_date,
                end_date=end_date,
            ),
            "fba_reimbursement_rows": (
                self.repo.fetch_fba_reimbursement_period_rows(
                    marketplace_id=marketplace_id,
                    start_date=start_date,
                    end_date=end_date,
                )
                if hasattr(self.repo, "fetch_fba_reimbursement_period_rows")
                else ()
            ),
        }

    def calculate_from_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        month: str,
        start_date: date,
        end_date: date,
        settlement_rows: Iterable[Mapping[str, Any]],
        finances_natural_month_rows: Iterable[Mapping[str, Any]] | None = None,
        sku_cost_rows: Iterable[Mapping[str, Any]] = (),
        inventory_cost_identity_rows: Iterable[Mapping[str, Any]] = (),
        orders_summary: Mapping[str, Any] | None = None,
        ads_summary: Mapping[str, Any] | None = None,
        sales_traffic_summary: Mapping[str, Any] | None = None,
        coupon_summary: Mapping[str, Any] | None = None,
        promotion_summary: Mapping[str, Any] | None = None,
        fba_reimbursement_summary: Mapping[str, Any] | None = None,
        fba_reimbursement_rows: Iterable[Mapping[str, Any]] = (),
        generated_at_utc: datetime | None = None,
    ) -> MonthlyFinancialCloseResult:
        _validate_period(start_date=start_date, end_date=end_date)
        if month != f"{start_date.year:04d}-{start_date.month:02d}":
            raise ValueError("month must match start_date")
        generated_at = generated_at_utc or datetime.now(UTC).replace(microsecond=0)
        settlement_row_mappings = list(settlement_rows)
        finance_natural_input_provided = finances_natural_month_rows is not None
        finance_natural_rows = list(finances_natural_month_rows or [])
        reimbursement_detail_rows = list(fba_reimbursement_rows or [])
        settlement_lines = [
            SettlementProfitLine.from_mapping(row) for row in settlement_row_mappings
        ]
        settlement_lines = [line for line in settlement_lines if not line.is_settlement_summary]
        sku_costs = [SkuCostRecord.from_mapping(row) for row in sku_cost_rows]
        cost_index = _build_cost_index(sku_costs)
        cost_identity_index = _build_cost_identity_index(inventory_cost_identity_rows)

        bucket_totals, bucket_counts = _sum_by_bucket(settlement_lines)
        category_totals, category_counts, category_bucket_index = _sum_by_category(settlement_lines)
        settlement_net_amount = _money(sum((line.amount for line in settlement_lines), ZERO))
        currency = _first_non_empty(line.currency for line in settlement_lines)
        product_sales_amount = _money(
            sum(
                (
                    line.amount
                    for line in settlement_lines
                    if line.amount_category in PRODUCT_SALES_CATEGORIES
                ),
                ZERO,
            )
        )

        sku_accumulator = _build_sku_accumulator(settlement_lines)
        sku_rows: list[MonthlySkuProfitRow] = []
        missing_cost_skus: set[str] = set()
        currency_mismatch_skus: set[str] = set()
        internal_cogs = ZERO
        product_cost_cogs = ZERO
        first_mile_cogs = ZERO
        packaging_cogs = ZERO
        other_unit_cogs = ZERO
        product_sales_units = 0
        costed_units = 0

        for seller_sku in sorted(sku_accumulator):
            sku_data = sku_accumulator[seller_sku]
            units = int(sku_data["units"])
            sku_product_sales = _money(sku_data["product_sales_amount"])
            sku_settlement_net = _money(sku_data["settlement_net_amount"])
            product_sales_units += units
            cogs = _calculate_sku_cogs(
                cost_index=cost_index,
                marketplace_id=marketplace_id,
                seller_sku=seller_sku,
                unit_events=sku_data["unit_events"],
                settlement_currency=currency,
            )
            if cogs.status == "missing_cost":
                missing_cost_skus.add(seller_sku)
            if cogs.status == "currency_mismatch":
                currency_mismatch_skus.add(seller_sku)
            costed_units += cogs.costed_units
            internal_cogs += cogs.total_cogs
            product_cost_cogs += cogs.product_cost_cogs
            first_mile_cogs += cogs.first_mile_cogs
            packaging_cogs += cogs.packaging_cogs
            other_unit_cogs += cogs.other_unit_cogs
            estimated_profit = _money(sku_settlement_net - cogs.total_cogs)
            sku_rows.append(
                MonthlySkuProfitRow(
                    seller_sku=seller_sku,
                    units=units,
                    product_sales_amount=sku_product_sales,
                    settlement_net_amount=sku_settlement_net,
                    unit_product_cost=cogs.unit_product_cost,
                    unit_first_mile_cost=cogs.unit_first_mile_cost,
                    unit_packaging_cost=cogs.unit_packaging_cost,
                    unit_other_cost=cogs.unit_other_cost,
                    unit_standard_cost=cogs.unit_standard_cost,
                    product_cost_cogs=cogs.product_cost_cogs,
                    first_mile_cogs=cogs.first_mile_cogs,
                    packaging_cogs=cogs.packaging_cogs,
                    other_unit_cogs=cogs.other_unit_cogs,
                    internal_cogs=cogs.total_cogs,
                    estimated_profit_after_cogs=estimated_profit,
                    profit_margin=_safe_ratio(estimated_profit, sku_product_sales),
                    revenue_share=_safe_ratio(sku_product_sales, product_sales_amount),
                    currency=currency,
                    cost_currency=cogs.cost_currency,
                    status=cogs.status,
                    notes=cogs.notes,
                )
            )

        product_cost_cogs = _money(product_cost_cogs)
        first_mile_cogs = _money(first_mile_cogs)
        packaging_cogs = _money(packaging_cogs)
        other_unit_cogs = _money(other_unit_cogs)
        internal_cogs = _money(internal_cogs)
        settlement_advertising_fee = _money(bucket_totals.get("advertising_cost", ZERO))
        settlement_advertising_fee_abs = _money(abs(settlement_advertising_fee))
        ads_api_report_date_spend = _money(_to_decimal((ads_summary or {}).get("ads_cost")))
        settlement_led_estimated_profit = _money(settlement_net_amount - internal_cogs)
        settlement_net_excluding_posted_ads = _money(
            settlement_net_amount - settlement_advertising_fee
        )
        ads_timing_difference = _money(ads_api_report_date_spend - settlement_advertising_fee_abs)
        natural_month_finance = _build_natural_month_financial_summary(
            finance_rows=finance_natural_rows,
            reimbursement_rows=reimbursement_detail_rows,
            cost_index=cost_index,
            cost_identity_index=cost_identity_index,
            marketplace_id=marketplace_id,
            ads_api_report_date_spend=ads_api_report_date_spend,
            fallback_currency=currency,
        )
        if natural_month_finance is not None:
            management_estimated_profit = natural_month_finance.management_operating_profit
            management_profit_margin = natural_month_finance.management_operating_margin
        else:
            management_estimated_profit = _money(
                settlement_net_excluding_posted_ads - ads_api_report_date_spend - internal_cogs
            )
            management_profit_margin = _safe_ratio(
                management_estimated_profit, product_sales_amount
            )
        financial_summary = MonthlyFinancialSummary(
            settlement_net_amount=settlement_net_amount,
            product_sales_amount=product_sales_amount,
            product_sales_units=product_sales_units,
            product_cost_cogs=product_cost_cogs,
            first_mile_cogs=first_mile_cogs,
            packaging_cogs=packaging_cogs,
            other_unit_cogs=other_unit_cogs,
            internal_cogs=internal_cogs,
            estimated_operating_profit=settlement_led_estimated_profit,
            profit_margin=_safe_ratio(settlement_led_estimated_profit, product_sales_amount),
            advertising_cost=settlement_advertising_fee,
            fba_fee=_money(bucket_totals.get("fba_fee", ZERO)),
            amazon_fee=_money(bucket_totals.get("amazon_fee", ZERO)),
            refund=_money(bucket_totals.get("refund", ZERO)),
            promotion_cost=_money(bucket_totals.get("promotion_cost", ZERO)),
            promotion_fee=_money(bucket_totals.get("promotion_fee", ZERO)),
            reimbursement=_money(bucket_totals.get("reimbursement", ZERO)),
            settlement_led_estimated_profit=settlement_led_estimated_profit,
            settlement_led_profit_margin=_safe_ratio(
                settlement_led_estimated_profit, product_sales_amount
            ),
            settlement_advertising_fee=settlement_advertising_fee,
            settlement_advertising_fee_abs=settlement_advertising_fee_abs,
            settlement_net_excluding_posted_ads=settlement_net_excluding_posted_ads,
            ads_api_report_date_spend=ads_api_report_date_spend,
            ads_timing_difference=ads_timing_difference,
            ads_timing_difference_pct=_safe_ratio(
                ads_timing_difference, settlement_advertising_fee_abs
            ),
            management_estimated_profit_report_date_ads=management_estimated_profit,
            management_profit_margin_report_date_ads=management_profit_margin,
        )
        bucket_rows = _build_bucket_rows(
            bucket_totals=bucket_totals,
            bucket_counts=bucket_counts,
            product_sales_amount=product_sales_amount,
            settlement_net_amount=settlement_net_amount,
        )
        category_rows = _build_category_rows(
            category_totals=category_totals,
            category_counts=category_counts,
            category_bucket_index=category_bucket_index,
        )
        operational_context = _build_operational_context(
            orders_summary=orders_summary or {},
            ads_summary=ads_summary or {},
            sales_traffic_summary=sales_traffic_summary or {},
            coupon_summary=coupon_summary or {},
            promotion_summary=promotion_summary or {},
            fba_reimbursement_summary=fba_reimbursement_summary or {},
            currency=currency,
        )
        reconciliation_checks = _build_reconciliation_checks(
            settlement_net_amount=settlement_net_amount,
            bucket_totals=bucket_totals,
            category_totals=category_totals,
            product_sales_amount=product_sales_amount,
            product_sales_units=product_sales_units,
            costed_units=costed_units,
            missing_cost_skus=missing_cost_skus,
            currency_mismatch_skus=currency_mismatch_skus,
            financial_summary=financial_summary,
            sales_traffic_summary=sales_traffic_summary or {},
            ads_summary=ads_summary or {},
            fba_reimbursement_summary=fba_reimbursement_summary or {},
        )
        if finance_natural_input_provided and natural_month_finance is None:
            reconciliation_checks.append(
                ReconciliationCheck(
                    check_name="finances_natural_month_coverage",
                    status="needs_review",
                    severity="error",
                    expected="non-empty natural-month ledger",
                    actual="0 rows",
                    message=(
                        "v1.90 Management P&L requires amazon_finance_transaction for the "
                        "marketplace-local month. Run ingest_finances_natural_month.py first."
                    ),
                )
            )
        elif natural_month_finance is not None and natural_month_finance.source_status != "ok":
            reconciliation_checks.append(
                ReconciliationCheck(
                    check_name="finances_natural_month_coverage",
                    status="needs_review",
                    severity="error",
                    expected="recognized lifecycle rows and full SKU cost coverage",
                    actual=natural_month_finance.source_status,
                    message=(
                        f"review_rows={natural_month_finance.review_required_count}; "
                        f"review_amount={natural_month_finance.review_required_amount}; "
                        f"costed_units={natural_month_finance.costed_units}; "
                        f"expected_units={natural_month_finance.product_sales_units + natural_month_finance.liquidation_units}; "
                        f"missing_cost_skus={list(natural_month_finance.missing_cost_skus)}; "
                        f"inventory_loss_status={natural_month_finance.inventory_loss_status}; "
                        f"inventory_loss_units={natural_month_finance.inventory_loss_units}; "
                        f"inventory_loss_costed_units={natural_month_finance.inventory_loss_costed_units}; "
                        f"inventory_loss_missing_cost_skus={list(natural_month_finance.inventory_loss_missing_cost_skus)}"
                    ),
                )
            )
        elif natural_month_finance is not None:
            reconciliation_checks.append(
                ReconciliationCheck(
                    check_name="finances_natural_month_coverage",
                    status="ok",
                    severity="info",
                    expected="ok",
                    actual="ok",
                    message=(
                        f"Natural-month ledger rows={natural_month_finance.ledger_row_count}; "
                        f"timezone={natural_month_finance.marketplace_timezone}."
                    ),
                )
            )
        if natural_month_finance is not None:
            inventory_loss_status = natural_month_finance.inventory_loss_status
            if inventory_loss_status == "needs_review":
                reconciliation_checks.append(
                    ReconciliationCheck(
                        check_name="warehouse_lost_inventory_writeoff",
                        status="needs_review",
                        severity="error",
                        expected="warehouse-lost reimbursement detail and effective landed cost fully reconciled",
                        actual=inventory_loss_status,
                        message=(
                            f"finances_reimbursement={natural_month_finance.warehouse_lost_reimbursement_amount}; "
                            f"reimbursement_report={natural_month_finance.warehouse_lost_reimbursement_report_amount}; "
                            f"loss_units={natural_month_finance.inventory_loss_units}; "
                            f"costed_units={natural_month_finance.inventory_loss_costed_units}; "
                            f"missing_cost_skus={list(natural_month_finance.inventory_loss_missing_cost_skus)}"
                        ),
                    )
                )
            elif inventory_loss_status == "ok":
                reconciliation_checks.append(
                    ReconciliationCheck(
                        check_name="warehouse_lost_inventory_writeoff",
                        status="ok",
                        severity="info",
                        expected="verified warehouse-lost inventory write-off",
                        actual=_decimal_to_string(natural_month_finance.inventory_loss_landed_cost),
                        message=(
                            f"warehouse_lost_reimbursement={natural_month_finance.warehouse_lost_reimbursement_amount}; "
                            f"reimbursement_report={natural_month_finance.warehouse_lost_reimbursement_report_amount}; "
                            f"loss_units={natural_month_finance.inventory_loss_units}; "
                            f"landed_cost_writeoff={natural_month_finance.inventory_loss_landed_cost}"
                        ),
                    )
                )
        warnings = _build_warnings(
            settlement_row_count=len(settlement_lines),
            missing_cost_skus=missing_cost_skus,
            currency_mismatch_skus=currency_mismatch_skus,
            product_sales_units=product_sales_units,
            financial_summary=financial_summary,
            ads_summary=ads_summary or {},
            reconciliation_checks=reconciliation_checks,
        )
        status = _result_status(
            settlement_row_count=len(settlement_lines),
            reconciliation_checks=reconciliation_checks,
        )
        raw_metadata = _build_raw_metadata(
            settlement_lines=settlement_lines,
            raw_settlement_rows=settlement_row_mappings,
            sku_costs=sku_costs,
            orders_summary=orders_summary or {},
            ads_summary=ads_summary or {},
            sales_traffic_summary=sales_traffic_summary or {},
            coupon_summary=coupon_summary or {},
            promotion_summary=promotion_summary or {},
            fba_reimbursement_summary=fba_reimbursement_summary or {},
        )
        raw_metadata["finances_natural_month_row_count"] = len(finance_natural_rows)
        raw_metadata["fba_reimbursement_detail_row_count"] = len(reimbursement_detail_rows)
        raw_metadata["management_pnl_source"] = (
            "finances_api_natural_month"
            if natural_month_finance is not None
            else "legacy_settlement_fallback"
        )
        return MonthlyFinancialCloseResult(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            month=month,
            start_date=start_date,
            end_date=end_date,
            generated_at_utc=generated_at,
            status=status,
            currency=currency,
            settlement_row_count=len(settlement_lines),
            financial_summary=financial_summary,
            settlement_bucket_breakdown=tuple(bucket_rows),
            amount_category_breakdown=tuple(category_rows),
            sku_profitability=tuple(sku_rows),
            operational_context=tuple(operational_context),
            reconciliation_checks=tuple(reconciliation_checks),
            warnings=tuple(warnings),
            raw_metadata=raw_metadata,
            natural_month_finance=natural_month_finance,
        )

    def write_report_files(
        self,
        *,
        result: MonthlyFinancialCloseResult,
        output_root: str | Path,
        recent_results: Sequence[MonthlyFinancialCloseResult] | None = None,
        finance_rows_by_month: Mapping[str, Sequence[Mapping[str, Any]]] | None = None,
    ) -> MonthlyFinancialCloseResult:
        output_dir = Path(output_root) / result.marketplace_id / result.month
        output_dir.mkdir(parents=True, exist_ok=True)
        filename_base = f"monthly_financial_close_{result.month}"
        json_path = output_dir / f"{filename_base}.json"
        legacy_xlsx_path = output_dir / f"{filename_base}.xlsx"
        operating_xlsx_path = output_dir / f"monthly_operating_report_{result.month}.xlsx"
        accounting_xlsx_path = output_dir / f"accountant_monthly_workbook_{result.month}.xlsx"
        output_files = {
            "json": str(json_path),
            "xlsx": str(legacy_xlsx_path),
            "operating_xlsx": str(operating_xlsx_path),
            "accounting_xlsx": str(accounting_xlsx_path),
        }
        result_with_paths = result.with_output_files(output_files)
        json_path.write_text(
            json.dumps(
                result_with_paths.to_dict(),
                ensure_ascii=False,
                indent=2,
                sort_keys=True,
            )
            + "\n",
            encoding="utf-8",
        )

        legacy_workbook = build_monthly_financial_close_workbook(result_with_paths)
        legacy_workbook.save(legacy_xlsx_path)

        operating_workbook = build_monthly_operating_report_workbook(
            result_with_paths,
            recent_results=recent_results,
            finance_rows_by_month=finance_rows_by_month,
        )
        operating_workbook.save(operating_xlsx_path)

        current_finance_rows: Sequence[Mapping[str, Any]] = ()
        if finance_rows_by_month:
            current_finance_rows = finance_rows_by_month.get(result.month, ())
        accounting_workbook = build_accountant_monthly_workbook(
            result_with_paths,
            current_finance_rows,
        )
        accounting_workbook.save(accounting_xlsx_path)
        return result_with_paths


def month_to_date_range(month: str) -> tuple[date, date]:
    try:
        year_text, month_text = month.split("-", 1)
        year = int(year_text)
        month_number = int(month_text)
        start = date(year, month_number, 1)
    except ValueError as exc:
        raise ValueError("month must be in YYYY-MM format") from exc
    if month_number == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month_number + 1, 1)
    return start, next_month - timedelta(days=1)


def _previous_month_keys(month: str, *, count: int) -> list[str]:
    start, _ = month_to_date_range(month)
    cursor = start
    keys: list[str] = []
    for _ in range(count):
        previous_day = cursor - timedelta(days=1)
        previous_start = date(previous_day.year, previous_day.month, 1)
        keys.append(f"{previous_start.year:04d}-{previous_start.month:02d}")
        cursor = previous_start
    return keys


def build_monthly_financial_close_workbook(result: MonthlyFinancialCloseResult) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_bilingual_readme_sheet(
        workbook,
        title_en="Monthly Financial Close Report",
        title_zh="月度财务结算报表",
        period=result.month,
        status=result.status,
        scope_en="Management P&L uses Finances API natural-month transactions + Ads API; Settlement remains the close/cash reconciliation source.",
        scope_zh="经营利润采用 Finances API 自然月交易 + Ads API；Settlement 保留为结算/现金对账主口径。",
    )
    _write_rows_sheet(workbook, "01_Summary", _summary_rows(result))
    _format_executive_metric_sheet(workbook["01_Summary"], summary=True)
    _write_rows_sheet(workbook, "02_Management_PnL", _management_pnl_rows(result))
    _format_executive_metric_sheet(workbook["02_Management_PnL"], summary=False)
    _write_rows_sheet(workbook, "03_Ads_Timing_Recon", _ads_timing_reconciliation_rows(result))
    _write_rows_sheet(
        workbook,
        "04_Settlement_Buckets",
        [row.to_dict() for row in result.settlement_bucket_breakdown],
    )
    _write_rows_sheet(
        workbook,
        "05_Amount_Categories",
        [row.to_dict() for row in result.amount_category_breakdown],
    )
    _write_rows_sheet(
        workbook,
        "06_SKU_Profit",
        [_flatten_sku_row(row) for row in result.sku_profitability],
    )
    _write_rows_sheet(
        workbook,
        "07_Operational_Context",
        [metric.to_dict() for metric in result.operational_context],
    )
    _write_rows_sheet(
        workbook,
        "08_Reconciliation_Checks",
        [check.to_dict() for check in result.reconciliation_checks],
    )
    _write_rows_sheet(workbook, "09_Warnings", [warning.to_dict() for warning in result.warnings])
    _write_rows_sheet(workbook, "10_Raw_Metadata", _metadata_rows(result))
    add_accountant_pack_sheets(workbook, result)
    workbook.active = 0
    return workbook


def _summary_rows(result: MonthlyFinancialCloseResult) -> list[dict[str, Any]]:
    fs = result.financial_summary
    nm = result.natural_month_finance
    management_sales = nm.product_sales_amount if nm else fs.product_sales_amount
    management_units = (
        nm.product_sales_units + nm.liquidation_units if nm else fs.product_sales_units
    )
    management_product_cogs = nm.product_cost_cogs if nm else fs.product_cost_cogs
    management_first_mile = nm.first_mile_cogs if nm else fs.first_mile_cogs
    management_packaging = nm.packaging_cogs if nm else fs.packaging_cogs
    management_other_unit = nm.other_unit_cogs if nm else fs.other_unit_cogs
    management_landed = nm.landed_cogs if nm else fs.internal_cogs
    return [
        _metric_row("Report Type", REPORT_TYPE, None, "Monthly Financial Close Report v1.5"),
        _metric_row("Marketplace ID", result.marketplace_id, None, ""),
        _metric_row(
            "Ads Profile ID",
            result.profile_id or "",
            None,
            "Used only for Ads API context.",
        ),
        _metric_row("Month", result.month, None, "Natural calendar month."),
        _metric_row("Period Start", result.start_date.isoformat(), None, "Inclusive."),
        _metric_row("Period End", result.end_date.isoformat(), None, "Inclusive."),
        _metric_row("Status", result.status, None, "ok / needs_review / no_data."),
        _metric_row("Currency", result.currency or "", None, "Settlement currency."),
        _metric_row("Settlement Rows", result.settlement_row_count, None, "Non-summary rows."),
        _metric_row(
            "Management Operating Profit",
            fs.management_estimated_profit_report_date_ads,
            result.currency,
            "Primary operating KPI: report-date Ads spend and total landed COGS included.",
        ),
        _metric_row(
            "Management Operating Margin",
            fs.management_profit_margin_report_date_ads,
            None,
            "Management operating profit / settlement product sales.",
        ),
        _metric_row(
            "Product Sales Amount",
            management_sales,
            result.currency,
            "Finances API marketplace-local natural-month product sales; management margin denominator.",
        ),
        _metric_row(
            "Product Sales Units",
            management_units,
            None,
            "Natural-month Shipment + liquidation units used for landed COGS allocation.",
        ),
        _metric_row(
            "Total Landed COGS",
            management_landed,
            result.currency,
            "Natural-month Product + first-mile + packaging + other unit costs from amazon_sku_cost.",
        ),
        _metric_row(
            "Product Cost COGS",
            management_product_cogs,
            result.currency,
            "Factory/product purchase cost allocated to natural-month units.",
        ),
        _metric_row(
            "First-Mile Freight COGS",
            management_first_mile,
            result.currency,
            "First-mile/ocean freight/customs/inbound allocation for natural-month units.",
        ),
        _metric_row(
            "Packaging COGS",
            management_packaging,
            result.currency,
            "Packaging unit cost allocated to natural-month units.",
        ),
        _metric_row(
            "Other Unit COGS",
            management_other_unit,
            result.currency,
            "Other stable unit cost allocated to natural-month units.",
        ),
        _metric_row(
            "Ads API Report-date Spend",
            fs.ads_api_report_date_spend,
            result.currency,
            "Ads API spend by report_date; management P&L ad cost.",
        ),
        _metric_row(
            "Settlement Close Profit",
            fs.settlement_led_estimated_profit,
            result.currency,
            "Accounting/close reference: settlement net minus total landed COGS.",
        ),
        _metric_row(
            "Settlement Net Amount",
            fs.settlement_net_amount,
            result.currency,
            "Posted-date settlement net; accounting/close source of truth.",
        ),
        _metric_row(
            "Settlement Advertising Fee",
            fs.settlement_advertising_fee,
            result.currency,
            "Posted-date advertising bucket; replaced by report-date Ads spend in management P&L.",
        ),
        _metric_row(
            "Ads Timing Difference",
            fs.ads_timing_difference,
            result.currency,
            "Ads API report-date spend minus absolute settlement advertising fee.",
        ),
        _metric_row("FBA Fee", fs.fba_fee, result.currency, "Settlement FBA fulfillment bucket."),
        _metric_row("Amazon Fee", fs.amazon_fee, result.currency, "Settlement Amazon fee bucket."),
        _metric_row("Refund", fs.refund, result.currency, "Settlement refund bucket."),
        _metric_row(
            "Promotion Cost",
            fs.promotion_cost,
            result.currency,
            "Settlement promotion discount bucket.",
        ),
        _metric_row(
            "Promotion Fee",
            fs.promotion_fee,
            result.currency,
            "Settlement coupon/deal fee bucket.",
        ),
        _metric_row(
            "Reimbursement",
            fs.reimbursement,
            result.currency,
            "Settlement reimbursement bucket.",
        ),
        _metric_row(
            "SKU Cost Coverage",
            _check_status(result, "sku_cost_coverage"),
            None,
            "ok means all product-sales units matched an effective amazon_sku_cost row.",
        ),
        _metric_row(
            "First-Mile Cost Included",
            "yes" if management_first_mile != ZERO else "no_or_zero",
            None,
            "Informational only: zero can be valid, but should be reviewed when first-mile is material.",
        ),
        _metric_row(
            "Unknown Settlement Classification",
            _unknown_classification_status(result),
            None,
            "ok means unknown/unclassified financial amounts are zero.",
        ),
        _metric_row(
            "Bank Payout Reconciliation",
            "pending_manual",
            None,
            "Settlement close does not yet prove WorldFirst/bank cash receipt; use 14_Payout_Recon.",
        ),
        _metric_row(
            "SKU Profit Table Scope",
            "Before account-level expense allocation",
            None,
            SKU_PROFIT_SCOPE_NOTE,
        ),
    ]


def _management_pnl_rows(result: MonthlyFinancialCloseResult) -> list[dict[str, Any]]:
    fs = result.financial_summary
    nm = result.natural_month_finance
    if nm is not None:
        contribution_after_ads = _money(
            nm.operating_net_before_ads_replacement - nm.ads_api_report_date_spend
        )
        return [
            _metric_row(
                "Product Sales Amount", nm.product_sales_amount, result.currency,
                "Finances API natural-month Shipment product charges in marketplace local time.",
            ),
            _metric_row("Order Total", nm.order_total, result.currency, "DEFERRED_RELEASED Shipment natural-month total."),
            _metric_row("Refund Total", nm.refund_total, result.currency, "RELEASED Refund natural-month total."),
            _metric_row("Liquidation Total", nm.liquidation_total, result.currency, "DEFERRED + DEFERRED_RELEASED RemovalShipment total."),
            _metric_row("Subscription Fee", nm.subscription_fee, result.currency, "Component of RELEASED ServiceFee."),
            _metric_row("Coupon Fees", nm.coupon_fee, result.currency, "Coupon participation/performance fees within RELEASED ServiceFee."),
            _metric_row("Deal Fees", nm.deal_fee, result.currency, "Deal participation/performance fees within RELEASED ServiceFee."),
            _metric_row("Storage Fee", nm.storage_fee, result.currency, "FBA storage fees within RELEASED ServiceFee."),
            _metric_row("Customer Returns Fee", nm.customer_return_fee, result.currency, "FBA customer-return/HRR fees within RELEASED ServiceFee."),
            _metric_row("Other Service Fees", nm.other_service_fee, result.currency, "Residual RELEASED ServiceFee amount not mapped to the named components."),
            _metric_row("Service Fees Total", nm.service_fee_total, result.currency, "Subtotal of all RELEASED ServiceFee transactions; component rows above are informational and not added again."),
            _metric_row("Reimbursements", nm.reimbursement_total, result.currency, "RELEASED FBAInventoryReimbursement total."),
            _metric_row("Other Adjustments", nm.adjustment_total, result.currency, "Recognized RELEASED MiscellaneousLedgerAdjustment total."),
            _metric_row(
                "Operating Net Before Ads Replacement",
                nm.operating_net_before_ads_replacement,
                result.currency,
                "Natural-month operating transactions; Transfer and Finances posted advertising excluded.",
            ),
            _metric_row(
                "Finances Posted Advertising Reference",
                nm.finances_ads_charge_reference,
                result.currency,
                "ProductAdsPayment reference only; excluded from Management P&L and replaced by Ads API spend.",
            ),
            _metric_row(
                "Less Ads API Report-date Spend",
                -nm.ads_api_report_date_spend,
                result.currency,
                "Calendar-month advertising cost by Ads API report_date.",
            ),
            _metric_row(
                "Contribution After Ads Before Landed COGS",
                contribution_after_ads,
                result.currency,
                "Natural-month operating net less Ads API spend.",
            ),
            _metric_row("Less Product Cost COGS", -nm.product_cost_cogs, result.currency, "Natural-month unit product cost."),
            _metric_row("Less First-Mile Freight COGS", -nm.first_mile_cogs, result.currency, "Natural-month unit first-mile allocation."),
            _metric_row("Less Packaging COGS", -nm.packaging_cogs, result.currency, "Natural-month packaging allocation."),
            _metric_row("Less Other Unit COGS", -nm.other_unit_cogs, result.currency, "Natural-month other unit-cost allocation."),
            _metric_row("Total Landed COGS", -nm.landed_cogs, result.currency, "Natural-month landed COGS."),
            _metric_row("Management Operating Profit", nm.management_operating_profit, result.currency, MANAGEMENT_PNL_POLICY_NOTE),
            _metric_row("Management Operating Margin", nm.management_operating_margin, None, "Management operating profit / natural-month product sales."),
            _metric_row(
                "Finances API Transfer Reference (API Sign)",
                nm.transfer_reference,
                result.currency,
                "Cash/settlement reference only; excluded from operating profit. API sign is preserved and may be opposite to Seller Central export presentation.",
            ),
            _metric_row("Settlement Close Profit", fs.settlement_led_estimated_profit, result.currency, "Separate posted-date accounting/close reference."),
        ]
    rows = [
        _metric_row(
            "Product Sales Amount",
            fs.product_sales_amount,
            result.currency,
            "Reference sales KPI; settlement bucket detail below reconciles to settlement net.",
        )
    ]

    bucket_map = {row.profit_bucket: row.amount for row in result.settlement_bucket_breakdown}
    preferred_buckets = [
        "revenue",
        "reimbursement",
        "liquidation",
        "other",
        "refund",
        "promotion_cost",
        "promotion_fee",
        "fba_fee",
        "fba_storage_fee",
        "amazon_fee",
        "amazon_fee_refund",
        "liquidation_fee",
        "tax_passthrough",
        "reconciliation",
    ]
    bucket_labels = {
        "revenue": "Settlement Revenue Bucket",
        "reimbursement": "Settlement Reimbursement",
        "liquidation": "Settlement Liquidation Income",
        "other": "Settlement Other",
        "refund": "Settlement Refund",
        "promotion_cost": "Settlement Promotion Discounts",
        "promotion_fee": "Settlement Promotion Fees",
        "fba_fee": "Settlement FBA Fulfillment Fees",
        "fba_storage_fee": "Settlement FBA Storage Fees",
        "amazon_fee": "Settlement Amazon Platform Fees",
        "amazon_fee_refund": "Settlement Amazon Fee Refunds",
        "liquidation_fee": "Settlement Liquidation Fees",
        "tax_passthrough": "Settlement Tax Passthrough",
        "reconciliation": "Settlement Reconciliation Items",
    }
    emitted: set[str] = set()
    for bucket in preferred_buckets:
        if bucket not in bucket_map:
            continue
        emitted.add(bucket)
        rows.append(
            _metric_row(
                bucket_labels[bucket],
                bucket_map[bucket],
                result.currency,
                f"Settlement posted-date bucket: {bucket}.",
            )
        )
    for bucket in sorted(bucket_map):
        if bucket in emitted or bucket == "advertising_cost":
            continue
        rows.append(
            _metric_row(
                f"Settlement Bucket: {bucket}",
                bucket_map[bucket],
                result.currency,
                "Additional settlement bucket retained for completeness.",
            )
        )

    contribution_after_ads = _money(
        fs.settlement_net_excluding_posted_ads - fs.ads_api_report_date_spend
    )
    rows.extend(
        [
            _metric_row(
                "Settlement Net Excluding Posted Ads",
                fs.settlement_net_excluding_posted_ads,
                result.currency,
                "All settlement buckets except posted-date advertising.",
            ),
            _metric_row(
                "Less Ads API Report-date Spend",
                -fs.ads_api_report_date_spend,
                result.currency,
                "Calendar-month advertising cost by Ads API report_date.",
            ),
            _metric_row(
                "Contribution After Ads Before Landed COGS",
                contribution_after_ads,
                result.currency,
                "Settlement net excluding posted ads, less report-date Ads spend.",
            ),
            _metric_row(
                "Less Product Cost COGS",
                -fs.product_cost_cogs,
                result.currency,
                "Factory/product purchase cost allocated to sold units.",
            ),
            _metric_row(
                "Less First-Mile Freight COGS",
                -fs.first_mile_cogs,
                result.currency,
                "First-mile/ocean freight/customs/inbound allocation for sold units.",
            ),
            _metric_row(
                "Less Packaging COGS",
                -fs.packaging_cogs,
                result.currency,
                "Packaging unit cost for sold units.",
            ),
            _metric_row(
                "Less Other Unit COGS",
                -fs.other_unit_cogs,
                result.currency,
                "Other stable unit cost for sold units.",
            ),
            _metric_row(
                "Total Landed COGS",
                -fs.internal_cogs,
                result.currency,
                "Product + first-mile + packaging + other unit costs.",
            ),
            _metric_row(
                "Management Operating Profit",
                fs.management_estimated_profit_report_date_ads,
                result.currency,
                MANAGEMENT_PNL_POLICY_NOTE,
            ),
            _metric_row(
                "Management Operating Margin",
                fs.management_profit_margin_report_date_ads,
                None,
                "Management operating profit / settlement product sales.",
            ),
            _metric_row(
                "Settlement Close Profit",
                fs.settlement_led_estimated_profit,
                result.currency,
                "Accounting/close reference using settlement posted-date advertising.",
            ),
        ]
    )
    return rows


def _check_status(result: MonthlyFinancialCloseResult, check_name: str) -> str:
    check = next(
        (item for item in result.reconciliation_checks if item.check_name == check_name),
        None,
    )
    return check.status if check is not None else "not_available"


def _unknown_classification_status(result: MonthlyFinancialCloseResult) -> str:
    names = {"unknown_bucket_amount", "unclassified_amount"}
    statuses = [
        check.status for check in result.reconciliation_checks if check.check_name in names
    ]
    if not statuses:
        return "not_available"
    return "ok" if all(status == "ok" for status in statuses) else "needs_review"

def _ads_timing_reconciliation_rows(result: MonthlyFinancialCloseResult) -> list[dict[str, Any]]:
    fs = result.financial_summary
    status = _ads_timing_status(
        ads_api_spend=fs.ads_api_report_date_spend,
        settlement_ads_abs=fs.settlement_advertising_fee_abs,
    )
    return [
        {
            "metric": bilingual_metric_label("Settlement Advertising Fee"),
            "value": _xlsx_value(fs.settlement_advertising_fee),
            "currency": result.currency,
            "source": "amazon_settlement_transaction",
            "notes": "Posted-date accounting/close advertising fee.",
        },
        {
            "metric": bilingual_metric_label("Settlement Advertising Fee Abs"),
            "value": _xlsx_value(fs.settlement_advertising_fee_abs),
            "currency": result.currency,
            "source": "amazon_settlement_transaction",
            "notes": "Absolute posted-date advertising fee used for timing comparison.",
        },
        {
            "metric": bilingual_metric_label("Ads API Report-date Spend"),
            "value": _xlsx_value(fs.ads_api_report_date_spend),
            "currency": result.currency,
            "source": "amazon_ads_sp_campaign_daily",
            "notes": "Report-date spend used by management P&L.",
        },
        {
            "metric": bilingual_metric_label("Ads Timing Difference"),
            "value": _xlsx_value(fs.ads_timing_difference),
            "currency": result.currency,
            "source": "derived",
            "notes": "Ads API spend minus absolute settlement advertising fee.",
        },
        {
            "metric": bilingual_metric_label("Ads Timing Difference Pct"),
            "value": _xlsx_value(fs.ads_timing_difference_pct),
            "currency": None,
            "source": "derived",
            "notes": "Difference divided by absolute settlement advertising fee.",
        },
        {
            "metric": bilingual_metric_label("Ads Timing Status"),
            "value": status,
            "currency": None,
            "source": "derived",
            "notes": "ok < $20 or <5%; otherwise warning because Settlement and Ads API use different timing semantics.",
        },
    ]


def _metric_row(metric: str, value: Any, currency: str | None, notes: str) -> dict[str, Any]:
    return {
        "metric": bilingual_metric_label(metric),
        "value": _xlsx_value(value),
        "currency": currency,
        "notes": notes,
    }


def _metadata_rows(result: MonthlyFinancialCloseResult) -> list[dict[str, Any]]:
    rows = [
        {"key": "report_type", "value": REPORT_TYPE},
        {"key": "version", "value": REPORT_VERSION},
        {"key": "marketplace_id", "value": result.marketplace_id},
        {"key": "profile_id", "value": result.profile_id or ""},
        {"key": "month", "value": result.month},
        {"key": "generated_at_utc", "value": result.generated_at_utc.isoformat()},
    ]
    for key, value in sorted(result.raw_metadata.items()):
        rows.append({"key": key, "value": _json_value(value)})
    return rows


def _flatten_sku_row(row: MonthlySkuProfitRow) -> dict[str, Any]:
    payload = row.to_dict()
    payload["notes"] = "; ".join(row.notes)
    payload["scope_note"] = SKU_PROFIT_SCOPE_NOTE
    return payload


def _write_rows_sheet(
    workbook: Workbook,
    sheet_name: str,
    rows: Sequence[Mapping[str, Any]],
) -> None:
    sheet = workbook.create_sheet(sheet_name)
    if rows:
        headers = list(rows[0].keys())
    else:
        headers = ["message"]
        rows = [{"message": "No rows / 无数据"}]
    sheet.append([xlsx_header_label(header) for header in headers])
    for row in rows:
        sheet.append([_xlsx_value(row.get(header)) for header in headers])
    _format_sheet(sheet)


def _format_executive_metric_sheet(sheet: Any, *, summary: bool) -> None:
    profit_fill = PatternFill("solid", fgColor="E2F0D9")
    landed_fill = PatternFill("solid", fgColor="D9EAF7")
    close_fill = PatternFill("solid", fgColor="FFF2CC")
    quality_fill = PatternFill("solid", fgColor="F2F2F2")

    for row in range(2, sheet.max_row + 1):
        metric = str(sheet.cell(row=row, column=1).value or "")
        fill = None
        if metric.startswith("Management Operating Profit") or metric.startswith(
            "Management Operating Margin"
        ):
            fill = profit_fill
        elif metric.startswith("Total Landed COGS") or metric.startswith(
            "Contribution After Ads Before Landed COGS"
        ):
            fill = landed_fill
        elif metric.startswith("Settlement Close Profit"):
            fill = close_fill
        elif summary and (
            metric.startswith("SKU Cost Coverage")
            or metric.startswith("First-Mile Cost Included")
            or metric.startswith("Unknown Settlement Classification")
            or metric.startswith("Bank Payout Reconciliation")
        ):
            fill = quality_fill
        if fill is None:
            continue
        for cell in sheet[row]:
            cell.fill = fill
        sheet.cell(row=row, column=1).font = Font(bold=True)
        sheet.cell(row=row, column=2).font = Font(bold=True)


def _format_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            header = sheet.cell(row=1, column=cell.column).value
            if isinstance(cell.value, int | float) and header:
                header_text = str(header)
                if "share" in header_text or "margin" in header_text or "pct" in header_text:
                    cell.number_format = "0.00%"
                elif "amount" in header_text or "cost" in header_text or "cogs" in header_text:
                    cell.number_format = "#,##0.00"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        header = str(column_cells[0].value or "")
        max_len = min(max(len(str(cell.value or "")) for cell in column_cells), 60)
        width = max(12, min(max_len + 2, 42))
        if header in {"message", "notes"}:
            width = 60
        if header in {"seller_sku", "metric", "check_name", "warning_code"}:
            width = max(width, 24)
        sheet.column_dimensions[column_cells[0].column_letter].width = width


def _build_natural_month_financial_summary(
    *,
    finance_rows: Sequence[Mapping[str, Any]],
    reimbursement_rows: Sequence[Mapping[str, Any]],
    cost_index: Mapping[tuple[str, str], Sequence[SkuCostRecord]],
    cost_identity_index: Mapping[tuple[str, str], Sequence[str]],
    marketplace_id: str,
    ads_api_report_date_spend: Decimal,
    fallback_currency: str | None,
) -> NaturalMonthFinancialSummary | None:
    if not finance_rows:
        return None

    operating_net = ZERO
    order_total = ZERO
    refund_total = ZERO
    liquidation_total = ZERO
    service_fee_total = ZERO
    subscription_fee = ZERO
    coupon_fee = ZERO
    deal_fee = ZERO
    storage_fee = ZERO
    customer_return_fee = ZERO
    other_service_fee = ZERO
    reimbursement_total = ZERO
    warehouse_lost_reimbursement_amount = ZERO
    adjustment_total = ZERO
    finances_ads_reference = ZERO
    transfer_reference = ZERO
    product_sales_amount = ZERO
    review_required_count = 0
    review_required_amount = ZERO
    timezone_name: str | None = None
    currency = fallback_currency
    events_by_sku: dict[str, list[dict[str, Any]]] = {}
    product_sales_units = 0
    liquidation_units = 0

    for row in finance_rows:
        amount = _money(_to_decimal(row.get("amount")))
        transaction_type = str(row.get("transaction_type") or "")
        management_include = _truthy(row.get("management_include"))
        replace_ads = _truthy(row.get("management_replace_with_ads_api"))
        if _truthy(row.get("review_required")):
            review_required_count += 1
            review_required_amount += amount
        timezone_name = timezone_name or _empty_to_none(row.get("marketplace_timezone"))
        currency = currency or _empty_to_none(row.get("currency"))

        if management_include:
            operating_net += amount
        if replace_ads:
            finances_ads_reference += amount
        if str(row.get("management_role") or "") == "cash_transfer_reference":
            transfer_reference += amount

        if management_include and transaction_type == "Shipment":
            order_total += amount
            product_sales_amount += _to_decimal(row.get("product_sales_amount"))
        elif management_include and transaction_type == "Refund":
            refund_total += amount
        elif management_include and transaction_type == "RemovalShipment":
            liquidation_total += amount
        elif management_include and transaction_type == "ServiceFee":
            service_fee_total += amount
            subscription_fee += _to_decimal(row.get("subscription_fee"))
            coupon_fee += _to_decimal(row.get("coupon_fee"))
            deal_fee += _to_decimal(row.get("deal_fee"))
            storage_fee += _to_decimal(row.get("storage_fee"))
            customer_return_fee += _to_decimal(row.get("customer_return_fee"))
            other_service_fee += _to_decimal(row.get("other_service_fee"))
        elif management_include and transaction_type == "FBAInventoryReimbursement":
            reimbursement_total += amount
            if _is_warehouse_lost_reason(row.get("description")):
                warehouse_lost_reimbursement_amount += amount
        elif management_include and transaction_type == "MiscellaneousLedgerAdjustment":
            adjustment_total += amount

        cogs_unit_include = management_include or (
            str(row.get("management_role") or "") == ZERO_VALUE_UNIT_COGS_ROLE
        )
        if not cogs_unit_include or transaction_type not in {"Shipment", "RemovalShipment"}:
            continue
        raw_events = row.get("unit_events_json")
        if isinstance(raw_events, str):
            try:
                raw_events = json.loads(raw_events)
            except json.JSONDecodeError:
                raw_events = []
        if not isinstance(raw_events, list):
            continue
        for event in raw_events:
            if not isinstance(event, Mapping):
                continue
            seller_sku = str(event.get("seller_sku") or "").strip()
            quantity = _optional_int(event.get("quantity")) or 0
            if not seller_sku or quantity <= 0:
                continue
            raw_event_date = event.get("posted_date")
            if isinstance(raw_event_date, date):
                event_date = raw_event_date
            else:
                try:
                    event_date = date.fromisoformat(str(raw_event_date)) if raw_event_date else None
                except ValueError:
                    event_date = None
            normalized = {"posted_date": event_date, "quantity": quantity}
            events_by_sku.setdefault(seller_sku, []).append(normalized)
            if transaction_type == "Shipment":
                product_sales_units += quantity
            else:
                liquidation_units += quantity

    product_cogs = ZERO
    first_mile_cogs = ZERO
    packaging_cogs = ZERO
    other_unit_cogs = ZERO
    landed_cogs = ZERO
    costed_units = 0
    missing_cost_skus: set[str] = set()
    cost_identity_resolutions: set[str] = set()
    for seller_sku, unit_events in sorted(events_by_sku.items()):
        resolved_sku, resolution_note = _resolve_cost_identity(
            cost_index=cost_index,
            cost_identity_index=cost_identity_index,
            marketplace_id=marketplace_id,
            source_sku=seller_sku,
        )
        cogs = _calculate_sku_cogs(
            cost_index=cost_index,
            marketplace_id=marketplace_id,
            seller_sku=resolved_sku or seller_sku,
            unit_events=unit_events,
            settlement_currency=currency,
        )
        if cogs.status != "ok":
            missing_cost_skus.add(seller_sku)
        elif resolution_note:
            cost_identity_resolutions.add(resolution_note)
        product_cogs += cogs.product_cost_cogs
        first_mile_cogs += cogs.first_mile_cogs
        packaging_cogs += cogs.packaging_cogs
        other_unit_cogs += cogs.other_unit_cogs
        landed_cogs += cogs.total_cogs
        costed_units += cogs.costed_units

    product_cogs = _money(product_cogs)
    first_mile_cogs = _money(first_mile_cogs)
    packaging_cogs = _money(packaging_cogs)
    other_unit_cogs = _money(other_unit_cogs)
    landed_cogs = _money(landed_cogs)
    operating_net = _money(operating_net)

    inventory_loss = _calculate_warehouse_lost_inventory_writeoff(
        finance_reimbursement_amount=_money(warehouse_lost_reimbursement_amount),
        reimbursement_rows=reimbursement_rows,
        cost_index=cost_index,
        cost_identity_index=cost_identity_index,
        marketplace_id=marketplace_id,
        settlement_currency=currency,
    )
    management_profit = _money(
        operating_net
        - ads_api_report_date_spend
        - landed_cogs
        - inventory_loss.landed_cost
    )
    expected_cost_units = product_sales_units + liquidation_units
    source_status = "ok"
    if (
        review_required_count
        or missing_cost_skus
        or costed_units != expected_cost_units
        or inventory_loss.status == "needs_review"
    ):
        source_status = "needs_review"

    return NaturalMonthFinancialSummary(
        source_status=source_status,
        marketplace_timezone=timezone_name,
        ledger_row_count=len(finance_rows),
        review_required_count=review_required_count,
        review_required_amount=_money(review_required_amount),
        product_sales_amount=_money(product_sales_amount),
        order_total=_money(order_total),
        refund_total=_money(refund_total),
        liquidation_total=_money(liquidation_total),
        service_fee_total=_money(service_fee_total),
        subscription_fee=_money(subscription_fee),
        coupon_fee=_money(coupon_fee),
        deal_fee=_money(deal_fee),
        storage_fee=_money(storage_fee),
        customer_return_fee=_money(customer_return_fee),
        other_service_fee=_money(other_service_fee),
        reimbursement_total=_money(reimbursement_total),
        warehouse_lost_reimbursement_amount=_money(warehouse_lost_reimbursement_amount),
        warehouse_lost_reimbursement_report_amount=inventory_loss.reimbursement_report_amount,
        adjustment_total=_money(adjustment_total),
        finances_ads_charge_reference=_money(finances_ads_reference),
        transfer_reference=_money(transfer_reference),
        operating_net_before_ads_replacement=operating_net,
        product_sales_units=product_sales_units,
        liquidation_units=liquidation_units,
        costed_units=costed_units,
        missing_cost_skus=tuple(sorted(missing_cost_skus)),
        cost_identity_resolutions=tuple(sorted(cost_identity_resolutions)),
        product_cost_cogs=product_cogs,
        first_mile_cogs=first_mile_cogs,
        packaging_cogs=packaging_cogs,
        other_unit_cogs=other_unit_cogs,
        landed_cogs=landed_cogs,
        inventory_loss_status=inventory_loss.status,
        inventory_loss_units=inventory_loss.inventory_loss_units,
        inventory_loss_costed_units=inventory_loss.costed_units,
        inventory_loss_missing_cost_skus=inventory_loss.missing_cost_skus,
        inventory_loss_cost_identity_resolutions=inventory_loss.cost_identity_resolutions,
        inventory_loss_product_cost=inventory_loss.product_cost,
        inventory_loss_first_mile_cost=inventory_loss.first_mile_cost,
        inventory_loss_packaging_cost=inventory_loss.packaging_cost,
        inventory_loss_other_unit_cost=inventory_loss.other_unit_cost,
        inventory_loss_landed_cost=inventory_loss.landed_cost,
        inventory_loss_details=inventory_loss.details,
        inventory_loss_notes=inventory_loss.notes,
        ads_api_report_date_spend=ads_api_report_date_spend,
        management_operating_profit=management_profit,
        management_operating_margin=_safe_ratio(management_profit, _money(product_sales_amount)),
    )


def _calculate_warehouse_lost_inventory_writeoff(
    *,
    finance_reimbursement_amount: Decimal,
    reimbursement_rows: Sequence[Mapping[str, Any]],
    cost_index: Mapping[tuple[str, str], Sequence[SkuCostRecord]],
    cost_identity_index: Mapping[tuple[str, str], Sequence[str]],
    marketplace_id: str,
    settlement_currency: str | None,
) -> InventoryLossCalculation:
    """Cost verified warehouse-lost inventory without guessing from quantity fields.

    Monetary recognition remains anchored to the Finances natural-month ledger. The
    FBA Reimbursements report is used only as a supporting identity/quantity source.
    Auto write-off is allowed when:

    - the Finances month contains a positive warehouse-lost reimbursement;
    - same-month FBA reimbursement detail contains warehouse-lost rows;
    - detail cash reimbursement amount reconciles to Finances within one cent;
    - each cash-reimbursed unit has an identifiable SKU/FNSKU and effective landed cost.

    Any ambiguity fails closed as ``needs_review`` and does not silently invent COGS.
    """

    finance_amount = _money(finance_reimbursement_amount)
    if finance_amount == ZERO:
        return InventoryLossCalculation(
            status="not_applicable",
            finance_reimbursement_amount=ZERO,
            reimbursement_report_amount=ZERO,
            inventory_loss_units=0,
            costed_units=0,
            missing_cost_skus=(),
            cost_identity_resolutions=(),
            product_cost=ZERO,
            first_mile_cost=ZERO,
            packaging_cost=ZERO,
            other_unit_cost=ZERO,
            landed_cost=ZERO,
            notes=(),
        )

    notes: list[str] = []
    if finance_amount < ZERO:
        return InventoryLossCalculation(
            status="needs_review",
            finance_reimbursement_amount=finance_amount,
            reimbursement_report_amount=ZERO,
            inventory_loss_units=0,
            costed_units=0,
            missing_cost_skus=(),
            cost_identity_resolutions=(),
            product_cost=ZERO,
            first_mile_cost=ZERO,
            packaging_cost=ZERO,
            other_unit_cost=ZERO,
            landed_cost=ZERO,
            notes=("warehouse-lost Finances reimbursement is negative; manual review required",),
        )

    loss_rows = [row for row in reimbursement_rows if _is_warehouse_lost_reason(row.get("reason"))]
    if not loss_rows:
        return InventoryLossCalculation(
            status="needs_review",
            finance_reimbursement_amount=finance_amount,
            reimbursement_report_amount=ZERO,
            inventory_loss_units=0,
            costed_units=0,
            missing_cost_skus=(),
            cost_identity_resolutions=(),
            product_cost=ZERO,
            first_mile_cost=ZERO,
            packaging_cost=ZERO,
            other_unit_cost=ZERO,
            landed_cost=ZERO,
            notes=(
                "Finances contains warehouse-lost reimbursement but FBA Reimbursements detail "
                "has no same-month warehouse-lost row",
            ),
        )

    detail_currencies = {
        str(row.get("currency") or "").strip().upper()
        for row in loss_rows
        if str(row.get("currency") or "").strip()
    }
    expected_currency = str(settlement_currency or "").strip().upper()
    if len(detail_currencies) > 1 or (
        expected_currency and detail_currencies and detail_currencies != {expected_currency}
    ):
        notes.append(
            "warehouse-lost reimbursement currency mismatch: "
            f"expected={expected_currency or 'unknown'} report={sorted(detail_currencies)}"
        )

    report_amount = _money(sum((_to_decimal(row.get("amount_total")) for row in loss_rows), ZERO))
    if abs(report_amount - finance_amount) > MONEY_QUANT:
        notes.append(
            "warehouse-lost reimbursement amount mismatch: "
            f"finances={finance_amount} report={report_amount}"
        )

    inventory_loss_units = 0
    invalid_detail = False
    details: list[dict[str, Any]] = []
    product_cost = ZERO
    first_mile_cost = ZERO
    packaging_cost = ZERO
    other_unit_cost = ZERO
    landed_cost = ZERO
    costed_units = 0
    missing_cost_skus: set[str] = set()
    resolutions: set[str] = set()

    for row in loss_rows:
        amount_total = _to_decimal(row.get("amount_total"))
        cash_units = _optional_int(row.get("quantity_reimbursed_cash")) or 0
        inventory_units = _optional_int(row.get("quantity_reimbursed_inventory")) or 0
        total_units = _optional_int(row.get("quantity_reimbursed_total")) or 0

        # Cash reimbursement removes a lost inventory asset. Inventory replacement
        # does not create the same P&L write-off because Amazon has replaced the unit.
        units = cash_units
        if units <= 0 and amount_total != ZERO and inventory_units <= 0:
            units = total_units
        if units <= 0:
            continue

        source_sku = str(row.get("seller_sku") or row.get("fnsku") or "").strip()
        detail: dict[str, Any] = {
            "reimbursement_id": row.get("reimbursement_id"),
            "reason": row.get("reason"),
            "approval_date": str(row.get("approval_date") or ""),
            "seller_sku": row.get("seller_sku"),
            "fnsku": row.get("fnsku"),
            "asin": row.get("asin"),
            "quantity": units,
            "reimbursement_amount": _money(amount_total),
            "resolved_cost_sku": None,
            "landed_cost_writeoff": ZERO,
            "status": "needs_review",
        }
        if not source_sku:
            invalid_detail = True
            notes.append(
                f"warehouse-lost reimbursement_id={row.get('reimbursement_id')} missing SKU/FNSKU"
            )
            details.append(detail)
            continue

        event_date = _coerce_date(row.get("approval_date"))
        if event_date is None:
            invalid_detail = True
            notes.append(
                f"warehouse-lost reimbursement_id={row.get('reimbursement_id')} missing approval date"
            )
            details.append(detail)
            continue
        inventory_loss_units += units
        resolved_sku, resolution_note = _resolve_cost_identity(
            cost_index=cost_index,
            cost_identity_index=cost_identity_index,
            marketplace_id=marketplace_id,
            source_sku=source_sku,
        )
        cogs = _calculate_sku_cogs(
            cost_index=cost_index,
            marketplace_id=marketplace_id,
            seller_sku=resolved_sku or source_sku,
            unit_events=({"posted_date": event_date, "quantity": units},),
            settlement_currency=settlement_currency,
        )
        detail["resolved_cost_sku"] = resolved_sku or source_sku
        detail["landed_cost_writeoff"] = cogs.total_cogs
        detail["status"] = cogs.status
        details.append(detail)
        if cogs.status != "ok":
            missing_cost_skus.add(source_sku)
        elif resolution_note:
            resolutions.add(resolution_note)
        product_cost += cogs.product_cost_cogs
        first_mile_cost += cogs.first_mile_cogs
        packaging_cost += cogs.packaging_cogs
        other_unit_cost += cogs.other_unit_cogs
        landed_cost += cogs.total_cogs
        costed_units += cogs.costed_units

    status = "ok"
    if (
        notes
        or invalid_detail
        or inventory_loss_units <= 0
        or missing_cost_skus
        or costed_units != inventory_loss_units
        or abs(report_amount - finance_amount) > MONEY_QUANT
    ):
        status = "needs_review"

    # Fail closed: candidate per-row costs remain visible in ``details`` for diagnosis,
    # but no inventory-loss cost is applied to Management P&L / accountant reference
    # profit until the complete warehouse-loss set reconciles.
    applied_product_cost = _money(product_cost) if status == "ok" else ZERO
    applied_first_mile_cost = _money(first_mile_cost) if status == "ok" else ZERO
    applied_packaging_cost = _money(packaging_cost) if status == "ok" else ZERO
    applied_other_unit_cost = _money(other_unit_cost) if status == "ok" else ZERO
    applied_landed_cost = _money(landed_cost) if status == "ok" else ZERO

    return InventoryLossCalculation(
        status=status,
        finance_reimbursement_amount=finance_amount,
        reimbursement_report_amount=report_amount,
        inventory_loss_units=inventory_loss_units,
        costed_units=costed_units,
        missing_cost_skus=tuple(sorted(missing_cost_skus)),
        cost_identity_resolutions=tuple(sorted(resolutions)),
        product_cost=applied_product_cost,
        first_mile_cost=applied_first_mile_cost,
        packaging_cost=applied_packaging_cost,
        other_unit_cost=applied_other_unit_cost,
        landed_cost=applied_landed_cost,
        details=tuple(details),
        notes=tuple(notes),
    )


def _is_warehouse_lost_reason(value: Any) -> bool:
    text = str(value or "").strip().upper()
    if not text:
        return False
    normalized = "".join(character if character.isalnum() else "_" for character in text)
    tokens = {token for token in normalized.split("_") if token}
    return "WAREHOUSE" in tokens and "LOST" in tokens


def _coerce_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value in (None, ""):
        return None
    text = str(value).strip()
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _truthy(value: Any) -> bool:
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y"}
    return bool(value)


def _build_sku_accumulator(
    settlement_lines: Sequence[SettlementProfitLine],
) -> dict[str, dict[str, Any]]:
    accumulator: dict[str, dict[str, Any]] = {}
    seen_unit_keys: set[tuple[Any, ...]] = set()
    for line in settlement_lines:
        if not line.seller_sku:
            continue
        data = accumulator.setdefault(
            line.seller_sku,
            {
                "units": 0,
                "product_sales_amount": ZERO,
                "settlement_net_amount": ZERO,
                "unit_events": [],
            },
        )
        data["settlement_net_amount"] += line.amount
        if line.amount_category in PRODUCT_SALES_CATEGORIES:
            data["product_sales_amount"] += line.amount
            unit_key = _unit_dedupe_key(line)
            if unit_key not in seen_unit_keys:
                seen_unit_keys.add(unit_key)
                quantity = max(0, line.quantity_purchased or 0)
                data["units"] += quantity
                if quantity > 0:
                    data["unit_events"].append(
                        {"posted_date": line.posted_date, "quantity": quantity}
                    )
    return accumulator


def _unit_dedupe_key(line: SettlementProfitLine) -> tuple[Any, ...]:
    if line.order_item_code:
        return (line.settlement_id, line.order_id, line.order_item_code, line.seller_sku)
    return (line.settlement_id, line.order_id, line.seller_sku, line.id)


def _calculate_sku_cogs(
    *,
    cost_index: Mapping[tuple[str, str], Sequence[SkuCostRecord]],
    marketplace_id: str,
    seller_sku: str,
    unit_events: Sequence[Mapping[str, Any]],
    settlement_currency: str | None,
) -> SkuCogsCalculation:
    candidates = cost_index.get((marketplace_id, seller_sku), ())
    if not candidates:
        return SkuCogsCalculation(
            total_cogs=ZERO,
            product_cost_cogs=ZERO,
            first_mile_cogs=ZERO,
            packaging_cogs=ZERO,
            other_unit_cogs=ZERO,
            cost_currency=None,
            unit_product_cost=None,
            unit_first_mile_cost=None,
            unit_packaging_cost=None,
            unit_other_cost=None,
            unit_standard_cost=None,
            status="missing_cost",
            notes=("no amazon_sku_cost row",),
            costed_units=0,
        )
    if not unit_events:
        return SkuCogsCalculation(
            total_cogs=ZERO,
            product_cost_cogs=ZERO,
            first_mile_cogs=ZERO,
            packaging_cogs=ZERO,
            other_unit_cogs=ZERO,
            cost_currency=candidates[0].currency,
            unit_product_cost=None,
            unit_first_mile_cost=None,
            unit_packaging_cost=None,
            unit_other_cost=None,
            unit_standard_cost=None,
            status="ok",
            notes=("no product-sales units",),
            costed_units=0,
        )

    product_cogs = ZERO
    first_mile_cogs = ZERO
    packaging_cogs = ZERO
    other_cogs = ZERO
    costed_units = 0
    matched_costs: list[SkuCostRecord] = []
    notes: list[str] = []
    status = "ok"
    cost_currency: str | None = None
    for event in unit_events:
        quantity = _optional_int(event.get("quantity")) or 0
        posted_date = event.get("posted_date")
        if posted_date is None:
            match = candidates[0]
            notes.append("missing posted_date; used latest cost row")
        else:
            match = next((cost for cost in candidates if cost.is_effective_on(posted_date)), None)
        if match is None:
            status = "missing_cost"
            notes.append(f"no effective cost for posted date {posted_date}")
            continue
        cost_currency = cost_currency or match.currency
        matched_costs.append(match)
        quantity_decimal = Decimal(quantity)
        product_cogs += match.product_cost * quantity_decimal
        first_mile_cogs += match.first_mile_cost * quantity_decimal
        packaging_cogs += match.packaging_cost * quantity_decimal
        other_cogs += match.other_unit_cost * quantity_decimal
        costed_units += quantity
        if settlement_currency and match.currency and match.currency != settlement_currency:
            status = "currency_mismatch"
            notes.append(
                f"cost currency {match.currency} != settlement currency {settlement_currency}"
            )

    total_cogs = _money(product_cogs + first_mile_cogs + packaging_cogs + other_cogs)
    product_cogs = _money(product_cogs)
    first_mile_cogs = _money(first_mile_cogs)
    packaging_cogs = _money(packaging_cogs)
    other_cogs = _money(other_cogs)
    component_total = product_cogs + first_mile_cogs + packaging_cogs + other_cogs
    # Keep the established total-cost rounding as source of truth while forcing the
    # displayed component breakdown to reconcile exactly to that total. Any cent-level
    # residual from independently rounded components is absorbed into product cost.
    product_cogs = _money(product_cogs + (total_cogs - component_total))

    if not matched_costs:
        return SkuCogsCalculation(
            total_cogs=ZERO,
            product_cost_cogs=ZERO,
            first_mile_cogs=ZERO,
            packaging_cogs=ZERO,
            other_unit_cogs=ZERO,
            cost_currency=cost_currency,
            unit_product_cost=None,
            unit_first_mile_cost=None,
            unit_packaging_cost=None,
            unit_other_cost=None,
            unit_standard_cost=None,
            status="missing_cost",
            notes=tuple(notes),
            costed_units=costed_units,
        )
    unique_cost_keys = {
        (
            cost.effective_from,
            cost.effective_to,
            cost.currency,
            cost.product_cost,
            cost.first_mile_cost,
            cost.packaging_cost,
            cost.other_unit_cost,
        )
        for cost in matched_costs
    }
    if len(unique_cost_keys) > 1:
        notes.append("multiple cost rows matched; weighted average unit costs shown")
    if status == "missing_cost":
        unit_values = (None, None, None, None, None)
    else:
        denominator = Decimal(costed_units) if costed_units else None
        unit_values = (
            _money(product_cogs / denominator) if denominator else None,
            _money(first_mile_cogs / denominator) if denominator else None,
            _money(packaging_cogs / denominator) if denominator else None,
            _money(other_cogs / denominator) if denominator else None,
            _money(total_cogs / denominator) if denominator else None,
        )
    return SkuCogsCalculation(
        total_cogs=total_cogs,
        product_cost_cogs=product_cogs,
        first_mile_cogs=first_mile_cogs,
        packaging_cogs=packaging_cogs,
        other_unit_cogs=other_cogs,
        cost_currency=cost_currency,
        unit_product_cost=unit_values[0],
        unit_first_mile_cost=unit_values[1],
        unit_packaging_cost=unit_values[2],
        unit_other_cost=unit_values[3],
        unit_standard_cost=unit_values[4],
        status=status,
        notes=tuple(notes),
        costed_units=costed_units,
    )

def _build_cost_identity_index(
    rows: Iterable[Mapping[str, Any]],
) -> dict[tuple[str, str], tuple[str, ...]]:
    candidates: dict[tuple[str, str], set[str]] = defaultdict(set)
    for row in rows:
        marketplace_id = str(row.get("marketplace_id") or "").strip()
        fnsku = str(row.get("fnsku") or "").strip()
        seller_sku = str(row.get("seller_sku") or "").strip()
        if not marketplace_id or not fnsku or not seller_sku:
            continue
        candidates[(marketplace_id, fnsku)].add(seller_sku)
    return {key: tuple(sorted(values)) for key, values in candidates.items()}


def _resolve_cost_identity(
    *,
    cost_index: Mapping[tuple[str, str], Sequence[SkuCostRecord]],
    cost_identity_index: Mapping[tuple[str, str], Sequence[str]],
    marketplace_id: str,
    source_sku: str,
) -> tuple[str | None, str | None]:
    # Direct Seller SKU cost always wins. FNSKU fallback is only used when the
    # transaction SKU itself has no cost row. This prevents an inventory alias
    # from silently overriding an explicitly maintained cost identity.
    if cost_index.get((marketplace_id, source_sku)):
        return source_sku, None

    aliases = tuple(cost_identity_index.get((marketplace_id, source_sku), ()))
    if len(aliases) != 1:
        return None, None

    resolved_sku = aliases[0]
    if not cost_index.get((marketplace_id, resolved_sku)):
        return None, None

    return resolved_sku, f"{source_sku}->{resolved_sku}"


def _build_cost_index(costs: Iterable[SkuCostRecord]) -> dict[tuple[str, str], list[SkuCostRecord]]:
    index: dict[tuple[str, str], list[SkuCostRecord]] = defaultdict(list)
    for cost in costs:
        index[(cost.marketplace_id, cost.seller_sku)].append(cost)
    for rows in index.values():
        rows.sort(key=lambda cost: cost.effective_from, reverse=True)
    return dict(index)


def _sum_by_bucket(
    settlement_lines: Iterable[SettlementProfitLine],
) -> tuple[dict[str, Decimal], dict[str, int]]:
    totals: dict[str, Decimal] = defaultdict(Decimal)
    counts: dict[str, int] = defaultdict(int)
    for line in settlement_lines:
        key = str(line.profit_bucket or "unknown")
        totals[key] += line.amount
        counts[key] += 1
    return dict(totals), dict(counts)


def _sum_by_category(
    settlement_lines: Iterable[SettlementProfitLine],
) -> tuple[dict[str, Decimal], dict[str, int], dict[str, set[str]]]:
    totals: dict[str, Decimal] = defaultdict(Decimal)
    counts: dict[str, int] = defaultdict(int)
    bucket_index: dict[str, set[str]] = defaultdict(set)
    for line in settlement_lines:
        key = str(line.amount_category or "unknown")
        totals[key] += line.amount
        counts[key] += 1
        bucket_index[key].add(str(line.profit_bucket or "unknown"))
    return dict(totals), dict(counts), dict(bucket_index)


def _build_bucket_rows(
    *,
    bucket_totals: Mapping[str, Decimal],
    bucket_counts: Mapping[str, int],
    product_sales_amount: Decimal,
    settlement_net_amount: Decimal,
) -> list[SettlementBucketBreakdownRow]:
    rows = []
    for bucket, amount in sorted(bucket_totals.items()):
        notes = "needs review" if bucket in REVIEW_BUCKETS and amount != ZERO else ""
        rows.append(
            SettlementBucketBreakdownRow(
                profit_bucket=bucket,
                amount=_money(amount),
                row_count=bucket_counts.get(bucket, 0),
                share_of_product_sales=_safe_ratio(amount, product_sales_amount),
                share_of_settlement_net=_safe_ratio(amount, settlement_net_amount),
                notes=notes,
            )
        )
    return rows


def _build_category_rows(
    *,
    category_totals: Mapping[str, Decimal],
    category_counts: Mapping[str, int],
    category_bucket_index: Mapping[str, set[str]],
) -> list[AmountCategoryBreakdownRow]:
    rows = []
    for category, amount in sorted(category_totals.items()):
        buckets = sorted(category_bucket_index.get(category, set()))
        bucket = buckets[0] if len(buckets) == 1 else "mixed"
        notes = "needs review" if category in REVIEW_CATEGORIES and amount != ZERO else ""
        rows.append(
            AmountCategoryBreakdownRow(
                amount_category=category,
                profit_bucket=bucket,
                amount=_money(amount),
                row_count=category_counts.get(category, 0),
                notes=notes,
            )
        )
    return rows


def _build_operational_context(
    *,
    orders_summary: Mapping[str, Any],
    ads_summary: Mapping[str, Any],
    sales_traffic_summary: Mapping[str, Any],
    coupon_summary: Mapping[str, Any],
    promotion_summary: Mapping[str, Any],
    fba_reimbursement_summary: Mapping[str, Any],
    currency: str | None,
) -> list[OperationalMetric]:
    order_currency = _empty_to_none(orders_summary.get("currency")) or currency
    sales_currency = (
        _empty_to_none(sales_traffic_summary.get("ordered_product_sales_currency")) or currency
    )
    coupon_currency = _empty_to_none(coupon_summary.get("coupon_currency")) or currency
    promotion_currency = _empty_to_none(promotion_summary.get("promotion_currency")) or currency
    reimbursement_currency = (
        _empty_to_none(fba_reimbursement_summary.get("reimbursement_currency")) or currency
    )
    metrics: list[OperationalMetric] = []

    def add(
        group: str,
        name: str,
        value: Decimal | int | str | None,
        metric_currency: str | None,
        source: str,
        notes: str = "",
    ) -> None:
        metrics.append(OperationalMetric(group, name, value, metric_currency, source, notes))

    add(
        "Orders",
        "order_count",
        _int_metric(orders_summary, "order_count"),
        None,
        "amazon_order_item",
    )
    add(
        "Orders",
        "order_item_rows",
        _int_metric(orders_summary, "order_item_rows"),
        None,
        "amazon_order_item",
    )
    add(
        "Orders",
        "ordered_units",
        _int_metric(orders_summary, "ordered_units"),
        None,
        "amazon_order_item",
    )
    add(
        "Orders",
        "ordered_item_sales_amount",
        _to_decimal(orders_summary.get("ordered_item_sales_amount")),
        order_currency,
        "amazon_order_item",
    )
    add(
        "Orders",
        "item_promotion_discount_amount",
        _to_decimal(orders_summary.get("item_promotion_discount_amount")),
        order_currency,
        "amazon_order_item",
    )
    add(
        "Orders",
        "ship_promotion_discount_amount",
        _to_decimal(orders_summary.get("ship_promotion_discount_amount")),
        order_currency,
        "amazon_order_item",
    )
    add(
        "Orders",
        "order_exception_count",
        _int_metric(orders_summary, "order_exception_count"),
        None,
        "amazon_order_item",
        "Cancelled/Canceled rows only in v1.",
    )

    add(
        "Sales & Traffic",
        "units_ordered",
        _int_metric(sales_traffic_summary, "units_ordered"),
        None,
        "amazon_sales_traffic_daily",
    )
    add(
        "Sales & Traffic",
        "ordered_product_sales_amount",
        _to_decimal(sales_traffic_summary.get("ordered_product_sales_amount")),
        sales_currency,
        "amazon_sales_traffic_daily",
    )
    add(
        "Sales & Traffic",
        "sessions",
        _int_metric(sales_traffic_summary, "sessions"),
        None,
        "amazon_sales_traffic_daily",
    )
    add(
        "Sales & Traffic",
        "page_views",
        _int_metric(sales_traffic_summary, "page_views"),
        None,
        "amazon_sales_traffic_daily",
    )
    add(
        "Sales & Traffic",
        "total_order_items",
        _int_metric(sales_traffic_summary, "total_order_items"),
        None,
        "amazon_sales_traffic_daily",
    )
    add(
        "Sales & Traffic",
        "units_refunded",
        _int_metric(sales_traffic_summary, "units_refunded"),
        None,
        "amazon_sales_traffic_daily",
    )
    add(
        "Sales & Traffic",
        "unit_session_rate",
        _safe_ratio(
            _to_decimal(sales_traffic_summary.get("units_ordered")),
            _to_decimal(sales_traffic_summary.get("sessions")),
        ),
        None,
        "derived",
        "units_ordered / sessions",
    )
    add(
        "Sales & Traffic",
        "sales_per_session",
        _safe_ratio(
            _to_decimal(sales_traffic_summary.get("ordered_product_sales_amount")),
            _to_decimal(sales_traffic_summary.get("sessions")),
        ),
        sales_currency,
        "derived",
        "ordered_product_sales_amount / sessions",
    )

    add(
        "Ads API",
        "ads_cost",
        _to_decimal(ads_summary.get("ads_cost")),
        currency,
        "amazon_ads_sp_campaign_daily",
        "Report-date Ads API context; not financial source of truth.",
    )
    add(
        "Ads API",
        "ads_sales_7d",
        _to_decimal(ads_summary.get("ads_sales_7d")),
        currency,
        "amazon_ads_sp_campaign_daily",
    )
    for name in ("ads_clicks", "ads_impressions", "ads_purchases_7d", "ads_row_count"):
        add("Ads API", name, _int_metric(ads_summary, name), None, "amazon_ads_sp_campaign_daily")

    add(
        "Coupon",
        "coupon_count",
        _int_metric(coupon_summary, "coupon_count"),
        None,
        "amazon_coupon_performance",
        "Overlapping campaign period.",
    )
    for name in ("coupon_clips", "coupon_redemptions"):
        add("Coupon", name, _int_metric(coupon_summary, name), None, "amazon_coupon_performance")
    for name in ("coupon_total_discount", "coupon_budget_spent", "coupon_sales"):
        add(
            "Coupon",
            name,
            _to_decimal(coupon_summary.get(name)),
            coupon_currency,
            "amazon_coupon_performance",
        )

    add(
        "Promotion",
        "promotion_count",
        _int_metric(promotion_summary, "promotion_count"),
        None,
        "amazon_promotion_performance",
        "Overlapping campaign period.",
    )
    for name in ("promotion_glance_views", "promotion_units_sold"):
        add(
            "Promotion",
            name,
            _int_metric(promotion_summary, name),
            None,
            "amazon_promotion_performance",
        )
    add(
        "Promotion",
        "promotion_revenue",
        _to_decimal(promotion_summary.get("promotion_revenue")),
        promotion_currency,
        "amazon_promotion_performance",
    )

    add(
        "FBA Reimbursements",
        "reimbursement_count",
        _int_metric(fba_reimbursement_summary, "reimbursement_count"),
        None,
        "amazon_fba_reimbursement",
    )
    add(
        "FBA Reimbursements",
        "reimbursement_report_amount",
        _to_decimal(fba_reimbursement_summary.get("reimbursement_report_amount")),
        reimbursement_currency,
        "amazon_fba_reimbursement",
    )
    for name in ("reimbursement_quantity", "reimbursement_reason_count"):
        add(
            "FBA Reimbursements",
            name,
            _int_metric(fba_reimbursement_summary, name),
            None,
            "amazon_fba_reimbursement",
        )
    return metrics


def _build_reconciliation_checks(
    *,
    settlement_net_amount: Decimal,
    bucket_totals: Mapping[str, Decimal],
    category_totals: Mapping[str, Decimal],
    product_sales_amount: Decimal,
    product_sales_units: int,
    costed_units: int,
    missing_cost_skus: set[str],
    currency_mismatch_skus: set[str],
    financial_summary: MonthlyFinancialSummary,
    sales_traffic_summary: Mapping[str, Any],
    ads_summary: Mapping[str, Any],
    fba_reimbursement_summary: Mapping[str, Any],
) -> list[ReconciliationCheck]:
    checks = []
    bucket_sum = _money(sum(bucket_totals.values(), ZERO))
    checks.append(
        _money_check("settlement_bucket_sum_matches_net", settlement_net_amount, bucket_sum)
    )
    category_sum = _money(sum(category_totals.values(), ZERO))
    checks.append(
        _money_check("amount_category_sum_matches_net", settlement_net_amount, category_sum)
    )
    checks.append(
        ReconciliationCheck(
            check_name="sku_cost_coverage",
            status="ok" if costed_units == product_sales_units else "needs_review",
            severity="error" if costed_units != product_sales_units else "info",
            expected=str(product_sales_units),
            actual=str(costed_units),
            diff=str(product_sales_units - costed_units),
            message="Product-sales units with matched amazon_sku_cost.",
        )
    )
    checks.append(
        ReconciliationCheck(
            check_name="missing_cost_skus",
            status="ok" if not missing_cost_skus else "needs_review",
            severity="error" if missing_cost_skus else "info",
            expected="none",
            actual=", ".join(sorted(missing_cost_skus)) if missing_cost_skus else "none",
            message="Missing SKU cost must be reviewed; report does not silently assume zero cost.",
        )
    )
    checks.append(
        ReconciliationCheck(
            check_name="currency_mismatch_skus",
            status="ok" if not currency_mismatch_skus else "needs_review",
            severity="error" if currency_mismatch_skus else "info",
            expected="none",
            actual=", ".join(sorted(currency_mismatch_skus)) if currency_mismatch_skus else "none",
            message="SKU cost currency should match settlement currency in v1.",
        )
    )
    unknown_bucket_amount = _money(sum(bucket_totals.get(key, ZERO) for key in REVIEW_BUCKETS))
    checks.append(_zero_review_check("unknown_bucket_amount", unknown_bucket_amount))
    unknown_category_amount = _money(
        sum(category_totals.get(key, ZERO) for key in REVIEW_CATEGORIES)
    )
    checks.append(_zero_review_check("unclassified_amount", unknown_category_amount))
    tax_passthrough = _money(bucket_totals.get("tax_passthrough", ZERO))
    checks.append(
        ReconciliationCheck(
            check_name="tax_passthrough_net",
            status="ok" if tax_passthrough == ZERO else "warning",
            severity="warning" if tax_passthrough != ZERO else "info",
            expected="0.00",
            actual=_decimal_to_string(tax_passthrough),
            diff=_decimal_to_string(tax_passthrough),
            message="Tax passthrough should normally be reviewed but does not change core formula.",
        )
    )
    sales_traffic_sales = _to_decimal(sales_traffic_summary.get("ordered_product_sales_amount"))
    checks.append(
        _context_diff_check(
            "settlement_product_sales_vs_sales_traffic",
            expected=product_sales_amount,
            actual=sales_traffic_sales,
            message=(
                "Settlement posted-date product sales versus Sales & Traffic report-date sales. "
                "Timing differences are expected."
            ),
        )
    )
    checks.append(_ads_timing_reconciliation_check(financial_summary))
    reimbursement_report_amount = _to_decimal(
        fba_reimbursement_summary.get("reimbursement_report_amount")
    )
    checks.append(
        _context_diff_check(
            "settlement_reimbursement_vs_fba_reimbursement_report",
            expected=financial_summary.reimbursement,
            actual=reimbursement_report_amount,
            message="Settlement reimbursement versus FBA Reimbursements report context.",
        )
    )
    return checks


def _money_check(name: str, expected: Decimal, actual: Decimal) -> ReconciliationCheck:
    diff = _money(actual - expected)
    return ReconciliationCheck(
        check_name=name,
        status="ok" if diff == ZERO else "needs_review",
        severity="error" if diff != ZERO else "info",
        expected=_decimal_to_string(expected),
        actual=_decimal_to_string(actual),
        diff=_decimal_to_string(diff),
        message="Settlement aggregation self-check.",
    )


def _zero_review_check(name: str, amount: Decimal) -> ReconciliationCheck:
    return ReconciliationCheck(
        check_name=name,
        status="ok" if amount == ZERO else "needs_review",
        severity="error" if amount != ZERO else "info",
        expected="0.00",
        actual=_decimal_to_string(amount),
        diff=_decimal_to_string(amount),
        message="Non-zero unknown/unclassified amount should be reviewed before sharing.",
    )


def _ads_timing_reconciliation_check(
    financial_summary: MonthlyFinancialSummary,
) -> ReconciliationCheck:
    settlement_ads_abs = financial_summary.settlement_advertising_fee_abs
    ads_api_spend = financial_summary.ads_api_report_date_spend
    diff = financial_summary.ads_timing_difference
    diff_pct = financial_summary.ads_timing_difference_pct
    status = _ads_timing_status(
        ads_api_spend=ads_api_spend,
        settlement_ads_abs=settlement_ads_abs,
    )
    severity = "info" if status == "ok" else "warning" if status == "warning" else "error"
    return ReconciliationCheck(
        check_name="settlement_ads_fee_vs_ads_api_spend",
        status=status,
        severity=severity,
        expected=_decimal_to_string(settlement_ads_abs),
        actual=_decimal_to_string(ads_api_spend),
        diff=_decimal_to_string(diff),
        diff_pct=_optional_ratio_to_string(diff_pct),
        message=(
            "Settlement advertising fee and Ads API report-date spend use different timing. "
            "Management P&L replaces posted-date settlement ads with report-date Ads API spend."
        ),
    )


def _ads_timing_status(*, ads_api_spend: Decimal, settlement_ads_abs: Decimal) -> str:
    diff_abs = abs(_money(ads_api_spend - settlement_ads_abs))
    if ads_api_spend == ZERO and settlement_ads_abs == ZERO:
        return "ok"
    denominator = settlement_ads_abs if settlement_ads_abs != ZERO else ads_api_spend
    diff_ratio = _safe_ratio(diff_abs, denominator)
    if diff_abs <= Decimal("20.00") or (diff_ratio is not None and diff_ratio <= Decimal("0.05")):
        return "ok"
    return "warning"


def _context_diff_check(
    name: str,
    *,
    expected: Decimal,
    actual: Decimal,
    message: str,
) -> ReconciliationCheck:
    diff = _money(actual - expected)
    diff_pct = _optional_ratio_to_string(_safe_ratio(diff, expected))
    if expected == ZERO and actual == ZERO:
        status = "ok"
        severity = "info"
    else:
        status = "warning"
        severity = "warning"
    return ReconciliationCheck(
        check_name=name,
        status=status,
        severity=severity,
        expected=_decimal_to_string(expected),
        actual=_decimal_to_string(actual),
        diff=_decimal_to_string(diff),
        diff_pct=diff_pct,
        message=message,
    )


def _build_warnings(
    *,
    settlement_row_count: int,
    missing_cost_skus: set[str],
    currency_mismatch_skus: set[str],
    product_sales_units: int,
    financial_summary: MonthlyFinancialSummary,
    ads_summary: Mapping[str, Any],
    reconciliation_checks: Sequence[ReconciliationCheck],
) -> list[WarningEntry]:
    warnings = []
    if settlement_row_count == 0:
        warnings.append(
            WarningEntry(
                "no_settlement_rows",
                "error",
                "No settlement rows found for this marketplace/month.",
                related_source="amazon_settlement_transaction",
            )
        )
    if product_sales_units == 0 and settlement_row_count > 0:
        warnings.append(
            WarningEntry(
                "no_product_sales_units",
                "warning",
                "No product-sales units found; SKU COGS may be zero for this period.",
                related_source="amazon_settlement_transaction",
            )
        )
    for sku in sorted(missing_cost_skus):
        warnings.append(
            WarningEntry(
                "missing_sku_cost",
                "error",
                "Missing amazon_sku_cost row or effective cost date.",
                related_sku=sku,
                related_source="amazon_sku_cost",
            )
        )
    for sku in sorted(currency_mismatch_skus):
        warnings.append(
            WarningEntry(
                "currency_mismatch",
                "error",
                "SKU cost currency differs from settlement currency.",
                related_sku=sku,
                related_source="amazon_sku_cost",
            )
        )
    ads_row_count = _optional_int(ads_summary.get("ads_row_count")) or 0
    ads_api_spend = _to_decimal(ads_summary.get("ads_cost"))
    if (
        abs(financial_summary.advertising_cost) != ZERO
        and ads_row_count == 0
        and ads_api_spend == ZERO
    ):
        warnings.append(
            WarningEntry(
                "ads_api_context_missing",
                "warning",
                (
                    "Settlement has advertising fees but Ads API campaign daily context has "
                    "zero rows for this period/profile. This affects operational context only; "
                    "financial profit remains Settlement-led."
                ),
                related_source="amazon_ads_sp_campaign_daily",
            )
        )
    for check in reconciliation_checks:
        if check.status == "needs_review" and check.check_name not in {
            "missing_cost_skus",
            "currency_mismatch_skus",
        }:
            warnings.append(
                WarningEntry(
                    check.check_name,
                    check.severity,
                    check.message,
                    related_source="reconciliation_checks",
                )
            )
    warnings.append(
        WarningEntry(
            "settlement_led_policy",
            "info",
            SETTLEMENT_LED_POLICY_NOTE,
        )
    )
    warnings.append(
        WarningEntry(
            "management_pnl_policy",
            "info",
            MANAGEMENT_PNL_POLICY_NOTE,
        )
    )
    return warnings


def _result_status(
    *,
    settlement_row_count: int,
    reconciliation_checks: Sequence[ReconciliationCheck],
) -> str:
    if settlement_row_count == 0:
        return "no_data"
    if any(check.status == "needs_review" for check in reconciliation_checks):
        return "needs_review"
    return "ok"


def _build_raw_metadata(
    *,
    settlement_lines: Sequence[SettlementProfitLine],
    raw_settlement_rows: Sequence[Mapping[str, Any]],
    sku_costs: Sequence[SkuCostRecord],
    orders_summary: Mapping[str, Any],
    ads_summary: Mapping[str, Any],
    sales_traffic_summary: Mapping[str, Any],
    coupon_summary: Mapping[str, Any],
    promotion_summary: Mapping[str, Any],
    fba_reimbursement_summary: Mapping[str, Any],
) -> dict[str, Any]:
    settlement_ids = sorted({line.settlement_id for line in settlement_lines if line.settlement_id})
    source_report_ids = sorted(
        {
            str(row.get("source_report_id"))
            for row in raw_settlement_rows
            if row.get("source_report_id")
        }
    )
    source_raw_file_paths = sorted(
        {
            str(row.get("source_raw_file_path"))
            for row in raw_settlement_rows
            if row.get("source_raw_file_path")
        }
    )
    return {
        "settlement_id_count": len(settlement_ids),
        "settlement_ids": settlement_ids,
        "settlement_source_report_id_count": len(source_report_ids),
        "settlement_source_report_ids": source_report_ids,
        "settlement_source_raw_file_path_count": len(source_raw_file_paths),
        "settlement_source_raw_file_paths": source_raw_file_paths,
        "sku_cost_rows_loaded": len(sku_costs),
        "orders_summary_row_count": _optional_int(orders_summary.get("order_item_rows")) or 0,
        "ads_summary_row_count": _optional_int(ads_summary.get("ads_row_count")) or 0,
        "sales_traffic_summary_row_count": _optional_int(
            sales_traffic_summary.get("sales_traffic_row_count")
        )
        or 0,
        "coupon_count": _optional_int(coupon_summary.get("coupon_count")) or 0,
        "promotion_count": _optional_int(promotion_summary.get("promotion_count")) or 0,
        "reimbursement_count": _optional_int(fba_reimbursement_summary.get("reimbursement_count"))
        or 0,
    }


def _validate_period(*, start_date: date, end_date: date) -> None:
    if end_date < start_date:
        raise ValueError("end_date must be >= start_date")


def _int_metric(values: Mapping[str, Any], key: str) -> int:
    return _optional_int(values.get(key)) or 0


def _to_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return ZERO
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).replace(",", ""))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Invalid decimal value: {value!r}") from exc


def _optional_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    return int(value)


def _empty_to_none(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _first_non_empty(values: Iterable[str | None]) -> str | None:
    for value in values:
        if value:
            return value
    return None


def _money(value: Decimal) -> Decimal:
    return value.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def _ratio(value: Decimal) -> Decimal:
    return value.quantize(RATIO_QUANT, rounding=ROUND_HALF_UP)


def _safe_ratio(numerator: Decimal, denominator: Decimal) -> Decimal | None:
    if denominator == ZERO:
        return None
    return _ratio(numerator / denominator)


def _decimal_to_string(value: Decimal) -> str:
    return str(_money(value))


def _optional_decimal_to_string(value: Decimal | None) -> str | None:
    return None if value is None else _decimal_to_string(value)


def _optional_ratio_to_string(value: Decimal | None) -> str | None:
    return None if value is None else str(_ratio(value))


def _format_money(value: Decimal, currency: str | None) -> str:
    prefix = f"{currency} " if currency else ""
    return f"{prefix}{_decimal_to_string(value)}"


def _format_ratio(value: Decimal) -> str:
    return f"{(value * Decimal('100')).quantize(Decimal('0.01'))}%"


def _json_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return _decimal_to_string(value)
    if isinstance(value, datetime | date):
        return value.isoformat()
    return value


def _json_safe_mapping(values: Mapping[str, Any]) -> dict[str, Any]:
    return {key: _json_value(value) for key, value in values.items()}


def _xlsx_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, list | tuple | set):
        return json.dumps(list(value), ensure_ascii=False)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


__all__ = [
    "DEFAULT_OUTPUT_ROOT",
    "MonthlyFinancialCloseResult",
    "MonthlyFinancialCloseService",
    "MonthlyFinancialSummary",
    "MonthlySkuProfitRow",
    "OperationalMetric",
    "ReconciliationCheck",
    "SettlementBucketBreakdownRow",
    "WarningEntry",
    "build_monthly_financial_close_workbook",
    "month_to_date_range",
]
