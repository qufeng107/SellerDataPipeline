from __future__ import annotations

import json
from collections import defaultdict
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass, field
from datetime import UTC, date, datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Protocol

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from seller_data_pipeline.services.report_bilingual import (
    add_bilingual_readme_sheet,
    bilingual_metric_label,
    xlsx_header_label,
)

MONEY_QUANT = Decimal("0.01")
RATIO_QUANT = Decimal("0.0001")
ZERO = Decimal("0")
REPORT_TYPE = "weekly_business_review"
REPORT_VERSION = "v1.1-operational-contribution-labels"
DEFAULT_OUTPUT_ROOT = "runtime/analysis_reports/weekly_business_review"
WBR_SCOPE_NOTE = (
    "Weekly Business Review is an operational report using a Saturday-Friday period. "
    "Sales & Traffic, Orders, and Ads API report-date spend drive weekly business metrics; "
    "Settlement is posted-date financial context only."
)
CONTRIBUTION_SCOPE_NOTE = (
    "Sales & Traffic ordered sales - estimated COGS - Ads API report-date spend; "
    "before full Amazon platform fees, refunds, storage, account-level fees, "
    "and not final net profit."
)
SETTLEMENT_PREVIEW_NOTE = (
    "Settlement preview is posted-date financial context, not final weekly profit."
)


@dataclass(frozen=True)
class WarningEntry:
    warning_code: str
    severity: str
    message: str
    related_sku: str | None = None
    related_source: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "warning_code": self.warning_code,
            "severity": self.severity,
            "message": self.message,
            "related_sku": self.related_sku,
            "related_source": self.related_source,
        }


@dataclass(frozen=True)
class ReconciliationCheck:
    check_name: str
    status: str
    severity: str
    expected: str | None
    actual: str | None
    diff: str | None = None
    diff_pct: str | None = None
    message: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "check_name": self.check_name,
            "status": self.status,
            "severity": self.severity,
            "expected": self.expected,
            "actual": self.actual,
            "diff": self.diff,
            "diff_pct": self.diff_pct,
            "message": self.message,
        }


@dataclass(frozen=True)
class AlertAction:
    severity: str
    area: str
    metric: str
    current_value: str | None
    previous_value: str | None
    threshold: str | None
    message: str
    recommended_action: str
    related_sku: str | None = None
    related_campaign: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "severity": self.severity,
            "area": self.area,
            "metric": self.metric,
            "current_value": self.current_value,
            "previous_value": self.previous_value,
            "threshold": self.threshold,
            "message": self.message,
            "recommended_action": self.recommended_action,
            "related_sku": self.related_sku,
            "related_campaign": self.related_campaign,
        }


@dataclass(frozen=True)
class KpiComparison:
    metric: str
    current_value: Decimal | int | None
    previous_value: Decimal | int | None
    absolute_change: Decimal | int | None
    percentage_change: Decimal | None
    change_label: str
    unit: str | None = None
    notes: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "metric": self.metric,
            "current_value": _json_value(self.current_value),
            "previous_value": _json_value(self.previous_value),
            "absolute_change": _json_value(self.absolute_change),
            "percentage_change": _optional_ratio_to_string(self.percentage_change),
            "change_label": self.change_label,
            "unit": self.unit,
            "notes": self.notes,
        }


@dataclass(frozen=True)
class SalesTrafficSummary:
    ordered_product_sales: Decimal = ZERO
    currency: str | None = None
    units_ordered: int = 0
    total_order_items: int = 0
    sessions: int = 0
    page_views: int = 0
    units_refunded: int = 0
    row_count: int = 0
    avg_selling_price: Decimal | None = None
    unit_session_percentage: Decimal | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "ordered_product_sales": _decimal_to_string(self.ordered_product_sales),
            "currency": self.currency,
            "units_ordered": self.units_ordered,
            "total_order_items": self.total_order_items,
            "sessions": self.sessions,
            "page_views": self.page_views,
            "units_refunded": self.units_refunded,
            "row_count": self.row_count,
            "avg_selling_price": _optional_decimal_to_string(self.avg_selling_price),
            "unit_session_percentage": _optional_ratio_to_string(self.unit_session_percentage),
        }


@dataclass(frozen=True)
class AdsSummary:
    row_count: int = 0
    campaign_count: int = 0
    impressions: int = 0
    clicks: int = 0
    spend: Decimal = ZERO
    sales_7d: Decimal = ZERO
    orders_7d: int = 0
    units_7d: int = 0
    ctr: Decimal | None = None
    cpc: Decimal | None = None
    ads_cvr: Decimal | None = None
    acos: Decimal | None = None
    roas: Decimal | None = None
    tacos: Decimal | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "ads_row_count": self.row_count,
            "campaign_count": self.campaign_count,
            "ads_impressions": self.impressions,
            "ads_clicks": self.clicks,
            "ads_spend": _decimal_to_string(self.spend),
            "ads_sales_7d": _decimal_to_string(self.sales_7d),
            "ads_orders_7d": self.orders_7d,
            "ads_units_7d": self.units_7d,
            "ctr": _optional_ratio_to_string(self.ctr),
            "cpc": _optional_decimal_to_string(self.cpc),
            "ads_cvr": _optional_ratio_to_string(self.ads_cvr),
            "acos": _optional_ratio_to_string(self.acos),
            "roas": _optional_decimal_to_string(self.roas),
            "tacos": _optional_ratio_to_string(self.tacos),
        }


@dataclass(frozen=True)
class SettlementPreview:
    settlement_net_preview: Decimal = ZERO
    settlement_product_sales: Decimal = ZERO
    settlement_advertising_fee: Decimal = ZERO
    settlement_fba_fee: Decimal = ZERO
    settlement_refund_amount: Decimal = ZERO
    settlement_promotion_amount: Decimal = ZERO
    settlement_row_count: int = 0
    currency: str | None = None
    note: str = SETTLEMENT_PREVIEW_NOTE

    def to_dict(self) -> dict[str, Any]:
        return {
            "settlement_net_preview": _decimal_to_string(self.settlement_net_preview),
            "settlement_product_sales": _decimal_to_string(self.settlement_product_sales),
            "settlement_advertising_fee": _decimal_to_string(self.settlement_advertising_fee),
            "settlement_fba_fee": _decimal_to_string(self.settlement_fba_fee),
            "settlement_refund_amount": _decimal_to_string(self.settlement_refund_amount),
            "settlement_promotion_amount": _decimal_to_string(self.settlement_promotion_amount),
            "settlement_row_count": self.settlement_row_count,
            "currency": self.currency,
            "note": self.note,
        }


@dataclass(frozen=True)
class DailyTrendRow:
    report_date: date
    ordered_product_sales: Decimal
    units_ordered: int
    total_order_items: int
    sessions: int
    page_views: int
    unit_session_percentage: Decimal | None
    avg_selling_price: Decimal | None
    ads_spend: Decimal
    ads_sales_7d: Decimal
    tacos: Decimal | None

    def to_dict(self) -> dict[str, Any]:
        return {
            "report_date": self.report_date.isoformat(),
            "ordered_product_sales": _decimal_to_string(self.ordered_product_sales),
            "units_ordered": self.units_ordered,
            "total_order_items": self.total_order_items,
            "sessions": self.sessions,
            "page_views": self.page_views,
            "unit_session_percentage": _optional_ratio_to_string(self.unit_session_percentage),
            "avg_selling_price": _optional_decimal_to_string(self.avg_selling_price),
            "ads_spend": _decimal_to_string(self.ads_spend),
            "ads_sales_7d": _decimal_to_string(self.ads_sales_7d),
            "tacos": _optional_ratio_to_string(self.tacos),
        }


@dataclass(frozen=True)
class SkuPerformanceRow:
    seller_sku: str
    asin: str | None
    product_name: str | None
    units_ordered: int
    order_item_sales: Decimal
    shipping_revenue: Decimal
    discount_total: Decimal
    order_net_sales_estimate: Decimal
    unit_standard_cost: Decimal | None
    estimated_cogs: Decimal
    gross_margin_before_ads: Decimal
    gross_margin_rate_before_ads: Decimal | None
    ads_spend: Decimal
    ads_sales_7d: Decimal
    ads_orders_7d: int
    contribution_after_ads: Decimal
    sku_tacos: Decimal | None
    ads_dependency_rate: Decimal | None
    fulfillable_quantity: int | None
    days_of_supply: Decimal | None
    inventory_risk: str | None
    status: str
    notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "seller_sku": self.seller_sku,
            "asin": self.asin,
            "product_name": self.product_name,
            "units_ordered": self.units_ordered,
            "order_item_sales": _decimal_to_string(self.order_item_sales),
            "shipping_revenue": _decimal_to_string(self.shipping_revenue),
            "discount_total": _decimal_to_string(self.discount_total),
            "order_net_sales_estimate": _decimal_to_string(self.order_net_sales_estimate),
            "unit_standard_cost": _optional_decimal_to_string(self.unit_standard_cost),
            "estimated_cogs": _decimal_to_string(self.estimated_cogs),
            "gross_margin_before_ads": _decimal_to_string(self.gross_margin_before_ads),
            "gross_margin_rate_before_ads": _optional_ratio_to_string(
                self.gross_margin_rate_before_ads
            ),
            "ads_spend": _decimal_to_string(self.ads_spend),
            "ads_sales_7d": _decimal_to_string(self.ads_sales_7d),
            "ads_orders_7d": self.ads_orders_7d,
            "contribution_after_ads": _decimal_to_string(self.contribution_after_ads),
            "contribution_after_cogs_ads_before_amazon_fees": _decimal_to_string(
                self.contribution_after_ads
            ),
            "sku_tacos": _optional_ratio_to_string(self.sku_tacos),
            "sku_ads_dependency_rate": _optional_ratio_to_string(self.ads_dependency_rate),
            "fulfillable_quantity": self.fulfillable_quantity,
            "days_of_supply": _optional_decimal_to_string(self.days_of_supply),
            "inventory_risk": self.inventory_risk,
            "status": self.status,
            "notes": "; ".join(self.notes),
        }


@dataclass(frozen=True)
class CampaignSummaryRow:
    campaign_id: str | None
    campaign_name: str | None
    campaign_status: str | None
    impressions: int
    clicks: int
    spend: Decimal
    sales_7d: Decimal
    orders_7d: int
    acos: Decimal | None
    roas: Decimal | None
    cpc: Decimal | None

    def to_dict(self) -> dict[str, Any]:
        return {
            "campaign_id": self.campaign_id,
            "campaign_name": self.campaign_name,
            "campaign_status": self.campaign_status,
            "impressions": self.impressions,
            "clicks": self.clicks,
            "campaign_spend": _decimal_to_string(self.spend),
            "campaign_sales_7d": _decimal_to_string(self.sales_7d),
            "campaign_orders_7d": self.orders_7d,
            "campaign_acos": _optional_ratio_to_string(self.acos),
            "campaign_roas": _optional_decimal_to_string(self.roas),
            "campaign_cpc": _optional_decimal_to_string(self.cpc),
        }


@dataclass(frozen=True)
class InventoryRiskRow:
    seller_sku: str
    asin: str | None
    product_name: str | None
    snapshot_date: date | None
    afn_fulfillable_quantity: int | None
    afn_reserved_quantity: int | None
    afn_unsellable_quantity: int | None
    afn_total_quantity: int | None
    weekly_units_ordered: int
    avg_daily_units_ordered_7d: Decimal | None
    days_of_supply: Decimal | None
    unit_standard_cost: Decimal | None
    inventory_value_at_cost: Decimal | None
    inventory_risk: str
    velocity_status: str
    notes: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "seller_sku": self.seller_sku,
            "asin": self.asin,
            "product_name": self.product_name,
            "snapshot_date": self.snapshot_date.isoformat() if self.snapshot_date else None,
            "afn_fulfillable_quantity": self.afn_fulfillable_quantity,
            "afn_reserved_quantity": self.afn_reserved_quantity,
            "afn_unsellable_quantity": self.afn_unsellable_quantity,
            "afn_total_quantity": self.afn_total_quantity,
            "weekly_units_ordered": self.weekly_units_ordered,
            "avg_daily_units_ordered_7d": _optional_decimal_to_string(
                self.avg_daily_units_ordered_7d
            ),
            "days_of_supply": _optional_decimal_to_string(self.days_of_supply),
            "unit_standard_cost": _optional_decimal_to_string(self.unit_standard_cost),
            "inventory_value_at_cost": _optional_decimal_to_string(self.inventory_value_at_cost),
            "inventory_risk": self.inventory_risk,
            "inventory_velocity_status": self.velocity_status,
            "notes": self.notes,
        }


@dataclass(frozen=True)
class WeeklyBusinessReviewResult:
    marketplace_id: str
    profile_id: str | None
    week_start: date
    week_end: date
    previous_week_start: date
    previous_week_end: date
    generated_at_utc: datetime
    status: str
    currency: str | None
    sales_traffic_summary: SalesTrafficSummary
    previous_sales_traffic_summary: SalesTrafficSummary
    ads_summary: AdsSummary
    previous_ads_summary: AdsSummary
    estimated_cogs: Decimal
    gross_margin_before_ads: Decimal
    contribution_after_ads: Decimal
    contribution_margin_after_ads: Decimal | None
    kpi_summary: list[KpiComparison]
    daily_trend: list[DailyTrendRow]
    sku_performance: list[SkuPerformanceRow]
    campaign_summary: list[CampaignSummaryRow]
    inventory_risk: list[InventoryRiskRow]
    settlement_finance_preview: SettlementPreview
    alerts: list[AlertAction]
    reconciliation_checks: list[ReconciliationCheck]
    warnings: list[WarningEntry]
    raw_metadata: dict[str, Any]
    output_files: dict[str, str] = field(default_factory=dict)

    def executive_summary(self) -> dict[str, Any]:
        sales_text = _format_money(
            self.sales_traffic_summary.ordered_product_sales,
            self.currency,
        )
        contribution_text = _format_money(self.contribution_after_ads, self.currency)
        headline = (
            f"{self.week_start.isoformat()}..{self.week_end.isoformat()} "
            f"ordered sales were {sales_text}, with contribution after COGS & Ads "
            f"before full Amazon fees of {contribution_text}."
        )
        warnings = [warning for warning in self.warnings if warning.severity != "info"]
        needs_review = [
            check for check in self.reconciliation_checks if check.status == "needs_review"
        ]
        return {
            "headline": headline,
            "key_points": [
                f"Report status: {self.status}.",
                f"Units ordered: {self.sales_traffic_summary.units_ordered}.",
                f"Sessions: {self.sales_traffic_summary.sessions}.",
                f"Ads spend: {_format_money(self.ads_summary.spend, self.currency)}; "
                f"TACOS: {_format_ratio(self.ads_summary.tacos)}.",
                f"Estimated COGS: {_format_money(self.estimated_cogs, self.currency)}.",
                f"Alerts: {len(self.alerts)}; non-info warnings: {len(warnings)}; "
                f"needs-review checks: {len(needs_review)}.",
                WBR_SCOPE_NOTE,
                CONTRIBUTION_SCOPE_NOTE,
            ],
        }

    def to_dict(self) -> dict[str, Any]:
        return {
            "report_type": REPORT_TYPE,
            "version": REPORT_VERSION,
            "marketplace_id": self.marketplace_id,
            "profile_id": self.profile_id,
            "period": {
                "week_start": self.week_start.isoformat(),
                "week_end": self.week_end.isoformat(),
                "previous_week_start": self.previous_week_start.isoformat(),
                "previous_week_end": self.previous_week_end.isoformat(),
            },
            "generated_at_utc": self.generated_at_utc.isoformat(),
            "status": self.status,
            "currency": self.currency,
            "executive_summary": self.executive_summary(),
            "kpi_summary": [row.to_dict() for row in self.kpi_summary],
            "daily_trend": [row.to_dict() for row in self.daily_trend],
            "sales_traffic_summary": self.sales_traffic_summary.to_dict(),
            "previous_sales_traffic_summary": self.previous_sales_traffic_summary.to_dict(),
            "sku_performance": [row.to_dict() for row in self.sku_performance],
            "ads_overview": {
                "summary": self.ads_summary.to_dict(),
                "previous_summary": self.previous_ads_summary.to_dict(),
                "campaign_summary": [row.to_dict() for row in self.campaign_summary],
            },
            "management_contribution": {
                "contribution_after_cogs_ads_before_amazon_fees": _decimal_to_string(
                    self.contribution_after_ads
                ),
                "contribution_margin_after_cogs_ads_before_amazon_fees": _optional_ratio_to_string(
                    self.contribution_margin_after_ads
                ),
                "scope_note": CONTRIBUTION_SCOPE_NOTE,
            },
            "inventory_risk": [row.to_dict() for row in self.inventory_risk],
            "settlement_finance_preview": self.settlement_finance_preview.to_dict(),
            "alerts": [alert.to_dict() for alert in self.alerts],
            "reconciliation_checks": [check.to_dict() for check in self.reconciliation_checks],
            "warnings": [warning.to_dict() for warning in self.warnings],
            "raw_metadata": _json_safe_mapping(self.raw_metadata),
            "output_files": self.output_files,
        }

    def with_output_files(self, output_files: Mapping[str, str]) -> WeeklyBusinessReviewResult:
        return WeeklyBusinessReviewResult(
            marketplace_id=self.marketplace_id,
            profile_id=self.profile_id,
            week_start=self.week_start,
            week_end=self.week_end,
            previous_week_start=self.previous_week_start,
            previous_week_end=self.previous_week_end,
            generated_at_utc=self.generated_at_utc,
            status=self.status,
            currency=self.currency,
            sales_traffic_summary=self.sales_traffic_summary,
            previous_sales_traffic_summary=self.previous_sales_traffic_summary,
            ads_summary=self.ads_summary,
            previous_ads_summary=self.previous_ads_summary,
            estimated_cogs=self.estimated_cogs,
            gross_margin_before_ads=self.gross_margin_before_ads,
            contribution_after_ads=self.contribution_after_ads,
            contribution_margin_after_ads=self.contribution_margin_after_ads,
            kpi_summary=self.kpi_summary,
            daily_trend=self.daily_trend,
            sku_performance=self.sku_performance,
            campaign_summary=self.campaign_summary,
            inventory_risk=self.inventory_risk,
            settlement_finance_preview=self.settlement_finance_preview,
            alerts=self.alerts,
            reconciliation_checks=self.reconciliation_checks,
            warnings=self.warnings,
            raw_metadata=self.raw_metadata,
            output_files=dict(output_files),
        )


class WeeklyBusinessReviewDataRepo(Protocol):
    def fetch_sales_traffic_daily_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_order_item_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_sku_cost_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_ads_campaign_daily_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_ads_advertised_product_daily_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_latest_inventory_rows(
        self,
        *,
        marketplace_id: str,
        as_of_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_latest_listing_rows(
        self,
        *,
        marketplace_id: str,
        as_of_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_settlement_preview_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...


class WeeklyBusinessReviewService:
    def __init__(self, repo: WeeklyBusinessReviewDataRepo | None = None) -> None:
        self.repo = repo

    def run(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        week_start: date,
        output_root: str | Path = DEFAULT_OUTPUT_ROOT,
        target_acos: Decimal = Decimal("0.30"),
        target_tacos: Decimal = Decimal("0.20"),
        low_stock_days: int = 14,
        watch_stock_days: int = 30,
        min_stable_lag_days: int = 2,
        ads_stable_lag_days: int = 3,
    ) -> WeeklyBusinessReviewResult:
        if self.repo is None:
            raise RuntimeError("repo is required for run(); use calculate_from_rows() in tests")
        week_end = week_start + timedelta(days=6)
        previous_week_start = week_start - timedelta(days=7)
        previous_week_end = week_start - timedelta(days=1)
        _validate_week_start(week_start)

        sales_rows = self.repo.fetch_sales_traffic_daily_rows(
            marketplace_id=marketplace_id,
            start_date=week_start,
            end_date=week_end,
        )
        previous_sales_rows = self.repo.fetch_sales_traffic_daily_rows(
            marketplace_id=marketplace_id,
            start_date=previous_week_start,
            end_date=previous_week_end,
        )
        order_rows = self.repo.fetch_order_item_rows(
            marketplace_id=marketplace_id,
            start_date=week_start,
            end_date=week_end,
        )
        previous_order_rows = self.repo.fetch_order_item_rows(
            marketplace_id=marketplace_id,
            start_date=previous_week_start,
            end_date=previous_week_end,
        )
        cost_rows = self.repo.fetch_sku_cost_rows(
            marketplace_id=marketplace_id,
            start_date=previous_week_start,
            end_date=week_end,
        )
        ads_rows = self.repo.fetch_ads_campaign_daily_rows(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            start_date=week_start,
            end_date=week_end,
        )
        previous_ads_rows = self.repo.fetch_ads_campaign_daily_rows(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            start_date=previous_week_start,
            end_date=previous_week_end,
        )
        ads_product_rows = self.repo.fetch_ads_advertised_product_daily_rows(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            start_date=week_start,
            end_date=week_end,
        )
        inventory_rows = self.repo.fetch_latest_inventory_rows(
            marketplace_id=marketplace_id,
            as_of_date=week_end,
        )
        listing_rows = self.repo.fetch_latest_listing_rows(
            marketplace_id=marketplace_id,
            as_of_date=week_end,
        )
        settlement_rows = self.repo.fetch_settlement_preview_rows(
            marketplace_id=marketplace_id,
            start_date=week_start,
            end_date=week_end,
        )
        result = self.calculate_from_rows(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            week_start=week_start,
            week_end=week_end,
            previous_week_start=previous_week_start,
            previous_week_end=previous_week_end,
            sales_traffic_rows=sales_rows,
            previous_sales_traffic_rows=previous_sales_rows,
            order_item_rows=order_rows,
            previous_order_item_rows=previous_order_rows,
            sku_cost_rows=cost_rows,
            ads_campaign_rows=ads_rows,
            previous_ads_campaign_rows=previous_ads_rows,
            ads_product_rows=ads_product_rows,
            inventory_rows=inventory_rows,
            listing_rows=listing_rows,
            settlement_rows=settlement_rows,
            target_acos=target_acos,
            target_tacos=target_tacos,
            low_stock_days=low_stock_days,
            watch_stock_days=watch_stock_days,
            min_stable_lag_days=min_stable_lag_days,
            ads_stable_lag_days=ads_stable_lag_days,
        )
        return self.write_report_files(result=result, output_root=output_root)

    def calculate_from_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        week_start: date,
        week_end: date | None = None,
        previous_week_start: date | None = None,
        previous_week_end: date | None = None,
        sales_traffic_rows: Sequence[Mapping[str, Any]] | None = None,
        previous_sales_traffic_rows: Sequence[Mapping[str, Any]] | None = None,
        order_item_rows: Sequence[Mapping[str, Any]] | None = None,
        previous_order_item_rows: Sequence[Mapping[str, Any]] | None = None,
        sku_cost_rows: Sequence[Mapping[str, Any]] | None = None,
        ads_campaign_rows: Sequence[Mapping[str, Any]] | None = None,
        previous_ads_campaign_rows: Sequence[Mapping[str, Any]] | None = None,
        ads_product_rows: Sequence[Mapping[str, Any]] | None = None,
        inventory_rows: Sequence[Mapping[str, Any]] | None = None,
        listing_rows: Sequence[Mapping[str, Any]] | None = None,
        settlement_rows: Sequence[Mapping[str, Any]] | None = None,
        target_acos: Decimal = Decimal("0.30"),
        target_tacos: Decimal = Decimal("0.20"),
        low_stock_days: int = 14,
        watch_stock_days: int = 30,
        min_stable_lag_days: int = 2,
        ads_stable_lag_days: int = 3,
        generated_at_utc: datetime | None = None,
    ) -> WeeklyBusinessReviewResult:
        _validate_week_start(week_start)
        week_end = week_end or week_start + timedelta(days=6)
        previous_week_start = previous_week_start or week_start - timedelta(days=7)
        previous_week_end = previous_week_end or week_start - timedelta(days=1)
        generated_at_utc = generated_at_utc or datetime.now(tz=UTC)

        sales_rows = list(sales_traffic_rows or [])
        previous_sales_rows = list(previous_sales_traffic_rows or [])
        orders = list(order_item_rows or [])
        previous_orders = list(previous_order_item_rows or [])
        costs = list(sku_cost_rows or [])
        ads_rows = list(ads_campaign_rows or [])
        previous_ads_rows = list(previous_ads_campaign_rows or [])
        ads_products = list(ads_product_rows or [])
        inventory = list(inventory_rows or [])
        listings = list(listing_rows or [])
        settlements = list(settlement_rows or [])

        sales_summary = _aggregate_sales_traffic(sales_rows)
        previous_sales_summary = _aggregate_sales_traffic(previous_sales_rows)
        ads_summary = _aggregate_ads_summary(
            ads_rows, ordered_product_sales=sales_summary.ordered_product_sales
        )
        previous_ads_summary = _aggregate_ads_summary(
            previous_ads_rows,
            ordered_product_sales=previous_sales_summary.ordered_product_sales,
        )
        cost_index = _build_cost_index(costs)
        ads_by_sku = _aggregate_ads_product_by_sku(ads_products)
        inventory_index = _build_inventory_index(inventory)
        listing_index = _build_listing_index(listings)
        sku_performance = _build_sku_performance(
            order_rows=orders,
            cost_index=cost_index,
            ads_by_sku=ads_by_sku,
            inventory_index=inventory_index,
            listing_index=listing_index,
            low_stock_days=low_stock_days,
            watch_stock_days=watch_stock_days,
        )
        estimated_cogs = _money(sum((row.estimated_cogs for row in sku_performance), ZERO))
        gross_margin_before_ads = _money(sales_summary.ordered_product_sales - estimated_cogs)
        contribution_after_ads = _money(gross_margin_before_ads - ads_summary.spend)
        contribution_margin_after_ads = _safe_ratio(
            contribution_after_ads, sales_summary.ordered_product_sales
        )
        daily_trend = _build_daily_trend(week_start, week_end, sales_rows, ads_rows)
        campaign_summary = _build_campaign_summary(ads_rows)
        inventory_risk = _build_inventory_risk(
            sku_performance=sku_performance,
            inventory_index=inventory_index,
            cost_index=cost_index,
            listing_index=listing_index,
            low_stock_days=low_stock_days,
            watch_stock_days=watch_stock_days,
        )
        settlement_preview = _build_settlement_preview(settlements)
        kpi_summary = _build_kpi_summary(
            current_sales=sales_summary,
            previous_sales=previous_sales_summary,
            current_ads=ads_summary,
            previous_ads=previous_ads_summary,
            current_cogs=estimated_cogs,
            current_gross_margin=gross_margin_before_ads,
            current_contribution=contribution_after_ads,
            current_contribution_margin=contribution_margin_after_ads,
            previous_order_rows=previous_orders,
            cost_index=cost_index,
            currency=sales_summary.currency,
        )
        warnings = _build_warnings(
            sales_rows=sales_rows,
            order_rows=orders,
            ads_rows=ads_rows,
            inventory_rows=inventory,
            sku_performance=sku_performance,
            week_start=week_start,
            week_end=week_end,
            generated_at_utc=generated_at_utc,
            min_stable_lag_days=min_stable_lag_days,
            ads_stable_lag_days=ads_stable_lag_days,
        )
        alerts = _build_alerts(
            sales_summary=sales_summary,
            previous_sales_summary=previous_sales_summary,
            ads_summary=ads_summary,
            previous_ads_summary=previous_ads_summary,
            sku_performance=sku_performance,
            inventory_risk=inventory_risk,
            warnings=warnings,
            target_acos=target_acos,
            target_tacos=target_tacos,
        )
        checks = _build_reconciliation_checks(
            sales_rows=sales_rows,
            order_rows=orders,
            ads_rows=ads_rows,
            inventory_rows=inventory,
            sku_performance=sku_performance,
            week_start=week_start,
            week_end=week_end,
            generated_at_utc=generated_at_utc,
            min_stable_lag_days=min_stable_lag_days,
            ads_stable_lag_days=ads_stable_lag_days,
        )
        status = _result_status(
            sales_summary=sales_summary,
            checks=checks,
            warnings=warnings,
        )
        currency = _first_non_empty(
            [
                sales_summary.currency,
                settlement_preview.currency,
                _first_non_empty(_empty_to_none(row.get("currency")) for row in orders),
            ]
        )
        raw_metadata = _build_raw_metadata(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            week_start=week_start,
            week_end=week_end,
            previous_week_start=previous_week_start,
            previous_week_end=previous_week_end,
            generated_at_utc=generated_at_utc,
            sales_rows=sales_rows,
            previous_sales_rows=previous_sales_rows,
            order_rows=orders,
            previous_order_rows=previous_orders,
            sku_cost_rows=costs,
            ads_rows=ads_rows,
            previous_ads_rows=previous_ads_rows,
            ads_product_rows=ads_products,
            inventory_rows=inventory,
            listing_rows=listings,
            settlement_rows=settlements,
        )

        return WeeklyBusinessReviewResult(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            week_start=week_start,
            week_end=week_end,
            previous_week_start=previous_week_start,
            previous_week_end=previous_week_end,
            generated_at_utc=generated_at_utc,
            status=status,
            currency=currency,
            sales_traffic_summary=sales_summary,
            previous_sales_traffic_summary=previous_sales_summary,
            ads_summary=ads_summary,
            previous_ads_summary=previous_ads_summary,
            estimated_cogs=estimated_cogs,
            gross_margin_before_ads=gross_margin_before_ads,
            contribution_after_ads=contribution_after_ads,
            contribution_margin_after_ads=contribution_margin_after_ads,
            kpi_summary=kpi_summary,
            daily_trend=daily_trend,
            sku_performance=sku_performance,
            campaign_summary=campaign_summary,
            inventory_risk=inventory_risk,
            settlement_finance_preview=settlement_preview,
            alerts=alerts,
            reconciliation_checks=checks,
            warnings=warnings,
            raw_metadata=raw_metadata,
        )

    def write_report_files(
        self,
        *,
        result: WeeklyBusinessReviewResult,
        output_root: str | Path = DEFAULT_OUTPUT_ROOT,
    ) -> WeeklyBusinessReviewResult:
        output_dir = (
            Path(output_root)
            / result.marketplace_id
            / f"{result.week_start.isoformat()}_{result.week_end.isoformat()}"
        )
        output_dir.mkdir(parents=True, exist_ok=True)
        period_key = f"{result.week_start.isoformat()}_{result.week_end.isoformat()}"
        filename_base = f"weekly_business_review_{period_key}"
        json_path = output_dir / f"{filename_base}.json"
        xlsx_path = output_dir / f"{filename_base}.xlsx"
        result_with_files = result.with_output_files(
            {"json": str(json_path), "xlsx": str(xlsx_path)}
        )
        json_path.write_text(
            json.dumps(result_with_files.to_dict(), ensure_ascii=False, indent=2, sort_keys=True)
            + "\n",
            encoding="utf-8",
        )
        workbook = build_weekly_business_review_workbook(result_with_files)
        workbook.save(xlsx_path)
        return result_with_files


def build_weekly_business_review_workbook(result: WeeklyBusinessReviewResult) -> Workbook:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    add_bilingual_readme_sheet(
        workbook,
        title_en="Weekly Business Review",
        title_zh="每周经营复盘",
        period=f"{result.week_start.isoformat()}_{result.week_end.isoformat()}",
        status=result.status,
        scope_en=(
            "Sales & Traffic, Orders, and Ads API report-date spend drive weekly business metrics; "
            "Settlement is finance context only. Contribution after COGS & Ads "
            "is not final net profit."
        ),
        scope_zh=(
            "Sales & Traffic、Orders 和 Ads API 发生日广告费用于周度经营指标；"
            "Settlement 仅作为财务参考。广告和货本后贡献不等于最终净利润。"
        ),
    )
    _write_rows_sheet(workbook, "01_Executive_Summary", _summary_rows(result))
    _write_rows_sheet(workbook, "02_Daily_Trend", [row.to_dict() for row in result.daily_trend])
    _write_rows_sheet(workbook, "03_Sales_Traffic", _sales_traffic_rows(result))
    _write_rows_sheet(
        workbook,
        "04_SKU_Performance",
        [_flatten_sku_row(row) for row in result.sku_performance],
    )
    _write_rows_sheet(workbook, "05_Ads_Overview", _ads_overview_rows(result))
    _write_rows_sheet(
        workbook,
        "06_Inventory_Risk",
        [row.to_dict() for row in result.inventory_risk],
    )
    _write_rows_sheet(workbook, "07_Alerts_Actions", [row.to_dict() for row in result.alerts])
    _write_rows_sheet(
        workbook,
        "08_Reconciliation_Checks",
        [row.to_dict() for row in result.reconciliation_checks],
    )
    _write_rows_sheet(workbook, "09_Raw_Metadata", _metadata_rows(result))
    return workbook


def _summary_rows(result: WeeklyBusinessReviewResult) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = [
        _metric_row("Report Status", result.status, None, WBR_SCOPE_NOTE),
        _metric_row("Week Start", result.week_start.isoformat(), None, ""),
        _metric_row("Week End", result.week_end.isoformat(), None, ""),
        _metric_row(
            "Ordered Product Sales",
            result.sales_traffic_summary.ordered_product_sales,
            result.currency,
            "Sales & Traffic report-date sales; main WBR sales metric.",
        ),
        _metric_row("Units Ordered", result.sales_traffic_summary.units_ordered, None, ""),
        _metric_row("Total Order Items", result.sales_traffic_summary.total_order_items, None, ""),
        _metric_row("Sessions", result.sales_traffic_summary.sessions, None, ""),
        _metric_row("Page Views", result.sales_traffic_summary.page_views, None, ""),
        _metric_row(
            "Unit Session Percentage",
            result.sales_traffic_summary.unit_session_percentage,
            None,
            "Recomputed from total units / total sessions.",
        ),
        _metric_row(
            "Ads Spend",
            result.ads_summary.spend,
            result.currency,
            "Ads API campaign daily.",
        ),
        _metric_row("Ads Sales 7d", result.ads_summary.sales_7d, result.currency, ""),
        _metric_row("ACOS", result.ads_summary.acos, None, "ads_spend / ads_sales_7d"),
        _metric_row("ROAS", result.ads_summary.roas, None, "ads_sales_7d / ads_spend"),
        _metric_row("TACOS", result.ads_summary.tacos, None, "ads_spend / sales"),
        _metric_row(
            "Estimated COGS",
            result.estimated_cogs,
            result.currency,
            "Orders units * SKU standard cost.",
        ),
        _metric_row(
            "Gross Margin Before Ads",
            result.gross_margin_before_ads,
            result.currency,
            "Sales - estimated COGS.",
        ),
        _metric_row(
            "Contribution After COGS & Ads (Before Amazon Fees)",
            result.contribution_after_ads,
            result.currency,
            CONTRIBUTION_SCOPE_NOTE,
        ),
        _metric_row(
            "Contribution Margin After COGS & Ads (Before Amazon Fees)",
            result.contribution_margin_after_ads,
            None,
            "Contribution / Sales & Traffic ordered sales; not final net margin.",
        ),
        _metric_row(
            "Settlement Net Preview",
            result.settlement_finance_preview.settlement_net_preview,
            result.currency,
            SETTLEMENT_PREVIEW_NOTE,
        ),
        _metric_row("Alert Count", len(result.alerts), None, ""),
    ]
    rows.append({"metric": "", "value": "", "currency": "", "notes": ""})
    for metric in result.kpi_summary:
        rows.append(
            {
                "metric": f"WoW - {metric.metric}",
                "value": _xlsx_value(metric.current_value),
                "previous_value": _xlsx_value(metric.previous_value),
                "absolute_change": _xlsx_value(metric.absolute_change),
                "percentage_change": _xlsx_value(metric.percentage_change),
                "change_label": metric.change_label,
                "currency": metric.unit,
                "notes": metric.notes,
            }
        )
    return rows


def _sales_traffic_rows(result: WeeklyBusinessReviewResult) -> list[dict[str, Any]]:
    rows = [
        {"metric": bilingual_metric_label(key), "value": value, "period": "current_week"}
        for key, value in result.sales_traffic_summary.to_dict().items()
    ]
    rows.extend(
        {"metric": bilingual_metric_label(key), "value": value, "period": "previous_week"}
        for key, value in result.previous_sales_traffic_summary.to_dict().items()
    )
    return rows


def _ads_overview_rows(result: WeeklyBusinessReviewResult) -> list[dict[str, Any]]:
    rows = [
        {"section": "summary", "metric": bilingual_metric_label(key), "value": value}
        for key, value in result.ads_summary.to_dict().items()
    ]
    rows.extend(
        {"section": "previous_summary", "metric": bilingual_metric_label(key), "value": value}
        for key, value in result.previous_ads_summary.to_dict().items()
    )
    rows.append({"section": "", "metric": "", "value": ""})
    rows.extend(
        {"section": "campaign", **campaign.to_dict()} for campaign in result.campaign_summary
    )
    return rows


def _metadata_rows(result: WeeklyBusinessReviewResult) -> list[dict[str, Any]]:
    rows = [{"key": key, "value": value} for key, value in result.raw_metadata.items()]
    rows.extend(
        {"key": f"output_{key}", "value": value} for key, value in result.output_files.items()
    )
    return rows


def _metric_row(metric: str, value: Any, currency: str | None, notes: str) -> dict[str, Any]:
    return {
        "metric": bilingual_metric_label(metric),
        "value": _xlsx_value(value),
        "currency": currency,
        "notes": notes,
    }


def _flatten_sku_row(row: SkuPerformanceRow) -> dict[str, Any]:
    data = row.to_dict()
    data["scope_note"] = CONTRIBUTION_SCOPE_NOTE
    return data


def _write_rows_sheet(workbook: Workbook, title: str, rows: Sequence[Mapping[str, Any]]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet.append([xlsx_header_label("message")])
        sheet.append(["No rows / 无数据"])
        _format_sheet(sheet)
        return
    headers = list(rows[0].keys())
    for row in rows[1:]:
        for key in row.keys():
            if key not in headers:
                headers.append(key)
    sheet.append([xlsx_header_label(header) for header in headers])
    for row in rows:
        sheet.append([_xlsx_value(row.get(header)) for header in headers])
    _format_sheet(sheet)


def _format_sheet(sheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def _aggregate_sales_traffic(rows: Sequence[Mapping[str, Any]]) -> SalesTrafficSummary:
    sales = ZERO
    units = 0
    order_items = 0
    sessions = 0
    page_views = 0
    units_refunded = 0
    currencies: list[str | None] = []
    for row in rows:
        sales += _to_decimal(row.get("ordered_product_sales_amount"))
        units += _int_metric(row, "units_ordered")
        order_items += _int_metric(row, "total_order_items")
        sessions += _int_metric(row, "sessions")
        page_views += _int_metric(row, "page_views")
        units_refunded += _int_metric(row, "units_refunded")
        currencies.append(_empty_to_none(row.get("ordered_product_sales_currency")))
    return SalesTrafficSummary(
        ordered_product_sales=_money(sales),
        currency=_first_non_empty(currencies),
        units_ordered=units,
        total_order_items=order_items,
        sessions=sessions,
        page_views=page_views,
        units_refunded=units_refunded,
        row_count=len(rows),
        avg_selling_price=_safe_money_ratio(_money(sales), Decimal(units)),
        unit_session_percentage=_safe_ratio(Decimal(units), Decimal(sessions)),
    )


def _aggregate_ads_summary(
    rows: Sequence[Mapping[str, Any]], *, ordered_product_sales: Decimal
) -> AdsSummary:
    impressions = sum(_int_metric(row, "impressions") for row in rows)
    clicks = sum(_int_metric(row, "clicks") for row in rows)
    spend = _money(sum((_to_decimal(row.get("cost")) for row in rows), ZERO))
    sales_7d = _money(sum((_to_decimal(row.get("sales_7d")) for row in rows), ZERO))
    orders_7d = sum(_int_metric(row, "purchases_7d") for row in rows)
    units_7d = sum(_int_metric(row, "units_sold_clicks_7d") for row in rows)
    campaign_keys = {
        _empty_to_none(row.get("campaign_id")) or _empty_to_none(row.get("campaign_name")) or "-"
        for row in rows
    }
    return AdsSummary(
        row_count=len(rows),
        campaign_count=len(campaign_keys) if rows else 0,
        impressions=impressions,
        clicks=clicks,
        spend=spend,
        sales_7d=sales_7d,
        orders_7d=orders_7d,
        units_7d=units_7d,
        ctr=_safe_ratio(Decimal(clicks), Decimal(impressions)),
        cpc=_safe_money_ratio(spend, Decimal(clicks)),
        ads_cvr=_safe_ratio(Decimal(orders_7d), Decimal(clicks)),
        acos=_safe_ratio(spend, sales_7d),
        roas=_safe_money_ratio(sales_7d, spend),
        tacos=_safe_ratio(spend, ordered_product_sales),
    )


def _aggregate_ads_product_by_sku(
    rows: Sequence[Mapping[str, Any]],
) -> dict[str, dict[str, Decimal | int]]:
    summary: dict[str, dict[str, Decimal | int]] = defaultdict(
        lambda: {
            "spend": ZERO,
            "sales_7d": ZERO,
            "orders_7d": 0,
            "impressions": 0,
            "clicks": 0,
        }
    )
    for row in rows:
        sku = _empty_to_none(row.get("advertised_sku"))
        if not sku:
            continue
        current = summary[sku]
        current["spend"] = _to_decimal(current["spend"]) + _to_decimal(row.get("cost"))
        current["sales_7d"] = _to_decimal(current["sales_7d"]) + _to_decimal(row.get("sales_7d"))
        current["orders_7d"] = int(current["orders_7d"]) + _int_metric(row, "purchases_7d")
        current["impressions"] = int(current["impressions"]) + _int_metric(row, "impressions")
        current["clicks"] = int(current["clicks"]) + _int_metric(row, "clicks")
    return dict(summary)


def _build_sku_performance(
    *,
    order_rows: Sequence[Mapping[str, Any]],
    cost_index: dict[str, list[Mapping[str, Any]]],
    ads_by_sku: Mapping[str, Mapping[str, Decimal | int]],
    inventory_index: Mapping[str, Mapping[str, Any]],
    listing_index: Mapping[str, Mapping[str, Any]],
    low_stock_days: int,
    watch_stock_days: int,
) -> list[SkuPerformanceRow]:
    accumulator: dict[str, dict[str, Any]] = defaultdict(_new_sku_accumulator)
    for row in order_rows:
        sku = _empty_to_none(row.get("seller_sku"))
        if not sku:
            continue
        acc = accumulator[sku]
        purchase_date = _as_date(row.get("purchase_date"))
        quantity = _int_metric(row, "quantity")
        item_price = _to_decimal(row.get("item_price"))
        shipping_price = _to_decimal(row.get("shipping_price"))
        item_discount = _to_decimal(row.get("item_promotion_discount"))
        ship_discount = _to_decimal(row.get("ship_promotion_discount"))
        discount_total = item_discount + ship_discount
        cost_record = _match_cost(cost_index, sku, purchase_date)
        unit_cost = _unit_cost(cost_record) if cost_record else None
        acc["seller_sku"] = sku
        acc["asin"] = acc["asin"] or _empty_to_none(row.get("asin"))
        acc["product_name"] = acc["product_name"] or _empty_to_none(row.get("product_name"))
        acc["units_ordered"] += quantity
        acc["order_item_sales"] += item_price
        acc["shipping_revenue"] += shipping_price
        acc["discount_total"] += discount_total
        if unit_cost is None:
            acc["missing_cost_units"] += quantity
        else:
            acc["estimated_cogs"] += unit_cost * Decimal(quantity)
            acc["unit_cost_total"] += unit_cost * Decimal(quantity)
            acc["unit_cost_units"] += quantity
            acc["cost_values"].add(unit_cost)
    for sku, acc in accumulator.items():
        ads = ads_by_sku.get(sku, {})
        inventory = inventory_index.get(sku, {})
        listing = listing_index.get(sku, {})
        if not acc["asin"]:
            acc["asin"] = _empty_to_none(inventory.get("asin")) or _empty_to_none(
                listing.get("asin")
            )
        if not acc["product_name"]:
            acc["product_name"] = _empty_to_none(inventory.get("product_name")) or _empty_to_none(
                listing.get("item_name")
            )
        acc["ads_spend"] = _money(_to_decimal(ads.get("spend")))
        acc["ads_sales_7d"] = _money(_to_decimal(ads.get("sales_7d")))
        acc["ads_orders_7d"] = int(ads.get("orders_7d") or 0)
        _attach_inventory_metrics(
            acc,
            inventory=inventory,
            low_stock_days=low_stock_days,
            watch_stock_days=watch_stock_days,
        )
    rows: list[SkuPerformanceRow] = []
    for sku, acc in accumulator.items():
        units = int(acc["units_ordered"])
        order_item_sales = _money(_to_decimal(acc["order_item_sales"]))
        shipping_revenue = _money(_to_decimal(acc["shipping_revenue"]))
        discount_total = _money(_to_decimal(acc["discount_total"]))
        order_net_sales = _money(order_item_sales + shipping_revenue - discount_total)
        estimated_cogs = _money(_to_decimal(acc["estimated_cogs"]))
        gross_margin = _money(order_item_sales - discount_total - estimated_cogs)
        ads_spend = _money(_to_decimal(acc["ads_spend"]))
        contribution = _money(gross_margin - ads_spend)
        unit_standard_cost = None
        if int(acc["unit_cost_units"]):
            unit_standard_cost = _safe_money_ratio(
                _to_decimal(acc["unit_cost_total"]), Decimal(acc["unit_cost_units"])
            )
        notes: list[str] = []
        status = "ok"
        if int(acc["missing_cost_units"]):
            status = "missing_cost"
            notes.append(f"missing cost for {acc['missing_cost_units']} units")
        if len(acc["cost_values"]) > 1:
            notes.append("multiple cost rows matched inside week")
        rows.append(
            SkuPerformanceRow(
                seller_sku=sku,
                asin=acc["asin"],
                product_name=acc["product_name"],
                units_ordered=units,
                order_item_sales=order_item_sales,
                shipping_revenue=shipping_revenue,
                discount_total=discount_total,
                order_net_sales_estimate=order_net_sales,
                unit_standard_cost=unit_standard_cost,
                estimated_cogs=estimated_cogs,
                gross_margin_before_ads=gross_margin,
                gross_margin_rate_before_ads=_safe_ratio(gross_margin, order_item_sales),
                ads_spend=ads_spend,
                ads_sales_7d=_money(_to_decimal(acc["ads_sales_7d"])),
                ads_orders_7d=int(acc["ads_orders_7d"]),
                contribution_after_ads=contribution,
                sku_tacos=_safe_ratio(ads_spend, order_item_sales),
                ads_dependency_rate=_safe_ratio(_to_decimal(acc["ads_sales_7d"]), order_item_sales),
                fulfillable_quantity=acc.get("fulfillable_quantity"),
                days_of_supply=acc.get("days_of_supply"),
                inventory_risk=acc.get("inventory_risk"),
                status=status,
                notes=notes,
            )
        )
    return sorted(rows, key=lambda row: row.order_item_sales, reverse=True)


def _new_sku_accumulator() -> dict[str, Any]:
    return {
        "seller_sku": None,
        "asin": None,
        "product_name": None,
        "units_ordered": 0,
        "order_item_sales": ZERO,
        "shipping_revenue": ZERO,
        "discount_total": ZERO,
        "estimated_cogs": ZERO,
        "unit_cost_total": ZERO,
        "unit_cost_units": 0,
        "missing_cost_units": 0,
        "cost_values": set(),
        "ads_spend": ZERO,
        "ads_sales_7d": ZERO,
        "ads_orders_7d": 0,
        "fulfillable_quantity": None,
        "days_of_supply": None,
        "inventory_risk": None,
    }


def _attach_inventory_metrics(
    acc: dict[str, Any],
    *,
    inventory: Mapping[str, Any],
    low_stock_days: int,
    watch_stock_days: int,
) -> None:
    fulfillable = _optional_int(inventory.get("afn_fulfillable_quantity"))
    acc["fulfillable_quantity"] = fulfillable
    units = int(acc["units_ordered"])
    if fulfillable is None:
        acc["inventory_risk"] = "missing_inventory"
        acc["days_of_supply"] = None
        return
    if units <= 0:
        acc["inventory_risk"] = "overstock_watch" if fulfillable > 0 else "no_recent_sales"
        acc["days_of_supply"] = None
        return
    avg_daily = Decimal(units) / Decimal(7)
    days_of_supply = _safe_money_ratio(Decimal(fulfillable), avg_daily)
    acc["days_of_supply"] = days_of_supply
    acc["inventory_risk"] = _inventory_risk_label(
        fulfillable=fulfillable,
        units_ordered=units,
        days_of_supply=days_of_supply,
        low_stock_days=low_stock_days,
        watch_stock_days=watch_stock_days,
    )


def _build_inventory_risk(
    *,
    sku_performance: Sequence[SkuPerformanceRow],
    inventory_index: Mapping[str, Mapping[str, Any]],
    cost_index: Mapping[str, list[Mapping[str, Any]]],
    listing_index: Mapping[str, Mapping[str, Any]],
    low_stock_days: int,
    watch_stock_days: int,
) -> list[InventoryRiskRow]:
    rows: list[InventoryRiskRow] = []
    order_skus = {row.seller_sku for row in sku_performance}
    all_skus = sorted(order_skus | set(inventory_index.keys()))
    sku_lookup = {row.seller_sku: row for row in sku_performance}
    for sku in all_skus:
        sku_row = sku_lookup.get(sku)
        inventory = inventory_index.get(sku, {})
        listing = listing_index.get(sku, {})
        units = sku_row.units_ordered if sku_row else 0
        fulfillable = _optional_int(inventory.get("afn_fulfillable_quantity"))
        total_qty = _optional_int(inventory.get("afn_total_quantity"))
        avg_daily = _safe_money_ratio(Decimal(units), Decimal(7)) if units else None
        days_of_supply = None
        if fulfillable is not None and avg_daily and avg_daily > ZERO:
            days_of_supply = _safe_money_ratio(Decimal(fulfillable), avg_daily)
        cost_record = _match_cost(cost_index, sku, None)
        unit_cost = _unit_cost(cost_record) if cost_record else None
        inventory_value = None
        if unit_cost is not None and total_qty is not None:
            inventory_value = _money(unit_cost * Decimal(total_qty))
        risk = _inventory_risk_label(
            fulfillable=fulfillable,
            units_ordered=units,
            days_of_supply=days_of_supply,
            low_stock_days=low_stock_days,
            watch_stock_days=watch_stock_days,
        )
        rows.append(
            InventoryRiskRow(
                seller_sku=sku,
                asin=(sku_row.asin if sku_row else None)
                or _empty_to_none(inventory.get("asin"))
                or _empty_to_none(listing.get("asin")),
                product_name=(sku_row.product_name if sku_row else None)
                or _empty_to_none(inventory.get("product_name"))
                or _empty_to_none(listing.get("item_name")),
                snapshot_date=_as_date(inventory.get("snapshot_date")),
                afn_fulfillable_quantity=fulfillable,
                afn_reserved_quantity=_optional_int(inventory.get("afn_reserved_quantity")),
                afn_unsellable_quantity=_optional_int(inventory.get("afn_unsellable_quantity")),
                afn_total_quantity=total_qty,
                weekly_units_ordered=units,
                avg_daily_units_ordered_7d=avg_daily,
                days_of_supply=days_of_supply,
                unit_standard_cost=unit_cost,
                inventory_value_at_cost=inventory_value,
                inventory_risk=risk,
                velocity_status="active" if units > 0 else "no_recent_sales",
                notes="latest inventory snapshot; not historical daily inventory",
            )
        )
    return sorted(rows, key=lambda row: (row.inventory_risk, row.seller_sku))


def _build_daily_trend(
    week_start: date,
    week_end: date,
    sales_rows: Sequence[Mapping[str, Any]],
    ads_rows: Sequence[Mapping[str, Any]],
) -> list[DailyTrendRow]:
    sales_by_date = {_as_date(row.get("report_date")): row for row in sales_rows}
    ads_by_date: dict[date, dict[str, Decimal]] = defaultdict(
        lambda: {"spend": ZERO, "sales": ZERO}
    )
    for row in ads_rows:
        report_date = _as_date(row.get("report_date"))
        if report_date is None:
            continue
        ads_by_date[report_date]["spend"] += _to_decimal(row.get("cost"))
        ads_by_date[report_date]["sales"] += _to_decimal(row.get("sales_7d"))
    output: list[DailyTrendRow] = []
    for offset in range((week_end - week_start).days + 1):
        current_date = week_start + timedelta(days=offset)
        sales = sales_by_date.get(current_date, {})
        sales_amount = _money(_to_decimal(sales.get("ordered_product_sales_amount")))
        units = _int_metric(sales, "units_ordered")
        sessions = _int_metric(sales, "sessions")
        ads = ads_by_date.get(current_date, {"spend": ZERO, "sales": ZERO})
        ads_spend = _money(_to_decimal(ads["spend"]))
        output.append(
            DailyTrendRow(
                report_date=current_date,
                ordered_product_sales=sales_amount,
                units_ordered=units,
                total_order_items=_int_metric(sales, "total_order_items"),
                sessions=sessions,
                page_views=_int_metric(sales, "page_views"),
                unit_session_percentage=_safe_ratio(Decimal(units), Decimal(sessions)),
                avg_selling_price=_safe_money_ratio(sales_amount, Decimal(units)),
                ads_spend=ads_spend,
                ads_sales_7d=_money(_to_decimal(ads["sales"])),
                tacos=_safe_ratio(ads_spend, sales_amount),
            )
        )
    return output


def _build_campaign_summary(rows: Sequence[Mapping[str, Any]]) -> list[CampaignSummaryRow]:
    buckets: dict[tuple[str | None, str | None], dict[str, Any]] = defaultdict(
        lambda: {
            "status": None,
            "impressions": 0,
            "clicks": 0,
            "spend": ZERO,
            "sales_7d": ZERO,
            "orders_7d": 0,
        }
    )
    for row in rows:
        key = (_empty_to_none(row.get("campaign_id")), _empty_to_none(row.get("campaign_name")))
        bucket = buckets[key]
        bucket["status"] = _empty_to_none(row.get("campaign_status")) or bucket["status"]
        bucket["impressions"] += _int_metric(row, "impressions")
        bucket["clicks"] += _int_metric(row, "clicks")
        bucket["spend"] += _to_decimal(row.get("cost"))
        bucket["sales_7d"] += _to_decimal(row.get("sales_7d"))
        bucket["orders_7d"] += _int_metric(row, "purchases_7d")
    output = []
    for (campaign_id, campaign_name), bucket in buckets.items():
        spend = _money(_to_decimal(bucket["spend"]))
        sales = _money(_to_decimal(bucket["sales_7d"]))
        clicks = int(bucket["clicks"])
        output.append(
            CampaignSummaryRow(
                campaign_id=campaign_id,
                campaign_name=campaign_name,
                campaign_status=bucket["status"],
                impressions=int(bucket["impressions"]),
                clicks=clicks,
                spend=spend,
                sales_7d=sales,
                orders_7d=int(bucket["orders_7d"]),
                acos=_safe_ratio(spend, sales),
                roas=_safe_money_ratio(sales, spend),
                cpc=_safe_money_ratio(spend, Decimal(clicks)),
            )
        )
    return sorted(output, key=lambda row: row.spend, reverse=True)


def _build_settlement_preview(rows: Sequence[Mapping[str, Any]]) -> SettlementPreview:
    net = ZERO
    product_sales = ZERO
    advertising = ZERO
    fba_fee = ZERO
    refund = ZERO
    promotion = ZERO
    currencies = []
    for row in rows:
        amount = _to_decimal(row.get("amount"))
        bucket = (_empty_to_none(row.get("profit_bucket")) or "").lower()
        transaction_type = (_empty_to_none(row.get("transaction_type")) or "").lower()
        net += amount
        currencies.append(_empty_to_none(row.get("currency")))
        if bucket == "product_sales":
            product_sales += amount
        if bucket == "advertising_cost":
            advertising += amount
        if bucket == "fba_fee":
            fba_fee += amount
        if bucket == "refund" or transaction_type == "refund":
            refund += amount
        if bucket in {"promotion_cost", "promotion_fee"}:
            promotion += amount
    return SettlementPreview(
        settlement_net_preview=_money(net),
        settlement_product_sales=_money(product_sales),
        settlement_advertising_fee=_money(advertising),
        settlement_fba_fee=_money(fba_fee),
        settlement_refund_amount=_money(refund),
        settlement_promotion_amount=_money(promotion),
        settlement_row_count=len(rows),
        currency=_first_non_empty(currencies),
    )


def _build_kpi_summary(
    *,
    current_sales: SalesTrafficSummary,
    previous_sales: SalesTrafficSummary,
    current_ads: AdsSummary,
    previous_ads: AdsSummary,
    current_cogs: Decimal,
    current_gross_margin: Decimal,
    current_contribution: Decimal,
    current_contribution_margin: Decimal | None,
    previous_order_rows: Sequence[Mapping[str, Any]],
    cost_index: Mapping[str, list[Mapping[str, Any]]],
    currency: str | None,
) -> list[KpiComparison]:
    previous_cogs = _estimate_orders_cogs(previous_order_rows, cost_index)
    previous_gross_margin = previous_sales.ordered_product_sales - previous_cogs
    previous_contribution = previous_gross_margin - previous_ads.spend
    previous_contribution_margin = _safe_ratio(
        previous_contribution, previous_sales.ordered_product_sales
    )
    return [
        _compare(
            "ordered_product_sales",
            current_sales.ordered_product_sales,
            previous_sales.ordered_product_sales,
            currency,
        ),
        _compare("units_ordered", current_sales.units_ordered, previous_sales.units_ordered, None),
        _compare("sessions", current_sales.sessions, previous_sales.sessions, None),
        _compare(
            "unit_session_percentage",
            current_sales.unit_session_percentage,
            previous_sales.unit_session_percentage,
            None,
        ),
        _compare("ads_spend", current_ads.spend, previous_ads.spend, currency),
        _compare("tacos", current_ads.tacos, previous_ads.tacos, None),
        _compare("estimated_cogs", current_cogs, previous_cogs, currency),
        _compare("gross_margin_before_ads", current_gross_margin, previous_gross_margin, currency),
        _compare("contribution_after_ads", current_contribution, previous_contribution, currency),
        _compare(
            "contribution_margin_after_ads",
            current_contribution_margin,
            previous_contribution_margin,
            None,
        ),
    ]


def _compare(
    metric: str,
    current_value: Decimal | int | None,
    previous_value: Decimal | int | None,
    unit: str | None,
) -> KpiComparison:
    if current_value is None and previous_value is None:
        return KpiComparison(metric, None, None, None, None, "no_data", unit)
    current_decimal = _to_decimal(current_value)
    previous_decimal = _to_decimal(previous_value)
    absolute_change = current_decimal - previous_decimal
    if previous_decimal == ZERO:
        if current_decimal > ZERO:
            percentage_change = None
            change_label = "new_activity"
        else:
            percentage_change = ZERO
            change_label = "flat"
    else:
        percentage_change = _ratio(absolute_change / abs(previous_decimal))
        if absolute_change > ZERO:
            change_label = "up"
        elif absolute_change < ZERO:
            change_label = "down"
        else:
            change_label = "flat"
    return KpiComparison(
        metric=metric,
        current_value=(
            _money(current_decimal) if isinstance(current_value, Decimal) else current_value
        ),
        previous_value=(
            _money(previous_decimal) if isinstance(previous_value, Decimal) else previous_value
        ),
        absolute_change=(
            _money(absolute_change) if isinstance(current_value, Decimal) else int(absolute_change)
        ),
        percentage_change=percentage_change,
        change_label=change_label,
        unit=unit,
    )


def _estimate_orders_cogs(
    order_rows: Sequence[Mapping[str, Any]],
    cost_index: Mapping[str, list[Mapping[str, Any]]],
) -> Decimal:
    total = ZERO
    for row in order_rows:
        sku = _empty_to_none(row.get("seller_sku"))
        if not sku:
            continue
        cost_record = _match_cost(cost_index, sku, _as_date(row.get("purchase_date")))
        if cost_record:
            total += _unit_cost(cost_record) * Decimal(_int_metric(row, "quantity"))
    return _money(total)


def _build_warnings(
    *,
    sales_rows: Sequence[Mapping[str, Any]],
    order_rows: Sequence[Mapping[str, Any]],
    ads_rows: Sequence[Mapping[str, Any]],
    inventory_rows: Sequence[Mapping[str, Any]],
    sku_performance: Sequence[SkuPerformanceRow],
    week_start: date,
    week_end: date,
    generated_at_utc: datetime,
    min_stable_lag_days: int,
    ads_stable_lag_days: int,
) -> list[WarningEntry]:
    warnings: list[WarningEntry] = [
        WarningEntry("wbr_scope_policy", "info", WBR_SCOPE_NOTE, related_source="policy")
    ]
    expected_dates = _date_set(week_start, week_end)
    sales_dates = {_as_date(row.get("report_date")) for row in sales_rows}
    missing_sales_dates = sorted(date_ for date_ in expected_dates - sales_dates if date_)
    if missing_sales_dates:
        warnings.append(
            WarningEntry(
                "sales_traffic_missing_dates",
                "warning",
                "Sales & Traffic is missing dates: "
                + ", ".join(date_.isoformat() for date_ in missing_sales_dates),
                related_source="amazon_sales_traffic_daily",
            )
        )
    if not order_rows and sales_rows:
        warnings.append(
            WarningEntry(
                "orders_context_missing",
                "warning",
                "Sales & Traffic exists but Orders has no matching rows; "
                "SKU/COGS section needs review.",
                related_source="amazon_order_item",
            )
        )
    if not ads_rows:
        warnings.append(
            WarningEntry(
                "ads_api_context_missing",
                "warning",
                "Ads API campaign daily context has zero rows for this week/profile. "
                "Sales, traffic and SKU sections can still be used.",
                related_source="amazon_ads_sp_campaign_daily",
            )
        )
    missing_cost = [row.seller_sku for row in sku_performance if row.status == "missing_cost"]
    for sku in missing_cost:
        warnings.append(
            WarningEntry(
                "missing_sku_cost",
                "critical",
                "SKU has ordered units but no effective standard cost.",
                related_sku=sku,
                related_source="amazon_sku_cost",
            )
        )
    if not inventory_rows:
        warnings.append(
            WarningEntry(
                "inventory_context_missing",
                "warning",
                "No inventory snapshot was found at or before week end; inventory risk is partial.",
                related_source="amazon_inventory_daily",
            )
        )
    else:
        snapshot_dates = [_as_date(row.get("snapshot_date")) for row in inventory_rows]
        latest_snapshot = max((snapshot for snapshot in snapshot_dates if snapshot), default=None)
        if latest_snapshot and (week_end - latest_snapshot).days > 7:
            warnings.append(
                WarningEntry(
                    "stale_inventory_snapshot",
                    "warning",
                    f"Latest inventory snapshot {latest_snapshot.isoformat()} is "
                    "more than 7 days before week end.",
                    related_source="amazon_inventory_daily",
                )
            )
    stable_sales_end = generated_at_utc.date() - timedelta(days=min_stable_lag_days)
    stable_ads_end = generated_at_utc.date() - timedelta(days=ads_stable_lag_days)
    if week_end > stable_sales_end:
        warnings.append(
            WarningEntry(
                "sales_week_unstable",
                "warning",
                "Week end is newer than the configured Sales/Orders stable cutoff.",
                related_source="stable_cutoff",
            )
        )
    if week_end > stable_ads_end:
        warnings.append(
            WarningEntry(
                "ads_week_unstable",
                "warning",
                "Week end is newer than the configured Ads stable cutoff.",
                related_source="stable_cutoff",
            )
        )
    return warnings


def _build_alerts(
    *,
    sales_summary: SalesTrafficSummary,
    previous_sales_summary: SalesTrafficSummary,
    ads_summary: AdsSummary,
    previous_ads_summary: AdsSummary,
    sku_performance: Sequence[SkuPerformanceRow],
    inventory_risk: Sequence[InventoryRiskRow],
    warnings: Sequence[WarningEntry],
    target_acos: Decimal,
    target_tacos: Decimal,
) -> list[AlertAction]:
    alerts: list[AlertAction] = []
    sales_change = _safe_ratio(
        sales_summary.ordered_product_sales - previous_sales_summary.ordered_product_sales,
        previous_sales_summary.ordered_product_sales,
    )
    if sales_change is not None and sales_change < Decimal("-0.20"):
        alerts.append(
            AlertAction(
                "warning",
                "sales",
                "ordered_product_sales",
                _decimal_to_string(sales_summary.ordered_product_sales),
                _decimal_to_string(previous_sales_summary.ordered_product_sales),
                "-20%",
                "本周销售额环比下降超过 20%。",
                "检查广告、价格、优惠券、库存和 Buy Box。",
            )
        )
    conversion_change = _safe_ratio(
        _to_decimal(sales_summary.unit_session_percentage)
        - _to_decimal(previous_sales_summary.unit_session_percentage),
        _to_decimal(previous_sales_summary.unit_session_percentage),
    )
    if conversion_change is not None and conversion_change < Decimal("-0.20"):
        alerts.append(
            AlertAction(
                "warning",
                "sales",
                "unit_session_percentage",
                _optional_ratio_to_string(sales_summary.unit_session_percentage),
                _optional_ratio_to_string(previous_sales_summary.unit_session_percentage),
                "-20%",
                "本周转化率环比下降超过 20%。",
                "检查 listing、价格、评论、优惠券和图片。",
            )
        )
    if ads_summary.acos is not None and ads_summary.acos > target_acos:
        alerts.append(
            AlertAction(
                "warning",
                "ads",
                "acos",
                _optional_ratio_to_string(ads_summary.acos),
                _optional_ratio_to_string(previous_ads_summary.acos),
                _format_ratio(target_acos),
                "ACOS 高于目标阈值。",
                "检查高花费 campaign / target，必要时降价或否词。",
            )
        )
    if ads_summary.tacos is not None and ads_summary.tacos > target_tacos:
        alerts.append(
            AlertAction(
                "warning",
                "ads",
                "tacos",
                _optional_ratio_to_string(ads_summary.tacos),
                _optional_ratio_to_string(previous_ads_summary.tacos),
                _format_ratio(target_tacos),
                "TACOS 高于目标阈值，广告压力偏大。",
                "控制低效广告预算，保留能带动自然销售的核心词。",
            )
        )
    if ads_summary.spend > ZERO and ads_summary.sales_7d == ZERO:
        alerts.append(
            AlertAction(
                "warning",
                "ads",
                "ads_spend_without_sales",
                _decimal_to_string(ads_summary.spend),
                _decimal_to_string(previous_ads_summary.spend),
                "ads_sales_7d = 0",
                "广告有花费但无归因销售。",
                "检查投放词和 campaign，必要时下调或暂停。",
            )
        )
    if ads_summary.spend > previous_ads_summary.spend * Decimal("1.30") and (
        sales_summary.ordered_product_sales <= previous_sales_summary.ordered_product_sales
    ):
        alerts.append(
            AlertAction(
                "warning",
                "ads",
                "spend_up_sales_not_up",
                _decimal_to_string(ads_summary.spend),
                _decimal_to_string(previous_ads_summary.spend),
                "+30% spend",
                "广告花费环比上升，但销售额未同步上升。",
                "检查新广告或预算调整是否带来有效订单。",
            )
        )
    for row in sku_performance:
        if row.contribution_after_ads < ZERO:
            alerts.append(
                AlertAction(
                    "warning",
                    "sku",
                    "sku_negative_contribution",
                    _decimal_to_string(row.contribution_after_ads),
                    None,
                    "< 0",
                    "SKU 广告后贡献为负。",
                    "检查价格、成本、广告和促销策略。",
                    related_sku=row.seller_sku,
                )
            )
        if row.status == "missing_cost":
            alerts.append(
                AlertAction(
                    "critical",
                    "data_quality",
                    "missing_sku_cost",
                    None,
                    None,
                    "cost required",
                    "SKU 缺少成本，无法输出可靠贡献。",
                    "先维护 amazon_sku_cost，再重新生成周报。",
                    related_sku=row.seller_sku,
                )
            )
    for row in inventory_risk:
        if row.inventory_risk in {"stockout", "urgent_low_stock", "overstock_watch"}:
            alerts.append(
                AlertAction(
                    "critical" if row.inventory_risk == "stockout" else "warning",
                    "inventory",
                    row.inventory_risk,
                    _optional_decimal_to_string(row.days_of_supply),
                    None,
                    "low/over stock rule",
                    f"SKU inventory risk: {row.inventory_risk}.",
                    "按库存风险调整广告、补货或清仓策略。",
                    related_sku=row.seller_sku,
                )
            )
    for warning in warnings:
        if warning.warning_code == "ads_api_context_missing":
            alerts.append(
                AlertAction(
                    "info",
                    "data_quality",
                    "ads_api_context_missing",
                    None,
                    None,
                    "ads_row_count = 0",
                    "Ads API context 缺失。",
                    "确认 core_rolling/weekly_full 是否覆盖该周；不影响销售主体。",
                )
            )
    return alerts


def _build_reconciliation_checks(
    *,
    sales_rows: Sequence[Mapping[str, Any]],
    order_rows: Sequence[Mapping[str, Any]],
    ads_rows: Sequence[Mapping[str, Any]],
    inventory_rows: Sequence[Mapping[str, Any]],
    sku_performance: Sequence[SkuPerformanceRow],
    week_start: date,
    week_end: date,
    generated_at_utc: datetime,
    min_stable_lag_days: int,
    ads_stable_lag_days: int,
) -> list[ReconciliationCheck]:
    expected_dates = _date_set(week_start, week_end)
    sales_dates = {_as_date(row.get("report_date")) for row in sales_rows}
    ads_dates = {_as_date(row.get("report_date")) for row in ads_rows}
    missing_sales = expected_dates - {item for item in sales_dates if item}
    missing_ads = expected_dates - {item for item in ads_dates if item}
    missing_cost_skus = [row.seller_sku for row in sku_performance if row.status == "missing_cost"]
    checks = [
        ReconciliationCheck(
            "sales_traffic_coverage",
            "ok" if not missing_sales else "needs_review",
            "critical" if missing_sales else "info",
            "7 dates",
            f"{len(expected_dates) - len(missing_sales)} dates",
            message=(
                "Sales & Traffic covers the full week."
                if not missing_sales
                else "Missing Sales & Traffic dates: "
                + ", ".join(sorted(date_.isoformat() for date_ in missing_sales))
            ),
        ),
        ReconciliationCheck(
            "orders_context_available",
            "ok" if order_rows else "needs_review",
            "critical" if not order_rows else "info",
            "> 0 rows when sales exist",
            f"{len(order_rows)} rows",
            message="Orders rows are required for SKU/COGS analysis.",
        ),
        ReconciliationCheck(
            "sku_cost_coverage",
            "ok" if not missing_cost_skus else "needs_review",
            "critical" if missing_cost_skus else "info",
            "all ordered SKUs have effective cost",
            ", ".join(missing_cost_skus) if missing_cost_skus else "all covered",
            message="Missing cost SKUs must be fixed before contribution is trusted.",
        ),
        ReconciliationCheck(
            "ads_campaign_coverage",
            "ok" if not missing_ads else "warning",
            "warning" if missing_ads else "info",
            "7 dates",
            f"{len(expected_dates) - len(missing_ads)} dates",
            message=(
                "Ads campaign daily covers the full week."
                if not missing_ads
                else "Ads campaign daily is partial/missing; operational ads context only."
            ),
        ),
        ReconciliationCheck(
            "inventory_snapshot_available",
            "ok" if inventory_rows else "warning",
            "warning" if not inventory_rows else "info",
            "> 0 latest snapshot rows",
            f"{len(inventory_rows)} rows",
            message=(
                "Inventory snapshot is used for stock risk; missing inventory makes "
                "that module partial."
            ),
        ),
    ]
    stable_sales_end = generated_at_utc.date() - timedelta(days=min_stable_lag_days)
    stable_ads_end = generated_at_utc.date() - timedelta(days=ads_stable_lag_days)
    stable_ok = week_end <= stable_sales_end and week_end <= stable_ads_end
    checks.append(
        ReconciliationCheck(
            "stable_cutoff_check",
            "ok" if stable_ok else "warning",
            "warning" if not stable_ok else "info",
            f"week_end <= {min(stable_sales_end, stable_ads_end).isoformat()}",
            week_end.isoformat(),
            message="Week should be generated after Sales/Orders/Ads stable cutoff.",
        )
    )
    return checks


def _result_status(
    *,
    sales_summary: SalesTrafficSummary,
    checks: Sequence[ReconciliationCheck],
    warnings: Sequence[WarningEntry],
) -> str:
    if sales_summary.row_count == 0:
        return "no_data"
    if any(check.status == "needs_review" for check in checks):
        return "needs_review"
    if any(warning.warning_code.endswith("_unstable") for warning in warnings):
        return "preview_unstable"
    if any(check.status == "warning" for check in checks):
        return "partial"
    if any(warning.severity == "warning" for warning in warnings):
        return "partial"
    return "ok"


def _build_raw_metadata(
    *,
    marketplace_id: str,
    profile_id: str | None,
    week_start: date,
    week_end: date,
    previous_week_start: date,
    previous_week_end: date,
    generated_at_utc: datetime,
    sales_rows: Sequence[Mapping[str, Any]],
    previous_sales_rows: Sequence[Mapping[str, Any]],
    order_rows: Sequence[Mapping[str, Any]],
    previous_order_rows: Sequence[Mapping[str, Any]],
    sku_cost_rows: Sequence[Mapping[str, Any]],
    ads_rows: Sequence[Mapping[str, Any]],
    previous_ads_rows: Sequence[Mapping[str, Any]],
    ads_product_rows: Sequence[Mapping[str, Any]],
    inventory_rows: Sequence[Mapping[str, Any]],
    listing_rows: Sequence[Mapping[str, Any]],
    settlement_rows: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    return {
        "report_type": REPORT_TYPE,
        "version": REPORT_VERSION,
        "marketplace_id": marketplace_id,
        "profile_id": profile_id,
        "week_start": week_start.isoformat(),
        "week_end": week_end.isoformat(),
        "previous_week_start": previous_week_start.isoformat(),
        "previous_week_end": previous_week_end.isoformat(),
        "generated_at_utc": generated_at_utc.isoformat(),
        "source_tables": (
            "amazon_sales_traffic_daily, amazon_order_item, amazon_sku_cost, "
            "amazon_ads_sp_campaign_daily, amazon_ads_sp_advertised_product_daily, "
            "amazon_inventory_daily, amazon_listing_snapshot, amazon_settlement_transaction"
        ),
        "sales_traffic_row_count": len(sales_rows),
        "previous_sales_traffic_row_count": len(previous_sales_rows),
        "order_item_row_count": len(order_rows),
        "previous_order_item_row_count": len(previous_order_rows),
        "sku_cost_row_count": len(sku_cost_rows),
        "ads_campaign_row_count": len(ads_rows),
        "previous_ads_campaign_row_count": len(previous_ads_rows),
        "ads_advertised_product_row_count": len(ads_product_rows),
        "inventory_row_count": len(inventory_rows),
        "listing_row_count": len(listing_rows),
        "settlement_row_count": len(settlement_rows),
        "note": WBR_SCOPE_NOTE,
    }


def _build_cost_index(rows: Iterable[Mapping[str, Any]]) -> dict[str, list[Mapping[str, Any]]]:
    index: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        sku = _empty_to_none(row.get("seller_sku"))
        if not sku:
            continue
        index[sku].append(row)
    for entries in index.values():
        entries.sort(key=lambda row: _as_date(row.get("effective_from")) or date.min, reverse=True)
    return dict(index)


def _match_cost(
    cost_index: Mapping[str, list[Mapping[str, Any]]],
    sku: str,
    effective_date: date | None,
) -> Mapping[str, Any] | None:
    entries = cost_index.get(sku) or []
    for row in entries:
        effective_from = _as_date(row.get("effective_from")) or date.min
        effective_to = _as_date(row.get("effective_to")) or date.max
        if effective_date is None or (effective_from <= effective_date <= effective_to):
            return row
    return None


def _unit_cost(row: Mapping[str, Any]) -> Decimal:
    return _money(
        _to_decimal(row.get("product_cost"))
        + _to_decimal(row.get("first_mile_cost"))
        + _to_decimal(row.get("packaging_cost"))
        + _to_decimal(row.get("other_unit_cost"))
    )


def _build_inventory_index(rows: Iterable[Mapping[str, Any]]) -> dict[str, Mapping[str, Any]]:
    return {sku: row for row in rows if (sku := _empty_to_none(row.get("seller_sku"))) is not None}


def _build_listing_index(rows: Iterable[Mapping[str, Any]]) -> dict[str, Mapping[str, Any]]:
    return {sku: row for row in rows if (sku := _empty_to_none(row.get("seller_sku"))) is not None}


def _inventory_risk_label(
    *,
    fulfillable: int | None,
    units_ordered: int,
    days_of_supply: Decimal | None,
    low_stock_days: int,
    watch_stock_days: int,
) -> str:
    if fulfillable is None:
        return "missing_inventory"
    if fulfillable == 0 and units_ordered > 0:
        return "stockout"
    if days_of_supply is None:
        return "overstock_watch" if fulfillable > 0 else "no_recent_sales"
    if days_of_supply < Decimal(low_stock_days):
        return "urgent_low_stock"
    if days_of_supply < Decimal(watch_stock_days):
        return "watch_low_stock"
    if days_of_supply > Decimal(120):
        return "overstock_watch"
    return "healthy"


def _date_set(start_date: date, end_date: date) -> set[date]:
    return {
        start_date + timedelta(days=offset) for offset in range((end_date - start_date).days + 1)
    }


def _validate_week_start(week_start: date) -> None:
    if week_start.weekday() not in {0, 5}:
        raise ValueError("week_start must be a Monday or Saturday")


def parse_week_start(value: str) -> date:
    try:
        week_start = date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError("week_start must use YYYY-MM-DD format") from exc
    _validate_week_start(week_start)
    return week_start


def _int_metric(values: Mapping[str, Any], key: str) -> int:
    return int(values.get(key) or 0)


def _to_decimal(value: Any) -> Decimal:
    if value is None:
        return ZERO
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return ZERO


def _optional_int(value: Any) -> int | None:
    if value is None:
        return None
    return int(value)


def _as_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _empty_to_none(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _first_non_empty(values: Iterable[str | None]) -> str | None:
    for value in values:
        if value:
            return value
    return None


def _money(value: Decimal) -> Decimal:
    return value.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def _ratio(value: Decimal) -> Decimal:
    return value.quantize(RATIO_QUANT, rounding=ROUND_HALF_UP)


def _safe_ratio(numerator: Decimal, denominator: Decimal) -> Decimal | None:
    if denominator == ZERO:
        return None
    return _ratio(numerator / denominator)


def _safe_money_ratio(numerator: Decimal, denominator: Decimal) -> Decimal | None:
    if denominator == ZERO:
        return None
    return _money(numerator / denominator)


def _decimal_to_string(value: Decimal) -> str:
    return format(_money(value), "f")


def _optional_decimal_to_string(value: Decimal | None) -> str | None:
    return None if value is None else _decimal_to_string(value)


def _optional_ratio_to_string(value: Decimal | None) -> str | None:
    return None if value is None else format(_ratio(value), "f")


def _format_money(value: Decimal, currency: str | None) -> str:
    prefix = f"{currency} " if currency else ""
    return f"{prefix}{_decimal_to_string(value)}"


def _format_ratio(value: Decimal | None) -> str:
    if value is None:
        return "-"
    return f"{(_ratio(value) * Decimal('100')).quantize(MONEY_QUANT)}%"


def _json_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return _decimal_to_string(value)
    if isinstance(value, date | datetime):
        return value.isoformat()
    if isinstance(value, dict):
        return _json_safe_mapping(value)
    if isinstance(value, list):
        return [_json_value(item) for item in value]
    return value


def _json_safe_mapping(values: Mapping[str, Any]) -> dict[str, Any]:
    return {key: _json_value(value) for key, value in values.items()}


def _xlsx_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, date | datetime):
        return value.isoformat()
    return value


__all__ = [
    "DEFAULT_OUTPUT_ROOT",
    "WeeklyBusinessReviewService",
    "WeeklyBusinessReviewResult",
    "build_weekly_business_review_workbook",
    "parse_week_start",
]
