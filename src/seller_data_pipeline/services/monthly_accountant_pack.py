from __future__ import annotations

from collections.abc import Mapping, Sequence
from datetime import date, datetime
from decimal import Decimal
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

if TYPE_CHECKING:
    from seller_data_pipeline.services.monthly_financial_close_service import (
        MonthlyFinancialCloseResult,
    )

ACCOUNTANT_TABLE_START_ROW = 7
ACCOUNTANT_DATA_START_ROW = ACCOUNTANT_TABLE_START_ROW + 1

SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="305496")
DESCRIPTION_KEY_FILL = PatternFill("solid", fgColor="E2F0D9")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")

ACCOUNTANT_SHEET_NOTES: dict[str, tuple[tuple[str, str], ...]] = {
    "09_Accounting_Summary": (
        (
            "Sheet Purpose / 本表用途",
            "按会计做账视角汇总本月 Amazon 经营数据，将 Settlement、广告费、FBA 费、"
            "退款、赔偿和内部 COGS 转换为会计可理解的项目。",
        ),
        (
            "How to Use / 使用方式",
            "会计可先检查每个会计项目的原币金额、建议科目、汇率和人民币金额，再决定"
            "是否直接入账或调整科目。黄色/待确认字段需要会计确认。",
        ),
        (
            "Main Data Sources / 主要数据来源",
            "Amazon Settlement Flat File V2 / Payments、amazon_sku_cost、Ads API 辅助指标、"
            "FBA Reimbursements 辅助指标。",
        ),
        (
            "Accounting Caveats / 会计注意事项",
            "本表为做账辅助底稿，不替代正式会计判断；Estimated Operating Profit 是管理"
            "口径利润，不等同于企业所得税应纳税所得额。",
        ),
        (
            "Required Accountant Confirmation / 需要会计确认",
            "收入是否含税/不含税、汇率口径、建议科目、发票/凭证是否齐全、COGS 是否可税前扣除。",
        ),
    ),
    "10_Journal_Entries": (
        (
            "Sheet Purpose / 本表用途",
            "提供月度 Amazon 业务的建议记账分录，减少会计手工拆分 Amazon 结算单的工作量。",
        ),
        (
            "How to Use / 使用方式",
            "会计逐行检查借方科目、贷方科目、原币金额、汇率、人民币金额和附件索引，"
            "确认后录入财务软件。",
        ),
        (
            "Main Data Sources / 主要数据来源",
            "09_Accounting_Summary、Amazon Settlement、SKU 成本、银行/万里汇到账核对、人工调整项。",
        ),
        (
            "Accounting Caveats / 会计注意事项",
            "本表是建议分录，不是强制分录；收入、成本、费用、汇兑损益和银行到账的最终"
            "分录方式由会计决定。",
        ),
        (
            "Required Accountant Confirmation / 需要会计确认",
            "凭证日期、摘要、借贷科目、汇率、附件、是否需要拆分多张凭证。",
        ),
    ),
    "11_Quarter_Rollup": (
        (
            "Sheet Purpose / 本表用途",
            "按季度汇总每月 Amazon 收入、退款、费用、成本、利润和待确认项，帮助会计做季度"
            "报税和内部对账准备。",
        ),
        (
            "How to Use / 使用方式",
            "第一版在单月月报中展示当季截至当前月份的可累计字段；会计可后续合并三个月月报复核季度累计。",
        ),
        (
            "Main Data Sources / 主要数据来源",
            "本月月报、同季度其他月份月报、09_Accounting_Summary、15_Adjustments。",
        ),
        (
            "Accounting Caveats / 会计注意事项",
            "季度汇总是申报辅助表，不等于税务申报表；季度申报口径需结合公司其他非 Amazon"
            "收入费用、发票、税务政策和会计账簿。",
        ),
        (
            "Required Accountant Confirmation / 需要会计确认",
            "季度所属月份是否齐全、汇率是否一致、是否存在跨月调整、非 Amazon 费用是否已纳入账簿。",
        ),
    ),
    "12_FX_Rates": (
        (
            "Sheet Purpose / 本表用途",
            "记录本月用于会计入账和季度汇总的汇率口径，驱动会计辅助表中的人民币金额换算。",
        ),
        (
            "How to Use / 使用方式",
            "会计填写或确认本月 USD/CNY 汇率；其他会计用表的 CNY 金额会按本表汇率计算。",
        ),
        (
            "Main Data Sources / 主要数据来源",
            "会计确认的记账汇率、中国人民银行/银行结汇单/公司会计政策，系统不自动替代会计判断。",
        ),
        (
            "Accounting Caveats / 会计注意事项",
            "不同税种、会计事项或实际结汇可能使用不同汇率；系统默认留空，最终以会计确认口径为准。",
        ),
        (
            "Required Accountant Confirmation / 需要会计确认",
            "汇率日期、汇率来源、适用范围、是否使用月末汇率/月平均汇率/实际结汇汇率。",
        ),
    ),
    "13_Source_Doc_Index": (
        (
            "Sheet Purpose / 本表用途",
            "列出会计做账和税务复核可能需要的 Amazon、银行、采购、物流和内部成本凭证，"
            "减少遗漏附件。",
        ),
        (
            "How to Use / 使用方式",
            "会计或运营逐项勾选凭证是否已提供，并用 Source Reference 查找原始文件或下载路径。",
        ),
        (
            "Main Data Sources / 主要数据来源",
            "pipeline_artifact_store raw reports、Amazon Settlement report、Amazon fee invoice、"
            "万里汇/银行流水、采购发票、物流发票、SKU 成本表。",
        ),
        (
            "Accounting Caveats / 会计注意事项",
            "系统能保存 Amazon raw report，但不一定自动获取所有发票、银行流水和采购/物流凭证；"
            "缺失凭证需要人工补充。",
        ),
        (
            "Required Accountant Confirmation / 需要会计确认",
            "附件是否满足入账和税务留存要求、凭证金额是否和做账金额一致。",
        ),
    ),
    "14_Payout_Recon": (
        (
            "Sheet Purpose / 本表用途",
            "核对 Amazon Settlement 净额与万里汇/银行实际到账，帮助会计确认应收账款、其他货币资金、"
            "银行存款和汇兑损益。",
        ),
        (
            "How to Use / 使用方式",
            "运营或会计填入实际到账流水，系统计算与 Amazon settlement net 的差异，"
            "并标记是否为手续费、汇率差、跨月未到账或待查差异。",
        ),
        (
            "Main Data Sources / 主要数据来源",
            "Amazon Settlement、Amazon Payments、WorldFirst/银行流水、12_FX_Rates。",
        ),
        (
            "Accounting Caveats / 会计注意事项",
            "Amazon settlement 日期、付款日期、万里汇到账日期和银行入账日期可能跨月；"
            "差异不一定是错误，需按会计政策处理。",
        ),
        (
            "Required Accountant Confirmation / 需要会计确认",
            "回款日期、到账账户、手续费、汇率、汇兑损益、跨月应收余额。",
        ),
    ),
    "15_Adjustments": (
        (
            "Sheet Purpose / 本表用途",
            "记录会计对系统月报的手工调整，保证最终入账口径与管理口径差异可追溯。",
        ),
        (
            "How to Use / 使用方式",
            "会计逐项填写调整原因、金额、科目、附件和确认状态；"
            "季度汇总时将已确认调整纳入季度 rollup。",
        ),
        (
            "Main Data Sources / 主要数据来源",
            "会计账簿、发票、银行流水、库存记录、采购付款记录、税务调整底稿。",
        ),
        (
            "Accounting Caveats / 会计注意事项",
            "系统不会自动判断所有税务和会计调整；所有手工调整必须保留原因和附件，避免后续无法追溯。",
        ),
        (
            "Required Accountant Confirmation / 需要会计确认",
            "调整金额、科目、汇率、附件、是否影响当月或跨月、是否影响季度申报。",
        ),
    ),
}

ACCOUNTING_SUMMARY_COLUMNS: tuple[tuple[str, str], ...] = (
    ("line_no", "Line No. / 行号"),
    ("accounting_item", "Accounting Item / 会计项目"),
    ("suggested_account", "Suggested Account / 建议会计科目"),
    ("statement_direction", "Statement Direction / 报表方向"),
    ("debit_or_credit", "Debit or Credit / 借贷方向"),
    ("amount_usd", "Amount USD / 美元金额"),
    ("currency", "Currency / 币种"),
    ("fx_rate", "FX Rate / 汇率"),
    ("amount_cny", "Amount CNY / 人民币金额"),
    ("source_bucket", "Source Bucket / 来源利润桶"),
    ("source_category", "Source Category / 来源金额类别"),
    ("source_sheet", "Source Sheet / 来源Sheet"),
    ("source_reference", "Source Reference / 来源引用"),
    ("auto_generated", "Auto Generated / 是否系统生成"),
    ("needs_accountant_review", "Needs Accountant Review / 是否需要会计复核"),
    ("accountant_confirmation", "Accountant Confirmation / 会计确认"),
    ("notes", "Notes / 说明"),
)

JOURNAL_ENTRY_COLUMNS: tuple[tuple[str, str], ...] = (
    ("voucher_date", "Voucher Date / 凭证日期"),
    ("voucher_group", "Voucher Group / 凭证组"),
    ("voucher_line_no", "Voucher Line No. / 凭证行号"),
    ("entry_type", "Entry Type / 分录类型"),
    ("summary", "Summary / 摘要"),
    ("debit_account", "Debit Account / 借方科目"),
    ("credit_account", "Credit Account / 贷方科目"),
    ("currency", "Currency / 币种"),
    ("amount_original", "Amount Original / 原币金额"),
    ("fx_rate", "FX Rate / 汇率"),
    ("amount_cny", "Amount CNY / 人民币金额"),
    ("source_sheet", "Source Sheet / 来源Sheet"),
    ("source_line_no", "Source Line No. / 来源行号"),
    ("source_document_id", "Source Document ID / 来源凭证ID"),
    ("attachment_required", "Attachment Required / 是否需要附件"),
    ("accountant_confirmation", "Accountant Confirmation / 会计确认"),
    ("notes", "Notes / 说明"),
)

QUARTER_ROLLUP_COLUMNS: tuple[tuple[str, str], ...] = (
    ("quarter", "Quarter / 季度"),
    ("month", "Month / 月份"),
    ("accounting_item", "Accounting Item / 会计项目"),
    ("suggested_account", "Suggested Account / 建议会计科目"),
    ("amount_usd", "Amount USD / 美元金额"),
    ("currency", "Currency / 币种"),
    ("fx_rate", "FX Rate / 汇率"),
    ("amount_cny", "Amount CNY / 人民币金额"),
    ("quarter_to_date_amount_cny", "Quarter-to-Date Amount CNY / 季度累计人民币金额"),
    ("source_report", "Source Report / 来源月报"),
    ("status", "Status / 状态"),
    ("needs_accountant_review", "Needs Accountant Review / 是否需要会计复核"),
    ("notes", "Notes / 说明"),
)

FX_RATE_COLUMNS: tuple[tuple[str, str], ...] = (
    ("currency_pair", "Currency Pair / 货币对"),
    ("period", "Period / 期间"),
    ("fx_rate_type", "FX Rate Type / 汇率类型"),
    ("fx_rate", "FX Rate / 汇率"),
    ("rate_date", "Rate Date / 汇率日期"),
    ("rate_source", "Rate Source / 汇率来源"),
    ("applies_to", "Applies To / 适用范围"),
    ("accountant_confirmation", "Accountant Confirmation / 会计确认"),
    ("notes", "Notes / 说明"),
)

SOURCE_DOC_COLUMNS: tuple[tuple[str, str], ...] = (
    ("document_type", "Document Type / 凭证类型"),
    ("period", "Period / 期间"),
    ("document_name", "Document Name / 文件名"),
    ("source_system", "Source System / 来源系统"),
    ("source_reference", "Source Reference / 来源引用"),
    ("related_amount_original", "Related Amount Original / 相关原币金额"),
    ("currency", "Currency / 币种"),
    ("related_amount_cny", "Related Amount CNY / 相关人民币金额"),
    ("provided_to_accountant", "Provided to Accountant / 是否已提供给会计"),
    ("required_for_booking", "Required for Booking / 是否做账必需"),
    ("required_for_tax_filing", "Required for Tax Filing / 是否报税必需"),
    ("status", "Status / 状态"),
    ("owner", "Owner / 负责人"),
    ("notes", "Notes / 说明"),
)

PAYOUT_RECON_COLUMNS: tuple[tuple[str, str], ...] = (
    ("settlement_id", "Settlement ID / 结算单ID"),
    ("settlement_period", "Settlement Period / 结算期间"),
    ("amazon_net_amount", "Amazon Net Amount / 亚马逊净结算金额"),
    ("amazon_currency", "Amazon Currency / 亚马逊币种"),
    ("expected_payout_date", "Expected Payout Date / 预计打款日期"),
    ("payout_account", "Payout Account / 到账账户"),
    ("actual_received_amount", "Actual Received Amount / 实际到账金额"),
    ("received_currency", "Received Currency / 到账币种"),
    ("bank_or_worldfirst_fee", "Bank or WorldFirst Fee / 银行或万里汇手续费"),
    ("fx_rate", "FX Rate / 汇率"),
    ("received_amount_cny", "Received Amount CNY / 到账人民币金额"),
    ("difference_original", "Difference Original / 原币差异"),
    ("difference_cny", "Difference CNY / 人民币差异"),
    ("difference_reason", "Difference Reason / 差异原因"),
    ("reconciliation_status", "Reconciliation Status / 核对状态"),
    ("accountant_confirmation", "Accountant Confirmation / 会计确认"),
    ("notes", "Notes / 说明"),
)

ADJUSTMENT_COLUMNS: tuple[tuple[str, str], ...] = (
    ("adjustment_id", "Adjustment ID / 调整ID"),
    ("adjustment_type", "Adjustment Type / 调整类型"),
    ("accounting_item", "Accounting Item / 会计项目"),
    ("suggested_account", "Suggested Account / 建议会计科目"),
    ("debit_or_credit", "Debit or Credit / 借贷方向"),
    ("currency", "Currency / 币种"),
    ("amount_original", "Amount Original / 原币金额"),
    ("fx_rate", "FX Rate / 汇率"),
    ("amount_cny", "Amount CNY / 人民币金额"),
    ("affects_month", "Affects Month / 影响月份"),
    ("affects_quarter", "Affects Quarter / 影响季度"),
    ("reason", "Reason / 调整原因"),
    ("source_document_id", "Source Document ID / 来源凭证ID"),
    ("accountant_confirmation", "Accountant Confirmation / 会计确认"),
    ("status", "Status / 状态"),
    ("notes", "Notes / 说明"),
)


def build_accountant_pack_payload(result: MonthlyFinancialCloseResult) -> dict[str, Any]:
    summary_rows = _build_accounting_summary_rows(result)
    return {
        "bookkeeping_summary": summary_rows,
        "suggested_journal_entries": _build_journal_entry_rows(result, summary_rows),
        "quarter_rollup": _build_quarter_rollup_rows(result, summary_rows),
        "fx_rates": _build_fx_rate_rows(result),
        "source_document_index": _build_source_doc_rows(result),
        "payout_reconciliation": _build_payout_recon_rows(result),
        "manual_adjustments": _build_adjustment_rows(result),
    }


def add_accountant_pack_sheets(workbook: Workbook, result: MonthlyFinancialCloseResult) -> None:
    pack = build_accountant_pack_payload(result)
    _write_accountant_sheet(
        workbook,
        "09_Accounting_Summary",
        ACCOUNTING_SUMMARY_COLUMNS,
        pack["bookkeeping_summary"],
        formula_config={
            "fx_rate": "default_fx_rate",
            "amount_cny": "amount_usd_times_fx_rate",
        },
    )
    _write_accountant_sheet(
        workbook,
        "10_Journal_Entries",
        JOURNAL_ENTRY_COLUMNS,
        pack["suggested_journal_entries"],
        formula_config={
            "fx_rate": "default_fx_rate",
            "amount_cny": "amount_original_times_fx_rate",
        },
    )
    _write_accountant_sheet(
        workbook,
        "11_Quarter_Rollup",
        QUARTER_ROLLUP_COLUMNS,
        pack["quarter_rollup"],
        formula_config={
            "fx_rate": "default_fx_rate",
            "amount_cny": "amount_usd_times_fx_rate",
            "quarter_to_date_amount_cny": "same_row_amount_cny",
        },
    )
    _write_accountant_sheet(workbook, "12_FX_Rates", FX_RATE_COLUMNS, pack["fx_rates"])
    _write_accountant_sheet(
        workbook,
        "13_Source_Doc_Index",
        SOURCE_DOC_COLUMNS,
        pack["source_document_index"],
        formula_config={"related_amount_cny": "related_amount_original_times_default_fx_rate"},
    )
    _write_accountant_sheet(
        workbook,
        "14_Payout_Recon",
        PAYOUT_RECON_COLUMNS,
        pack["payout_reconciliation"],
        formula_config={
            "fx_rate": "default_fx_rate",
            "received_amount_cny": "actual_received_amount_times_fx_rate",
            "difference_original": "actual_received_minus_amazon_net",
            "difference_cny": "difference_original_times_fx_rate",
        },
    )
    _write_accountant_sheet(
        workbook,
        "15_Adjustments",
        ADJUSTMENT_COLUMNS,
        pack["manual_adjustments"],
        formula_config={
            "fx_rate": "default_fx_rate",
            "amount_cny": "amount_original_times_fx_rate",
        },
    )


def _build_accounting_summary_rows(result: MonthlyFinancialCloseResult) -> list[dict[str, Any]]:
    fs = result.financial_summary
    currency = result.currency or "USD"
    rows = [
        _accounting_summary_row(
            1,
            "Product Sales Revenue / 商品销售收入",
            "Main Business Revenue / 主营业务收入",
            "Revenue / 收入",
            "Credit / 贷方",
            fs.product_sales_amount,
            currency,
            "revenue",
            "product_sales; liquidation_revenue",
            "04_Settlement_Buckets; 05_Amount_Categories",
            "Settlement product-sales categories",
            needs_review=True,
            notes="会计确认收入是否按不含税口径入账。",
        ),
        _accounting_summary_row(
            2,
            "Refunds and Sales Returns / 退款及销售退回",
            "Sales Returns or Sales Allowance / 主营业务收入-销售退回或销售折让",
            "Revenue Reduction / 收入冲减",
            "Debit / 借方",
            abs(fs.refund),
            currency,
            "refund",
            "refund_revenue; refund_fee",
            "04_Settlement_Buckets; 05_Amount_Categories",
            "Settlement refund bucket",
            needs_review=True,
            notes="可作为收入冲减或单独销售退回科目，最终由会计确认。",
        ),
        _accounting_summary_row(
            3,
            "Referral and Platform Fees / 平台佣金及服务费",
            "Selling Expenses - Platform Fees / 销售费用-平台服务费",
            "Expense / 费用",
            "Debit / 借方",
            abs(fs.amazon_fee),
            currency,
            "amazon_fee",
            "referral_fee; subscription_fee; other_platform_fee",
            "04_Settlement_Buckets; 05_Amount_Categories",
            "Settlement Amazon fee bucket",
            needs_review=True,
            notes="需配合 Amazon fee invoice 或结算单。",
        ),
        _accounting_summary_row(
            4,
            "FBA Fulfillment Fees / FBA配送履约费",
            "Selling Expenses - Fulfillment Fees / 销售费用-FBA履约费",
            "Expense / 费用",
            "Debit / 借方",
            abs(fs.fba_fee),
            currency,
            "fba_fee",
            "fba_fulfillment_fee; fba_storage_fee; inbound_fee",
            "04_Settlement_Buckets; 05_Amount_Categories",
            "Settlement FBA bucket",
            needs_review=True,
            notes="Amazon 代履约、仓储或入仓相关费用，可按公司科目细分。",
        ),
        _accounting_summary_row(
            5,
            "Advertising Fees / 广告费",
            "Selling Expenses - Advertising / 销售费用-广告费",
            "Expense / 费用",
            "Debit / 借方",
            abs(fs.advertising_cost),
            currency,
            "advertising_cost",
            "advertising_fee",
            "04_Settlement_Buckets; 07_Operational_Context",
            "Settlement advertising fee; Ads API context only",
            needs_review=True,
            notes="财务扣费以 Settlement 为主，Ads API 仅用于投放解释。",
        ),
        _accounting_summary_row(
            6,
            "Promotion Discounts and Fees / 促销折扣及活动费用",
            "Selling Expenses - Promotion / 销售费用-促销费",
            "Expense / 费用",
            "Debit / 借方",
            abs(fs.promotion_cost) + abs(fs.promotion_fee),
            currency,
            "promotion_cost; promotion_fee",
            "promotion_discount; coupon_fee; promotion_fee",
            "04_Settlement_Buckets; 07_Operational_Context",
            "Settlement promotion buckets",
            needs_review=True,
            notes="Coupon/Promotion performance 只做解释，不替代实扣。",
        ),
        _accounting_summary_row(
            7,
            "Reimbursements / 亚马逊赔偿",
            "Other Income or Expense Offset / 其他收益或费用冲减",
            "Income or Offset / 收益或冲减",
            "Credit / 贷方" if fs.reimbursement >= 0 else "Debit / 借方",
            abs(fs.reimbursement),
            currency,
            "reimbursement",
            "reimbursement",
            "04_Settlement_Buckets; 07_Operational_Context",
            "Settlement reimbursement bucket",
            needs_review=True,
            notes="赔偿类科目需会计确认，可计入其他收益或冲减相关成本费用。",
        ),
        _accounting_summary_row(
            8,
            "Total Landed COGS / 到岸COGS合计",
            "Cost of Goods Sold / 主营业务成本",
            "Cost / 成本",
            "Debit / 借方",
            fs.internal_cogs,
            currency,
            "internal_cogs",
            "amazon_sku_cost",
            "06_SKU_Profit",
            "amazon_sku_cost × product sales units",
            needs_review=True,
            notes="管理成本口径，需会计结合采购发票、库存结转、头程物流和税前扣除要求确认。",
        ),
        _accounting_summary_row(
            9,
            "Marketplace Facilitator Tax / 平台代收代缴税费",
            "Pass-through Tax or Memo Only / 代收代缴税费或备查",
            "Informational / 备查",
            "Memo / 备查",
            _tax_passthrough_amount(result),
            currency,
            "tax_passthrough",
            "sales_tax; marketplace_facilitator_tax",
            "04_Settlement_Buckets; 05_Amount_Categories",
            "Settlement tax passthrough categories",
            needs_review=True,
            notes="通常不作为公司收入或费用；最终展示和入账方式由会计确认。",
        ),
        _accounting_summary_row(
            10,
            "Settlement Close Profit / Settlement结算口径利润",
            "Management KPI Only / 管理口径指标",
            "Informational / 备查",
            "Memo / 备查",
            fs.settlement_led_estimated_profit,
            currency,
            "derived",
            "settlement_net_minus_landed_cogs",
            "01_Summary",
            "Settlement net amount minus total landed COGS",
            needs_review=True,
            notes=(
                "Settlement posted-date 月结参考利润，不等同于经营发生口径利润、"
                "企业所得税应纳税所得额或法定利润。"
            ),
        ),
    ]
    return rows


def _accounting_summary_row(
    line_no: int,
    accounting_item: str,
    suggested_account: str,
    statement_direction: str,
    debit_or_credit: str,
    amount: Decimal,
    currency: str,
    source_bucket: str,
    source_category: str,
    source_sheet: str,
    source_reference: str,
    *,
    needs_review: bool,
    notes: str,
) -> dict[str, Any]:
    return {
        "line_no": line_no,
        "accounting_item": accounting_item,
        "suggested_account": suggested_account,
        "statement_direction": statement_direction,
        "debit_or_credit": debit_or_credit,
        "amount_usd": _money_float(amount),
        "currency": currency,
        "fx_rate": None,
        "amount_cny": None,
        "source_bucket": source_bucket,
        "source_category": source_category,
        "source_sheet": source_sheet,
        "source_reference": source_reference,
        "auto_generated": "Yes / 是",
        "needs_accountant_review": "Yes / 是" if needs_review else "No / 否",
        "accountant_confirmation": "Pending / 待确认",
        "notes": notes,
    }


def _build_journal_entry_rows(
    result: MonthlyFinancialCloseResult,
    summary_rows: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    currency = result.currency or "USD"
    voucher_date = result.end_date.isoformat()

    def summary_amount(line_no: int) -> Any:
        for row in summary_rows:
            if row.get("line_no") == line_no:
                return row.get("amount_usd")
        return None

    def add(
        group: str,
        line_no: int,
        entry_type: str,
        summary: str,
        debit_account: str,
        credit_account: str,
        amount: Any,
        source_line_no: int,
        notes: str,
        *,
        attachment_required: str = "Yes / 是",
    ) -> None:
        rows.append(
            {
                "voucher_date": voucher_date,
                "voucher_group": group,
                "voucher_line_no": line_no,
                "entry_type": entry_type,
                "summary": summary,
                "debit_account": debit_account,
                "credit_account": credit_account,
                "currency": currency,
                "amount_original": amount,
                "fx_rate": None,
                "amount_cny": None,
                "source_sheet": "09_Accounting_Summary",
                "source_line_no": source_line_no,
                "source_document_id": _settlement_reference(result),
                "attachment_required": attachment_required,
                "accountant_confirmation": "Pending / 待确认",
                "notes": notes,
            }
        )

    add(
        "Revenue recognition / 确认销售收入与退款冲减",
        1,
        "Revenue / 收入",
        (
            f"Recognize Amazon product sales for {result.month} / "
            f"确认 {result.month} Amazon 商品销售收入"
        ),
        "Accounts Receivable - Amazon / 应收账款-Amazon",
        "Main Business Revenue / 主营业务收入",
        summary_amount(1),
        1,
        "收入是否按不含税口径入账需会计确认。",
    )
    add(
        "Revenue recognition / 确认销售收入与退款冲减",
        2,
        "Sales return / 销售退回",
        f"Recognize Amazon refunds for {result.month} / 确认 {result.month} Amazon 退款及销售退回",
        "Sales Returns or Sales Allowance / 主营业务收入-销售退回或销售折让",
        "Accounts Receivable - Amazon / 应收账款-Amazon",
        summary_amount(2),
        2,
        "可作为收入冲减或单独销售退回科目，最终由会计确认。",
    )
    add(
        "Amazon fees recognition / 确认 Amazon 费用",
        3,
        "Platform fee / 平台费",
        f"Recognize Amazon platform fees for {result.month} / 确认 {result.month} Amazon 平台费用",
        "Selling Expenses - Platform Fees / 销售费用-平台服务费",
        "Accounts Receivable - Amazon / 应收账款-Amazon",
        summary_amount(3),
        3,
        "需配 Amazon fee invoice 或 Settlement 作为附件。",
    )
    add(
        "Amazon fees recognition / 确认 Amazon 费用",
        4,
        "FBA fee / FBA费用",
        f"Recognize FBA fulfillment fees for {result.month} / 确认 {result.month} FBA 履约费用",
        "Selling Expenses - Fulfillment Fees / 销售费用-FBA履约费",
        "Accounts Receivable - Amazon / 应收账款-Amazon",
        summary_amount(4),
        4,
        "Amazon 代履约、仓储或入仓相关费用，可按公司科目细分。",
    )
    add(
        "Amazon fees recognition / 确认 Amazon 费用",
        5,
        "Advertising fee / 广告费",
        f"Recognize Amazon advertising fees for {result.month} / 确认 {result.month} Amazon 广告费",
        "Selling Expenses - Advertising / 销售费用-广告费",
        "Accounts Receivable - Amazon / 应收账款-Amazon",
        summary_amount(5),
        5,
        "财务扣费以 Settlement 为主，Ads API 仅用于投放解释。",
    )
    add(
        "Amazon fees recognition / 确认 Amazon 费用",
        6,
        "Promotion fee / 促销费",
        f"Recognize Amazon promotion fees for {result.month} / 确认 {result.month} Amazon 促销费用",
        "Selling Expenses - Promotion / 销售费用-促销费",
        "Accounts Receivable - Amazon / 应收账款-Amazon",
        summary_amount(6),
        6,
        "Coupon/Promotion performance 只做解释，不替代实扣。",
    )
    add(
        "COGS recognition / 确认主营业务成本",
        7,
        "COGS / 主营业务成本",
        f"Recognize internal COGS for {result.month} / 确认 {result.month} 内部商品成本",
        "Cost of Goods Sold / 主营业务成本",
        "Inventory Goods / 库存商品",
        summary_amount(8),
        8,
        "需结合采购发票、库存结转和头程物流成本分摊确认。",
    )
    add(
        "Reimbursement recognition / 确认赔偿",
        8,
        "Reimbursement / 赔偿",
        f"Recognize Amazon reimbursements for {result.month} / 确认 {result.month} Amazon 赔偿",
        "Accounts Receivable - Amazon / 应收账款-Amazon",
        "Other Income or Expense Offset / 其他收益或费用冲减",
        summary_amount(7),
        7,
        "赔偿类科目需会计确认。",
    )
    add(
        "Payout clearing / 回款清账",
        9,
        "Payout clearing / 回款清账",
        f"Clear Amazon payout for {result.month} / 核对并清账 {result.month} Amazon 回款",
        "Bank Deposit or WorldFirst / 银行存款或万里汇",
        "Accounts Receivable - Amazon / 应收账款-Amazon",
        _money_float(result.financial_summary.settlement_net_amount),
        0,
        "实际到账金额、手续费和汇率需在 14_Payout_Recon 中确认。",
    )
    return rows


def _build_quarter_rollup_rows(
    result: MonthlyFinancialCloseResult,
    summary_rows: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    source_report = f"monthly_financial_close_{result.month}.xlsx"
    quarter = _quarter_label(result.start_date)
    rows = []
    for row in summary_rows:
        rows.append(
            {
                "quarter": quarter,
                "month": result.month,
                "accounting_item": row["accounting_item"],
                "suggested_account": row["suggested_account"],
                "amount_usd": row["amount_usd"],
                "currency": row["currency"],
                "fx_rate": None,
                "amount_cny": None,
                "quarter_to_date_amount_cny": None,
                "source_report": source_report,
                "status": "Current month only / 当前仅本月",
                "needs_accountant_review": row["needs_accountant_review"],
                "notes": "第一版仅展示本月数据；季度包可后续合并三个月月报。",
            }
        )
    return rows


def _build_fx_rate_rows(result: MonthlyFinancialCloseResult) -> list[dict[str, Any]]:
    currency = result.currency or "USD"
    return [
        {
            "currency_pair": f"{currency}/CNY",
            "period": result.month,
            "fx_rate_type": "Accountant Manual Rate / 会计手工确认汇率",
            "fx_rate": None,
            "rate_date": "",
            "rate_source": "Accountant input required / 需要会计填写",
            "applies_to": "Monthly accountant pack / 月度会计做账包",
            "accountant_confirmation": "Pending / 待确认",
            "notes": "请在 FX Rate / 汇率 列填写本月记账汇率；其他会计 sheet 会自动换算 CNY。",
        }
    ]


def _build_source_doc_rows(result: MonthlyFinancialCloseResult) -> list[dict[str, Any]]:
    currency = result.currency or "USD"
    settlement_reference = _settlement_reference(result)
    raw_paths = result.raw_metadata.get("settlement_source_raw_file_paths") or []
    document_name = "; ".join(raw_paths) if raw_paths else settlement_reference
    rows = [
        _source_doc_row(
            "Amazon Settlement Flat File V2 / 亚马逊结算明细",
            result.month,
            document_name,
            "Amazon SP-API / Seller Central",
            settlement_reference,
            result.financial_summary.settlement_net_amount,
            currency,
            "Yes / 是",
            "Yes / 是",
            "Yes / 是",
            "System saved / 系统已保存",
            "Operations / 运营",
            "财务主口径，需要作为做账和审计底稿留存。",
        ),
        _source_doc_row(
            "Amazon Seller Fee Invoice / 亚马逊服务费发票",
            result.month,
            "To be downloaded manually / 待人工下载",
            "Seller Central",
            "",
            None,
            currency,
            "No / 否",
            "Yes / 是",
            "Yes / 是",
            "Pending / 待补充",
            "Operations or Accountant / 运营或会计",
            "用于平台佣金、FBA费用等费用凭证。",
        ),
        _source_doc_row(
            "Amazon Advertising Invoice or Statement / 亚马逊广告费账单",
            result.month,
            "To be downloaded manually / 待人工下载",
            "Amazon Ads / Seller Central",
            "07_Operational_Context",
            abs(result.financial_summary.advertising_cost),
            currency,
            "No / 否",
            "Yes / 是",
            "Yes / 是",
            "Pending / 待补充",
            "Operations / 运营",
            "财务扣费以 Settlement 为主，账单用于凭证留存。",
        ),
        _source_doc_row(
            "WorldFirst Payout Statement / 万里汇到账流水",
            result.month,
            "To be uploaded manually / 待人工上传",
            "WorldFirst / Bank",
            "14_Payout_Recon",
            result.financial_summary.settlement_net_amount,
            currency,
            "No / 否",
            "Yes / 是",
            "Yes / 是",
            "Pending / 待补充",
            "Operations or Accountant / 运营或会计",
            "用于 Amazon settlement 与实际到账核对。",
        ),
        _source_doc_row(
            "Supplier Purchase Invoice / 供应商采购发票",
            result.month,
            "To be uploaded manually / 待人工上传",
            "Supplier / 供应商",
            "06_SKU_Profit; amazon_sku_cost",
            result.financial_summary.internal_cogs,
            currency,
            "No / 否",
            "Yes / 是",
            "Yes / 是",
            "Pending / 待补充",
            "Operations or Accountant / 运营或会计",
            "用于确认 COGS 税前扣除和库存成本。",
        ),
        _source_doc_row(
            "First-Mile Freight Invoice / 头程物流发票",
            result.month,
            "To be uploaded manually / 待人工上传",
            "Freight Forwarder / 货代",
            "amazon_sku_cost first_mile_cost",
            result.financial_summary.first_mile_cogs,
            currency,
            "No / 否",
            "Yes / 是",
            "Yes / 是",
            "Pending / 待补充",
            "Operations or Accountant / 运营或会计",
            "金额为月报按已售件数确认的头程成本；仍需货代发票/付款资料支持会计归档。",
        ),
        _source_doc_row(
            "SKU Cost Sheet / SKU成本表",
            result.month,
            "amazon_sku_cost",
            "SellerDataPipeline Azure SQL",
            "06_SKU_Profit",
            result.financial_summary.internal_cogs,
            currency,
            "Yes / 是",
            "Yes / 是",
            "No / 否",
            "System generated / 系统生成",
            "Operations / 运营",
            "管理成本口径，仍需发票和库存资料支持。",
        ),
    ]
    return rows


def _source_doc_row(
    document_type: str,
    period: str,
    document_name: str,
    source_system: str,
    source_reference: str,
    related_amount: Decimal | None,
    currency: str,
    provided: str,
    booking_required: str,
    tax_required: str,
    status: str,
    owner: str,
    notes: str,
) -> dict[str, Any]:
    return {
        "document_type": document_type,
        "period": period,
        "document_name": document_name,
        "source_system": source_system,
        "source_reference": source_reference,
        "related_amount_original": None if related_amount is None else _money_float(related_amount),
        "currency": currency,
        "related_amount_cny": None,
        "provided_to_accountant": provided,
        "required_for_booking": booking_required,
        "required_for_tax_filing": tax_required,
        "status": status,
        "owner": owner,
        "notes": notes,
    }


def _build_payout_recon_rows(result: MonthlyFinancialCloseResult) -> list[dict[str, Any]]:
    currency = result.currency or "USD"
    settlement_ids = result.raw_metadata.get("settlement_ids") or []
    settlement_id = (
        "; ".join(str(value) for value in settlement_ids)
        or "All settlements / 全部结算单"
    )
    return [
        {
            "settlement_id": settlement_id,
            "settlement_period": f"{result.start_date.isoformat()}..{result.end_date.isoformat()}",
            "amazon_net_amount": _money_float(result.financial_summary.settlement_net_amount),
            "amazon_currency": currency,
            "expected_payout_date": "",
            "payout_account": "WorldFirst or bank / 万里汇或银行",
            "actual_received_amount": None,
            "received_currency": currency,
            "bank_or_worldfirst_fee": None,
            "fx_rate": None,
            "received_amount_cny": None,
            "difference_original": None,
            "difference_cny": None,
            "difference_reason": "Pending Payout / 尚未核对到账",
            "reconciliation_status": "Pending / 待核对",
            "accountant_confirmation": "Pending / 待确认",
            "notes": "请填入实际到账金额、到账日期、手续费和账户，系统将计算差异。",
        }
    ]


def _build_adjustment_rows(result: MonthlyFinancialCloseResult) -> list[dict[str, Any]]:
    currency = result.currency or "USD"
    quarter = _quarter_label(result.start_date)
    adjustment_types = [
        "Purchase Invoice Adjustment / 采购发票调整",
        "Inventory Cost Adjustment / 库存成本调整",
        "First-Mile Freight Allocation / 头程物流分摊",
        "Bank Fee / 银行手续费",
        "FX Gain or Loss / 汇兑损益",
        "Tax Adjustment / 税务调整",
        "Non-Amazon Expense / 非Amazon费用",
        "Prior Period Adjustment / 前期调整",
        "Other / 其他",
    ]
    rows = []
    for index, adjustment_type in enumerate(adjustment_types, start=1):
        rows.append(
            {
                "adjustment_id": f"ADJ-{index:03d}",
                "adjustment_type": adjustment_type,
                "accounting_item": "",
                "suggested_account": "",
                "debit_or_credit": "",
                "currency": currency,
                "amount_original": None,
                "fx_rate": None,
                "amount_cny": None,
                "affects_month": result.month,
                "affects_quarter": quarter,
                "reason": "",
                "source_document_id": "",
                "accountant_confirmation": "Pending / 待确认",
                "status": "Blank template / 空白模板",
                "notes": "如无该类调整，可留空。",
            }
        )
    return rows


def _write_accountant_sheet(
    workbook: Workbook,
    sheet_name: str,
    columns: Sequence[tuple[str, str]],
    rows: Sequence[Mapping[str, Any]],
    *,
    formula_config: Mapping[str, str] | None = None,
) -> None:
    sheet = workbook.create_sheet(sheet_name)
    _write_description_block(sheet, ACCOUNTANT_SHEET_NOTES[sheet_name])
    header_row = ACCOUNTANT_TABLE_START_ROW
    for column_index, (_, label) in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, row in enumerate(rows, start=ACCOUNTANT_DATA_START_ROW):
        for column_index, (key, _) in enumerate(columns, start=1):
            value = _xlsx_value(row.get(key))
            formula = _formula_for_cell(
                formula_config or {}, key, row_index=row_index, columns=columns
            )
            sheet.cell(row=row_index, column=column_index, value=formula if formula else value)
    _format_accountant_sheet(sheet, columns, len(rows))


def _write_description_block(sheet: Any, notes: Sequence[tuple[str, str]]) -> None:
    for row_index, (key, value) in enumerate(notes, start=1):
        sheet.cell(row=row_index, column=1, value=key)
        sheet.cell(row=row_index, column=2, value=value)
        sheet.cell(row=row_index, column=1).font = Font(bold=True)
        sheet.cell(row=row_index, column=1).fill = DESCRIPTION_KEY_FILL
        sheet.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.cell(row=ACCOUNTANT_TABLE_START_ROW - 1, column=1, value="Data Table / 数据表")
    sheet.cell(row=ACCOUNTANT_TABLE_START_ROW - 1, column=1).font = Font(bold=True)
    sheet.cell(row=ACCOUNTANT_TABLE_START_ROW - 1, column=1).fill = SECTION_FILL


def _format_accountant_sheet(
    sheet: Any,
    columns: Sequence[tuple[str, str]],
    row_count: int,
) -> None:
    max_row = ACCOUNTANT_DATA_START_ROW + max(row_count, 1) - 1
    max_col = len(columns)
    sheet.freeze_panes = f"A{ACCOUNTANT_DATA_START_ROW}"
    filter_end = sheet.cell(max_row, max_col).coordinate
    sheet.auto_filter.ref = f"A{ACCOUNTANT_TABLE_START_ROW}:{filter_end}"
    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            label = str(sheet.cell(row=ACCOUNTANT_TABLE_START_ROW, column=cell.column).value or "")
            if _looks_like_money_column(label):
                cell.number_format = "#,##0.00"
            if "FX Rate / 汇率" in label:
                cell.number_format = "0.000000"
            if "Confirmation / 会计确认" in label or "Review / 是否需要会计复核" in label:
                cell.fill = REVIEW_FILL
    for column_index, (_, label) in enumerate(columns, start=1):
        letter = sheet.cell(row=ACCOUNTANT_TABLE_START_ROW, column=column_index).column_letter
        if "Notes / 说明" in label or "Source Reference / 来源引用" in label:
            width = 42
        elif "Account /" in label or "Accounting Item /" in label or "Summary /" in label:
            width = 36
        elif "Document" in label or "凭证" in label:
            width = 34
        elif "Confirmation /" in label:
            width = 24
        else:
            width = 18
        sheet.column_dimensions[letter].width = width
    sheet.column_dimensions["B"].width = max(sheet.column_dimensions["B"].width or 0, 72)
    for row_index in range(1, ACCOUNTANT_TABLE_START_ROW):
        sheet.row_dimensions[row_index].height = 36 if row_index <= 5 else 18


def _formula_for_cell(
    formula_config: Mapping[str, str],
    key: str,
    *,
    row_index: int,
    columns: Sequence[tuple[str, str]],
) -> str | None:
    formula_type = formula_config.get(key)
    if formula_type is None:
        return None
    if formula_type == "default_fx_rate":
        return "='12_FX_Rates'!$D$8"
    col = _column_letter(columns, row_index)
    if formula_type == "amount_usd_times_fx_rate":
        return f'=IFERROR({col("amount_usd")}*{col("fx_rate")},"")'
    if formula_type == "amount_original_times_fx_rate":
        return f'=IFERROR({col("amount_original")}*{col("fx_rate")},"")'
    if formula_type == "related_amount_original_times_default_fx_rate":
        return f'=IFERROR({col("related_amount_original")}*\'12_FX_Rates\'!$D$8,"")'
    if formula_type == "same_row_amount_cny":
        return f'=IFERROR({col("amount_cny")},"")'
    if formula_type == "actual_received_amount_times_fx_rate":
        return f'=IFERROR({col("actual_received_amount")}*{col("fx_rate")},"")'
    if formula_type == "actual_received_minus_amazon_net":
        return (
            f'=IF(AND({col("actual_received_amount")}<>"",{col("amazon_net_amount")}<>""),'
            f'{col("actual_received_amount")}-{col("amazon_net_amount")},"")'
        )
    if formula_type == "difference_original_times_fx_rate":
        return f'=IFERROR({col("difference_original")}*{col("fx_rate")},"")'
    return None


def _column_letter(columns: Sequence[tuple[str, str]], row_index: int):
    index = {key: position + 1 for position, (key, _) in enumerate(columns)}

    def ref(key: str) -> str:
        column_number = index[key]
        letters = ""
        while column_number:
            column_number, remainder = divmod(column_number - 1, 26)
            letters = chr(65 + remainder) + letters
        return f"{letters}{row_index}"

    return ref


def _looks_like_money_column(label: str) -> bool:
    keywords = (
        "Amount",
        "金额",
        "CNY",
        "人民币",
        "USD",
        "美元",
        "Fee",
        "费用",
        "Difference",
        "差异",
    )
    return any(keyword in label for keyword in keywords)


def _tax_passthrough_amount(result: MonthlyFinancialCloseResult) -> Decimal:
    for row in result.settlement_bucket_breakdown:
        if row.profit_bucket == "tax_passthrough":
            return abs(row.amount)
    return Decimal("0.00")


def _settlement_reference(result: MonthlyFinancialCloseResult) -> str:
    report_ids = result.raw_metadata.get("settlement_source_report_ids") or []
    settlement_ids = result.raw_metadata.get("settlement_ids") or []
    if report_ids:
        return "; ".join(str(value) for value in report_ids)
    if settlement_ids:
        return "; ".join(str(value) for value in settlement_ids)
    return f"{result.marketplace_id}:{result.month}"


def _quarter_label(value: date) -> str:
    quarter = (value.month - 1) // 3 + 1
    return f"{value.year}-Q{quarter}"


def _money_float(value: Decimal | int | float | None) -> float | None:
    if value is None:
        return None
    return float(value)


def _xlsx_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, list | tuple | set):
        return "; ".join(str(item) for item in value)
    if isinstance(value, dict):
        return str(value)
    return value


__all__ = ["add_accountant_pack_sheets", "build_accountant_pack_payload"]
