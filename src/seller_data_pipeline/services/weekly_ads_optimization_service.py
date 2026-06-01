from __future__ import annotations

import json
from collections import defaultdict
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass, field, replace
from datetime import UTC, date, datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Protocol

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from seller_data_pipeline.services.report_bilingual import (
    add_action_translation_columns,
    add_bilingual_readme_sheet,
    bilingual_metric_label,
    xlsx_header_label,
)

MONEY_QUANT = Decimal("0.01")
RATIO_QUANT = Decimal("0.0001")
ZERO = Decimal("0")
REPORT_TYPE = "weekly_ads_optimization"
REPORT_VERSION = "v1.1-active-negative-snapshot"
DEFAULT_OUTPUT_ROOT = "runtime/analysis_reports/weekly_ads_optimization"
WAOR_SCOPE_NOTE = (
    "Weekly Ads Optimization Report is an operational ads action report. Ads API campaign "
    "daily is the ads optimization source of truth; Settlement advertising fees are financial "
    "context only and must not overwrite Ads API spend."
)
DO_NOT_AUTO_APPLY_NOTE = (
    "Candidate only. Do not auto-apply; manually review relevance, inventory, current bids, "
    "budgets and existing negative keywords in Amazon Ads Console."
)
ACTIVE_CAMPAIGN_STATUSES = {"enabled", "delivering", "out_of_budget"}
INACTIVE_CAMPAIGN_STATUSES = {"paused", "archived", "ended", "campaign_paused", "deleted"}


@dataclass(frozen=True)
class WeeklyAdsOptimizationThresholds:
    target_acos: Decimal = Decimal("0.30")
    watch_acos: Decimal = Decimal("0.40")
    target_tacos: Decimal = Decimal("0.20")
    no_sale_cost_threshold: Decimal = Decimal("10.00")
    no_order_click_threshold: int = 12
    min_purchases_to_scale: int = 2
    min_sales_to_scale: Decimal = Decimal("40.00")
    low_ctr_threshold: Decimal = Decimal("0.002")
    low_cvr_threshold: Decimal = Decimal("0.03")
    high_cpc_multiplier: Decimal = Decimal("1.5")
    stable_lag_days: int = 3

    def to_dict(self) -> dict[str, Any]:
        return {
            "target_acos": _optional_ratio_to_string(self.target_acos),
            "watch_acos": _optional_ratio_to_string(self.watch_acos),
            "target_tacos": _optional_ratio_to_string(self.target_tacos),
            "no_sale_cost_threshold": _decimal_to_string(self.no_sale_cost_threshold),
            "no_order_click_threshold": self.no_order_click_threshold,
            "min_purchases_to_scale": self.min_purchases_to_scale,
            "min_sales_to_scale": _decimal_to_string(self.min_sales_to_scale),
            "low_ctr_threshold": _optional_ratio_to_string(self.low_ctr_threshold),
            "low_cvr_threshold": _optional_ratio_to_string(self.low_cvr_threshold),
            "high_cpc_multiplier": _optional_decimal_to_string(self.high_cpc_multiplier),
            "stable_lag_days": self.stable_lag_days,
        }


@dataclass(frozen=True)
class WarningEntry:
    warning_code: str
    severity: str
    message: str
    related_entity_type: str | None = None
    related_entity_id: str | None = None
    related_source: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "warning_code": self.warning_code,
            "severity": self.severity,
            "message": self.message,
            "related_entity_type": self.related_entity_type,
            "related_entity_id": self.related_entity_id,
            "related_source": self.related_source,
        }


@dataclass(frozen=True)
class ReconciliationCheck:
    check_name: str
    status: str
    severity: str
    expected: str | None
    actual: str | None
    diff: str | None = None
    diff_pct: str | None = None
    message: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "check_name": self.check_name,
            "status": self.status,
            "severity": self.severity,
            "expected": self.expected,
            "actual": self.actual,
            "diff": self.diff,
            "diff_pct": self.diff_pct,
            "message": self.message,
        }


@dataclass(frozen=True)
class OverallSummary:
    ads_spend: Decimal = ZERO
    ads_sales_7d: Decimal = ZERO
    ads_purchases_7d: int = 0
    ads_units_7d: int = 0
    impressions: int = 0
    clicks: int = 0
    campaign_count: int = 0
    campaign_row_count: int = 0
    ctr: Decimal | None = None
    cpc: Decimal | None = None
    cvr: Decimal | None = None
    acos: Decimal | None = None
    roas: Decimal | None = None
    ordered_product_sales: Decimal = ZERO
    units_ordered: int = 0
    sessions: int = 0
    unit_session_percentage: Decimal | None = None
    tacos: Decimal | None = None
    ads_sales_share: Decimal | None = None
    ads_spend_per_unit_ordered: Decimal | None = None
    settlement_advertising_fee_abs: Decimal = ZERO
    settlement_advertising_fee: Decimal = ZERO
    currency: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "ads_spend": _decimal_to_string(self.ads_spend),
            "ads_sales_7d": _decimal_to_string(self.ads_sales_7d),
            "ads_purchases_7d": self.ads_purchases_7d,
            "ads_units_7d": self.ads_units_7d,
            "impressions": self.impressions,
            "clicks": self.clicks,
            "campaign_count": self.campaign_count,
            "campaign_row_count": self.campaign_row_count,
            "ctr": _optional_ratio_to_string(self.ctr),
            "cpc": _optional_decimal_to_string(self.cpc),
            "cvr": _optional_ratio_to_string(self.cvr),
            "acos": _optional_ratio_to_string(self.acos),
            "roas": _optional_decimal_to_string(self.roas),
            "ordered_product_sales": _decimal_to_string(self.ordered_product_sales),
            "units_ordered": self.units_ordered,
            "sessions": self.sessions,
            "unit_session_percentage": _optional_ratio_to_string(self.unit_session_percentage),
            "tacos": _optional_ratio_to_string(self.tacos),
            "ads_sales_share": _optional_ratio_to_string(self.ads_sales_share),
            "ads_spend_per_unit_ordered": _optional_decimal_to_string(
                self.ads_spend_per_unit_ordered
            ),
            "settlement_advertising_fee_abs": _decimal_to_string(
                self.settlement_advertising_fee_abs
            ),
            "settlement_advertising_fee": _decimal_to_string(self.settlement_advertising_fee),
            "currency": self.currency,
        }


@dataclass(frozen=True)
class DailyTrendRow:
    report_date: date
    ads_spend: Decimal
    ads_sales_7d: Decimal
    ads_purchases_7d: int
    ads_units_7d: int
    impressions: int
    clicks: int
    ctr: Decimal | None
    cpc: Decimal | None
    cvr: Decimal | None
    acos: Decimal | None
    roas: Decimal | None
    ordered_product_sales: Decimal
    tacos: Decimal | None

    def to_dict(self) -> dict[str, Any]:
        return {
            "report_date": self.report_date.isoformat(),
            "ads_spend": _decimal_to_string(self.ads_spend),
            "ads_sales_7d": _decimal_to_string(self.ads_sales_7d),
            "ads_purchases_7d": self.ads_purchases_7d,
            "ads_units_7d": self.ads_units_7d,
            "impressions": self.impressions,
            "clicks": self.clicks,
            "ctr": _optional_ratio_to_string(self.ctr),
            "cpc": _optional_decimal_to_string(self.cpc),
            "cvr": _optional_ratio_to_string(self.cvr),
            "acos": _optional_ratio_to_string(self.acos),
            "roas": _optional_decimal_to_string(self.roas),
            "ordered_product_sales": _decimal_to_string(self.ordered_product_sales),
            "tacos": _optional_ratio_to_string(self.tacos),
        }


@dataclass(frozen=True)
class AdsEntityPerformanceRow:
    entity_type: str
    campaign_id: str | None
    campaign_name: str | None
    campaign_status: str | None = None
    ad_group_id: str | None = None
    ad_group_name: str | None = None
    keyword_id: str | None = None
    keyword: str | None = None
    match_type: str | None = None
    targeting: str | None = None
    keyword_text: str | None = None
    search_term: str | None = None
    advertised_sku: str | None = None
    advertised_asin: str | None = None
    spend: Decimal = ZERO
    sales_7d: Decimal = ZERO
    purchases_7d: int = 0
    units_7d: int = 0
    impressions: int = 0
    clicks: int = 0
    ctr: Decimal | None = None
    cpc: Decimal | None = None
    cvr: Decimal | None = None
    acos: Decimal | None = None
    roas: Decimal | None = None
    spend_share: Decimal | None = None
    sales_share: Decimal | None = None
    waste_cost: Decimal = ZERO
    potential_sales_efficiency: Decimal | None = None
    unit_standard_cost: Decimal | None = None
    estimated_ads_cogs: Decimal | None = None
    ads_contribution_proxy: Decimal | None = None
    cost_status: str | None = None
    action_label: str = "keep_observing"
    action_priority: str = "low"
    action_reason: str = "Data volume or performance is not strong enough for a firm action."
    manual_review_note: str = DO_NOT_AUTO_APPLY_NOTE
    already_negative: bool = False
    negative_scope: str | None = None
    negative_match_type: str | None = None
    recommended_action: str | None = None
    is_active_campaign: bool = True
    action_bucket: str = "active"

    def to_campaign_dict(self) -> dict[str, Any]:
        return {
            "campaign_id": self.campaign_id,
            "campaign_name": self.campaign_name,
            "campaign_status": self.campaign_status,
            "spend": _decimal_to_string(self.spend),
            "sales_7d": _decimal_to_string(self.sales_7d),
            "purchases_7d": self.purchases_7d,
            "units_7d": self.units_7d,
            "impressions": self.impressions,
            "clicks": self.clicks,
            "ctr": _optional_ratio_to_string(self.ctr),
            "cpc": _optional_decimal_to_string(self.cpc),
            "cvr": _optional_ratio_to_string(self.cvr),
            "acos": _optional_ratio_to_string(self.acos),
            "roas": _optional_decimal_to_string(self.roas),
            "spend_share": _optional_ratio_to_string(self.spend_share),
            "sales_share": _optional_ratio_to_string(self.sales_share),
            "action_label": self.action_label,
            "action_priority": self.action_priority,
            "action_reason": self.action_reason,
        }

    def to_targeting_dict(self) -> dict[str, Any]:
        return {
            "campaign_id": self.campaign_id,
            "campaign_name": self.campaign_name,
            "campaign_status": self.campaign_status,
            "is_active_campaign": self.is_active_campaign,
            "action_bucket": self.action_bucket,
            "ad_group_id": self.ad_group_id,
            "ad_group_name": self.ad_group_name,
            "keyword_id": self.keyword_id,
            "keyword": self.keyword,
            "match_type": self.match_type,
            "targeting": self.targeting,
            "keyword_text": self.keyword_text,
            "spend": _decimal_to_string(self.spend),
            "sales_7d": _decimal_to_string(self.sales_7d),
            "purchases_7d": self.purchases_7d,
            "units_7d": self.units_7d,
            "impressions": self.impressions,
            "clicks": self.clicks,
            "ctr": _optional_ratio_to_string(self.ctr),
            "cpc": _optional_decimal_to_string(self.cpc),
            "cvr": _optional_ratio_to_string(self.cvr),
            "acos": _optional_ratio_to_string(self.acos),
            "roas": _optional_decimal_to_string(self.roas),
            "action_label": self.action_label,
            "action_priority": self.action_priority,
            "action_reason": self.action_reason,
            "manual_review_note": self.manual_review_note,
        }

    def to_search_term_dict(self) -> dict[str, Any]:
        return {
            "campaign_id": self.campaign_id,
            "campaign_name": self.campaign_name,
            "campaign_status": self.campaign_status,
            "is_active_campaign": self.is_active_campaign,
            "action_bucket": self.action_bucket,
            "ad_group_id": self.ad_group_id,
            "ad_group_name": self.ad_group_name,
            "keyword_id": self.keyword_id,
            "keyword": self.keyword,
            "match_type": self.match_type,
            "targeting": self.targeting,
            "search_term": self.search_term,
            "spend": _decimal_to_string(self.spend),
            "sales_7d": _decimal_to_string(self.sales_7d),
            "purchases_7d": self.purchases_7d,
            "units_7d": self.units_7d,
            "impressions": self.impressions,
            "clicks": self.clicks,
            "ctr": _optional_ratio_to_string(self.ctr),
            "cpc": _optional_decimal_to_string(self.cpc),
            "cvr": _optional_ratio_to_string(self.cvr),
            "acos": _optional_ratio_to_string(self.acos),
            "roas": _optional_decimal_to_string(self.roas),
            "waste_cost": _decimal_to_string(self.waste_cost),
            "potential_sales_efficiency": _optional_decimal_to_string(
                self.potential_sales_efficiency
            ),
            "already_negative": self.already_negative,
            "negative_scope": self.negative_scope,
            "negative_match_type": self.negative_match_type,
            "recommended_action": self.recommended_action,
            "action_label": self.action_label,
            "action_reason": self.action_reason,
        }

    def to_search_term_action_dict(self) -> dict[str, Any]:
        return {
            "action_type": self.action_label,
            "priority": self.action_priority,
            "campaign_name": self.campaign_name,
            "campaign_status": self.campaign_status,
            "is_active_campaign": self.is_active_campaign,
            "action_bucket": self.action_bucket,
            "already_negative": self.already_negative,
            "negative_scope": self.negative_scope,
            "negative_match_type": self.negative_match_type,
            "recommended_action": self.recommended_action,
            "ad_group_name": self.ad_group_name,
            "keyword": self.keyword,
            "match_type": self.match_type,
            "search_term": self.search_term,
            "spend": _decimal_to_string(self.spend),
            "sales_7d": _decimal_to_string(self.sales_7d),
            "purchases_7d": self.purchases_7d,
            "clicks": self.clicks,
            "impressions": self.impressions,
            "acos": _optional_ratio_to_string(self.acos),
            "roas": _optional_decimal_to_string(self.roas),
            "reason": self.action_reason,
            "suggested_manual_action": _suggested_manual_action(self.action_label),
            "manual_review_note": self.manual_review_note,
            "do_not_auto_apply": True,
        }

    def to_advertised_product_dict(self) -> dict[str, Any]:
        return {
            "advertised_sku": self.advertised_sku,
            "advertised_asin": self.advertised_asin,
            "ads_spend": _decimal_to_string(self.spend),
            "ads_sales_7d": _decimal_to_string(self.sales_7d),
            "ads_purchases_7d": self.purchases_7d,
            "ads_units_7d": self.units_7d,
            "ads_acos": _optional_ratio_to_string(self.acos),
            "ads_roas": _optional_decimal_to_string(self.roas),
            "unit_standard_cost": _optional_decimal_to_string(self.unit_standard_cost),
            "estimated_ads_cogs": _optional_decimal_to_string(self.estimated_ads_cogs),
            "ads_contribution_proxy": _optional_decimal_to_string(self.ads_contribution_proxy),
            "cost_status": self.cost_status,
            "action_label": self.action_label,
            "action_reason": self.action_reason,
        }


@dataclass(frozen=True)
class NegativeKeywordSnapshotRow:
    scope: str
    campaign_id: str | None
    campaign_name: str | None
    ad_group_id: str | None
    ad_group_name: str | None
    keyword_text: str
    match_type: str | None
    state: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "scope": self.scope,
            "campaign_id": self.campaign_id,
            "campaign_name": self.campaign_name,
            "ad_group_id": self.ad_group_id,
            "ad_group_name": self.ad_group_name,
            "keyword_text": self.keyword_text,
            "match_type": self.match_type,
            "state": self.state,
            "normalized_keyword_text": _normalize_keyword_text(self.keyword_text),
        }


@dataclass(frozen=True)
class ActionItem:
    priority: str
    action_type: str
    entity_type: str
    campaign_name: str | None
    ad_group_name: str | None
    entity_id: str | None
    entity_text: str | None
    metric_summary: str
    reason: str
    suggested_manual_action: str
    manual_review_note: str = DO_NOT_AUTO_APPLY_NOTE
    do_not_auto_apply: bool = True
    action_bucket: str = "active"
    campaign_status: str | None = None
    already_negative: bool = False
    negative_scope: str | None = None
    negative_match_type: str | None = None
    recommended_action: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "priority": self.priority,
            "action_type": self.action_type,
            "entity_type": self.entity_type,
            "campaign_name": self.campaign_name,
            "ad_group_name": self.ad_group_name,
            "entity_id": self.entity_id,
            "entity_text": self.entity_text,
            "metric_summary": self.metric_summary,
            "reason": self.reason,
            "suggested_manual_action": self.suggested_manual_action,
            "manual_review_note": self.manual_review_note,
            "do_not_auto_apply": self.do_not_auto_apply,
            "action_bucket": self.action_bucket,
            "campaign_status": self.campaign_status,
            "already_negative": self.already_negative,
            "negative_scope": self.negative_scope,
            "negative_match_type": self.negative_match_type,
            "recommended_action": self.recommended_action,
        }


@dataclass(frozen=True)
class WeeklyAdsOptimizationResult:
    marketplace_id: str
    profile_id: str | None
    week_start: date
    week_end: date
    generated_at_utc: datetime
    status: str
    currency: str | None
    thresholds: WeeklyAdsOptimizationThresholds
    overall_summary: OverallSummary
    daily_trend: list[DailyTrendRow]
    campaign_performance: list[AdsEntityPerformanceRow]
    targeting_performance: list[AdsEntityPerformanceRow]
    search_term_performance: list[AdsEntityPerformanceRow]
    search_term_action_candidates: list[AdsEntityPerformanceRow]
    advertised_product_performance: list[AdsEntityPerformanceRow]
    action_items: list[ActionItem]
    historical_action_items: list[ActionItem] = field(default_factory=list)
    negative_keyword_snapshot: list[NegativeKeywordSnapshotRow] = field(default_factory=list)
    reconciliation_checks: list[ReconciliationCheck] = field(default_factory=list)
    warnings: list[WarningEntry] = field(default_factory=list)
    raw_metadata: dict[str, Any] = field(default_factory=dict)
    output_files: dict[str, str] = field(default_factory=dict)

    def executive_summary(self) -> dict[str, Any]:
        action_count = len(self.action_items)
        historical_action_count = len(self.historical_action_items)
        non_info_warnings = [warning for warning in self.warnings if warning.severity != "info"]
        headline = (
            f"{self.week_start.isoformat()}..{self.week_end.isoformat()} ads spend was "
            f"{_format_money(self.overall_summary.ads_spend, self.currency)}, with ACOS "
            f"{_format_ratio(self.overall_summary.acos)} and {action_count} active action candidates "
            f"({historical_action_count} historical/paused lessons)."
        )
        negative_count = sum(
            1 for row in self.search_term_action_candidates if "negative" in row.action_label
        )
        already_negative_count = sum(1 for row in self.search_term_performance if row.already_negative)
        harvest_count = sum(
            1
            for row in self.search_term_action_candidates
            if row.action_label == "harvest_to_exact_candidate"
        )
        return {
            "headline": headline,
            "key_points": [
                f"Report status: {self.status}.",
                f"Ads sales 7d: {_format_money(self.overall_summary.ads_sales_7d, self.currency)}.",
                f"ROAS: {_format_decimal(self.overall_summary.roas)}; TACOS: "
                f"{_format_ratio(self.overall_summary.tacos)}.",
                f"Negative candidates: {negative_count}; already-negative search terms skipped: {already_negative_count}; harvest candidates: {harvest_count}.",
                f"Active action items: {action_count}; historical/paused lessons: {historical_action_count}.",
                f"Non-info warnings: {len(non_info_warnings)}.",
                WAOR_SCOPE_NOTE,
            ],
            "recommended_next_steps": [
                "Review 08_Active_Action_Items before changing Ads Console settings.",
                "Use 09_Historical_Paused_Lessons for context only unless a campaign is re-enabled.",
                "Prioritize high-priority negative candidates and high-efficiency exact harvests.",
                "Do not force Ads API spend to equal Settlement advertising fee.",
            ],
        }

    def to_dict(self) -> dict[str, Any]:
        return {
            "report_type": REPORT_TYPE,
            "version": REPORT_VERSION,
            "marketplace_id": self.marketplace_id,
            "profile_id": self.profile_id,
            "period": {
                "week_start": self.week_start.isoformat(),
                "week_end": self.week_end.isoformat(),
            },
            "generated_at_utc": self.generated_at_utc.isoformat(),
            "status": self.status,
            "currency": self.currency,
            "thresholds": self.thresholds.to_dict(),
            "executive_summary": self.executive_summary(),
            "overall_summary": self.overall_summary.to_dict(),
            "daily_trend": [row.to_dict() for row in self.daily_trend],
            "campaign_performance": [row.to_campaign_dict() for row in self.campaign_performance],
            "targeting_performance": [
                row.to_targeting_dict() for row in self.targeting_performance
            ],
            "search_term_performance": [
                row.to_search_term_dict() for row in self.search_term_performance
            ],
            "search_term_action_candidates": [
                row.to_search_term_action_dict() for row in self.search_term_action_candidates
            ],
            "advertised_product_performance": [
                row.to_advertised_product_dict() for row in self.advertised_product_performance
            ],
            "action_items": [row.to_dict() for row in self.action_items],
            "active_action_items": [row.to_dict() for row in self.action_items],
            "historical_action_items": [row.to_dict() for row in self.historical_action_items],
            "negative_keyword_snapshot": [row.to_dict() for row in self.negative_keyword_snapshot],
            "reconciliation_checks": [check.to_dict() for check in self.reconciliation_checks],
            "warnings": [warning.to_dict() for warning in self.warnings],
            "raw_metadata": _json_safe_mapping(self.raw_metadata),
            "output_files": self.output_files,
        }

    def with_output_files(self, output_files: Mapping[str, str]) -> WeeklyAdsOptimizationResult:
        return WeeklyAdsOptimizationResult(
            marketplace_id=self.marketplace_id,
            profile_id=self.profile_id,
            week_start=self.week_start,
            week_end=self.week_end,
            generated_at_utc=self.generated_at_utc,
            status=self.status,
            currency=self.currency,
            thresholds=self.thresholds,
            overall_summary=self.overall_summary,
            daily_trend=self.daily_trend,
            campaign_performance=self.campaign_performance,
            targeting_performance=self.targeting_performance,
            search_term_performance=self.search_term_performance,
            search_term_action_candidates=self.search_term_action_candidates,
            advertised_product_performance=self.advertised_product_performance,
            action_items=self.action_items,
            historical_action_items=self.historical_action_items,
            negative_keyword_snapshot=self.negative_keyword_snapshot,
            reconciliation_checks=self.reconciliation_checks,
            warnings=self.warnings,
            raw_metadata=self.raw_metadata,
            output_files=dict(output_files),
        )


class WeeklyAdsOptimizationDataRepo(Protocol):
    def fetch_campaign_daily_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_targeting_daily_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_search_term_daily_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_advertised_product_daily_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_sales_traffic_daily_rows(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_sku_cost_rows(
        self,
        *,
        marketplace_id: str,
        as_of_date: date,
    ) -> list[dict[str, Any]]: ...

    def fetch_settlement_advertising_summary(
        self,
        *,
        marketplace_id: str,
        start_date: date,
        end_date: date,
    ) -> dict[str, Any]: ...

    def fetch_negative_keyword_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        start_date: date,
        end_date: date,
    ) -> list[dict[str, Any]]: ...


class WeeklyAdsOptimizationService:
    def __init__(self, repo: WeeklyAdsOptimizationDataRepo | None = None) -> None:
        self.repo = repo

    def run(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        week_start: date,
        output_root: str | Path = DEFAULT_OUTPUT_ROOT,
        thresholds: WeeklyAdsOptimizationThresholds | None = None,
        negative_keyword_rows: Sequence[Mapping[str, Any]] | None = None,
    ) -> WeeklyAdsOptimizationResult:
        if self.repo is None:
            raise RuntimeError("repo is required for run(); use calculate_from_rows() in tests")
        _validate_week_start(week_start)
        week_end = week_start + timedelta(days=6)
        thresholds = thresholds or WeeklyAdsOptimizationThresholds()
        campaign_rows = self.repo.fetch_campaign_daily_rows(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            start_date=week_start,
            end_date=week_end,
        )
        targeting_rows = self.repo.fetch_targeting_daily_rows(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            start_date=week_start,
            end_date=week_end,
        )
        search_rows = self.repo.fetch_search_term_daily_rows(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            start_date=week_start,
            end_date=week_end,
        )
        product_rows = self.repo.fetch_advertised_product_daily_rows(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            start_date=week_start,
            end_date=week_end,
        )
        sales_rows = self.repo.fetch_sales_traffic_daily_rows(
            marketplace_id=marketplace_id,
            start_date=week_start,
            end_date=week_end,
        )
        cost_rows = self.repo.fetch_sku_cost_rows(
            marketplace_id=marketplace_id,
            as_of_date=week_end,
        )
        settlement_summary = self.repo.fetch_settlement_advertising_summary(
            marketplace_id=marketplace_id,
            start_date=week_start,
            end_date=week_end,
        )
        fetch_negative_rows = getattr(self.repo, "fetch_negative_keyword_rows", None)
        repo_negative_keyword_rows = (
            fetch_negative_rows(
                marketplace_id=marketplace_id,
                profile_id=profile_id,
                start_date=week_start,
                end_date=week_end,
            )
            if callable(fetch_negative_rows)
            else []
        )
        negative_keyword_rows = [
            *list(repo_negative_keyword_rows or []),
            *list(negative_keyword_rows or []),
        ]
        result = self.calculate_from_rows(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            week_start=week_start,
            week_end=week_end,
            campaign_rows=campaign_rows,
            targeting_rows=targeting_rows,
            search_term_rows=search_rows,
            advertised_product_rows=product_rows,
            sales_traffic_rows=sales_rows,
            sku_cost_rows=cost_rows,
            settlement_advertising_summary=settlement_summary,
            negative_keyword_rows=negative_keyword_rows,
            thresholds=thresholds,
        )
        return self.write_report_files(result=result, output_root=output_root)

    def calculate_from_rows(
        self,
        *,
        marketplace_id: str,
        profile_id: str | None,
        week_start: date,
        week_end: date | None = None,
        campaign_rows: Sequence[Mapping[str, Any]] | None = None,
        targeting_rows: Sequence[Mapping[str, Any]] | None = None,
        search_term_rows: Sequence[Mapping[str, Any]] | None = None,
        advertised_product_rows: Sequence[Mapping[str, Any]] | None = None,
        sales_traffic_rows: Sequence[Mapping[str, Any]] | None = None,
        sku_cost_rows: Sequence[Mapping[str, Any]] | None = None,
        settlement_advertising_summary: Mapping[str, Any] | None = None,
        negative_keyword_rows: Sequence[Mapping[str, Any]] | None = None,
        thresholds: WeeklyAdsOptimizationThresholds | None = None,
        generated_at_utc: datetime | None = None,
    ) -> WeeklyAdsOptimizationResult:
        _validate_week_start(week_start)
        week_end = week_end or week_start + timedelta(days=6)
        generated_at_utc = generated_at_utc or datetime.now(tz=UTC)
        thresholds = thresholds or WeeklyAdsOptimizationThresholds()
        campaigns = list(campaign_rows or [])
        targets = list(targeting_rows or [])
        searches = list(search_term_rows or [])
        products = list(advertised_product_rows or [])
        sales_rows = list(sales_traffic_rows or [])
        costs = list(sku_cost_rows or [])
        settlement = dict(settlement_advertising_summary or {})
        negative_rows = list(negative_keyword_rows or [])

        sales_context = _aggregate_sales_context(sales_rows)
        settlement_fee = _money(_to_decimal(settlement.get("settlement_advertising_fee")))
        currency = _first_non_empty(
            [
                sales_context["currency"],
                _empty_to_none(settlement.get("currency")),
                _first_non_empty(_empty_to_none(row.get("currency")) for row in costs),
            ]
        )
        overall = _aggregate_overall(
            campaign_rows=campaigns,
            sales_context=sales_context,
            settlement_advertising_fee=settlement_fee,
            currency=currency,
        )
        avg_cpc = overall.cpc or ZERO
        campaign_performance = _build_campaign_performance(
            campaigns,
            thresholds=thresholds,
            total_spend=overall.ads_spend,
            total_sales=overall.ads_sales_7d,
        )
        campaign_status_index = _campaign_status_index(campaign_performance)
        negative_keyword_snapshot = _build_negative_keyword_snapshot(negative_rows)
        targeting_performance = _build_targeting_performance(
            targets,
            thresholds=thresholds,
            avg_cpc=avg_cpc,
            campaign_status_index=campaign_status_index,
        )
        search_term_performance = _build_search_term_performance(
            searches,
            thresholds=thresholds,
            avg_cpc=avg_cpc,
            campaign_status_index=campaign_status_index,
            negative_keyword_snapshot=negative_keyword_snapshot,
        )
        search_term_action_candidates = _search_term_action_candidates(search_term_performance)
        cost_index = _build_cost_index(costs)
        advertised_product_performance = _build_advertised_product_performance(
            products,
            thresholds=thresholds,
            cost_index=cost_index,
            as_of_date=week_end,
        )
        daily_trend = _build_daily_trend(
            week_start=week_start,
            week_end=week_end,
            campaign_rows=campaigns,
            sales_rows=sales_rows,
        )
        checks = _build_reconciliation_checks(
            week_start=week_start,
            week_end=week_end,
            generated_at_utc=generated_at_utc,
            thresholds=thresholds,
            campaign_rows=campaigns,
            targeting_rows=targets,
            search_term_rows=searches,
            advertised_product_rows=products,
            sales_rows=sales_rows,
            overall=overall,
        )
        warnings = _build_warnings(checks=checks, overall=overall)
        action_items, historical_action_items = _build_action_items(
            campaign_performance=campaign_performance,
            targeting_performance=targeting_performance,
            search_term_candidates=search_term_action_candidates,
            advertised_product_performance=advertised_product_performance,
        )
        status = _result_status(
            campaign_rows=campaigns,
            checks=checks,
            warnings=warnings,
        )
        raw_metadata = _build_raw_metadata(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            week_start=week_start,
            week_end=week_end,
            generated_at_utc=generated_at_utc,
            campaign_rows=campaigns,
            targeting_rows=targets,
            search_term_rows=searches,
            advertised_product_rows=products,
            sales_rows=sales_rows,
            sku_cost_rows=costs,
            negative_keyword_rows=negative_rows,
            settlement=settlement,
            thresholds=thresholds,
        )
        return WeeklyAdsOptimizationResult(
            marketplace_id=marketplace_id,
            profile_id=profile_id,
            week_start=week_start,
            week_end=week_end,
            generated_at_utc=generated_at_utc,
            status=status,
            currency=currency,
            thresholds=thresholds,
            overall_summary=overall,
            daily_trend=daily_trend,
            campaign_performance=campaign_performance,
            targeting_performance=targeting_performance,
            search_term_performance=search_term_performance,
            search_term_action_candidates=search_term_action_candidates,
            advertised_product_performance=advertised_product_performance,
            action_items=action_items,
            historical_action_items=historical_action_items,
            negative_keyword_snapshot=negative_keyword_snapshot,
            reconciliation_checks=checks,
            warnings=warnings,
            raw_metadata=raw_metadata,
        )

    def write_report_files(
        self,
        *,
        result: WeeklyAdsOptimizationResult,
        output_root: str | Path = DEFAULT_OUTPUT_ROOT,
    ) -> WeeklyAdsOptimizationResult:
        profile_part = result.profile_id or "no_profile"
        output_dir = (
            Path(output_root)
            / profile_part
            / f"{result.week_start.isoformat()}_{result.week_end.isoformat()}"
        )
        output_dir.mkdir(parents=True, exist_ok=True)
        period_key = f"{result.week_start.isoformat()}_{result.week_end.isoformat()}"
        filename_base = f"weekly_ads_optimization_{period_key}"
        json_path = output_dir / f"{filename_base}.json"
        xlsx_path = output_dir / f"{filename_base}.xlsx"
        result_with_files = result.with_output_files(
            {"json": str(json_path), "xlsx": str(xlsx_path)}
        )
        json_path.write_text(
            json.dumps(result_with_files.to_dict(), ensure_ascii=False, indent=2, sort_keys=True)
            + "\n",
            encoding="utf-8",
        )
        workbook = build_weekly_ads_optimization_workbook(result_with_files)
        workbook.save(xlsx_path)
        return result_with_files


def build_weekly_ads_optimization_workbook(result: WeeklyAdsOptimizationResult) -> Workbook:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    add_bilingual_readme_sheet(
        workbook,
        title_en="Weekly Ads Optimization Report",
        title_zh="每周广告优化报表",
        period=f"{result.week_start.isoformat()}_{result.week_end.isoformat()}",
        status=result.status,
        scope_en=(
            "Ads API campaign daily is the ads optimization source of truth; "
            "Settlement advertising fees are finance context only."
        ),
        scope_zh="Ads API campaign daily 是广告优化主口径；Settlement 广告费仅作为财务参考。",
    )
    _write_rows_sheet(workbook, "01_Executive_Summary", _summary_rows(result))
    _write_rows_sheet(workbook, "02_Daily_Trend", [row.to_dict() for row in result.daily_trend])
    _write_rows_sheet(
        workbook,
        "03_Campaigns",
        [row.to_campaign_dict() for row in result.campaign_performance],
    )
    _write_rows_sheet(
        workbook,
        "04_Targeting",
        [row.to_targeting_dict() for row in result.targeting_performance],
    )
    _write_rows_sheet(
        workbook,
        "05_Search_Terms",
        [
            add_action_translation_columns(row.to_search_term_dict())
            for row in result.search_term_performance
        ],
    )
    _write_rows_sheet(
        workbook,
        "06_Search_Term_Actions",
        [
            add_action_translation_columns(row.to_search_term_action_dict())
            for row in result.search_term_action_candidates
        ],
    )
    _write_rows_sheet(
        workbook,
        "07_Advertised_Products",
        [row.to_advertised_product_dict() for row in result.advertised_product_performance],
    )
    _write_rows_sheet(
        workbook,
        "08_Active_Action_Items",
        [add_action_translation_columns(row.to_dict()) for row in result.action_items],
    )
    _write_rows_sheet(
        workbook,
        "09_Historical_Paused_Lessons",
        [add_action_translation_columns(row.to_dict()) for row in result.historical_action_items],
    )
    _write_rows_sheet(
        workbook,
        "10_Reconciliation_Checks",
        [row.to_dict() for row in result.reconciliation_checks],
    )
    _write_rows_sheet(workbook, "11_Warnings", [row.to_dict() for row in result.warnings])
    _write_rows_sheet(
        workbook,
        "12_Negative_Snapshot",
        [row.to_dict() for row in result.negative_keyword_snapshot],
    )
    _write_rows_sheet(workbook, "13_Raw_Metadata", _metadata_rows(result))
    return workbook


def _summary_rows(result: WeeklyAdsOptimizationResult) -> list[dict[str, Any]]:
    overall = result.overall_summary
    negative_count = sum(
        1 for row in result.search_term_action_candidates if "negative" in row.action_label
    )
    harvest_count = sum(
        1
        for row in result.search_term_action_candidates
        if row.action_label == "harvest_to_exact_candidate"
    )
    rows = [
        _metric_row("status", "report_status", result.status, "status", WAOR_SCOPE_NOTE),
        _metric_row(
            "overall", "ads_spend", overall.ads_spend, result.currency, "Campaign table cost."
        ),
        _metric_row(
            "overall",
            "ads_sales_7d",
            overall.ads_sales_7d,
            result.currency,
            "7-day attributed sales.",
        ),
        _metric_row("overall", "ads_purchases_7d", overall.ads_purchases_7d, "count", ""),
        _metric_row("overall", "acos", overall.acos, "ratio", "ads_spend / ads_sales_7d"),
        _metric_row("overall", "roas", overall.roas, "ratio", "ads_sales_7d / ads_spend"),
        _metric_row(
            "sales_context",
            "ordered_product_sales",
            overall.ordered_product_sales,
            result.currency,
            "Sales & Traffic report-date sales.",
        ),
        _metric_row(
            "sales_context", "tacos", overall.tacos, "ratio", "ads_spend / ordered_product_sales"
        ),
        _metric_row(
            "financial_context",
            "settlement_advertising_fee_abs",
            overall.settlement_advertising_fee_abs,
            result.currency,
            "Settlement posted-date context only.",
        ),
        _metric_row("actions", "campaign_count", overall.campaign_count, "count", ""),
        _metric_row("actions", "active_action_item_count", len(result.action_items), "count", ""),
        _metric_row(
            "actions",
            "historical_paused_lesson_count",
            len(result.historical_action_items),
            "count",
            "Actionable context from inactive/paused campaigns; not primary execution list.",
        ),
        _metric_row(
            "actions",
            "negative_snapshot_count",
            len(result.negative_keyword_snapshot),
            "count",
            "Existing negative keywords loaded for de-duplication.",
        ),
        _metric_row("actions", "negative_candidate_count", negative_count, "count", ""),
        _metric_row("actions", "harvest_candidate_count", harvest_count, "count", ""),
        _metric_row(
            "status",
            "warning_count",
            len([w for w in result.warnings if w.severity != "info"]),
            "count",
            "",
        ),
    ]
    for key, value in result.thresholds.to_dict().items():
        rows.append(_metric_row("thresholds", key, value, None, "CLI/default threshold."))
    return rows


def _metric_row(
    metric_group: str,
    metric_name: str,
    value: Any,
    unit: str | None,
    notes: str,
) -> dict[str, Any]:
    return {
        "metric_group": metric_group,
        "metric_name": bilingual_metric_label(metric_name),
        "value": _xlsx_value(value),
        "unit": unit,
        "notes": notes,
    }


def _metadata_rows(result: WeeklyAdsOptimizationResult) -> list[dict[str, Any]]:
    rows = [{"key": key, "value": value} for key, value in result.raw_metadata.items()]
    rows.extend(
        {"key": f"output_{key}", "value": value} for key, value in result.output_files.items()
    )
    return rows


def _write_rows_sheet(workbook: Workbook, title: str, rows: Sequence[Mapping[str, Any]]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet.append([xlsx_header_label("message")])
        sheet.append(["No rows / 无数据"])
        _format_sheet(sheet)
        return
    headers = list(rows[0].keys())
    for row in rows[1:]:
        for key in row.keys():
            if key not in headers:
                headers.append(key)
    sheet.append([xlsx_header_label(header) for header in headers])
    for row in rows:
        sheet.append([_xlsx_value(row.get(header)) for header in headers])
    _format_sheet(sheet)


def _format_sheet(sheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="7030A0")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def _aggregate_sales_context(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    sales = sum((_to_decimal(row.get("ordered_product_sales_amount")) for row in rows), ZERO)
    units = sum(_int_metric(row, "units_ordered") for row in rows)
    sessions = sum(_int_metric(row, "sessions") for row in rows)
    currency = _first_non_empty(
        _empty_to_none(row.get("ordered_product_sales_currency")) for row in rows
    )
    return {
        "ordered_product_sales": _money(sales),
        "units_ordered": units,
        "sessions": sessions,
        "unit_session_percentage": _safe_ratio(Decimal(units), Decimal(sessions)),
        "currency": currency,
    }


def _aggregate_overall(
    *,
    campaign_rows: Sequence[Mapping[str, Any]],
    sales_context: Mapping[str, Any],
    settlement_advertising_fee: Decimal,
    currency: str | None,
) -> OverallSummary:
    impressions = sum(_int_metric(row, "impressions") for row in campaign_rows)
    clicks = sum(_int_metric(row, "clicks") for row in campaign_rows)
    spend = _money(sum((_to_decimal(row.get("cost")) for row in campaign_rows), ZERO))
    sales_7d = _money(sum((_to_decimal(row.get("sales_7d")) for row in campaign_rows), ZERO))
    purchases = sum(_int_metric(row, "purchases_7d") for row in campaign_rows)
    units = sum(_int_metric(row, "units_sold_clicks_7d") for row in campaign_rows)
    campaigns = {
        _empty_to_none(row.get("campaign_id")) or _empty_to_none(row.get("campaign_name")) or "-"
        for row in campaign_rows
    }
    ordered_product_sales = _money(_to_decimal(sales_context.get("ordered_product_sales")))
    units_ordered = int(sales_context.get("units_ordered") or 0)
    return OverallSummary(
        ads_spend=spend,
        ads_sales_7d=sales_7d,
        ads_purchases_7d=purchases,
        ads_units_7d=units,
        impressions=impressions,
        clicks=clicks,
        campaign_count=len(campaigns) if campaign_rows else 0,
        campaign_row_count=len(campaign_rows),
        ctr=_safe_ratio(Decimal(clicks), Decimal(impressions)),
        cpc=_safe_money_ratio(spend, Decimal(clicks)),
        cvr=_safe_ratio(Decimal(purchases), Decimal(clicks)),
        acos=_safe_ratio(spend, sales_7d),
        roas=_safe_money_ratio(sales_7d, spend),
        ordered_product_sales=ordered_product_sales,
        units_ordered=units_ordered,
        sessions=int(sales_context.get("sessions") or 0),
        unit_session_percentage=sales_context.get("unit_session_percentage"),
        tacos=_safe_ratio(spend, ordered_product_sales),
        ads_sales_share=_safe_ratio(sales_7d, ordered_product_sales),
        ads_spend_per_unit_ordered=_safe_money_ratio(spend, Decimal(units_ordered)),
        settlement_advertising_fee_abs=_money(abs(settlement_advertising_fee)),
        settlement_advertising_fee=settlement_advertising_fee,
        currency=currency,
    )


def _build_campaign_performance(
    rows: Sequence[Mapping[str, Any]],
    *,
    thresholds: WeeklyAdsOptimizationThresholds,
    total_spend: Decimal,
    total_sales: Decimal,
) -> list[AdsEntityPerformanceRow]:
    buckets = _aggregate_ads_rows(
        rows,
        key_fields=("campaign_id", "campaign_name"),
        passthrough_fields=("campaign_status",),
    )
    output: list[AdsEntityPerformanceRow] = []
    for key, bucket in buckets.items():
        metrics = _metrics_from_bucket(bucket)
        action = _campaign_action(metrics=metrics, thresholds=thresholds)
        campaign_status = _empty_to_none(bucket.get("campaign_status"))
        output.append(
            AdsEntityPerformanceRow(
                entity_type="campaign",
                campaign_id=key[0],
                campaign_name=key[1],
                campaign_status=campaign_status,
                is_active_campaign=_is_active_campaign_status(campaign_status),
                action_bucket=_action_bucket_for_status(campaign_status),
                spend=metrics["spend"],
                sales_7d=metrics["sales"],
                purchases_7d=int(metrics["purchases"]),
                units_7d=int(metrics["units"]),
                impressions=int(metrics["impressions"]),
                clicks=int(metrics["clicks"]),
                ctr=metrics["ctr"],
                cpc=metrics["cpc"],
                cvr=metrics["cvr"],
                acos=metrics["acos"],
                roas=metrics["roas"],
                spend_share=_safe_ratio(metrics["spend"], total_spend),
                sales_share=_safe_ratio(metrics["sales"], total_sales),
                action_label=action["label"],
                action_priority=action["priority"],
                action_reason=action["reason"],
            )
        )
    return sorted(output, key=lambda row: row.spend, reverse=True)


def _campaign_status_index(rows: Sequence[AdsEntityPerformanceRow]) -> dict[str, str | None]:
    index: dict[str, str | None] = {}
    for row in rows:
        status = row.campaign_status
        if row.campaign_id:
            index[f"id:{row.campaign_id}"] = status
        if row.campaign_name:
            index[f"name:{row.campaign_name}"] = status
    return index


def _campaign_status_for(
    index: Mapping[str, str | None],
    campaign_id: str | None,
    campaign_name: str | None,
) -> str | None:
    if campaign_id and f"id:{campaign_id}" in index:
        return index[f"id:{campaign_id}"]
    if campaign_name and f"name:{campaign_name}" in index:
        return index[f"name:{campaign_name}"]
    return None


def _is_active_campaign_status(status: str | None) -> bool:
    normalized = _normalize_status(status)
    if normalized in INACTIVE_CAMPAIGN_STATUSES:
        return False
    if normalized in ACTIVE_CAMPAIGN_STATUSES:
        return True
    return True


def _action_bucket_for_status(status: str | None) -> str:
    return "active" if _is_active_campaign_status(status) else "historical_or_paused"


def _normalize_status(status: str | None) -> str:
    return str(status or "").strip().lower().replace(" ", "_").replace("-", "_")


def _build_negative_keyword_snapshot(
    rows: Sequence[Mapping[str, Any]],
) -> list[NegativeKeywordSnapshotRow]:
    output: list[NegativeKeywordSnapshotRow] = []
    for row in rows:
        state = _empty_to_none(
            row.get("state") or row.get("status") or row.get("negative_keyword_status")
        )
        if _normalize_status(state) in {"archived", "deleted"}:
            continue
        keyword_text = _first_non_empty(
            [
                _empty_to_none(row.get("keyword_text")),
                _empty_to_none(row.get("negative_keyword_text")),
                _empty_to_none(row.get("negative_keyword")),
                _empty_to_none(row.get("keyword")),
                _empty_to_none(row.get("search_term")),
            ]
        )
        if not keyword_text:
            continue
        scope = _empty_to_none(row.get("scope") or row.get("negative_scope"))
        if not scope:
            scope = "ad_group" if _empty_to_none(row.get("ad_group_id") or row.get("ad_group_name")) else "campaign"
        output.append(
            NegativeKeywordSnapshotRow(
                scope=str(scope).strip().lower().replace(" ", "_"),
                campaign_id=_empty_to_none(row.get("campaign_id")),
                campaign_name=_empty_to_none(row.get("campaign_name")),
                ad_group_id=_empty_to_none(row.get("ad_group_id")),
                ad_group_name=_empty_to_none(row.get("ad_group_name")),
                keyword_text=str(keyword_text).strip(),
                match_type=_empty_to_none(row.get("match_type") or row.get("negative_match_type")),
                state=state,
            )
        )
    return output


def _apply_negative_snapshot(
    row: AdsEntityPerformanceRow,
    snapshot: Sequence[NegativeKeywordSnapshotRow],
) -> AdsEntityPerformanceRow:
    match = _matching_negative_keyword(row, snapshot)
    if match is None:
        return replace(row, recommended_action=row.recommended_action or row.action_label)
    action_label = row.action_label
    action_priority = row.action_priority
    action_reason = row.action_reason
    recommended_action = row.action_label
    if "negative" in row.action_label:
        action_label = "already_negative"
        action_priority = "low"
        action_reason = (
            "Search term is already covered by existing negative keyword snapshot; "
            "do not recommend duplicate negative action."
        )
        recommended_action = "already_done"
    return replace(
        row,
        already_negative=True,
        negative_scope=match.scope,
        negative_match_type=match.match_type,
        action_label=action_label,
        action_priority=action_priority,
        action_reason=action_reason,
        recommended_action=recommended_action,
    )


def _matching_negative_keyword(
    row: AdsEntityPerformanceRow,
    snapshot: Sequence[NegativeKeywordSnapshotRow],
) -> NegativeKeywordSnapshotRow | None:
    search = _normalize_keyword_text(row.search_term)
    if not search:
        return None
    for negative in snapshot:
        if not _negative_scope_matches(row, negative):
            continue
        negative_text = _normalize_keyword_text(negative.keyword_text)
        if not negative_text:
            continue
        match_type = str(negative.match_type or "").lower()
        if "phrase" in match_type:
            if negative_text in search:
                return negative
        elif search == negative_text:
            return negative
    return None


def _negative_scope_matches(row: AdsEntityPerformanceRow, negative: NegativeKeywordSnapshotRow) -> bool:
    if negative.campaign_id and row.campaign_id and negative.campaign_id != row.campaign_id:
        return False
    if negative.campaign_name and row.campaign_name and negative.campaign_name != row.campaign_name:
        return False
    if negative.scope in {"ad_group", "adgroup", "ad_group_negative"}:
        if negative.ad_group_id and row.ad_group_id and negative.ad_group_id != row.ad_group_id:
            return False
        if negative.ad_group_name and row.ad_group_name and negative.ad_group_name != row.ad_group_name:
            return False
    return True


def _normalize_keyword_text(value: str | None) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _build_targeting_performance(
    rows: Sequence[Mapping[str, Any]],
    *,
    thresholds: WeeklyAdsOptimizationThresholds,
    avg_cpc: Decimal,
    campaign_status_index: Mapping[str, str | None],
) -> list[AdsEntityPerformanceRow]:
    buckets = _aggregate_ads_rows(
        rows,
        key_fields=(
            "campaign_id",
            "campaign_name",
            "ad_group_id",
            "ad_group_name",
            "keyword_id",
            "keyword",
            "match_type",
            "targeting",
        ),
    )
    output: list[AdsEntityPerformanceRow] = []
    for key, bucket in buckets.items():
        metrics = _metrics_from_bucket(bucket)
        action = _targeting_action(metrics=metrics, thresholds=thresholds, avg_cpc=avg_cpc)
        keyword_text = _empty_to_none(key[5]) or _empty_to_none(key[7])
        campaign_status = _campaign_status_for(campaign_status_index, key[0], key[1])
        is_active = _is_active_campaign_status(campaign_status)
        output.append(
            AdsEntityPerformanceRow(
                entity_type="targeting",
                campaign_id=key[0],
                campaign_name=key[1],
                campaign_status=campaign_status,
                is_active_campaign=is_active,
                action_bucket=_action_bucket_for_status(campaign_status),
                ad_group_id=key[2],
                ad_group_name=key[3],
                keyword_id=key[4],
                keyword=key[5],
                match_type=key[6],
                targeting=key[7],
                keyword_text=keyword_text,
                spend=metrics["spend"],
                sales_7d=metrics["sales"],
                purchases_7d=int(metrics["purchases"]),
                units_7d=int(metrics["units"]),
                impressions=int(metrics["impressions"]),
                clicks=int(metrics["clicks"]),
                ctr=metrics["ctr"],
                cpc=metrics["cpc"],
                cvr=metrics["cvr"],
                acos=metrics["acos"],
                roas=metrics["roas"],
                action_label=action["label"],
                action_priority=action["priority"],
                action_reason=action["reason"],
            )
        )
    return sorted(output, key=lambda row: row.spend, reverse=True)


def _build_search_term_performance(
    rows: Sequence[Mapping[str, Any]],
    *,
    thresholds: WeeklyAdsOptimizationThresholds,
    avg_cpc: Decimal,
    campaign_status_index: Mapping[str, str | None],
    negative_keyword_snapshot: Sequence[NegativeKeywordSnapshotRow],
) -> list[AdsEntityPerformanceRow]:
    buckets = _aggregate_ads_rows(
        rows,
        key_fields=(
            "campaign_id",
            "campaign_name",
            "ad_group_id",
            "ad_group_name",
            "keyword_id",
            "keyword",
            "match_type",
            "targeting",
            "search_term",
        ),
    )
    output: list[AdsEntityPerformanceRow] = []
    for key, bucket in buckets.items():
        metrics = _metrics_from_bucket(bucket)
        action = _search_term_action(
            metrics=metrics,
            thresholds=thresholds,
            avg_cpc=avg_cpc,
            keyword=key[5],
            search_term=key[8],
            match_type=key[6],
        )
        sales = metrics["sales"]
        spend = metrics["spend"]
        campaign_status = _campaign_status_for(campaign_status_index, key[0], key[1])
        is_active = _is_active_campaign_status(campaign_status)
        row = AdsEntityPerformanceRow(
            entity_type="search_term",
            campaign_id=key[0],
            campaign_name=key[1],
            campaign_status=campaign_status,
            is_active_campaign=is_active,
            action_bucket=_action_bucket_for_status(campaign_status),
            ad_group_id=key[2],
            ad_group_name=key[3],
            keyword_id=key[4],
            keyword=key[5],
            match_type=key[6],
            targeting=key[7],
            search_term=key[8],
            spend=spend,
            sales_7d=sales,
            purchases_7d=int(metrics["purchases"]),
            units_7d=int(metrics["units"]),
            impressions=int(metrics["impressions"]),
            clicks=int(metrics["clicks"]),
            ctr=metrics["ctr"],
            cpc=metrics["cpc"],
            cvr=metrics["cvr"],
            acos=metrics["acos"],
            roas=metrics["roas"],
            waste_cost=spend if sales == ZERO else ZERO,
            potential_sales_efficiency=_money(sales - spend),
            action_label=action["label"],
            action_priority=action["priority"],
            action_reason=action["reason"],
        )
        output.append(_apply_negative_snapshot(row, negative_keyword_snapshot))
    return sorted(output, key=lambda row: row.spend, reverse=True)


def _search_term_action_candidates(
    rows: Sequence[AdsEntityPerformanceRow],
) -> list[AdsEntityPerformanceRow]:
    actionable = [
        row
        for row in rows
        if row.action_label not in {"keep_monitoring", "keep_observing", "already_negative"}
        and not row.already_negative
    ]
    priority_rank = {"high": 0, "medium": 1, "low": 2}
    return sorted(
        actionable,
        key=lambda row: (
            priority_rank.get(row.action_priority, 9),
            -float(row.waste_cost if "negative" in row.action_label else row.sales_7d),
            -row.clicks,
        ),
    )


def _build_advertised_product_performance(
    rows: Sequence[Mapping[str, Any]],
    *,
    thresholds: WeeklyAdsOptimizationThresholds,
    cost_index: Mapping[str, list[Mapping[str, Any]]],
    as_of_date: date,
) -> list[AdsEntityPerformanceRow]:
    buckets = _aggregate_ads_rows(
        rows,
        key_fields=("advertised_sku", "advertised_asin"),
    )
    output: list[AdsEntityPerformanceRow] = []
    for key, bucket in buckets.items():
        metrics = _metrics_from_bucket(bucket)
        sku = key[0]
        cost_record = _match_cost(cost_index, sku, as_of_date) if sku else None
        unit_cost = _unit_cost(cost_record) if cost_record else None
        estimated_cogs = None
        contribution = None
        cost_status = "missing_cost" if sku and unit_cost is None else "ok"
        if unit_cost is not None:
            estimated_cogs = _money(unit_cost * Decimal(int(metrics["units"])))
            contribution = _money(metrics["sales"] - estimated_cogs - metrics["spend"])
        action = _advertised_product_action(metrics=metrics, thresholds=thresholds)
        output.append(
            AdsEntityPerformanceRow(
                entity_type="advertised_product",
                campaign_id=None,
                campaign_name=None,
                advertised_sku=sku,
                advertised_asin=key[1],
                spend=metrics["spend"],
                sales_7d=metrics["sales"],
                purchases_7d=int(metrics["purchases"]),
                units_7d=int(metrics["units"]),
                impressions=int(metrics["impressions"]),
                clicks=int(metrics["clicks"]),
                ctr=metrics["ctr"],
                cpc=metrics["cpc"],
                cvr=metrics["cvr"],
                acos=metrics["acos"],
                roas=metrics["roas"],
                unit_standard_cost=unit_cost,
                estimated_ads_cogs=estimated_cogs,
                ads_contribution_proxy=contribution,
                cost_status=cost_status,
                action_label=action["label"],
                action_priority=action["priority"],
                action_reason=action["reason"],
            )
        )
    return sorted(output, key=lambda row: row.spend, reverse=True)


def _build_daily_trend(
    *,
    week_start: date,
    week_end: date,
    campaign_rows: Sequence[Mapping[str, Any]],
    sales_rows: Sequence[Mapping[str, Any]],
) -> list[DailyTrendRow]:
    campaign_by_date: dict[date, dict[str, Any]] = defaultdict(_new_metric_bucket)
    for row in campaign_rows:
        report_date = _as_date(row.get("report_date"))
        if report_date is not None:
            _add_to_bucket(campaign_by_date[report_date], row)
    sales_by_date = {_as_date(row.get("report_date")): row for row in sales_rows}
    output: list[DailyTrendRow] = []
    for offset in range((week_end - week_start).days + 1):
        current_date = week_start + timedelta(days=offset)
        metrics = _metrics_from_bucket(campaign_by_date.get(current_date, _new_metric_bucket()))
        sales = _money(
            _to_decimal(sales_by_date.get(current_date, {}).get("ordered_product_sales_amount"))
        )
        output.append(
            DailyTrendRow(
                report_date=current_date,
                ads_spend=metrics["spend"],
                ads_sales_7d=metrics["sales"],
                ads_purchases_7d=int(metrics["purchases"]),
                ads_units_7d=int(metrics["units"]),
                impressions=int(metrics["impressions"]),
                clicks=int(metrics["clicks"]),
                ctr=metrics["ctr"],
                cpc=metrics["cpc"],
                cvr=metrics["cvr"],
                acos=metrics["acos"],
                roas=metrics["roas"],
                ordered_product_sales=sales,
                tacos=_safe_ratio(metrics["spend"], sales),
            )
        )
    return output


def _build_reconciliation_checks(
    *,
    week_start: date,
    week_end: date,
    generated_at_utc: datetime,
    thresholds: WeeklyAdsOptimizationThresholds,
    campaign_rows: Sequence[Mapping[str, Any]],
    targeting_rows: Sequence[Mapping[str, Any]],
    search_term_rows: Sequence[Mapping[str, Any]],
    advertised_product_rows: Sequence[Mapping[str, Any]],
    sales_rows: Sequence[Mapping[str, Any]],
    overall: OverallSummary,
) -> list[ReconciliationCheck]:
    expected_dates = _date_set(week_start, week_end)
    checks: list[ReconciliationCheck] = []
    table_specs = [
        ("ads_campaign_coverage", "amazon_ads_sp_campaign_daily", campaign_rows, "critical"),
        (
            "ads_search_term_coverage",
            "amazon_ads_sp_search_term_daily",
            search_term_rows,
            "critical",
        ),
        ("ads_targeting_coverage", "amazon_ads_sp_targeting_daily", targeting_rows, "warning"),
        (
            "ads_advertised_product_coverage",
            "amazon_ads_sp_advertised_product_daily",
            advertised_product_rows,
            "warning",
        ),
    ]
    for check_name, source, rows, severity_if_missing in table_specs:
        row_dates = {_as_date(row.get("report_date")) for row in rows}
        available = {item for item in row_dates if item}
        missing = expected_dates - available
        status = "ok"
        severity = "info"
        if missing:
            status = "needs_review" if severity_if_missing == "critical" else "warning"
            severity = severity_if_missing
        checks.append(
            ReconciliationCheck(
                check_name,
                status,
                severity,
                "7 dates",
                f"{len(expected_dates) - len(missing)} dates",
                message=(
                    f"{source} covers full week."
                    if not missing
                    else f"{source} missing dates: "
                    + ", ".join(sorted(day.isoformat() for day in missing))
                ),
            )
        )
    checks.extend(
        _spend_sanity_checks(
            campaign_rows=campaign_rows,
            targeting_rows=targeting_rows,
            search_term_rows=search_term_rows,
            advertised_product_rows=advertised_product_rows,
        )
    )
    checks.append(
        ReconciliationCheck(
            "sales_traffic_context_available",
            "ok" if sales_rows else "warning",
            "warning" if not sales_rows else "info",
            "> 0 rows",
            f"{len(sales_rows)} rows",
            message="Sales & Traffic is required for TACOS and total sales context.",
        )
    )
    if overall.ordered_product_sales > ZERO:
        ratio = _safe_ratio(overall.ads_sales_7d, overall.ordered_product_sales)
        status = "warning" if ratio is not None and ratio > Decimal("1.20") else "ok"
        checks.append(
            ReconciliationCheck(
                "ads_sales_vs_total_sales",
                status,
                "warning" if status == "warning" else "info",
                "ads_sales_7d <= ordered_product_sales * 1.20",
                _optional_ratio_to_string(ratio),
                message=(
                    "Ads attributed sales and Sales & Traffic use different timing windows."
                    if status == "warning"
                    else "Ads attributed sales is within expected context range."
                ),
            )
        )
    else:
        checks.append(
            ReconciliationCheck(
                "ads_sales_vs_total_sales",
                "warning",
                "warning",
                "ordered_product_sales > 0",
                "0",
                message="Total sales context is missing/zero; TACOS cannot be trusted.",
            )
        )
    settlement_diff = overall.settlement_advertising_fee_abs - overall.ads_spend
    settlement_diff_pct = _safe_ratio(settlement_diff, overall.ads_spend)
    checks.append(
        ReconciliationCheck(
            "ads_api_spend_vs_settlement_advertising_fee",
            "ok",
            "info",
            _decimal_to_string(overall.ads_spend),
            _decimal_to_string(overall.settlement_advertising_fee_abs),
            diff=_decimal_to_string(settlement_diff),
            diff_pct=_optional_ratio_to_string(settlement_diff_pct),
            message=(
                "Ads API spend and Settlement advertising fee use different timing; "
                "do not force tie."
            ),
        )
    )
    stable_end = generated_at_utc.date() - timedelta(days=thresholds.stable_lag_days)
    checks.append(
        ReconciliationCheck(
            "ads_stable_cutoff_check",
            "ok" if week_end <= stable_end else "warning",
            "warning" if week_end > stable_end else "info",
            f"week_end <= {stable_end.isoformat()}",
            week_end.isoformat(),
            message="Generate WAOR after Ads attribution has had a stable lag window.",
        )
    )
    return checks


def _spend_sanity_checks(
    *,
    campaign_rows: Sequence[Mapping[str, Any]],
    targeting_rows: Sequence[Mapping[str, Any]],
    search_term_rows: Sequence[Mapping[str, Any]],
    advertised_product_rows: Sequence[Mapping[str, Any]],
) -> list[ReconciliationCheck]:
    campaign_spend = _sum_cost(campaign_rows)
    specs = [
        ("campaign_vs_targeting_spend", targeting_rows),
        ("campaign_vs_search_term_spend", search_term_rows),
        ("campaign_vs_advertised_product_spend", advertised_product_rows),
    ]
    output: list[ReconciliationCheck] = []
    for name, rows in specs:
        other = _sum_cost(rows)
        diff = other - campaign_spend
        diff_ratio = _safe_ratio(abs(diff), campaign_spend)
        status = "ok"
        severity = "info"
        if diff_ratio is not None and diff_ratio > Decimal("0.10"):
            status = "needs_review"
            severity = "critical"
        elif diff_ratio is not None and diff_ratio > Decimal("0.02"):
            status = "warning"
            severity = "warning"
        if campaign_spend == ZERO and other > ZERO:
            status = "needs_review"
            severity = "critical"
        output.append(
            ReconciliationCheck(
                name,
                status,
                severity,
                _decimal_to_string(campaign_spend),
                _decimal_to_string(other),
                diff=_decimal_to_string(diff),
                diff_pct=_optional_ratio_to_string(diff_ratio),
                message=(
                    "Cross-table spend sanity check; tables are different dimensions, not additive."
                ),
            )
        )
    return output


def _build_warnings(
    *,
    checks: Sequence[ReconciliationCheck],
    overall: OverallSummary,
) -> list[WarningEntry]:
    warnings = [
        WarningEntry(
            "waor_scope_policy",
            "info",
            WAOR_SCOPE_NOTE,
            related_source="policy",
        )
    ]
    if overall.campaign_row_count == 0:
        warnings.append(
            WarningEntry(
                "ads_campaign_data_missing",
                "critical",
                "No Ads campaign daily rows were found. Backfill/ingest Ads before using WAOR.",
                related_source="amazon_ads_sp_campaign_daily",
            )
        )
    for check in checks:
        if check.status == "needs_review":
            warnings.append(
                WarningEntry(
                    f"{check.check_name}_needs_review",
                    "critical",
                    check.message,
                    related_source=check.check_name,
                )
            )
        elif check.status == "warning":
            warnings.append(
                WarningEntry(
                    f"{check.check_name}_warning",
                    "warning",
                    check.message,
                    related_source=check.check_name,
                )
            )
    return warnings


def _result_status(
    *,
    campaign_rows: Sequence[Mapping[str, Any]],
    checks: Sequence[ReconciliationCheck],
    warnings: Sequence[WarningEntry],
) -> str:
    if not campaign_rows:
        return "no_ads_data"
    if any(
        check.status == "needs_review"
        and check.check_name in {"ads_campaign_coverage", "ads_search_term_coverage"}
        for check in checks
    ):
        return "needs_backfill"
    if any(check.status == "needs_review" for check in checks):
        return "reviewable_with_warnings"
    if any(check.status == "warning" for check in checks):
        return "reviewable_with_warnings"
    if any(warning.severity in {"warning", "critical"} for warning in warnings):
        return "reviewable_with_warnings"
    return "ok"


def _build_action_items(
    *,
    campaign_performance: Sequence[AdsEntityPerformanceRow],
    targeting_performance: Sequence[AdsEntityPerformanceRow],
    search_term_candidates: Sequence[AdsEntityPerformanceRow],
    advertised_product_performance: Sequence[AdsEntityPerformanceRow],
) -> tuple[list[ActionItem], list[ActionItem]]:
    active_rows: list[ActionItem] = []
    historical_rows: list[ActionItem] = []

    def append_item(item: ActionItem) -> None:
        if item.action_bucket == "historical_or_paused":
            historical_rows.append(item)
        else:
            active_rows.append(item)

    for row in search_term_candidates:
        if row.already_negative:
            continue
        append_item(
            _action_item_from_entity(
                row,
                entity_type="search_term",
                entity_id=row.search_term,
                entity_text=row.search_term,
            )
        )
    for row in campaign_performance:
        if row.action_label != "keep_observing":
            append_item(
                _action_item_from_entity(
                    row,
                    entity_type="campaign",
                    entity_id=row.campaign_id,
                    entity_text=row.campaign_name,
                )
            )
    for row in targeting_performance:
        if row.action_label not in {"keep_observing", "low_relevance_check"}:
            append_item(
                _action_item_from_entity(
                    row,
                    entity_type="targeting",
                    entity_id=row.keyword_id or row.targeting or row.keyword,
                    entity_text=row.keyword_text,
                )
            )
    for row in advertised_product_performance:
        if row.action_label in {"sku_ads_efficiency_review", "sku_ads_scale_candidate"}:
            append_item(
                _action_item_from_entity(
                    row,
                    entity_type="advertised_product",
                    entity_id=row.advertised_sku or row.advertised_asin,
                    entity_text=row.advertised_sku or row.advertised_asin,
                )
            )
    priority_rank = {"high": 0, "medium": 1, "low": 2}
    sort_key = lambda row: priority_rank.get(row.priority, 9)
    return sorted(active_rows, key=sort_key)[:200], sorted(historical_rows, key=sort_key)[:200]


def _action_item_from_entity(
    row: AdsEntityPerformanceRow,
    *,
    entity_type: str,
    entity_id: str | None,
    entity_text: str | None,
) -> ActionItem:
    metric_summary = (
        f"spend={_decimal_to_string(row.spend)}, sales_7d={_decimal_to_string(row.sales_7d)}, "
        f"orders={row.purchases_7d}, clicks={row.clicks}, ACOS={_format_ratio(row.acos)}"
    )
    return ActionItem(
        priority=row.action_priority,
        action_type=row.action_label,
        entity_type=entity_type,
        campaign_name=row.campaign_name,
        ad_group_name=row.ad_group_name,
        entity_id=entity_id,
        entity_text=entity_text,
        metric_summary=metric_summary,
        reason=row.action_reason,
        suggested_manual_action=_suggested_manual_action(row.action_label),
        action_bucket=row.action_bucket,
        campaign_status=row.campaign_status,
        already_negative=row.already_negative,
        negative_scope=row.negative_scope,
        negative_match_type=row.negative_match_type,
        recommended_action=row.recommended_action or row.action_label,
    )


def _campaign_action(
    *,
    metrics: Mapping[str, Any],
    thresholds: WeeklyAdsOptimizationThresholds,
) -> dict[str, str]:
    spend = _to_decimal(metrics["spend"])
    sales = _to_decimal(metrics["sales"])
    purchases = int(metrics["purchases"])
    clicks = int(metrics["clicks"])
    impressions = int(metrics["impressions"])
    acos = metrics["acos"]
    ctr = metrics["ctr"]
    if (
        purchases >= thresholds.min_purchases_to_scale
        and sales >= thresholds.min_sales_to_scale
        and acos is not None
        and acos <= thresholds.target_acos
    ):
        return _action("scale_candidate", "high", "Efficient campaign with enough sales/orders.")
    if sales > ZERO and acos is not None and acos > thresholds.watch_acos:
        return _action(
            "reduce_budget_or_bid_review",
            "high",
            "Campaign has sales but ACOS is above watch threshold.",
        )
    if sales == ZERO and spend >= thresholds.no_sale_cost_threshold:
        return _action("waste_review", "high", "Campaign spent above no-sale threshold.")
    if impressions >= 1000 and ctr is not None and ctr < thresholds.low_ctr_threshold:
        return _action("relevance_review", "medium", "Campaign has high impressions but low CTR.")
    if clicks >= thresholds.no_order_click_threshold and purchases == 0:
        return _action("conversion_review", "medium", "Campaign has clicks but no purchases.")
    return _action("keep_observing", "low", "No firm campaign action from v1 rules.")


def _targeting_action(
    *,
    metrics: Mapping[str, Any],
    thresholds: WeeklyAdsOptimizationThresholds,
    avg_cpc: Decimal,
) -> dict[str, str]:
    spend = _to_decimal(metrics["spend"])
    sales = _to_decimal(metrics["sales"])
    purchases = int(metrics["purchases"])
    clicks = int(metrics["clicks"])
    impressions = int(metrics["impressions"])
    acos = metrics["acos"]
    ctr = metrics["ctr"]
    cvr = metrics["cvr"]
    cpc = metrics["cpc"]
    if (
        purchases >= thresholds.min_purchases_to_scale
        and acos is not None
        and acos <= thresholds.target_acos
    ):
        return _action("increase_bid_review", "high", "Target is converting within target ACOS.")
    if sales > ZERO and acos is not None and acos > thresholds.watch_acos:
        return _action("decrease_bid_review", "high", "Target has sales but inefficient ACOS.")
    if (
        purchases == 0
        and spend >= thresholds.no_sale_cost_threshold
        and clicks >= thresholds.no_order_click_threshold
    ):
        return _action(
            "pause_or_negative_review", "high", "Target spent and clicked without purchase."
        )
    if (
        clicks >= thresholds.no_order_click_threshold
        and cvr is not None
        and cvr < thresholds.low_cvr_threshold
    ):
        return _action("listing_check", "medium", "Clicks exist but conversion rate is weak.")
    if impressions >= 1000 and ctr is not None and ctr < thresholds.low_ctr_threshold:
        return _action("low_relevance_check", "medium", "High impressions with low CTR.")
    if avg_cpc > ZERO and cpc is not None and cpc > avg_cpc * thresholds.high_cpc_multiplier:
        return _action("high_cpc_review", "medium", "CPC is high versus account average.")
    return _action("keep_observing", "low", "No firm targeting action from v1 rules.")


def _search_term_action(
    *,
    metrics: Mapping[str, Any],
    thresholds: WeeklyAdsOptimizationThresholds,
    avg_cpc: Decimal,
    keyword: str | None,
    search_term: str | None,
    match_type: str | None,
) -> dict[str, str]:
    spend = _to_decimal(metrics["spend"])
    sales = _to_decimal(metrics["sales"])
    purchases = int(metrics["purchases"])
    clicks = int(metrics["clicks"])
    impressions = int(metrics["impressions"])
    acos = metrics["acos"]
    ctr = metrics["ctr"]
    cvr = metrics["cvr"]
    cpc = metrics["cpc"]
    if sales == ZERO and purchases == 0 and spend >= thresholds.no_sale_cost_threshold:
        return _action(
            "negative_candidate", "high", "Search term spent above threshold with no sales/orders."
        )
    if purchases == 0 and clicks >= thresholds.no_order_click_threshold:
        return _action(
            "negative_candidate_clicks", "medium", "Search term has clicks but no purchases."
        )
    keyword_normalized = _normalize_text(keyword)
    search_normalized = _normalize_text(search_term)
    if (
        purchases >= thresholds.min_purchases_to_scale
        and acos is not None
        and acos <= thresholds.target_acos
        and search_normalized
        and search_normalized != keyword_normalized
    ):
        return _action(
            "harvest_to_exact_candidate",
            "high",
            "Search term converts efficiently and differs from parent keyword.",
        )
    if (
        purchases >= thresholds.min_purchases_to_scale
        and acos is not None
        and acos <= thresholds.target_acos
        and _normalize_text(match_type) == "exact"
    ):
        return _action(
            "increase_bid_candidate", "medium", "Exact match term converts within target ACOS."
        )
    if sales > ZERO and acos is not None and acos > thresholds.watch_acos:
        return _action("reduce_bid_candidate", "medium", "Search term has sales but ACOS is high.")
    if impressions >= 1000 and ctr is not None and ctr < thresholds.low_ctr_threshold:
        return _action("relevance_review", "medium", "High impressions with low CTR.")
    if (
        clicks >= thresholds.no_order_click_threshold
        and cvr is not None
        and cvr < thresholds.low_cvr_threshold
    ):
        return _action("conversion_review", "medium", "Clicks exist but conversion is weak.")
    if avg_cpc > ZERO and cpc is not None and cpc > avg_cpc * thresholds.high_cpc_multiplier:
        return _action("high_cpc_review", "low", "CPC is high versus account average.")
    return _action("keep_monitoring", "low", "No firm search-term action from v1 rules.")


def _advertised_product_action(
    *,
    metrics: Mapping[str, Any],
    thresholds: WeeklyAdsOptimizationThresholds,
) -> dict[str, str]:
    sales = _to_decimal(metrics["sales"])
    purchases = int(metrics["purchases"])
    spend = _to_decimal(metrics["spend"])
    acos = metrics["acos"]
    if (
        purchases >= thresholds.min_purchases_to_scale
        and acos is not None
        and acos <= thresholds.target_acos
    ):
        return _action(
            "sku_ads_scale_candidate", "medium", "Advertised SKU/ASIN converts within target ACOS."
        )
    if sales > ZERO and acos is not None and acos > thresholds.watch_acos:
        return _action("sku_ads_efficiency_review", "medium", "Advertised SKU/ASIN has high ACOS.")
    if sales == ZERO and spend >= thresholds.no_sale_cost_threshold:
        return _action(
            "sku_ads_waste_review", "medium", "Advertised SKU/ASIN spent with no attributed sales."
        )
    return _action("keep_observing", "low", "No firm advertised-product action from v1 rules.")


def _aggregate_ads_rows(
    rows: Sequence[Mapping[str, Any]],
    *,
    key_fields: tuple[str, ...],
    passthrough_fields: tuple[str, ...] = (),
) -> dict[tuple[str | None, ...], dict[str, Any]]:
    buckets: dict[tuple[str | None, ...], dict[str, Any]] = defaultdict(_new_metric_bucket)
    for row in rows:
        key = tuple(_empty_to_none(row.get(field)) for field in key_fields)
        bucket = buckets[key]
        _add_to_bucket(bucket, row)
        for field_name in passthrough_fields:
            bucket[field_name] = _empty_to_none(row.get(field_name)) or bucket.get(field_name)
    return dict(buckets)


def _new_metric_bucket() -> dict[str, Any]:
    return {
        "impressions": 0,
        "clicks": 0,
        "spend": ZERO,
        "sales": ZERO,
        "purchases": 0,
        "units": 0,
    }


def _add_to_bucket(bucket: dict[str, Any], row: Mapping[str, Any]) -> None:
    bucket["impressions"] += _int_metric(row, "impressions")
    bucket["clicks"] += _int_metric(row, "clicks")
    bucket["spend"] += _to_decimal(row.get("cost"))
    bucket["sales"] += _to_decimal(row.get("sales_7d"))
    bucket["purchases"] += _int_metric(row, "purchases_7d")
    bucket["units"] += _int_metric(row, "units_sold_clicks_7d")


def _metrics_from_bucket(bucket: Mapping[str, Any]) -> dict[str, Any]:
    spend = _money(_to_decimal(bucket.get("spend")))
    sales = _money(_to_decimal(bucket.get("sales")))
    clicks = int(bucket.get("clicks") or 0)
    impressions = int(bucket.get("impressions") or 0)
    purchases = int(bucket.get("purchases") or 0)
    return {
        "spend": spend,
        "sales": sales,
        "purchases": purchases,
        "units": int(bucket.get("units") or 0),
        "impressions": impressions,
        "clicks": clicks,
        "ctr": _safe_ratio(Decimal(clicks), Decimal(impressions)),
        "cpc": _safe_money_ratio(spend, Decimal(clicks)),
        "cvr": _safe_ratio(Decimal(purchases), Decimal(clicks)),
        "acos": _safe_ratio(spend, sales),
        "roas": _safe_money_ratio(sales, spend),
    }


def _sum_cost(rows: Sequence[Mapping[str, Any]]) -> Decimal:
    return _money(sum((_to_decimal(row.get("cost")) for row in rows), ZERO))


def _build_cost_index(rows: Iterable[Mapping[str, Any]]) -> dict[str, list[Mapping[str, Any]]]:
    index: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        sku = _empty_to_none(row.get("seller_sku"))
        if sku:
            index[sku].append(row)
    for entries in index.values():
        entries.sort(key=lambda row: _as_date(row.get("effective_from")) or date.min, reverse=True)
    return dict(index)


def _match_cost(
    cost_index: Mapping[str, list[Mapping[str, Any]]], sku: str, effective_date: date
) -> Mapping[str, Any] | None:
    entries = cost_index.get(sku) or []
    for row in entries:
        effective_from = _as_date(row.get("effective_from")) or date.min
        effective_to = _as_date(row.get("effective_to")) or date.max
        if effective_from <= effective_date <= effective_to:
            return row
    return None


def _unit_cost(row: Mapping[str, Any]) -> Decimal:
    return _money(
        _to_decimal(row.get("product_cost"))
        + _to_decimal(row.get("first_mile_cost"))
        + _to_decimal(row.get("packaging_cost"))
        + _to_decimal(row.get("other_unit_cost"))
    )


def _build_raw_metadata(
    *,
    marketplace_id: str,
    profile_id: str | None,
    week_start: date,
    week_end: date,
    generated_at_utc: datetime,
    campaign_rows: Sequence[Mapping[str, Any]],
    targeting_rows: Sequence[Mapping[str, Any]],
    search_term_rows: Sequence[Mapping[str, Any]],
    advertised_product_rows: Sequence[Mapping[str, Any]],
    sales_rows: Sequence[Mapping[str, Any]],
    sku_cost_rows: Sequence[Mapping[str, Any]],
    negative_keyword_rows: Sequence[Mapping[str, Any]],
    settlement: Mapping[str, Any],
    thresholds: WeeklyAdsOptimizationThresholds,
) -> dict[str, Any]:
    return {
        "report_type": REPORT_TYPE,
        "version": REPORT_VERSION,
        "marketplace_id": marketplace_id,
        "profile_id": profile_id,
        "week_start": week_start.isoformat(),
        "week_end": week_end.isoformat(),
        "generated_at_utc": generated_at_utc.isoformat(),
        "source_tables": (
            "amazon_ads_sp_campaign_daily, amazon_ads_sp_targeting_daily, "
            "amazon_ads_sp_search_term_daily, amazon_ads_sp_advertised_product_daily, "
            "amazon_sales_traffic_daily, amazon_sku_cost, amazon_settlement_transaction, "
            "optional_negative_keyword_snapshot"
        ),
        "ads_campaign_row_count": len(campaign_rows),
        "ads_targeting_row_count": len(targeting_rows),
        "ads_search_term_row_count": len(search_term_rows),
        "ads_advertised_product_row_count": len(advertised_product_rows),
        "sales_traffic_row_count": len(sales_rows),
        "sku_cost_row_count": len(sku_cost_rows),
        "negative_keyword_snapshot_row_count": len(negative_keyword_rows),
        "settlement_row_count": settlement.get("settlement_row_count"),
        "thresholds": thresholds.to_dict(),
        "note": WAOR_SCOPE_NOTE,
    }


def _action(label: str, priority: str, reason: str) -> dict[str, str]:
    return {"label": label, "priority": priority, "reason": reason}


def _suggested_manual_action(action_label: str) -> str:
    suggestions = {
        "negative_candidate": "Review and consider adding as negative exact/phrase.",
        "negative_candidate_clicks": "Review query relevance before adding as negative.",
        "harvest_to_exact_candidate": "Consider adding as exact keyword in the relevant ad group.",
        "increase_bid_candidate": "Review current bid and consider a controlled bid increase.",
        "reduce_bid_candidate": "Review current bid and consider lowering bid or budget.",
        "scale_candidate": "Review budget headroom and consider increasing budget/bids carefully.",
        "reduce_budget_or_bid_review": "Review campaign budget/bids and reduce exposure if needed.",
        "waste_review": "Review campaign relevance and budget leakage.",
        "increase_bid_review": "Review target bid and consider controlled increase.",
        "decrease_bid_review": "Review target bid and consider controlled decrease.",
        "pause_or_negative_review": "Review whether to pause target or add negative targeting.",
        "listing_check": "Check product page, price, review count and main image.",
        "conversion_review": "Check listing and search-term relevance before bid changes.",
        "relevance_review": "Check keyword relevance and match type.",
        "low_relevance_check": "Consider narrowing match type or reducing exposure.",
        "high_cpc_review": "Review bid pressure and CPC competitiveness.",
        "sku_ads_scale_candidate": "Review SKU stock and margins before scaling ads.",
        "sku_ads_efficiency_review": "Review SKU ad efficiency and listing conversion.",
        "sku_ads_waste_review": "Review whether this SKU should continue receiving ad spend.",
        "already_negative": "Already covered by existing negative keyword snapshot; no duplicate action needed.",
    }
    return suggestions.get(action_label, "Keep monitoring; no immediate manual change suggested.")


def _date_set(start_date: date, end_date: date) -> set[date]:
    return {
        start_date + timedelta(days=offset) for offset in range((end_date - start_date).days + 1)
    }


def _validate_week_start(week_start: date) -> None:
    if week_start.weekday() not in {0, 5}:
        raise ValueError("week_start must be a Monday or Saturday")


def parse_week_start(value: str) -> date:
    try:
        week_start = date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError("week_start must use YYYY-MM-DD format") from exc
    _validate_week_start(week_start)
    return week_start


def _int_metric(values: Mapping[str, Any], key: str) -> int:
    return int(values.get(key) or 0)


def _to_decimal(value: Any) -> Decimal:
    if value is None:
        return ZERO
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return ZERO


def _as_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _empty_to_none(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _first_non_empty(values: Iterable[str | None]) -> str | None:
    for value in values:
        if value:
            return value
    return None


def _normalize_text(value: str | None) -> str:
    return " ".join((value or "").strip().lower().split())


def _money(value: Decimal) -> Decimal:
    return value.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def _ratio(value: Decimal) -> Decimal:
    return value.quantize(RATIO_QUANT, rounding=ROUND_HALF_UP)


def _safe_ratio(numerator: Decimal, denominator: Decimal) -> Decimal | None:
    if denominator == ZERO:
        return None
    return _ratio(numerator / denominator)


def _safe_money_ratio(numerator: Decimal, denominator: Decimal) -> Decimal | None:
    if denominator == ZERO:
        return None
    return _money(numerator / denominator)


def _decimal_to_string(value: Decimal) -> str:
    return format(_money(value), "f")


def _optional_decimal_to_string(value: Decimal | None) -> str | None:
    return None if value is None else _decimal_to_string(value)


def _optional_ratio_to_string(value: Decimal | None) -> str | None:
    return None if value is None else format(_ratio(value), "f")


def _format_money(value: Decimal, currency: str | None) -> str:
    prefix = f"{currency} " if currency else ""
    return f"{prefix}{_decimal_to_string(value)}"


def _format_ratio(value: Decimal | None) -> str:
    if value is None:
        return "-"
    return f"{(_ratio(value) * Decimal('100')).quantize(MONEY_QUANT)}%"


def _format_decimal(value: Decimal | None) -> str:
    return "-" if value is None else _decimal_to_string(value)


def _json_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return _decimal_to_string(value)
    if isinstance(value, date | datetime):
        return value.isoformat()
    if isinstance(value, dict):
        return _json_safe_mapping(value)
    if isinstance(value, list):
        return [_json_value(item) for item in value]
    return value


def _json_safe_mapping(values: Mapping[str, Any]) -> dict[str, Any]:
    return {key: _json_value(value) for key, value in values.items()}


def _xlsx_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, date | datetime):
        return value.isoformat()
    if isinstance(value, dict):
        return json.dumps(_json_safe_mapping(value), ensure_ascii=False, sort_keys=True)
    return value


__all__ = [
    "DEFAULT_OUTPUT_ROOT",
    "WeeklyAdsOptimizationResult",
    "WeeklyAdsOptimizationService",
    "WeeklyAdsOptimizationThresholds",
    "build_weekly_ads_optimization_workbook",
    "parse_week_start",
]
