from __future__ import annotations

from typing import Any, Mapping

from openpyxl.styles import Alignment, Font, PatternFill

HEADER_ZH: dict[str, str] = {
    "metric": "指标",
    "metric_group": "指标组",
    "metric_name": "指标名",
    "value": "值",
    "unit": "单位",
    "currency": "币种",
    "notes": "说明",
    "status": "状态",
    "severity": "严重级别",
    "warning_code": "警告代码",
    "message": "信息",
    "key": "键",
    "period": "周期",
    "section": "部分",
    "date": "日期",
    "report_date": "报表日期",
    "seller_sku": "卖家SKU",
    "asin": "ASIN",
    "sku": "SKU",
    "product_name": "商品名称",
    "units": "件数",
    "units_ordered": "订购件数",
    "ordered_units": "订购件数",
    "ordered_product_sales": "订购销售额",
    "ordered_product_sales_amount": "订购销售额",
    "sessions": "会话数",
    "page_views": "页面浏览量",
    "unit_session_percentage": "单位会话率",
    "campaign_id": "广告活动ID",
    "campaign_name": "广告活动名称",
    "ad_group_id": "广告组ID",
    "ad_group_name": "广告组名称",
    "keyword_id": "关键词ID",
    "keyword": "关键词",
    "match_type": "匹配类型",
    "targeting": "投放目标",
    "search_term": "搜索词",
    "spend": "花费",
    "cost": "成本/花费",
    "sales_7d": "7天归因销售额",
    "purchases_7d": "7天归因订单数",
    "units_7d": "7天归因件数",
    "impressions": "曝光量",
    "clicks": "点击量",
    "ctr": "点击率",
    "cpc": "单次点击成本",
    "cvr": "转化率",
    "acos": "ACOS",
    "roas": "ROAS",
    "tacos": "TACOS",
    "action_label": "动作标签",
    "action_label_zh": "动作标签中文",
    "action_type": "动作类型",
    "action_type_zh": "动作类型中文",
    "priority": "优先级",
    "priority_zh": "优先级中文",
    "reason": "原因",
    "reason_zh": "原因中文",
    "action_reason": "动作原因",
    "action_reason_zh": "动作原因中文",
    "suggested_manual_action": "建议人工动作",
    "suggested_manual_action_zh": "建议人工动作中文",
    "manual_review_note": "人工复核说明",
    "manual_review_note_zh": "人工复核说明中文",
    "do_not_auto_apply": "不要自动执行",
    "profit_bucket": "利润桶",
    "amount_category": "金额类别",
    "amount": "金额",
    "absolute_amount": "绝对金额",
    "share_of_product_sales": "销售额占比",
    "share_of_settlement_net": "结算净额占比",
    "row_count": "行数",
    "settlement_net_amount": "结算净额",
    "product_sales_amount": "商品销售额",
    "product_sales_units": "商品销售件数",
    "unit_standard_cost": "单位标准成本",
    "internal_cogs": "内部COGS",
    "estimated_profit_after_cogs": "扣COGS后估算利润",
    "estimated_operating_profit": "估算经营利润",
    "profit_margin": "利润率",
    "revenue_share": "收入占比",
    "check_name": "检查项",
    "expected": "预期值",
    "actual": "实际值",
    "diff": "差异",
    "diff_pct": "差异百分比",
    "scope_note": "口径说明",
}

METRIC_ZH: dict[str, str] = {
    "Report Type": "报表类型",
    "Marketplace ID": "市场ID",
    "Ads Profile ID": "广告Profile ID",
    "Month": "月份",
    "Period Start": "周期开始",
    "Period End": "周期结束",
    "Week Start": "周开始",
    "Week End": "周结束",
    "Status": "状态",
    "Report Status": "报表状态",
    "Currency": "币种",
    "Settlement Rows": "Settlement行数",
    "Settlement Net Amount": "Settlement净额",
    "Product Sales Amount": "商品销售额",
    "Product Sales Units": "商品销售件数",
    "Internal COGS": "内部COGS",
    "Estimated Operating Profit": "估算经营利润",
    "Profit Margin": "利润率",
    "Advertising Cost": "广告费用",
    "FBA Fee": "FBA费用",
    "Amazon Fee": "亚马逊平台费用",
    "Refund": "退款",
    "Promotion Cost": "促销成本",
    "Promotion Fee": "促销费用",
    "Reimbursement": "赔偿",
    "SKU Profit Table Scope": "SKU利润表口径",
    "Ordered Product Sales": "订购商品销售额",
    "Units Ordered": "订购件数",
    "Total Order Items": "订单商品项数",
    "Sessions": "会话数",
    "Page Views": "页面浏览量",
    "Unit Session Percentage": "单位会话率",
    "Ads Spend": "广告花费",
    "Ads Sales 7d": "7天归因广告销售额",
    "Ads Purchases 7d": "7天归因广告订单数",
    "ACOS": "ACOS广告销售成本比",
    "ROAS": "ROAS广告投入产出比",
    "TACOS": "TACOS总广告成本比",
    "Estimated COGS": "估算COGS",
    "Gross Margin Before Ads": "广告前毛利",
    "Contribution After Ads": "扣广告后贡献利润",
    "Contribution Margin After Ads": "扣广告后贡献利润率",
    "Settlement Net Preview": "Settlement净额预览",
    "Alert Count": "告警数量",
    "ads_spend": "广告花费",
    "ads_sales_7d": "7天归因广告销售额",
    "ads_purchases_7d": "7天归因广告订单数",
    "clicks": "点击量",
    "impressions": "曝光量",
    "acos": "ACOS广告销售成本比",
    "roas": "ROAS广告投入产出比",
    "tacos": "TACOS总广告成本比",
    "ordered_product_sales": "订购商品销售额",
    "settlement_advertising_fee_abs": "Settlement广告费绝对值",
    "campaign_count": "广告活动数量",
    "action_item_count": "动作候选数量",
    "negative_candidate_count": "否词候选数量",
    "harvest_candidate_count": "收词候选数量",
    "warning_count": "警告数量",
    "target_acos": "目标ACOS",
    "watch_acos": "观察ACOS阈值",
    "target_tacos": "目标TACOS",
    "no_sale_cost_threshold": "无销售花费阈值",
    "no_order_click_threshold": "无订单点击阈值",
    "min_purchases_to_scale": "加价最小订单阈值",
    "min_sales_to_scale": "加价最小销售额阈值",
    "low_ctr_threshold": "低CTR阈值",
    "low_cvr_threshold": "低CVR阈值",
    "high_cpc_multiplier": "高CPC倍数阈值",
    "stable_lag_days": "稳定延迟天数",
}

ACTION_ZH: dict[str, str] = {
    "harvest_to_exact_candidate": "建议收词到精准匹配候选",
    "negative_candidate": "否定关键词候选",
    "negative_candidate_clicks": "点击无转化否词候选",
    "reduce_bid_candidate": "建议降价候选",
    "decrease_bid_review": "建议复核并下调竞价",
    "increase_bid_review": "建议复核并上调竞价",
    "pause_or_negative_review": "建议暂停或否定复核",
    "high_cpc_review": "高CPC复核",
    "keep_monitoring": "继续观察",
}

PRIORITY_ZH = {"high": "高", "medium": "中", "low": "低", "info": "信息"}

REASON_ZH: dict[str, str] = {
    "Search term converts efficiently and differs from parent keyword.": "搜索词转化效率高，且不同于父关键词，可考虑收词。",
    "Search term spent above threshold with no sales/orders.": "搜索词花费超过阈值但没有销售/订单。",
    "Search term has clicks but no purchases.": "搜索词有点击但没有购买，需要复核相关性。",
    "Search term has sales but ACOS is high.": "搜索词有销售但ACOS偏高。",
    "CPC is high versus account average.": "CPC相对账户平均水平偏高。",
    "No firm search-term action from v1 rules.": "v1规则暂未给出明确搜索词动作。",
    "Target has sales but inefficient ACOS.": "投放目标有销售但ACOS效率偏低。",
    "Target spent and clicked without purchase.": "投放目标有花费和点击但没有购买。",
    "Target is converting within target ACOS.": "投放目标在目标ACOS范围内转化。",
}

SUGGESTED_ACTION_ZH: dict[str, str] = {
    "Consider adding as exact keyword in the relevant ad group.": "考虑添加为相关广告组的精准关键词。",
    "Review and consider adding as negative exact/phrase.": "复核后考虑添加为精准/词组否定关键词。",
    "Review current bid and consider lowering bid or budget.": "复核当前竞价，并考虑降低竞价或预算。",
    "Review query relevance before adding as negative.": "添加否词前先复核搜索词相关性。",
    "Review bid pressure and CPC competitiveness.": "复核竞价压力和CPC竞争水平。",
    "Review target bid and consider controlled decrease.": "复核目标竞价，并考虑小幅可控下调。",
    "Review target bid and consider controlled increase.": "复核目标竞价，并考虑小幅可控上调。",
    "Review whether to pause target or add negative targeting.": "复核是否暂停该目标或添加否定投放。",
}

MANUAL_REVIEW_NOTE_ZH = (
    "仅作为候选建议。不要自动执行；请先在 Amazon Ads Console 中人工复核相关性、库存、"
    "当前竞价、预算和已有否词。"
)


def xlsx_header_label(key: str) -> str:
    zh = HEADER_ZH.get(key)
    return f"{key} / {zh}" if zh else key


def bilingual_metric_label(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    zh = METRIC_ZH.get(value)
    return f"{value} / {zh}" if zh else value


def action_label_zh(value: Any) -> str:
    if not isinstance(value, str) or not value:
        return ""
    return ACTION_ZH.get(value, value)


def priority_zh(value: Any) -> str:
    if not isinstance(value, str) or not value:
        return ""
    return PRIORITY_ZH.get(value.lower(), value)


def reason_zh(value: Any) -> str:
    if not isinstance(value, str) or not value:
        return ""
    return REASON_ZH.get(value, value)


def suggested_action_zh(value: Any) -> str:
    if not isinstance(value, str) or not value:
        return ""
    return SUGGESTED_ACTION_ZH.get(value, value)


def manual_review_note_zh(value: Any) -> str:
    if not isinstance(value, str) or not value:
        return ""
    if "Do not auto-apply" in value:
        return MANUAL_REVIEW_NOTE_ZH
    return value


def add_action_translation_columns(row: Mapping[str, Any]) -> dict[str, Any]:
    payload = dict(row)
    action_value = payload.get("action_type", payload.get("action_label"))
    if action_value is not None:
        if "action_type" in payload:
            payload.setdefault("action_type_zh", action_label_zh(action_value))
        if "action_label" in payload:
            payload.setdefault("action_label_zh", action_label_zh(action_value))
    if "priority" in payload:
        payload.setdefault("priority_zh", priority_zh(payload.get("priority")))
    if "reason" in payload:
        payload.setdefault("reason_zh", reason_zh(payload.get("reason")))
    if "action_reason" in payload:
        payload.setdefault("action_reason_zh", reason_zh(payload.get("action_reason")))
    if "suggested_manual_action" in payload:
        payload.setdefault(
            "suggested_manual_action_zh",
            suggested_action_zh(payload.get("suggested_manual_action")),
        )
    if "manual_review_note" in payload:
        payload.setdefault(
            "manual_review_note_zh",
            manual_review_note_zh(payload.get("manual_review_note")),
        )
    return payload


def add_bilingual_readme_sheet(
    workbook: Any,
    *,
    title_en: str,
    title_zh: str,
    period: str,
    status: str,
    scope_en: str,
    scope_zh: str,
) -> None:
    sheet = workbook.create_sheet("00_Readme_说明", 0)
    rows = [
        ("Report / 报表", f"{title_en} / {title_zh}"),
        ("Period / 周期", period),
        ("Status / 状态", status),
        ("Scope / 口径", f"{scope_en}\n{scope_zh}"),
        (
            "How to use / 使用方式",
            "Use JSON as the machine-readable source of truth; use this XLSX for manual review.\n"
            "JSON 是机器可读的标准结果；本 XLSX 用于人工复核、筛选和沟通。",
        ),
        (
            "Language / 语言",
            "Fixed report labels and notes are bilingual. Amazon-native campaign names, "
            "search terms, SKU/ASIN and raw identifiers are kept as source data.\n"
            "固定报表标签和说明为中英文双语；Amazon 原始广告活动名称、搜索词、SKU/ASIN "
            "和原始ID保持源数据原样。",
        ),
    ]
    sheet.append(["field / 字段", "value / 值"])
    for row in rows:
        sheet.append(list(row))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 96
    sheet.freeze_panes = "A2"
