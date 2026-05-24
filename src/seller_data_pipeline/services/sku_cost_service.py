from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from seller_data_pipeline.db.repositories.sku_cost_repo import (
    SkuCandidateRecord,
    SkuCostRecord,
    SkuCostRepo,
    SkuCostWriteRecord,
)

SKU_COST_INPUT_SHEET = "sku_cost_input"
README_SHEET = "README"
DEFAULT_CURRENCY = "USD"
TEMPLATE_COLUMNS = (
    "marketplace_id",
    "seller_sku",
    "asin",
    "product_name",
    "sku_sources",
    "latest_source_date",
    "current_product_cost",
    "current_first_mile_cost",
    "current_packaging_cost",
    "current_other_unit_cost",
    "current_currency",
    "current_effective_from",
    "current_effective_to",
    "current_remark",
    "current_updated_at",
    "new_product_cost",
    "new_first_mile_cost",
    "new_packaging_cost",
    "new_other_unit_cost",
    "new_currency",
    "new_effective_from",
    "new_effective_to",
    "purchase_or_batch_note",
    "new_remark",
)
EDITABLE_COLUMNS = {
    "new_product_cost",
    "new_first_mile_cost",
    "new_packaging_cost",
    "new_other_unit_cost",
    "new_currency",
    "new_effective_from",
    "new_effective_to",
    "purchase_or_batch_note",
    "new_remark",
}
COST_COLUMNS = (
    "new_product_cost",
    "new_first_mile_cost",
    "new_packaging_cost",
    "new_other_unit_cost",
)
EDITABLE_TRIGGER_COLUMNS = EDITABLE_COLUMNS - {"new_currency"}


@dataclass(frozen=True)
class SkuCostTemplateRow:
    marketplace_id: str
    seller_sku: str
    asin: str | None
    product_name: str | None
    sku_sources: str
    latest_source_date: date | None
    current_cost: SkuCostRecord | None = None


@dataclass(frozen=True)
class SkuCostTemplateExportResult:
    output_path: Path
    marketplace_id: str
    row_count: int
    generated_at_utc: datetime


@dataclass(frozen=True)
class SkuCostImportIssue:
    row_number: int
    seller_sku: str | None
    severity: str
    message: str


@dataclass(frozen=True)
class SkuCostImportCandidate:
    row_number: int
    record: SkuCostWriteRecord


@dataclass(frozen=True)
class SkuCostImportResult:
    workbook_path: Path
    dry_run: bool
    candidate_rows: int
    inserted_rows: int
    updated_rows: int
    skipped_existing_rows: int
    closed_previous_rows: int
    issues: list[SkuCostImportIssue]

    @property
    def has_errors(self) -> bool:
        return any(issue.severity == "error" for issue in self.issues)

    @property
    def status(self) -> str:
        if self.has_errors:
            return "blocked"
        if self.candidate_rows == 0:
            return "no_rows"
        return "dry_run_ok" if self.dry_run else "executed"


class SkuCostWorkbookService:
    """Manual Excel workflow for SKU standard cost maintenance."""

    def export_template(
        self,
        *,
        repo: SkuCostRepo,
        marketplace_id: str,
        output_path: Path,
        currency: str = DEFAULT_CURRENCY,
        delete_existing: bool = True,
    ) -> SkuCostTemplateExportResult:
        generated_at = datetime.now(UTC).replace(microsecond=0)
        candidates = repo.fetch_sku_candidates(marketplace_id=marketplace_id)
        current_costs = repo.fetch_latest_sku_costs(marketplace_id=marketplace_id)
        rows = build_template_rows(candidates=candidates, current_costs=current_costs)
        if delete_existing and output_path.exists():
            output_path.unlink()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = build_sku_cost_workbook(
            rows=rows,
            marketplace_id=marketplace_id,
            generated_at_utc=generated_at,
            default_currency=currency,
        )
        workbook.save(output_path)
        return SkuCostTemplateExportResult(
            output_path=output_path,
            marketplace_id=marketplace_id,
            row_count=len(rows),
            generated_at_utc=generated_at,
        )

    def import_workbook(
        self,
        *,
        repo: SkuCostRepo,
        workbook_path: Path,
        dry_run: bool = True,
        update_existing: bool = False,
        close_previous_open: bool = True,
    ) -> SkuCostImportResult:
        candidates, issues = read_sku_cost_workbook(workbook_path)
        if any(issue.severity == "error" for issue in issues):
            return SkuCostImportResult(
                workbook_path=workbook_path,
                dry_run=dry_run,
                candidate_rows=len(candidates),
                inserted_rows=0,
                updated_rows=0,
                skipped_existing_rows=0,
                closed_previous_rows=0,
                issues=issues,
            )

        inserted = 0
        updated = 0
        skipped = 0
        closed_previous = 0
        for candidate in candidates:
            record = candidate.record
            exists = repo.sku_cost_exists(
                marketplace_id=record.marketplace_id,
                seller_sku=record.seller_sku,
                effective_from=record.effective_from,
            )
            if exists and not update_existing:
                skipped += 1
                issues.append(
                    SkuCostImportIssue(
                        row_number=candidate.row_number,
                        seller_sku=record.seller_sku,
                        severity="warning",
                        message=(
                            "Skipped existing cost row with the same "
                            "marketplace_id + seller_sku + effective_from."
                        ),
                    )
                )
                continue

            if not dry_run:
                if exists:
                    repo.update_sku_cost(record)
                    updated += 1
                else:
                    if close_previous_open:
                        closed_previous += repo.close_previous_open_cost(
                            marketplace_id=record.marketplace_id,
                            seller_sku=record.seller_sku,
                            new_effective_from=record.effective_from,
                        )
                    repo.insert_sku_cost(record)
                    inserted += 1
            else:
                if exists:
                    updated += 1 if update_existing else 0
                else:
                    inserted += 1

        if not dry_run:
            repo.commit()
        return SkuCostImportResult(
            workbook_path=workbook_path,
            dry_run=dry_run,
            candidate_rows=len(candidates),
            inserted_rows=inserted,
            updated_rows=updated,
            skipped_existing_rows=skipped,
            closed_previous_rows=closed_previous,
            issues=issues,
        )


def build_template_rows(
    *,
    candidates: list[SkuCandidateRecord],
    current_costs: dict[str, SkuCostRecord],
) -> list[SkuCostTemplateRow]:
    rows = []
    seen = set()
    for candidate in sorted(candidates, key=lambda item: item.seller_sku.lower()):
        if candidate.seller_sku in seen:
            continue
        seen.add(candidate.seller_sku)
        current_cost = current_costs.get(candidate.seller_sku)
        rows.append(
            SkuCostTemplateRow(
                marketplace_id=candidate.marketplace_id,
                seller_sku=candidate.seller_sku,
                asin=candidate.asin or (current_cost.asin if current_cost else None),
                product_name=candidate.product_name,
                sku_sources=candidate.sku_sources,
                latest_source_date=candidate.latest_source_date,
                current_cost=current_cost,
            )
        )
    return rows


def build_sku_cost_workbook(
    *,
    rows: list[SkuCostTemplateRow],
    marketplace_id: str,
    generated_at_utc: datetime,
    default_currency: str,
) -> Workbook:
    workbook = Workbook()
    input_sheet = workbook.active
    input_sheet.title = SKU_COST_INPUT_SHEET
    _write_input_sheet(
        input_sheet,
        rows=rows,
        marketplace_id=marketplace_id,
        generated_at_utc=generated_at_utc,
        default_currency=default_currency,
    )
    readme = workbook.create_sheet(README_SHEET)
    _write_readme_sheet(readme, marketplace_id=marketplace_id, generated_at_utc=generated_at_utc)
    workbook.active = 0
    return workbook


def read_sku_cost_workbook(
    workbook_path: Path,
) -> tuple[list[SkuCostImportCandidate], list[SkuCostImportIssue]]:
    workbook = load_workbook(workbook_path, data_only=True)
    if SKU_COST_INPUT_SHEET not in workbook.sheetnames:
        return [], [
            SkuCostImportIssue(
                row_number=0,
                seller_sku=None,
                severity="error",
                message=f"Workbook does not contain sheet {SKU_COST_INPUT_SHEET!r}.",
            )
        ]
    sheet = workbook[SKU_COST_INPUT_SHEET]
    header_values = [_cell_text(cell.value) for cell in sheet[1]]
    header_map = {value: index + 1 for index, value in enumerate(header_values) if value}
    missing_headers = [column for column in TEMPLATE_COLUMNS if column not in header_map]
    if missing_headers:
        return [], [
            SkuCostImportIssue(
                row_number=1,
                seller_sku=None,
                severity="error",
                message="Missing required column(s): " + ", ".join(missing_headers),
            )
        ]

    candidates: list[SkuCostImportCandidate] = []
    issues: list[SkuCostImportIssue] = []
    for row_number in range(2, sheet.max_row + 1):
        row = {
            column: sheet.cell(row=row_number, column=header_map[column]).value
            for column in TEMPLATE_COLUMNS
        }
        if _is_blank_row(row):
            continue
        if not _has_editable_input(row):
            continue
        candidate, row_issues = _parse_import_row(row, row_number=row_number)
        issues.extend(row_issues)
        if candidate is not None:
            candidates.append(candidate)
    return candidates, issues


def _write_input_sheet(
    sheet: Any,
    *,
    rows: list[SkuCostTemplateRow],
    marketplace_id: str,
    generated_at_utc: datetime,
    default_currency: str,
) -> None:
    sheet.append(TEMPLATE_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for template_row in rows:
        current = template_row.current_cost
        sheet.append(
            [
                template_row.marketplace_id,
                template_row.seller_sku,
                template_row.asin,
                template_row.product_name,
                template_row.sku_sources,
                _date_value(template_row.latest_source_date),
                _decimal_value(current.product_cost if current else None),
                _decimal_value(current.first_mile_cost if current else None),
                _decimal_value(current.packaging_cost if current else None),
                _decimal_value(current.other_unit_cost if current else None),
                current.currency if current else None,
                _date_value(current.effective_from if current else None),
                _date_value(current.effective_to if current else None),
                current.remark if current else None,
                _datetime_text(current.updated_at if current else None),
                None,
                None,
                None,
                None,
                default_currency,
                None,
                None,
                None,
                None,
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _set_column_widths(sheet)
    _shade_columns(sheet)
    _add_validations(sheet, max_row=max(2, sheet.max_row))
    _add_metadata_comment(sheet, marketplace_id=marketplace_id, generated_at_utc=generated_at_utc)


def _write_readme_sheet(sheet: Any, *, marketplace_id: str, generated_at_utc: datetime) -> None:
    rows = [
        ("SKU Cost Template", ""),
        ("marketplace_id", marketplace_id),
        ("generated_at_utc", generated_at_utc.isoformat()),
        (
            "Workflow",
            "1) Export template. 2) Fill only new_* columns. 3) Import with dry-run. "
            "4) Re-run with --execute after review.",
        ),
        (
            "Idempotency",
            "Import skips rows where marketplace_id + seller_sku + new_effective_from "
            "already exists, unless --update-existing is used.",
        ),
        (
            "Cost date",
            "new_effective_from is the cost effective date. For current practice, "
            "use the purchase/import/batch start date you want profit reports to match.",
        ),
        (
            "Currency",
            "Use the marketplace financial currency, for example USD for Amazon US, "
            "unless a later FX feature is added.",
        ),
        (
            "Required for import",
            "marketplace_id, seller_sku, new_product_cost, new_currency, new_effective_from.",
        ),
        (
            "Optional cost components",
            "Blank new_first_mile_cost/new_packaging_cost/new_other_unit_cost are imported as 0.",
        ),
        (
            "Safety",
            "Importer does not read current_* values as new costs; "
            "those columns are reference only.",
        ),
    ]
    for row in rows:
        sheet.append(row)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 120
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _shade_columns(sheet: Any) -> None:
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    reference_fill = PatternFill("solid", fgColor="D9EAD3")
    for index, column in enumerate(TEMPLATE_COLUMNS, start=1):
        fill = editable_fill if column in EDITABLE_COLUMNS else reference_fill
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=index).fill = fill


def _add_validations(sheet: Any, *, max_row: int) -> None:
    currency_col = _column_letter("new_currency")
    currency_validation = DataValidation(
        type="list",
        formula1='"USD,CNY,GBP,EUR,CAD,MXN"',
        allow_blank=False,
    )
    sheet.add_data_validation(currency_validation)
    currency_validation.add(f"{currency_col}2:{currency_col}{max_row}")

    for column in ("new_effective_from", "new_effective_to"):
        date_col = _column_letter(column)
        date_validation = DataValidation(type="date", operator="greaterThan", formula1="1900-01-01")
        sheet.add_data_validation(date_validation)
        date_validation.add(f"{date_col}2:{date_col}{max_row}")

    for column in COST_COLUMNS:
        cost_col = _column_letter(column)
        cost_validation = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
        )
        sheet.add_data_validation(cost_validation)
        cost_validation.add(f"{cost_col}2:{cost_col}{max_row}")


def _add_metadata_comment(sheet: Any, *, marketplace_id: str, generated_at_utc: datetime) -> None:
    sheet.cell(row=1, column=1).comment = None
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddHeader.center.text = (
        f"SKU cost template - {marketplace_id} - {generated_at_utc.isoformat()}"
    )


def _set_column_widths(sheet: Any) -> None:
    widths = {
        "marketplace_id": 18,
        "seller_sku": 24,
        "asin": 14,
        "product_name": 45,
        "sku_sources": 28,
        "latest_source_date": 16,
        "current_remark": 32,
        "current_updated_at": 22,
        "purchase_or_batch_note": 32,
        "new_remark": 40,
    }
    for index, column in enumerate(TEMPLATE_COLUMNS, start=1):
        letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[letter].width = widths.get(column, 18)
        if column.endswith("cost"):
            sheet.column_dimensions[letter].width = 20
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _parse_import_row(
    row: dict[str, Any],
    *,
    row_number: int,
) -> tuple[SkuCostImportCandidate | None, list[SkuCostImportIssue]]:
    issues: list[SkuCostImportIssue] = []
    marketplace_id = _cell_text(row.get("marketplace_id"))
    seller_sku = _cell_text(row.get("seller_sku"))
    asin = _optional_text(row.get("asin"))
    currency = _cell_text(row.get("new_currency")).upper()
    effective_from = _parse_date(row.get("new_effective_from"))
    effective_to = _parse_date(row.get("new_effective_to"))
    product_cost = _parse_decimal(row.get("new_product_cost"))
    first_mile_cost = _parse_optional_decimal(row.get("new_first_mile_cost"))
    packaging_cost = _parse_optional_decimal(row.get("new_packaging_cost"))
    other_unit_cost = _parse_optional_decimal(row.get("new_other_unit_cost"))

    if not marketplace_id:
        issues.append(_error(row_number, seller_sku, "marketplace_id is required."))
    if not seller_sku:
        issues.append(_error(row_number, seller_sku, "seller_sku is required."))
    if product_cost is None:
        issues.append(_error(row_number, seller_sku, "new_product_cost is required."))
    if not currency:
        issues.append(_error(row_number, seller_sku, "new_currency is required."))
    if effective_from is None:
        issues.append(_error(row_number, seller_sku, "new_effective_from is required."))
    if effective_from and effective_to and effective_to < effective_from:
        issues.append(
            _error(
                row_number,
                seller_sku,
                "new_effective_to cannot be earlier than new_effective_from.",
            )
        )

    costs = {
        "new_product_cost": product_cost,
        "new_first_mile_cost": first_mile_cost,
        "new_packaging_cost": packaging_cost,
        "new_other_unit_cost": other_unit_cost,
    }
    for column, value in costs.items():
        if value is not None and value < Decimal("0"):
            issues.append(_error(row_number, seller_sku, f"{column} cannot be negative."))

    if issues:
        return None, issues

    remark_parts = []
    purchase_note = _optional_text(row.get("purchase_or_batch_note"))
    new_remark = _optional_text(row.get("new_remark"))
    if purchase_note:
        remark_parts.append(f"purchase_or_batch_note={purchase_note}")
    if new_remark:
        remark_parts.append(new_remark)
    remark = " | ".join(remark_parts) or None
    record = SkuCostWriteRecord(
        marketplace_id=marketplace_id,
        seller_sku=seller_sku,
        asin=asin,
        product_cost=product_cost or Decimal("0"),
        first_mile_cost=first_mile_cost or Decimal("0"),
        packaging_cost=packaging_cost or Decimal("0"),
        other_unit_cost=other_unit_cost or Decimal("0"),
        currency=currency,
        effective_from=effective_from or date.min,
        effective_to=effective_to,
        remark=remark,
    )
    return SkuCostImportCandidate(row_number=row_number, record=record), issues


def _is_blank_row(row: dict[str, Any]) -> bool:
    return all(value in (None, "") for value in row.values())


def _has_editable_input(row: dict[str, Any]) -> bool:
    # new_currency is prefilled by the exporter, so it must not make an otherwise
    # untouched template row importable.
    return any(row.get(column) not in (None, "") for column in EDITABLE_TRIGGER_COLUMNS)


def _parse_decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.0001"))
    except (InvalidOperation, ValueError):
        return None


def _parse_optional_decimal(value: Any) -> Decimal | None:
    return Decimal("0") if value in (None, "") else _parse_decimal(value)


def _parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip()[:10])
    except ValueError:
        return None


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _optional_text(value: Any) -> str | None:
    text = _cell_text(value)
    return text or None


def _error(row_number: int, seller_sku: str | None, message: str) -> SkuCostImportIssue:
    return SkuCostImportIssue(
        row_number=row_number,
        seller_sku=seller_sku,
        severity="error",
        message=message,
    )


def _date_value(value: date | None) -> date | None:
    return value


def _datetime_text(value: Any | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    return str(value)


def _decimal_value(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def _column_letter(column_name: str) -> str:
    index = TEMPLATE_COLUMNS.index(column_name) + 1
    # Works up to Z for the current template and keeps the function dependency-free.
    if index <= 26:
        return chr(ord("A") + index - 1)
    first = chr(ord("A") + ((index - 1) // 26) - 1)
    second = chr(ord("A") + ((index - 1) % 26))
    return first + second


__all__ = [
    "SKU_COST_INPUT_SHEET",
    "SkuCostImportCandidate",
    "SkuCostImportIssue",
    "SkuCostImportResult",
    "SkuCostTemplateExportResult",
    "SkuCostTemplateRow",
    "SkuCostWorkbookService",
    "build_sku_cost_workbook",
    "build_template_rows",
    "read_sku_cost_workbook",
]
