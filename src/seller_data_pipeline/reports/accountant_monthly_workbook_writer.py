from __future__ import annotations

from collections.abc import Mapping, Sequence
from decimal import Decimal
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from seller_data_pipeline.reports.monthly_reporting_common import (
    DARK_BLUE,
    GREEN_FILL,
    HEADER_BLUE,
    LIGHT_BLUE,
    MID_BLUE,
    MONEY_FORMAT,
    WHITE,
    YELLOW_FILL,
    as_bool,
    as_decimal,
    summarize_finance_rows,
)

if TYPE_CHECKING:
    from seller_data_pipeline.services.monthly_financial_close_service import (
        MonthlyFinancialCloseResult,
    )

ZERO = Decimal("0")
KNOWN_ACCOUNTING_TYPES = {
    "Shipment",
    "Refund",
    "RemovalShipment",
    "ServiceFee",
    "FBAInventoryReimbursement",
    "MiscellaneousLedgerAdjustment",
    "ProductAdsPayment",
    "Transfer",
    "Retrocharge",
}


def build_accountant_monthly_workbook(
    result: MonthlyFinancialCloseResult,
    finance_rows: Sequence[Mapping[str, Any]],
) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("01_会计汇总")
    classified = workbook.create_sheet("02_分类明细")
    source = workbook.create_sheet("03_源交易明细")
    checks = workbook.create_sheet("04_核验与说明")

    components = summarize_finance_rows(finance_rows)
    classified_rows = [_classify_row(row) for row in finance_rows]
    _write_summary(summary, result, components, finance_rows)
    _write_classified(classified, classified_rows)
    _write_source(source, finance_rows)
    _write_checks(checks, result, finance_rows, classified_rows, components)
    workbook.active = 0
    return workbook


def _write_summary(
    sheet: Any,
    result: MonthlyFinancialCloseResult,
    components: Mapping[str, Decimal],
    finance_rows: Sequence[Mapping[str, Any]],
) -> None:
    nm = result.natural_month_finance
    sheet.merge_cells("A1:F1")
    sheet["A1"] = f"{result.month} Amazon 会计月度底稿 / Accountant Monthly Workbook"
    _title(sheet["A1"], 16)
    sheet.merge_cells("A2:F2")
    sheet["A2"] = (
        "主数据源 / Primary source：Amazon Finances API 美国站本地自然月 normalized ledger。"
        "本表用于会计做账辅助、分类核对和源交易追溯；不是法定财务报表、税务申报表或正式记账凭证。"
        " / This workbook supports bookkeeping, classification and traceability; it is not a statutory financial statement or tax filing."
    )
    _fill_range(sheet, "A2:F2", LIGHT_BLUE)
    sheet["A2"].alignment = Alignment(wrap_text=True)

    metadata = (
        ("月份 / Month", result.month),
        ("站点 / Marketplace", result.marketplace_id),
        ("Finances API生命周期源记录数 / Lifecycle source rows", len(finance_rows)),
        ("源币种 / Source currency", result.currency or "USD"),
        ("记账汇率 USD/CNY / Bookkeeping FX rate", None),
        ("商品销售件数 / Product sales units", nm.product_sales_units if nm else 0),
        ("成本计量件数 / Costed units (sales + liquidation)", nm.costed_units if nm else 0),
        ("仓库丢失计量件数 / Warehouse-lost units", nm.inventory_loss_units if nm else 0),
    )
    for row, (label, value) in enumerate(metadata, 4):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
    _header_pair(sheet, "A4:B4")
    fx_row = 8
    sheet.cell(fx_row, 2).fill = PatternFill("solid", fgColor=YELLOW_FILL)
    sheet.cell(fx_row, 2).alignment = Alignment(horizontal="center")
    sheet.cell(fx_row, 2).value = None

    status_rows = (
        ("Released + Deferred / 生命周期", "已按自然月生命周期规则去重纳入 / De-duplicated", "不可只筛 Released；Deferred 也可能属于当月。 / Do not filter to Released only."),
        ("Transfer / 资金划转", "排除损益 / Excluded from P&L", "仅作为 Amazon 余额转出、银行回款和现金对账参考。 / Cash reconciliation only."),
        ("商品成本 / Product cost", "SKU生效日成本 / Effective-date cost", "使用 amazon_sku_cost 的商品成本、头程、包装和其他单位成本；不再使用固定30 RMB/件历史算法。"),
        ("普通赔偿 / Normal reimbursement", "不自动重复扣COGS / No duplicate COGS", "退货相关赔偿等不会仅因 quantity 字段存在就再次扣成本。 / Quantity alone never creates extra COGS."),
        ("仓库丢失 / Warehouse lost", nm.inventory_loss_status if nm else "missing", "Finances 中 WAREHOUSE_LOST 必须与 FBA Reimbursements 明细、SKU和有效成本核对；通过后单独核销库存成本。"),
        ("广告 / Advertising", "账单与经营分口径 / Two timing views", "会计主表使用 Finances posted-date 广告账单；经营月报使用 Ads API report_date 当月实际消耗。"),
        ("Settlement", "仅结算参考 / Close reference", "Settlement Close 用于结算、现金、回款核验，不替代自然月会计交易分类。"),
        ("状态 / Status", "数据校验状态 / Data validation", "OK 仅表示自动数据与核验规则通过，不代表会计师已完成正式审核。"),
    )
    sheet["D4"] = "关键口径 / Key policy"
    sheet["E4"] = "状态 / Status"
    sheet["F4"] = "说明 / Explanation"
    _header_row(sheet, 4, 4, 6)
    for row, values in enumerate(status_rows, 5):
        for col, value in enumerate(values, 4):
            sheet.cell(row, col, value)
        sheet.cell(row, 6).alignment = Alignment(wrap_text=True)

    summary_section_row = 14
    _section(sheet, summary_section_row, "会计项目汇总 / Accounting Summary", 6)
    headers = (
        "会计项目 / Accounting item",
        "金额USD / Amount USD",
        "金额CNY / Amount CNY",
        "方向 / Direction",
        "建议科目 / Suggested account",
        "说明 / Explanation",
    )
    for col, value in enumerate(headers, 1):
        sheet.cell(summary_section_row + 1, col, value)
    _header_row(sheet, summary_section_row + 1, 1, 6)

    if nm is None:
        summary_rows: tuple[tuple[str, Decimal, str, str, str], ...] = ()
    else:
        posted_ads = components.get("posted_ads_reference", nm.finances_ads_charge_reference)
        amazon_net = nm.operating_net_before_ads_replacement + posted_ads
        reference_profit = amazon_net - nm.landed_cogs - nm.inventory_loss_landed_cost
        summary_rows = (
            ("商品销售收入 / Product sales revenue", nm.product_sales_amount, "收入 / Revenue", "主营业务收入 / Sales revenue", "Finances Shipment 的商品销售金额。 / Product sales from natural-month Shipment transactions."),
            ("运费收入 / Shipping revenue", components.get("shipping_income", ZERO), "收入 / Revenue", "主营业务收入-运费 / Shipping revenue", "Shipment / Sales/Shipping。"),
            ("订单促销折扣 / Order promotional rebates", components.get("order_promotion", ZERO), "收入冲减/费用 / Contra revenue", "销售折让或促销费 / Discounts", "订单层面的 PromoRebates。 / Shipment promotional rebates."),
            ("FBA单件履约费 / FBA fulfillment fee", components.get("fba_fulfillment", ZERO), "费用 / Expense", "销售费用-FBA履约费 / FBA fulfillment", "按订单件数收取的 FBA 履约费。 / Per-unit FBA fulfillment fee."),
            ("运费扣回 / Shipping Chargeback", components.get("shipping_chargeback", ZERO), "费用 / Expense", "销售费用-FBA履约费 / FBA fulfillment", "Amazon 订单运费相关扣回。 / Shipment shipping chargeback."),
            ("净销售佣金 / Net sales commission", components.get("net_commission", ZERO), "费用 / Expense", "销售费用-平台佣金 / Marketplace commission", "Commission/Base + Commission/Promo 的真实净额。 / Net commission after promo offset."),
            ("退款净额 / Refund net", nm.refund_total, "收入冲减 / Contra revenue", "销售退回/销售折让 / Sales returns", "自然月 Refund 交易净额。 / Natural-month refund transaction net."),
            ("广告账单净额（入账日） / Advertising billing (posted-date)", posted_ads, "费用 / Expense", "销售费用-广告费 / Advertising expense", "会计入账使用 Finances ProductAdsPayment；与 Ads API 当月投放发生额可能不同。"),
            ("店铺订阅费 / Subscription fee", nm.subscription_fee, "费用 / Expense", "销售费用-平台服务费 / Platform service fee", "Professional selling plan 等订阅费。 / Subscription fee."),
            ("Coupon费用 / Coupon fees", nm.coupon_fee, "费用 / Expense", "销售费用-促销费 / Promotion expense", "Coupon participation + performance fees。"),
            ("Deal费用 / Deal fees", nm.deal_fee, "费用 / Expense", "销售费用-促销费 / Promotion expense", "Deal participation + performance fees；活动报名费也在此分类。"),
            ("仓储费 / FBA storage fee", nm.storage_fee, "费用 / Expense", "销售费用-FBA仓储费 / Storage expense", "FBA Storage fee。"),
            ("客户退货处理费 / Customer return processing fee", nm.customer_return_fee, "费用 / Expense", "销售费用-FBA服务费 / FBA service expense", "Customer Return / HRR 等退货处理费用。"),
            ("其他平台服务费 / Other Service Fee", nm.other_service_fee, "费用 / Expense", "销售费用-其他平台费 / Other platform fees", "未归入上述类别的 ServiceFee。 / Other ServiceFee."),
            ("FBA赔偿收入净额 / FBA reimbursement income", nm.reimbursement_total, "其他收益 / Other income", "其他收益或费用冲减 / Other income", "包括退货相关赔偿和仓库丢失赔偿；赔偿收入与库存成本核销分开记录。"),
            ("其中：仓库丢失赔偿 / Of which: warehouse-lost reimbursement", nm.warehouse_lost_reimbursement_amount, "其中/信息 / Included above", "信息行 / Informational", "已包含在 FBA赔偿收入净额 中，不可再次加总。 / Included in reimbursement income above; do not double count."),
            ("其他账务调整 / Other ledger adjustments", nm.adjustment_total, "其他收益/调整 / Other", "其他收益或费用冲减 / Other income/adjustment", "MiscellaneousLedgerAdjustment。"),
            ("库存清算净额 / Liquidation net", nm.liquidation_total, "其他收入 / Other income", "其他业务收入/清算 / Liquidation", "Liquidation revenue - processing/referral fees。"),
            ("Amazon交易净额（不含Transfer） / Amazon transaction net excl. Transfer", amazon_net, "小计 / Subtotal", "Amazon往来净额 / Amazon clearing", "自然月经营交易 + posted-date 广告账单；Transfer 不计入。"),
            ("商品采购成本 / Product cost COGS", -nm.product_cost_cogs, "成本 / Cost", "主营业务成本 / Cost of sales", "销售+清算件数按 SKU effective-date product cost 计量。"),
            ("头程成本 / First-mile freight COGS", -nm.first_mile_cogs, "成本 / Cost", "主营业务成本-头程 / First-mile freight", "销售+清算件数按 SKU effective-date first-mile cost 计量。"),
            ("包装成本 / Packaging COGS", -nm.packaging_cogs, "成本 / Cost", "主营业务成本 / Cost of sales", "SKU effective-date packaging cost。"),
            ("其他单位成本 / Other unit COGS", -nm.other_unit_cogs, "成本 / Cost", "主营业务成本 / Cost of sales", "SKU effective-date other unit cost。"),
            ("仓库丢失库存成本核销 / Warehouse-lost inventory write-off", -nm.inventory_loss_landed_cost, "库存损失 / Inventory loss", "存货损失/主营业务成本 / Inventory loss", "仅在 Finances WAREHOUSE_LOST 与 FBA Reimbursements 明细、SKU数量和有效成本全部核验通过后自动核销。"),
            ("账单月参考利润 / Posted-month reference profit", reference_profit, "参考 / Reference", "管理/会计辅助 / Bookkeeping support", "Amazon交易净额 - 销售/清算到岸COGS - 已核验仓库丢失库存成本；非税务法定利润。"),
            ("管理经营利润（备查） / Management operating profit (reference)", nm.management_operating_profit, "备查 / Reference", "不入会计主表 / Management only", "经营月报口径：广告按 Ads API report_date，另扣已核验仓库丢失库存成本。"),
            ("Ads API当月广告消耗（备查） / Ads API monthly spend (reference)", -nm.ads_api_report_date_spend, "备查 / Reference", "不替代账单凭证 / Not bookkeeping bill", "仅用于解释广告投放发生额与 Amazon 实际扣款月份差异。"),
            (
                "Transfer（备查，不计损益） / Cash transfer (reference only)",
                _cash_transfer_display_amount(nm.transfer_reference),
                "备查 / Reference",
                "银行/收款账户往来 / Cash clearing",
                "按 Amazon 账户资金流方向显示；打款转出为负数，不是收入或费用。 / Payout outflow is shown negative and excluded from P&L.",
            ),
        )

    start = summary_section_row + 2
    for row_index, (label, amount, direction, account, note) in enumerate(summary_rows, start):
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, float(amount)).number_format = MONEY_FORMAT
        # CNY formula remains blank until the accountant explicitly enters B8 FX.
        sheet.cell(row_index, 3, f'=IF($B${fx_row}="","",B{row_index}*$B${fx_row})')
        sheet.cell(row_index, 3).number_format = MONEY_FORMAT
        sheet.cell(row_index, 4, direction)
        sheet.cell(row_index, 5, account)
        sheet.cell(row_index, 6, note).alignment = Alignment(wrap_text=True)
        if label.startswith(("Amazon交易净额", "账单月参考利润", "管理经营利润（备查）")):
            _fill_range(sheet, f"A{row_index}:F{row_index}", GREEN_FILL)
            for cell in sheet[row_index][:6]:
                cell.font = Font(bold=True)

    _apply_widths(sheet, {"A": 34, "B": 17, "C": 17, "D": 18, "E": 30, "F": 48})
    guide_row = start + len(summary_rows) + 2
    _section(sheet, guide_row, "会计使用说明 / Bookkeeping Instructions", 6)
    instructions = (
        "1. USD 金额是 Amazon 源交易主币种；如需人民币记账，请会计在黄色的“记账汇率 USD/CNY”单元格填写当月采用汇率，CNY列会自动换算。",
        "2. 商品销售件数只统计正常销售；成本计量件数 = 销售件数 + 清算件数。仓库丢失件数单独列示并单独核销，不混入商品毛利率。",
        "3. 广告账单净额使用 Finances posted-date，适合做账；Ads API 当月广告消耗只作经营分析备查，两者因扣款时点不同无需相等。",
        "4. Transfer 是 Amazon 余额向银行/收款账户的资金划转，不是收入或费用，禁止计入利润。",
        "5. FBA赔偿不会按 quantity 一律重复扣 COGS；只有 WAREHOUSE_LOST 等明确库存损失且明细/成本核验通过时才产生库存成本核销。",
        "6. 03_源交易明细保留 Finances API 生命周期记录，因此记录数可能多于 Seller Central Monthly Transaction；这不代表重复计入损益，是否计入由 management_role / lifecycle policy 控制。",
    )
    for offset, instruction in enumerate(instructions, 1):
        sheet.merge_cells(start_row=guide_row + offset, start_column=1, end_row=guide_row + offset, end_column=6)
        cell = sheet.cell(guide_row + offset, 1, instruction)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[guide_row + offset].height = 34

    sheet.freeze_panes = f"A{start}"


def _write_classified(sheet: Any, rows: Sequence[Mapping[str, Any]]) -> None:
    columns = (
        ("财务分类", "财务分类 / Accounting category"),
        ("建议会计项目", "建议会计项目 / Suggested account"),
        ("计入损益", "计入损益 / P&L inclusion"),
        ("处理说明", "处理说明 / Treatment note"),
        ("posted_date_local", "本地入账日期 / posted_date_local"),
        ("transaction_status", "交易状态 / transaction_status"),
        ("transaction_type", "交易类型 / transaction_type"),
        ("transaction_id", "交易ID / transaction_id"),
        ("order_id", "订单ID / order_id"),
        ("description", "交易描述 / description"),
        ("amount", "金额 / amount"),
        ("currency", "币种 / currency"),
        ("management_role", "生命周期角色 / management_role"),
        ("source_trace", "源追溯 / source_trace"),
    )
    for col, (_, label) in enumerate(columns, 1):
        sheet.cell(1, col, label)
    _header_row(sheet, 1, 1, len(columns), dark=True)
    for row_index, row in enumerate(rows, 2):
        for col, (key, _) in enumerate(columns, 1):
            value = row.get(key)
            if key == "amount":
                value = float(as_decimal(value))
            sheet.cell(row_index, col, value)
        sheet.cell(row_index, 11).number_format = MONEY_FORMAT
        sheet.cell(row_index, 4).alignment = Alignment(wrap_text=True)
        sheet.cell(row_index, 10).alignment = Alignment(wrap_text=True)
        if row.get("计入损益") == "人工复核":
            _fill_range(sheet, f"A{row_index}:N{row_index}", YELLOW_FILL)
    _apply_widths(
        sheet,
        {
            "A": 22,
            "B": 28,
            "C": 14,
            "D": 42,
            "E": 18,
            "F": 20,
            "G": 24,
            "H": 42,
            "I": 24,
            "J": 42,
            "K": 16,
            "L": 12,
            "M": 28,
            "N": 46,
        },
    )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:N{max(sheet.max_row, 1)}"


def _write_source(sheet: Any, rows: Sequence[Mapping[str, Any]]) -> None:
    columns = (
        ("marketplace_id", "站点ID / marketplace_id"),
        ("transaction_id", "交易ID / transaction_id"),
        ("transaction_status", "交易状态 / transaction_status"),
        ("transaction_type", "交易类型 / transaction_type"),
        ("description", "交易描述 / description"),
        ("posted_at_utc", "UTC入账时间 / posted_at_utc"),
        ("posted_at_local", "本地入账时间 / posted_at_local"),
        ("posted_date_local", "本地入账日期 / posted_date_local"),
        ("marketplace_timezone", "站点时区 / marketplace_timezone"),
        ("amount", "金额 / amount"),
        ("currency", "币种 / currency"),
        ("settlement_id", "结算ID / settlement_id"),
        ("order_id", "订单ID / order_id"),
        ("deferred_transaction_id", "Deferred交易ID / deferred_transaction_id"),
        ("release_transaction_id", "Release交易ID / release_transaction_id"),
        ("management_role", "生命周期角色 / management_role"),
        ("management_include", "计入经营 / management_include"),
        ("management_replace_with_ads_api", "广告由Ads API替换 / management_replace_with_ads_api"),
        ("review_required", "需人工复核 / review_required"),
        ("product_sales_amount", "商品销售额 / product_sales_amount"),
        ("shipping_amount", "运费 / shipping_amount"),
        ("promotion_amount", "促销金额 / promotion_amount"),
        ("fba_fulfillment_fee", "FBA履约费 / fba_fulfillment_fee"),
        ("shipping_chargeback", "运费扣回 / shipping_chargeback"),
        ("refund_product_amount", "退款商品金额 / refund_product_amount"),
        ("refund_shipping_amount", "退款运费 / refund_shipping_amount"),
        ("refund_promotion_amount", "退款促销返还 / refund_promotion_amount"),
        ("liquidation_revenue", "清算收入 / liquidation_revenue"),
        ("liquidation_fee", "清算费用 / liquidation_fee"),
        ("subscription_fee", "订阅费 / subscription_fee"),
        ("coupon_fee", "Coupon费用 / coupon_fee"),
        ("deal_fee", "Deal费用 / deal_fee"),
        ("storage_fee", "仓储费 / storage_fee"),
        ("customer_return_fee", "客户退货处理费 / customer_return_fee"),
        ("other_service_fee", "其他服务费 / other_service_fee"),
        ("unit_events_json", "成本数量事件 / unit_events_json"),
        ("related_identifiers_json", "关联标识 / related_identifiers_json"),
        ("raw_transaction_hash", "原始交易哈希 / raw_transaction_hash"),
        ("business_key_hash", "业务幂等键 / business_key_hash"),
    )
    for col, (_, label) in enumerate(columns, 1):
        sheet.cell(1, col, label)
    _header_row(sheet, 1, 1, len(columns), dark=True)
    money_headers = {
        "amount",
        "product_sales_amount",
        "shipping_amount",
        "promotion_amount",
        "fba_fulfillment_fee",
        "shipping_chargeback",
        "refund_product_amount",
        "refund_shipping_amount",
        "refund_promotion_amount",
        "liquidation_revenue",
        "liquidation_fee",
        "subscription_fee",
        "coupon_fee",
        "deal_fee",
        "storage_fee",
        "customer_return_fee",
        "other_service_fee",
    }
    for row_index, row in enumerate(rows, 2):
        for col, (key, _) in enumerate(columns, 1):
            value = row.get(key)
            if key in money_headers:
                value = float(as_decimal(value))
            elif key in {"management_include", "management_replace_with_ads_api", "review_required"}:
                value = bool(as_bool(value))
            elif value is not None and not isinstance(value, (str, int, float, bool)):
                value = str(value)
            sheet.cell(row_index, col, value)
            if key in money_headers:
                sheet.cell(row_index, col).number_format = MONEY_FORMAT
    for col in range(1, len(columns) + 1):
        letter = sheet.cell(1, col).column_letter
        sheet.column_dimensions[letter].width = 18
    sheet.column_dimensions["B"].width = 42
    sheet.column_dimensions["E"].width = 42
    sheet.column_dimensions["P"].width = 30
    sheet.column_dimensions["AJ"].width = 48
    sheet.column_dimensions["AK"].width = 48
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(columns)).column_letter}{max(sheet.max_row, 1)}"


def _write_checks(
    sheet: Any,
    result: MonthlyFinancialCloseResult,
    finance_rows: Sequence[Mapping[str, Any]],
    classified_rows: Sequence[Mapping[str, Any]],
    components: Mapping[str, Decimal],
) -> None:
    nm = result.natural_month_finance
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "核验与口径说明 / Reconciliation & Accounting Policies"
    _title(sheet["A1"], 16)
    headers = (
        "检查项 / Check",
        "结果 / Result",
        "实际值 / Actual",
        "期望/规则 / Expected rule",
        "状态 / Status",
        "说明 / Explanation",
    )
    for col, value in enumerate(headers, 1):
        sheet.cell(3, col, value)
    _header_row(sheet, 3, 1, 6)

    classified_pnl = sum(
        (as_decimal(row.get("amount")) for row in classified_rows if row.get("计入损益") == "是"),
        ZERO,
    )
    unknown_nonzero = [
        row for row in classified_rows if row.get("计入损益") == "人工复核" and as_decimal(row.get("amount")) != ZERO
    ]
    expected_accounting = components.get("accounting_transaction_net", ZERO)
    cost_expected = (nm.product_sales_units + nm.liquidation_units) if nm else 0
    cost_actual = nm.costed_units if nm else 0
    inventory_loss_expected = nm.inventory_loss_units if nm else 0
    inventory_loss_actual = nm.inventory_loss_costed_units if nm else 0

    checks = (
        ("Finances生命周期源记录 / Lifecycle source rows", len(finance_rows), len(classified_rows), "source rows = classified rows", "通过" if len(finance_rows) == len(classified_rows) else "需复核", "Finances API 会保留 Deferred/Released 等生命周期记录，记录数可能多于 Seller Central Monthly Transaction；分类明细不得丢行。"),
        ("会计损益分类金额 / Classified P&L", classified_pnl, expected_accounting, "classified P&L = operating rows + posted ads reference", "通过" if abs(classified_pnl - expected_accounting) < Decimal("0.01") else "需复核", "Transfer 和历史 release reference 不进入损益；分类金额必须与会计交易净额闭合。"),
        (
            "Transfer / 资金划转",
            _cash_transfer_display_amount(components.get("transfer_reference", ZERO)),
            "排除损益 / Excluded",
            "cash reference only; payout outflow shown negative",
            "通过",
            "按 Amazon 账户资金流方向显示；转出为负数，不作为收入或费用。 / Cash movement only.",
        ),
        ("未知非零分类 / Unknown non-zero classification", len(unknown_nonzero), 0, "must be zero", "通过" if not unknown_nonzero else "需复核", "新 transaction type/status 不得静默忽略；出现未知非零金额必须人工复核并阻止正常分享。"),
        ("销售+清算COGS覆盖 / Sales+liquidation COGS coverage", cost_actual, cost_expected, "costed units = sales + liquidation units", "通过" if nm and cost_actual == cost_expected and not nm.missing_cost_skus else "需复核", f"销售件数={nm.product_sales_units if nm else 0}; 清算件数={nm.liquidation_units if nm else 0}; missing_cost_skus={list(nm.missing_cost_skus) if nm else []}"),
        ("仓库丢失库存核销 / Warehouse-lost inventory write-off", inventory_loss_actual, inventory_loss_expected, "WAREHOUSE_LOST detail + SKU + effective landed cost must reconcile", "通过" if nm and nm.inventory_loss_status in {"ok", "not_applicable"} and inventory_loss_actual == inventory_loss_expected else "需复核", f"status={nm.inventory_loss_status if nm else 'missing'}; finances_reimbursement={nm.warehouse_lost_reimbursement_amount if nm else 0}; reimbursement_report={getattr(nm, 'warehouse_lost_reimbursement_report_amount', 0) if nm else 0}; landed_writeoff={nm.inventory_loss_landed_cost if nm else 0}; missing={list(nm.inventory_loss_missing_cost_skus) if nm else []}"),
        ("Seller Central人工核验 / Manual Seller Central reconciliation", str(result.raw_metadata.get("seller_central_monthly_transaction_reconciliation_status") or "not_provided"), "optional", "manual export missing does not block", "参考", "收到官方 Monthly Transaction 时可逐项核对；它是外部官方核验源，不是自动生成依赖。"),
        ("广告口径 / Advertising timing", components.get("posted_ads_reference", ZERO), -(nm.ads_api_report_date_spend if nm else ZERO), "posted billing != Ads API spend is allowed", "参考", "会计主表使用 Amazon 实际入账/扣款时点；经营月报使用 Ads API report_date 当月发生额。"),
        ("负数显示 / Negative values", "显式负号 / Explicit minus", "显式负号 / Explicit minus", "-$114.89 / -7.85%", "通过", "负数统一用“-”号，不依赖红色或括号表达。"),
    )
    for row_index, values in enumerate(checks, 4):
        for col, value in enumerate(values, 1):
            if isinstance(value, Decimal):
                value = float(value)
            sheet.cell(row_index, col, value)
        if isinstance(values[1], Decimal):
            sheet.cell(row_index, 2).number_format = MONEY_FORMAT
        if isinstance(values[2], Decimal):
            sheet.cell(row_index, 3).number_format = MONEY_FORMAT
        sheet.cell(row_index, 6).alignment = Alignment(wrap_text=True)
        status = values[4]
        fill = GREEN_FILL if status == "通过" else YELLOW_FILL
        sheet.cell(row_index, 5).fill = PatternFill("solid", fgColor=fill)
        sheet.cell(row_index, 5).font = Font(bold=True)

    note_row = 4 + len(checks) + 2
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 3, end_column=6)
    sheet.cell(note_row, 1, (
        "会计底稿原则 / Accounting principles：USD 为 Amazon 源交易主币种；CNY 仅在会计明确填写月度记账汇率后生成。"
        "销售和清算成本使用 amazon_sku_cost effective-date 数据，不使用历史固定 30 RMB/件 + 2.5 RMB/件算法。"
        "普通 Adjustment/Reimbursement 不会因为 quantity 字段存在就自动重复扣 COGS；WAREHOUSE_LOST 只有在 FBA Reimbursements 明细、SKU数量和有效成本全部核验通过后才单独核销库存成本。"
        "本底稿用于做账辅助和追溯，不替代会计对科目、税务、发票、汇率和当地准则的专业判断。"
    ))
    sheet.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    _fill_range(sheet, f"A{note_row}:F{note_row + 3}", LIGHT_BLUE)

    detail_start = note_row + 5
    _section(sheet, detail_start, "仓库丢失库存核销明细 / Warehouse-Lost Inventory Write-off Detail", 7)
    detail_headers = (
        "赔偿ID / Reimbursement ID",
        "日期 / Date",
        "SKU/FNSKU",
        "数量 / Qty",
        "赔偿金额 / Reimbursement",
        "候选/已核销到岸成本 / Candidate/applied landed cost",
        "状态 / Status",
    )
    for col, value in enumerate(detail_headers, 1):
        sheet.cell(detail_start + 1, col, value)
    _header_row(sheet, detail_start + 1, 1, 7)
    details = tuple(nm.inventory_loss_details) if nm else ()
    if not details:
        sheet.merge_cells(
            start_row=detail_start + 2,
            start_column=1,
            end_row=detail_start + 2,
            end_column=7,
        )
        sheet.cell(
            detail_start + 2,
            1,
            "本月无需要自动核销的 WAREHOUSE_LOST 库存损失。 / No verified warehouse-lost inventory write-off for this month.",
        )
    else:
        for offset, detail in enumerate(details, detail_start + 2):
            sku = detail.get("seller_sku") or detail.get("fnsku") or "-"
            values = (
                detail.get("reimbursement_id"),
                detail.get("approval_date"),
                f"{sku} -> {detail.get('resolved_cost_sku') or '-'}",
                detail.get("quantity"),
                float(as_decimal(detail.get("reimbursement_amount"))),
                -float(as_decimal(detail.get("landed_cost_writeoff"))),
                "已应用 / Applied" if detail.get("status") == "ok" else "候选-未应用 / Candidate - not applied",
            )
            for col, value in enumerate(values, 1):
                sheet.cell(offset, col, value)
            sheet.cell(offset, 5).number_format = MONEY_FORMAT
            sheet.cell(offset, 6).number_format = MONEY_FORMAT
            if detail.get("status") != "ok":
                _fill_range(sheet, f"A{offset}:G{offset}", YELLOW_FILL)

    _apply_widths(sheet, {"A": 28, "B": 22, "C": 24, "D": 14, "E": 18, "F": 28, "G": 32})
    sheet.freeze_panes = "A4"


def _cash_transfer_display_amount(amount: Decimal) -> Decimal:
    """Display Amazon payout transfers using Seller Central cash-flow direction.

    The normalized Finances ledger keeps Transfer as a positive cash-reference amount,
    while Seller Central Monthly Transaction represents a payout from the Amazon
    balance as a negative amount. Accounting-facing views use the latter convention
    so the direction is explicit; the source sheet continues to preserve the raw
    normalized ledger amount.
    """

    if amount == ZERO:
        return ZERO
    return -abs(amount)


def _classify_row(row: Mapping[str, Any]) -> dict[str, Any]:
    transaction_type = str(row.get("transaction_type") or "")
    amount = as_decimal(row.get("amount"))
    management_include = as_bool(row.get("management_include"))
    replace_ads = as_bool(row.get("management_replace_with_ads_api"))
    role = str(row.get("management_role") or "")
    classification = "参考/非损益 / Reference / non-P&L"
    accounting_item = "不计损益 / Non-P&L"
    include = "否"
    note = "仅作追溯或对账参考，不计入本月损益。 / Reference row only; excluded from current-month P&L."

    if management_include:
        include = "是"
        if transaction_type == "Shipment":
            classification = "订单销售 / Shipment sales"
            accounting_item = "销售收入及订单级费用 / Sales and order-level charges"
            note = "按美国站本地自然月和生命周期规则去重后的 Shipment 交易。 / Natural-month de-duplicated Shipment transaction."
        elif transaction_type == "Refund":
            classification = "退款 / Refund"
            accounting_item = "销售退回/销售折让 / Sales returns and allowances"
            note = "本地自然月退款交易净额。 / Natural-month refund transaction net."
        elif transaction_type == "RemovalShipment":
            classification = "库存清算 / Liquidation"
            accounting_item = "其他业务收入/清算 / Other income - liquidation"
            note = "库存清算收入扣除清算处理费/推荐费后的净额。 / Liquidation revenue net of related fees."
        elif transaction_type == "ServiceFee":
            classification, accounting_item = _service_fee_classification(row)
            note = "按已抽取的订阅、Coupon、Deal、仓储、退货处理等费用组件分类。 / ServiceFee classified from extracted fee components."
        elif transaction_type == "FBAInventoryReimbursement":
            if _is_warehouse_lost_description(row.get("description")):
                classification = "仓库丢失赔偿 / Warehouse-lost reimbursement"
                accounting_item = "其他收益-库存损失赔偿 / Other income"
                note = (
                    "赔偿收入计入本月；对应库存成本只有在 FBA Reimbursements 明细、SKU数量和"
                    "effective-date landed cost 核验通过后才单独核销。 / Reimbursement income is "
                    "recognized here; inventory cost is written off separately only after verification."
                )
            else:
                classification = "普通FBA赔偿 / Normal FBA reimbursement"
                accounting_item = "其他收益或费用冲减 / Other income"
                note = (
                    "普通退货/追回类赔偿不因 quantity 字段自动重复扣 COGS。 / No automatic "
                    "duplicate COGS for normal reimbursement quantity."
                )
        elif transaction_type == "MiscellaneousLedgerAdjustment":
            classification = "账务调整 / Ledger adjustment"
            accounting_item = "其他收益/调整 / Other income or adjustment"
            note = "Amazon 其他账务调整；非零未知组合仍需人工复核。 / Miscellaneous ledger adjustment; unknown non-zero combinations require review."
        else:
            classification = "未分类 / Unclassified"
            accounting_item = "待复核 / Review required"
            include = "人工复核"
            note = "经营口径纳入了未知非零交易类型，必须人工复核，不得静默忽略。 / Unknown non-zero management-included transaction; manual review required."
    elif replace_ads:
        classification = "广告账单 / Advertising billing"
        accounting_item = "销售费用-广告费 / Advertising expense"
        include = "是"
        note = "按 Amazon 入账日记录的 ProductAdsPayment；会计底稿使用该账单金额，经营月报改用 Ads API report_date 当月实际消耗。 / Posted-date ProductAdsPayment for bookkeeping; management reporting uses Ads API spend."
    elif role == "cash_transfer_reference" or transaction_type == "Transfer":
        classification = "资金划转 / Cash transfer"
        accounting_item = "银行/收款账户往来 / Bank or receiving-account clearing"
        note = "Amazon 余额转出/回款，仅用于现金和银行对账，排除损益。 / Cash transfer reference only; excluded from P&L."
    elif role == "zero_value_unit_cogs_reference":
        classification = "COGS数量参考 / COGS quantity reference"
        accounting_item = "不计损益 / Non-P&L"
        note = "零金额但有实物数量的事件，仅用于补全成本计量件数，不产生收入。 / Zero-value unit event used only for COGS quantity completeness."
    elif transaction_type == "Retrocharge" or role in {"non_operating_reference", "prior_period_release_reference"}:
        classification = "历史/追溯参考 / Historical trace reference"
        accounting_item = "不计损益 / Non-P&L"
        note = "生命周期规则排除的历史/释放参考记录，仅作追溯，避免重复计入。 / Excluded lifecycle reference row retained for traceability to avoid double counting."
    elif transaction_type not in KNOWN_ACCOUNTING_TYPES and amount != ZERO:
        classification = "未分类 / Unclassified"
        accounting_item = "待复核 / Review required"
        include = "人工复核"
        note = "未知非零交易类型，必须人工复核并阻止正常分享。 / Unknown non-zero transaction type; do not silently ignore."

    source_trace = "; ".join(
        part
        for part in (
            f"transaction_id={row.get('transaction_id')}" if row.get("transaction_id") else "",
            f"raw_hash={row.get('raw_transaction_hash')}" if row.get("raw_transaction_hash") else "",
            f"business_key={row.get('business_key_hash')}" if row.get("business_key_hash") else "",
        )
        if part
    )
    display_amount = (
        _cash_transfer_display_amount(amount)
        if role == "cash_transfer_reference" or transaction_type == "Transfer"
        else amount
    )

    return {
        "财务分类": classification,
        "建议会计项目": accounting_item,
        "计入损益": include,
        "处理说明": note,
        "posted_date_local": str(row.get("posted_date_local") or ""),
        "transaction_status": row.get("transaction_status"),
        "transaction_type": transaction_type,
        "transaction_id": row.get("transaction_id"),
        "order_id": row.get("order_id"),
        "description": row.get("description"),
        "amount": display_amount,
        "currency": row.get("currency"),
        "management_role": role,
        "source_trace": source_trace,
    }


def _is_warehouse_lost_description(value: Any) -> bool:
    text = str(value or "").strip().upper()
    if not text:
        return False
    normalized = "".join(character if character.isalnum() else "_" for character in text)
    tokens = {token for token in normalized.split("_") if token}
    return "WAREHOUSE" in tokens and "LOST" in tokens


def _service_fee_classification(row: Mapping[str, Any]) -> tuple[str, str]:
    candidates = (
        ("subscription_fee", "店铺订阅费 / Subscription fee", "销售费用-平台服务费 / Platform service fee"),
        ("coupon_fee", "Coupon费用 / Coupon fee", "销售费用-促销费 / Promotion expense"),
        ("deal_fee", "Deal费用 / Deal fee", "销售费用-促销费 / Promotion expense"),
        ("storage_fee", "FBA仓储费 / FBA storage fee", "销售费用-FBA仓储费 / Storage expense"),
        ("customer_return_fee", "客户退货处理费 / Customer return fee", "销售费用-FBA服务费 / FBA service expense"),
        ("other_service_fee", "其他平台服务费 / Other Service Fee", "销售费用-其他平台费 / Other platform expense"),
    )
    for key, classification, account in candidates:
        if as_decimal(row.get(key)) != ZERO:
            return classification, account
    return "其他平台服务费 / Other Service Fee", "销售费用-其他平台费 / Other platform expense"


def _title(cell: Any, size: int) -> None:
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    cell.font = Font(bold=True, color=WHITE, size=size)


def _section(sheet: Any, row: int, title: str, columns: int) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
    sheet.cell(row, 1, title)
    sheet.cell(row, 1).fill = PatternFill("solid", fgColor=MID_BLUE)
    sheet.cell(row, 1).font = Font(bold=True, color=WHITE)


def _header_pair(sheet: Any, cell_range: str) -> None:
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=HEADER_BLUE)
            cell.font = Font(bold=True)


def _header_row(sheet: Any, row: int, start_col: int, end_col: int, *, dark: bool = False) -> None:
    fill = DARK_BLUE if dark else HEADER_BLUE
    color = WHITE if dark else "000000"
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _fill_range(sheet: Any, cell_range: str, color: str) -> None:
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=color)


def _apply_widths(sheet: Any, widths: Mapping[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


__all__ = ["build_accountant_monthly_workbook"]
