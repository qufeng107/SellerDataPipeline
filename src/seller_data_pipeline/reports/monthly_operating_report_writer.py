from __future__ import annotations

from collections.abc import Mapping, Sequence
from decimal import Decimal
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from seller_data_pipeline.reports.monthly_reporting_common import (
    DARK_BLUE,
    GRAY_FILL,
    GREEN_FILL,
    HEADER_BLUE,
    INTEGER_FORMAT,
    LIGHT_BLUE,
    MID_BLUE,
    MONEY_FORMAT,
    PERCENT_FORMAT,
    WHITE,
    YELLOW_FILL,
    as_decimal,
    month_finance_value,
    operational_metric,
    safe_ratio,
    sorted_recent_results,
)

if TYPE_CHECKING:
    from seller_data_pipeline.services.monthly_financial_close_service import (
        MonthlyFinancialCloseResult,
    )

ZERO = Decimal("0")


def build_monthly_operating_report_workbook(
    result: MonthlyFinancialCloseResult,
    *,
    recent_results: Sequence[MonthlyFinancialCloseResult] | None = None,
    finance_rows_by_month: Mapping[str, Sequence[Mapping[str, Any]]] | None = None,
) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    overview = workbook.create_sheet("01_月度经营总览")
    pnl = workbook.create_sheet("02_经营损益")
    checks = workbook.create_sheet("03_核验与口径")

    recent = sorted_recent_results(result, recent_results)
    _write_overview(overview, result, recent, finance_rows_by_month)
    _write_pnl(pnl, result, recent, finance_rows_by_month)
    _write_checks(checks, result)
    workbook.active = 0
    return workbook


def _write_overview(
    sheet: Any,
    result: MonthlyFinancialCloseResult,
    recent: Sequence[MonthlyFinancialCloseResult],
    finance_rows_by_month: Mapping[str, Sequence[Mapping[str, Any]]] | None,
) -> None:
    nm = result.natural_month_finance
    if nm is None:
        _write_no_data_sheet(sheet, result, "Natural-month Finances summary is unavailable.")
        return

    sheet.merge_cells("A1:H1")
    sheet["A1"] = f"Amazon 美国站｜{result.month} 月度经营报告"
    _title(sheet["A1"], size=18)
    sheet.merge_cells("A2:H2")
    sheet["A2"] = (
        "经营主口径：marketplace-local 自然月 Finances + Ads API report_date + 到岸商品成本；"
        "Settlement / Transfer 仅用于结算与回款核验。"
    )
    _fill_range(sheet, "A2:H2", LIGHT_BLUE)
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")

    _section(sheet, 4, "本月核心结果", 8)
    gross_margin = safe_ratio(nm.product_sales_amount - nm.landed_cogs, nm.product_sales_amount)
    cards = (
        ("A5:B5", "A6:B7", "商品销售额", nm.product_sales_amount, MONEY_FORMAT),
        ("C5:D5", "C6:D7", "经营利润", nm.management_operating_profit, MONEY_FORMAT),
        ("E5:F5", "E6:F7", "经营利润率", nm.management_operating_margin, PERCENT_FORMAT),
        ("G5:H5", "G6:H7", "商品毛利率", gross_margin, PERCENT_FORMAT),
    )
    for title_range, value_range, label, value, number_format in cards:
        sheet.merge_cells(title_range)
        sheet.merge_cells(value_range)
        title_cell = sheet[title_range.split(":", 1)[0]]
        value_cell = sheet[value_range.split(":", 1)[0]]
        title_cell.value = label
        value_cell.value = _xlsx_number(value)
        _fill_range(sheet, title_range, HEADER_BLUE)
        title_cell.font = Font(bold=True, color="000000")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.font = Font(bold=True, size=16, color="000000")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = number_format

    _section(sheet, 9, "最近三个月趋势", 8)
    trend_headers = ["指标"] + [item.month for item in recent] + ["本月环比", "怎么读"]
    sheet.append([]) if sheet.max_row < 9 else None
    for col, value in enumerate(trend_headers, 1):
        sheet.cell(10, col, value)
    _header_row(sheet, 10, len(trend_headers))

    metrics = (
        ("商品销售额", "sales", MONEY_FORMAT, "本月业务规模"),
        ("管理口径销量", "units", INTEGER_FORMAT, "销售+清算的成本计量件数"),
        ("商品毛利率", "gross_margin", PERCENT_FORMAT, "只扣到岸商品成本，观察产品本身毛利空间"),
        ("经营利润", "profit", MONEY_FORMAT, "所有经营收入/费用和商品成本后的最终结果"),
        ("经营利润率", "margin", PERCENT_FORMAT, "最终经营利润占商品销售额比例"),
        ("广告费", "ads", MONEY_FORMAT, "Ads API 当月实际消耗"),
        ("广告费率", "ads_ratio", PERCENT_FORMAT, "广告花费占商品销售额；越高负担越重"),
        ("Sessions", "sessions", INTEGER_FORMAT, "流量规模"),
        ("Sales & Traffic转化率", "conversion", PERCENT_FORMAT, "官方 units_ordered / sessions"),
    )
    start_row = 11
    values_by_key: dict[str, list[Decimal | int | None]] = {key: [] for _, key, _, _ in metrics}
    for item in recent:
        values = _trend_values(item)
        for key in values_by_key:
            values_by_key[key].append(values.get(key))

    for offset, (label, key, number_format, note) in enumerate(metrics):
        row = start_row + offset
        sheet.cell(row, 1, label)
        for idx, value in enumerate(values_by_key[key], 2):
            cell = sheet.cell(row, idx, _xlsx_number(value))
            cell.number_format = number_format
        change_col = 2 + len(recent)
        change = _mom_change(values_by_key[key])
        sheet.cell(row, change_col, _xlsx_number(change)).number_format = PERCENT_FORMAT
        sheet.cell(row, change_col + 1, note).alignment = Alignment(wrap_text=True)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "月度商品销售额与经营利润"
    chart.y_axis.title = result.currency or "USD"
    chart.x_axis.title = "Month"
    sales_row = start_row
    profit_row = start_row + 3
    month_count = len(recent)
    chart.add_data(Reference(sheet, min_col=2, max_col=1 + month_count, min_row=sales_row, max_row=sales_row), titles_from_data=False)
    chart.add_data(Reference(sheet, min_col=2, max_col=1 + month_count, min_row=profit_row, max_row=profit_row), titles_from_data=False)
    chart.set_categories(Reference(sheet, min_col=2, max_col=1 + month_count, min_row=10, max_row=10))
    chart.height = 7
    chart.width = 12
    sheet.add_chart(chart, "F10")

    current_components = month_finance_value(finance_rows_by_month, result.month)
    _section(sheet, 21, "本月主要支出结构", 8)
    expense_rows = _expense_rows(result, current_components)
    headers = ["支出项目", "金额", "占商品销售额", "本月解读"]
    for col, value in enumerate(headers, 1):
        sheet.cell(22, col, value)
    _header_row(sheet, 22, 4)
    for idx, (label, amount, note) in enumerate(expense_rows, 23):
        sheet.cell(idx, 1, label)
        sheet.cell(idx, 2, float(amount)).number_format = MONEY_FORMAT
        ratio = safe_ratio(amount, nm.product_sales_amount)
        sheet.cell(idx, 3, _xlsx_number(ratio)).number_format = PERCENT_FORMAT
        sheet.cell(idx, 4, note).alignment = Alignment(wrap_text=True)

    observations = _build_observations(result, recent, current_components)
    conclusion_row = 23 + len(expense_rows) + 1
    _section(sheet, conclusion_row, "本月结论", 8)
    for index, (headline, detail) in enumerate(observations, 1):
        row = conclusion_row + index
        sheet.cell(row, 1, index)
        sheet.cell(row, 2, headline).font = Font(bold=True)
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        sheet.cell(row, 3, detail).alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[row].height = 38

    guide_row = conclusion_row + len(observations) + 2
    _section(sheet, guide_row, "核心指标怎么读", 8)
    guide = (
        ("商品毛利率", "(商品销售额-到岸商品成本)/商品销售额", "产品本身毛利空间", "不等于最终利润率"),
        ("经营利润率", "经营利润/商品销售额", "整个Amazon业务最终赚钱能力", "负数代表经营亏损"),
        ("广告费率", "广告费/商品销售额", "广告负担", "显著上升通常需要优先复盘"),
        ("转化率", "Sales & Traffic units_ordered / Sessions", "流量成交效率", "使用官方Sales & Traffic口径"),
    )
    guide_header_row = guide_row + 1
    for col, value in enumerate(("指标", "计算逻辑", "用途", "阅读提示"), 1):
        sheet.cell(guide_header_row, col, value)
    _header_row(sheet, guide_header_row, 4)
    for index, values in enumerate(guide, guide_header_row + 1):
        for col, value in enumerate(values, 1):
            sheet.cell(index, col, value)
        sheet.cell(index, 2).alignment = Alignment(wrap_text=True)
        sheet.cell(index, 3).alignment = Alignment(wrap_text=True)
        sheet.cell(index, 4).alignment = Alignment(wrap_text=True)
        sheet.row_dimensions[index].height = 38

    footer_row = guide_header_row + len(guide) + 2
    sheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=8)
    sheet.cell(footer_row, 1, "数字显示规则：所有负数均显式显示“-”号，不依赖红色或括号表达负数。")
    _fill_range(sheet, f"A{footer_row}:H{footer_row}", GRAY_FILL)
    sheet.cell(footer_row, 1).font = Font(italic=True)

    widths = {"A": 30, "B": 22, "C": 26, "D": 44, "E": 16, "F": 34, "G": 18, "H": 18}
    _apply_widths(sheet, widths)
    sheet.freeze_panes = "A4"


def _write_pnl(
    sheet: Any,
    result: MonthlyFinancialCloseResult,
    recent: Sequence[MonthlyFinancialCloseResult],
    finance_rows_by_month: Mapping[str, Sequence[Mapping[str, Any]]] | None,
) -> None:
    current = result
    previous = recent[-2] if len(recent) >= 2 else None
    current_nm = current.natural_month_finance
    previous_nm = previous.natural_month_finance if previous else None
    if current_nm is None:
        _write_no_data_sheet(sheet, result, "Natural-month Finances summary is unavailable.")
        return

    current_components = month_finance_value(finance_rows_by_month, current.month)
    previous_components = (
        month_finance_value(finance_rows_by_month, previous.month) if previous else {}
    )

    sheet.merge_cells("A1:F1")
    sheet["A1"] = f"{current.month} 经营损益（与上月对比）"
    _title(sheet["A1"], size=16)
    headers = ["经营项目", previous.month if previous else "上月", current.month, "环比金额", "本月占商品销售额", "说明"]
    for col, value in enumerate(headers, 1):
        sheet.cell(3, col, value)
    _header_row(sheet, 3, 6)

    row_specs = (
        ("【收入与退款】", None, ""),
        ("商品销售收入", "product_sales", "核心商品销售收入"),
        ("运费收入", "shipping_income", "Shipment / Sales/Shipping"),
        ("退款：商品金额", "refund_product", "商品销售退款"),
        ("退款：运费", "refund_shipping", "运费退款"),
        ("退款时促销返还", "refund_promotion", "退款发生时冲回的促销成本"),
        ("库存清算/残值收入", "liquidation_revenue", "FBA库存清算残值"),
        ("【Amazon订单级费用】", None, ""),
        ("净销售佣金", "net_commission", "Commission/Base + Commission/Promo真实净额"),
        ("FBA单件履约费", "fba_fulfillment", "FBA per-unit fulfillment fee"),
        ("Shipping Chargeback", "shipping_chargeback", "订单运费相关扣减"),
        ("订单促销折扣", "order_promotion", "Shipment promo rebates"),
        ("库存清算相关费用", "liquidation_fee", "Liquidation processing/referral fee"),
        ("【账户级收入与费用】", None, ""),
        ("FBA库存赔偿净额", "reimbursement", "库存赔偿及追回净额"),
        ("其他账务调整", "adjustment", "Miscellaneous ledger adjustment"),
        ("订阅费", "subscription", "Professional selling plan等"),
        ("Coupon费用", "coupon", "参与费 + 绩效费"),
        ("Deal费用", "deal", "参与费 + 绩效费"),
        ("仓储费", "storage", "FBA storage"),
        ("客户退货处理费", "customer_return", "Customer return / HRR fee"),
        ("其他服务费", "other_service", "其他ServiceFee"),
        ("【广告与商品成本】", None, ""),
        ("广告费", "ads", "Ads API report_date当月实际消耗"),
        ("商品采购成本", "product_cost", "SKU effective-date product cost"),
        ("头程成本", "first_mile", "SKU effective-date first-mile cost"),
        ("包装成本", "packaging", "SKU effective-date packaging cost"),
        ("其他单位成本", "other_unit", "SKU effective-date other unit cost"),
        ("人工/工资成本", "payroll", "当前无工资成本，显式为0"),
        ("【最终结果】", None, ""),
        ("经营利润", "profit", "Natural-month Management Operating Profit"),
        ("经营利润率", "margin", "经营利润 / 商品销售额"),
    )

    previous_values = _pnl_values(previous_nm, previous_components) if previous_nm else {}
    current_values = _pnl_values(current_nm, current_components)
    row = 4
    for label, key, note in row_specs:
        sheet.cell(row, 1, label)
        sheet.cell(row, 6, note).alignment = Alignment(wrap_text=True)
        if key is None:
            _fill_range(sheet, f"A{row}:F{row}", MID_BLUE)
            for cell in sheet[row][:6]:
                cell.font = Font(bold=True, color=WHITE)
        else:
            previous_value = previous_values.get(key)
            current_value = current_values.get(key)
            sheet.cell(row, 2, _xlsx_number(previous_value))
            sheet.cell(row, 3, _xlsx_number(current_value))
            if key == "margin":
                diff = _difference(current_value, previous_value)
                sheet.cell(row, 4, _xlsx_number(diff))
                sheet.cell(row, 5, _xlsx_number(current_value))
                for col in range(2, 6):
                    sheet.cell(row, col).number_format = PERCENT_FORMAT
            else:
                diff = _difference(current_value, previous_value)
                ratio = safe_ratio(abs(as_decimal(current_value)), current_nm.product_sales_amount)
                sheet.cell(row, 4, _xlsx_number(diff))
                sheet.cell(row, 5, _xlsx_number(ratio))
                for col in range(2, 5):
                    sheet.cell(row, col).number_format = MONEY_FORMAT
                sheet.cell(row, 5).number_format = PERCENT_FORMAT
        row += 1

    _apply_widths(sheet, {"A": 32, "B": 16, "C": 16, "D": 16, "E": 20, "F": 42})
    sheet.freeze_panes = "A4"


def _write_checks(sheet: Any, result: MonthlyFinancialCloseResult) -> None:
    nm = result.natural_month_finance
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "数据核验与口径"
    _title(sheet["A1"], size=16)
    headers = ("核验项", "状态", "本月结果", "数据源", "是否影响发送", "说明")
    for col, value in enumerate(headers, 1):
        sheet.cell(3, col, value)
    _header_row(sheet, 3, 6)

    source_status = nm.source_status if nm else "missing"
    cost_expected = (nm.product_sales_units + nm.liquidation_units) if nm else 0
    cost_actual = nm.costed_units if nm else 0
    missing_cost = ", ".join(nm.missing_cost_skus) if nm and nm.missing_cost_skus else "-"
    rows = (
        ("自然月Finances完整性", "通过" if source_status == "ok" else "需复核", f"source_status={source_status}", "amazon_finance_transaction", "是", "经营财务主链"),
        ("Review-required", "通过" if nm and nm.review_required_count == 0 else "需复核", f"count={nm.review_required_count if nm else '-'}; amount={nm.review_required_amount if nm else '-'}", "amazon_finance_transaction", "是", "存在非零未识别生命周期/分类时必须阻塞"),
        ("成本覆盖", "通过" if nm and cost_actual == cost_expected and not nm.missing_cost_skus else "需复核", f"costed={cost_actual}; expected={cost_expected}; missing={missing_cost}", "amazon_sku_cost", "是", "销售+清算成本必须完整"),
        ("Marketplace timezone", "通过" if nm and nm.marketplace_timezone else "需复核", nm.marketplace_timezone if nm else "-", "marketplace metadata", "是", "自然月边界使用站点本地时区"),
        ("广告费", "通过", f"{nm.ads_api_report_date_spend if nm else 0} {result.currency or ''}", "Ads API report_date", "是", "经营口径使用当月实际广告发生额"),
        ("Seller Central核验", "可选外部核验", str(result.raw_metadata.get("seller_central_monthly_transaction_reconciliation_status") or "not_provided"), "Monthly Transaction", "否", "手工CSV缺失不阻塞已通过Finances gate的自动月报"),
        ("Settlement / Transfer", "参考", "不计入Management P&L主口径", "Settlement / Finances", "否", "用于结算、现金和银行回款核验"),
        ("Released + Deferred", "必须包含", "由natural-month lifecycle policy去重纳入", "Finances normalized ledger", "是", "不可只筛Released"),
    )
    for row_index, values in enumerate(rows, 4):
        for col, value in enumerate(values, 1):
            sheet.cell(row_index, col, value)
        sheet.cell(row_index, 6).alignment = Alignment(wrap_text=True)
        fill = GREEN_FILL if values[1] == "通过" else YELLOW_FILL
        sheet.cell(row_index, 2).fill = PatternFill("solid", fgColor=fill)
        sheet.cell(row_index, 2).font = Font(bold=True)

    note_row = 4 + len(rows) + 2
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 2, end_column=6)
    sheet.cell(note_row, 1, (
        "经营月报是管理视角，不是法定财务报表。商品毛利率只扣到岸商品成本；"
        "经营利润率包含Amazon费用、退款、促销、广告和商品成本。"
        "Finances广告账单扣款与Ads API广告发生月份可能不同，两者不得强行相等。"
    ))
    sheet.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    _fill_range(sheet, f"A{note_row}:F{note_row + 2}", LIGHT_BLUE)

    _apply_widths(sheet, {"A": 28, "B": 18, "C": 38, "D": 28, "E": 16, "F": 48})
    sheet.freeze_panes = "A4"


def _trend_values(result: MonthlyFinancialCloseResult) -> dict[str, Decimal | int | None]:
    nm = result.natural_month_finance
    if nm is None:
        return {key: None for key in ("sales", "units", "gross_margin", "profit", "margin", "ads", "ads_ratio", "sessions", "conversion")}
    sessions = operational_metric(result, "Sales & Traffic", "sessions")
    conversion = operational_metric(result, "Sales & Traffic", "unit_session_rate")
    gross_margin = safe_ratio(nm.product_sales_amount - nm.landed_cogs, nm.product_sales_amount)
    ads_ratio = safe_ratio(nm.ads_api_report_date_spend, nm.product_sales_amount)
    return {
        "sales": nm.product_sales_amount,
        "units": nm.costed_units,
        "gross_margin": gross_margin,
        "profit": nm.management_operating_profit,
        "margin": nm.management_operating_margin,
        "ads": nm.ads_api_report_date_spend,
        "ads_ratio": ads_ratio,
        "sessions": int(sessions) if sessions not in (None, "") else None,
        "conversion": as_decimal(conversion) if conversion is not None else None,
    }


def _pnl_values(nm: Any, components: Mapping[str, Decimal]) -> dict[str, Decimal | None]:
    if nm is None:
        return {}
    return {
        "product_sales": nm.product_sales_amount,
        "shipping_income": components.get("shipping_income", ZERO),
        "refund_product": components.get("refund_product", ZERO),
        "refund_shipping": components.get("refund_shipping", ZERO),
        "refund_promotion": components.get("refund_promotion", ZERO),
        "liquidation_revenue": components.get("liquidation_revenue", ZERO),
        "net_commission": components.get("net_commission", ZERO),
        "fba_fulfillment": components.get("fba_fulfillment", ZERO),
        "shipping_chargeback": components.get("shipping_chargeback", ZERO),
        "order_promotion": components.get("order_promotion", ZERO),
        "liquidation_fee": components.get("liquidation_fee", ZERO),
        "reimbursement": nm.reimbursement_total,
        "adjustment": nm.adjustment_total,
        "subscription": nm.subscription_fee,
        "coupon": nm.coupon_fee,
        "deal": nm.deal_fee,
        "storage": nm.storage_fee,
        "customer_return": nm.customer_return_fee,
        "other_service": nm.other_service_fee,
        "ads": -nm.ads_api_report_date_spend,
        "product_cost": -nm.product_cost_cogs,
        "first_mile": -nm.first_mile_cogs,
        "packaging": -nm.packaging_cogs,
        "other_unit": -nm.other_unit_cogs,
        "payroll": ZERO,
        "profit": nm.management_operating_profit,
        "margin": nm.management_operating_margin,
    }


def _expense_rows(result: MonthlyFinancialCloseResult, components: Mapping[str, Decimal]) -> tuple[tuple[str, Decimal, str], ...]:
    nm = result.natural_month_finance
    if nm is None:
        return ()
    return (
        ("广告费", abs(nm.ads_api_report_date_spend), "销售下降但广告未同步下降时，应优先复盘广告效率"),
        ("到岸商品成本", abs(nm.landed_cogs), "采购+头程+包装+其他单位成本；观察产品本身成本结构"),
        ("FBA履约费", abs(components.get("fba_fulfillment", ZERO)), "随销量变化的核心Amazon履约费用"),
        ("退款净额", abs(nm.refund_total), "销售退回对利润的直接冲减"),
        ("订单促销折扣", abs(components.get("order_promotion", ZERO)), "订单侧促销让利"),
        ("账户级费用", abs(nm.service_fee_total), "订阅+Coupon+Deal+仓储+退货处理+其他ServiceFee"),
    )


def _build_observations(
    current: MonthlyFinancialCloseResult,
    recent: Sequence[MonthlyFinancialCloseResult],
    components: Mapping[str, Decimal],
) -> list[tuple[str, str]]:
    nm = current.natural_month_finance
    if nm is None:
        return [("数据不足", "Natural-month Finances summary unavailable.")]
    observations: list[tuple[str, str]] = []
    previous = recent[-2] if len(recent) >= 2 else None
    previous_nm = previous.natural_month_finance if previous else None

    if nm.management_operating_profit < ZERO:
        observations.append(("本月经营亏损", f"经营利润为 {_money_text(nm.management_operating_profit)}，经营利润率 {_pct_text(nm.management_operating_margin)}。"))

    if previous_nm and previous_nm.product_sales_amount != ZERO:
        sales_change = nm.product_sales_amount / previous_nm.product_sales_amount - 1
        if abs(sales_change) >= Decimal("0.10"):
            observations.append(("销售规模显著变化", f"商品销售额较上月 {_signed_pct_text(sales_change)}，从 {_money_text(previous_nm.product_sales_amount)} 变为 {_money_text(nm.product_sales_amount)}。"))

        gross_now = safe_ratio(nm.product_sales_amount - nm.landed_cogs, nm.product_sales_amount)
        gross_prev = safe_ratio(previous_nm.product_sales_amount - previous_nm.landed_cogs, previous_nm.product_sales_amount)
        if gross_now is not None and gross_prev is not None:
            gross_diff = gross_now - gross_prev
            if abs(gross_diff) < Decimal("0.02"):
                observations.append(("商品毛利保持稳定", f"商品毛利率本月 {_pct_text(gross_now)}，上月 {_pct_text(gross_prev)}；产品成本结构不是主要异常来源。"))
            else:
                observations.append(("商品毛利率明显变化", f"商品毛利率由 {_pct_text(gross_prev)} 变为 {_pct_text(gross_now)}，变化 {_signed_pp_text(gross_diff)}。"))

        current_sessions = as_decimal(operational_metric(current, "Sales & Traffic", "sessions"))
        previous_sessions = as_decimal(operational_metric(previous, "Sales & Traffic", "sessions"))
        current_conv = as_decimal(operational_metric(current, "Sales & Traffic", "unit_session_rate"))
        previous_conv = as_decimal(operational_metric(previous, "Sales & Traffic", "unit_session_rate"))
        if previous_sessions > ZERO and previous_conv > ZERO:
            session_change = current_sessions / previous_sessions - 1
            conv_change = current_conv / previous_conv - 1
            if conv_change <= Decimal("-0.15"):
                observations.append(("转化效率恶化", f"Sessions环比 {_signed_pct_text(session_change)}，但转化率从 {_pct_text(previous_conv)} 降至 {_pct_text(current_conv)}；需要优先检查流量质量、Listing和价格/促销。"))

        ads_now = safe_ratio(nm.ads_api_report_date_spend, nm.product_sales_amount)
        ads_prev = safe_ratio(previous_nm.ads_api_report_date_spend, previous_nm.product_sales_amount)
        if ads_now is not None and ads_prev is not None and ads_now - ads_prev >= Decimal("0.05"):
            observations.append(("广告负担明显上升", f"广告费率从 {_pct_text(ads_prev)} 升至 {_pct_text(ads_now)}，增加 {_signed_pp_text(ads_now - ads_prev)}。"))

    if abs(nm.deal_fee) >= Decimal("50"):
        observations.append(("Deal费用需要单独关注", f"本月Deal费用 {_money_text(abs(nm.deal_fee))}，属于一次性/活动型费用，应与活动增量销售一起评估。"))

    if not observations:
        observations.append(("本月总体稳定", "未触发主要异常阈值，建议继续关注利润率、广告费率和转化率趋势。"))
    return observations[:5]


def _mom_change(values: Sequence[Decimal | int | None]) -> Decimal | None:
    if len(values) < 2:
        return None
    previous = as_decimal(values[-2])
    current = as_decimal(values[-1])
    if previous == ZERO:
        return None
    return current / previous - 1


def _difference(current: Any, previous: Any) -> Decimal | None:
    if current is None or previous is None:
        return None
    return as_decimal(current) - as_decimal(previous)


def _money_text(value: Decimal) -> str:
    sign = "-" if value < ZERO else ""
    return f"{sign}${abs(value):,.2f}"


def _pct_text(value: Decimal | None) -> str:
    if value is None:
        return "-"
    sign = "-" if value < ZERO else ""
    return f"{sign}{abs(value) * 100:.2f}%"


def _signed_pct_text(value: Decimal) -> str:
    sign = "+" if value >= ZERO else "-"
    return f"{sign}{abs(value) * 100:.1f}%"


def _signed_pp_text(value: Decimal) -> str:
    sign = "+" if value >= ZERO else "-"
    return f"{sign}{abs(value) * 100:.2f}个百分点"


def _xlsx_number(value: Any) -> float | int | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    return value


def _section(sheet: Any, row: int, title: str, columns: int) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
    cell = sheet.cell(row, 1, title)
    cell.fill = PatternFill("solid", fgColor=MID_BLUE)
    cell.font = Font(bold=True, color=WHITE)


def _header_row(sheet: Any, row: int, columns: int) -> None:
    for cell in sheet[row][:columns]:
        cell.fill = PatternFill("solid", fgColor=HEADER_BLUE)
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _title(cell: Any, *, size: int) -> None:
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    cell.font = Font(bold=True, color=WHITE, size=size)
    cell.alignment = Alignment(vertical="center")


def _fill_range(sheet: Any, cell_range: str, color: str) -> None:
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=color)


def _apply_widths(sheet: Any, widths: Mapping[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _write_no_data_sheet(sheet: Any, result: MonthlyFinancialCloseResult, message: str) -> None:
    sheet["A1"] = f"{result.month} 月度经营报告"
    _title(sheet["A1"], size=16)
    sheet["A3"] = message
    sheet["A3"].fill = PatternFill("solid", fgColor=YELLOW_FILL)


__all__ = ["build_monthly_operating_report_workbook"]
